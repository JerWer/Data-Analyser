import sys
import datetime
import os
import platform
from pathlib import Path
import numpy as np
from numpy import array, linspace, asarray, arange, abs as npabs, int as npint, mat as npmat, linalg as nplinalg, concatenate as npconcatenate, convolve as npconvolve
import sqlite3
import shutil
import math
from math import factorial
from scipy.interpolate import interp1d, UnivariateSpline

#%%######################################################################################################
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use("Qt5Agg")

# print(matplotlib.get_configdir())
#%%######################################################################################################
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal
# from PyQt5.QtCore.QElapsedTimer import timer
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QAction, QTableWidgetItem, QColorDialog
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
import copy
import xlsxwriter
import openpyxl as Opxl
import xlrd
from scipy import integrate
from operator import itemgetter
from itertools import groupby, combinations
from PIL import Image as ImageTk
from matplotlib.ticker import MaxNLocator, AutoMinorLocator
import pickle
import six
import os.path
from dateutil import parser
from scipy.interpolate import interp1d
from PyQt5.uic import loadUiType
from functools import partial
from datetime import date
from database_Tables_seris import CreateAllTables, criteriaexclusionlist, dropdowncriteria, fromtocriteria, timecriteria
# from MainWindowApp import MainWindowAllApps

#%%############# Global variable definition
file_path=[]
resultsdirectory=''
testdata = []
DATA = {} #{sampleID: {"SampleName":, "CellNumber": , "MeasDayTime":, "CellSurface":, "Voc":, "Jsc":, "FF":, "Eff":, "Pmpp":, "Vmpp":, "Jmpp":, "Roc":, "Rsc":, "VocFF":, "RscJsc":, "NbPoints":, "Delay":, "IntegTime":, "Vstart":, "Vend":, "Illumination":, "ScanDirection":, "ImaxComp":, "Isenserange":,"AreaJV":, "Operator":, "MeasComment":, "IVData":}]
DATAJVforexport=[]
DATAJVtabforexport=[]
DATAmppforexport=[]
DATAgroupforexport=[]
DATAHistforexport=[]
DATAcompforexport=[]
DATAtimeevolforexport={}#key: [[realtimeF, relativetimeF, valueF, normalizedvaluetot0F, realtimeR, relativetimeR, valueR, normalizedvaluetot0R]]
takenforplot=[]
takenforplotmpp=[]
takenforplotTime=[]

DATAMPP = {}
DATAdark = []
DATAFV=[]
DATAFFloss={}

numbLightfiles=0
numbDarkfiles=0

DBisconnected=0
instructionsofsearch=''

IVlegendMod = []
colorstylelist = ['black', 'red', 'blue', 'brown', 'green','cyan','magenta','olive','navy','orange','gray','aliceblue','antiquewhite','aqua','aquamarine','azure','beige','bisque','blanchedalmond','blue','blueviolet','brown','burlywood','cadetblue','chartreuse','chocolate','coral','cornflowerblue','cornsilk','crimson','darkblue','darkcyan','darkgoldenrod','darkgray','darkgreen','darkkhaki','darkmagenta','darkolivegreen','darkorange','darkorchid','darkred','darksalmon','darkseagreen','darkslateblue','darkslategray','darkturquoise','darkviolet','deeppink','deepskyblue','dimgray','dodgerblue','firebrick','floralwhite','forestgreen','fuchsia','gainsboro','ghostwhite','gold','goldenrod','greenyellow','honeydew','hotpink','indianred','indigo','ivory','khaki','lavender','lavenderblush','lawngreen','lemonchiffon','lightblue','lightcoral','lightcyan','lightgoldenrodyellow','lightgreen','lightgray','lightpink','lightsalmon','lightseagreen','lightskyblue','lightslategray','lightsteelblue','lightyellow','lime','limegreen','linen','magenta','maroon','mediumaquamarine','mediumblue','mediumorchid','mediumpurple','mediumseagreen','mediumslateblue','mediumspringgreen','mediumturquoise','mediumvioletred','midnightblue','mintcream','mistyrose','moccasin','navajowhite','navy','oldlace','olive','olivedrab','orange','orangered','orchid','palegoldenrod','palegreen','paleturquoise','palevioletred','papayawhip','peachpuff','peru','pink','plum','powderblue','purple','red','rosybrown','royalblue','saddlebrown','salmon','sandybrown','seagreen','seashell','sienna','silver','skyblue','slateblue','slategray','snow','springgreen','steelblue','tan','teal','thistle','tomato','turquoise','violet','wheat','white','whitesmoke','yellow','yellowgreen']
colormapname="jet"

MPPlegendMod = []
MPPlinestyle = []

titIV =0
titmpp=0
titStat=0
samplesgroups=["Default group"]
samplesgroups12=[]
groupstoplotcomp=["Default group"]
groupstoplot=["Default group"]
groupcomment={}
groupSubgroups={}

listofanswer=[]
listoflinestyle=[]
listofcolorstyle=[]
listoflinewidthstyle=[]

listofanswermpp=[]
listoflinestylempp=[]
listofcolorstylempp=[]
listoflinewidthstylempp=[]

#%%#############
def modification_date(path_to_file):
    return datetime.datetime.fromtimestamp(os.path.getmtime(path_to_file)).strftime("%Y-%m-%d %H:%M:%S")

def creation_date(path_to_file):
    """
    Try to get the date that a file was created, falling back to when it was
    last modified if that isn't possible.
    See http://stackoverflow.com/a/39501288/1709587 for explanation.
    """
    if platform.system() == 'Windows':
        return os.path.getctime(path_to_file)
    else:
        stat = os.stat(path_to_file)
        try:
            return stat.st_birthtime
        except AttributeError:
            # We're probably on Linux. No easy way to get creation dates here,
            # so we'll settle for when its content was last modified.
            return stat.st_mtime

def sumofsquaredev(datalist):
    average=sum(datalist)/len(datalist)
    deviations=[x-average for x in datalist]
    return sum([x**2 for x in deviations])

# Ui_MainWindow, QMainWindow = loadUiType('IVpyqt5gui.ui')

# Ui_MainWindow, QMainWindow = loadUiType(r'C:\Users\serjw\Documents\nBox\PythonScripts\SERIS-pythonscripts\executables\All\IVpyqt5gui.ui')
from IVpyqt5gui import Ui_MainWindow
exedirectory=str(Path(os.path.abspath(__file__)).parent)


#%%#############
class FixFigureCanvas(FigureCanvas):
    def resizeEvent(self, event):
        if event.size().width() <= 0 or event.size().height() <= 0:
            return
        super(FixFigureCanvas, self).resizeEvent(event)
        
class IVapp(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        global DATA
        QtWidgets.QMainWindow.__init__(self, parent)
        
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        self.fig1 = Figure(constrained_layout=True)
        self.JVgraph = self.fig1.add_subplot(111)
        self.JVgraph.set_xlabel('Voltage (mV)')
        self.JVgraph.set_ylabel('Current density (mA/cm'+'\xb2'+')')
        self.JVgraph.axhline(y=0, color='k')
        self.JVgraph.axvline(x=0, color='k')
        self.addmpl(self.fig1,self.ui.gridLayout_5, self.ui.widget_JVGraph)
        
        self.fig2 = Figure(constrained_layout=True)
        self.MPPgraph = self.fig2.add_subplot(111)
        self.MPPgraph.set_ylabel('Power (mW/cm'+'\xb2'+')')
        self.MPPgraph.set_xlabel('Time (s)')
        self.addmpl(self.fig2,self.ui.gridLayout_2, self.ui.widget_Mpp)
        
        self.fig31 = Figure()
        self.fig31.subplots_adjust(hspace=0, wspace=0)
        self.addmpl(self.fig31,self.ui.gridLayout_37, self.ui.widget_BoxPlot_2)
        
        self.fig4 = Figure(constrained_layout=True)
        self.TimeEvolgraph = self.fig4.add_subplot(111)
        self.addmpl(self.fig4,self.ui.gridLayout_10, self.ui.widget_PVPTime)
        
        self.fig5 = Figure(constrained_layout=True)
        self.ParamParamgraph = self.fig5.add_subplot(111)
        self.addmpl(self.fig5,self.ui.gridLayout_11, self.ui.widget_PVPGraph)
        
        self.fig6 = Figure(constrained_layout=True)
        self.Histgraph = self.fig6.add_subplot(111)
        self.addmpl(self.fig6,self.ui.gridLayout_13, self.ui.widget_HistoGraph)
        
        self.fig7 = Figure(constrained_layout=True)
        self.DBgraph = self.fig7.add_subplot(111)
        self.DBgraph.set_xlabel('Date/Time')
        self.DBgraph.set_ylabel('X (-)')
        # self.DBgraph.axhline(y=0, color='k')
        # self.DBgraph.axvline(x=0, color='k')
        self.addmpl(self.fig7,self.ui.gridLayout_mplDB, self.ui.widget_DB)
        
        self.fig8 = Figure(constrained_layout=True)
        self.FFlossgraph = self.fig8.add_subplot(111)
        self.addmpl(self.fig8,self.ui.gridLayout_35, self.ui.widget_FFloss)
        
        finish = QAction("Quit", self)
        finish.triggered.connect(lambda: self.closeEvent(1))
        
        
        self.ui.actionReinitialize_the_app.triggered.connect(self.Reinitializetheapp)
        self.ui.actionLoad.triggered.connect(self.LoadSession)
        self.ui.actionSave.triggered.connect(self.SaveSession)
        self.ui.actionLoad_2.triggered.connect(lambda: self.loadconfigsgui(''))
        self.ui.actionSave_2.triggered.connect(lambda: self.saveconfigsgui(''))
        self.ui.pushButton_chooseDB.clicked.connect(self.chooseandconnectDB)
        self.ui.pushButton_AddtoDB.clicked.connect(self.loadtoDB)
        
        self.ui.pushButton_chooseDB.setStyleSheet("background-color: red")
        self.ui.pushButton_connecttoDBread.clicked.connect(self.connectToDBreading)
        self.ui.comboBox_DBTime.currentTextChanged.connect(self.on_combobox_DBTime_changed)
        self.ui.comboBox_DBTimeYaxis.currentTextChanged.connect(self.SearchAndPlot)
        self.ui.comboBox_DBTime_restrictions.currentTextChanged.connect(self.on_combobox_DBTimeRestrictions_changed)
        self.ui.pushButton_DBTime_Add.clicked.connect(self.Addrestriction)
        self.ui.pushButton_DBTime_remove.clicked.connect(self.Removerestriction)
        self.ui.pushButton_ExportDBreadResults.clicked.connect(self.ExportDBsearchresults)
        self.ui.pushButton_AutoGraph.clicked.connect(self.AutoGraphExport)
        
        self.ui.pushButton_ImportData.clicked.connect(self.startimporting)
        self.ui.pushButton_DeleteRows.clicked.connect(self.DeleteRows)
        self.ui.pushButton_DefineGroup.clicked.connect(self.Confirmgroupnamechanges)
        self.ui.pushButton_DefineGroup2.clicked.connect(self.Confirmgroup2namechanges)
        self.ui.ConfirmCommentChanges.clicked.connect(self.ConfirmCommentChanges)
        self.ui.pushButton_UpdateTable.clicked.connect(lambda: self.updateTable(DATA))
        self.ui.pushButton_confirmarea.clicked.connect(self.ConfirmAreachanges)
        self.ui.pushButton_mppChangeArea.clicked.connect(self.ConfirmMPPAreachanges)
        self.ui.pushButton_confirmSunInten.clicked.connect(self.ConfirmSunIntenchanges)
        self.ui.pushButton_confirmNameChange.clicked.connect(self.ConfirmSampleNameChanges)
        self.ui.pushButton_switchD2L.clicked.connect(self.switchD2L)
        
        self.ui.checkBox_minorticksBoxPlot.toggled.connect(self.UpdateBoxGraph2)
        
        self.ui.checkBox_MinorTicksJV.toggled.connect(self.PlotIV)
        self.ui.radioButton_JVtopleft.toggled.connect(self.PlotIV)
        self.ui.radioButton_JVtopright.toggled.connect(self.PlotIV)
        self.ui.radioButton_JVbottomleft.toggled.connect(self.PlotIV)
        self.ui.radioButton_JVbottomright.toggled.connect(self.PlotIV)
        self.ui.radioButton_JVoutside.toggled.connect(self.PlotIV)
        self.ui.radioButton_JVBest.toggled.connect(self.PlotIV)
        self.ui.checkBox_JVLegend.toggled.connect(self.PlotIV)
        self.ui.pushButton_PlotJV.clicked.connect(self.PlotIV)
        self.ui.pushButton_SaveJV.clicked.connect(self.GraphJVsave_as)
        self.ui.spinBox_JVfontsize.valueChanged.connect(self.PlotIV)
        self.ui.comboBox_MPPT_TrueOrRelative.currentTextChanged.connect(self.PlotMPP)
        self.ui.comboBox_MPPT_TimeUnit.currentTextChanged.connect(self.PlotMPP)
        self.ui.pushButton_deletemppdata.clicked.connect(self.DeleteMPP)
        self.ui.radioButton_MppTopleft.toggled.connect(self.PlotMPP)
        self.ui.radioButton_MppTopright.toggled.connect(self.PlotMPP)
        self.ui.radioButton_MppBottomleft.toggled.connect(self.PlotMPP)
        self.ui.radioButton_MppBottomright.toggled.connect(self.PlotMPP)
        self.ui.radioButton_MppOutside.toggled.connect(self.PlotMPP)
        self.ui.radioButton_MppBest.toggled.connect(self.PlotMPP)
        self.ui.checkBox_MppLegend.toggled.connect(self.PlotMPP)
        self.ui.listWidget_MppSamples.itemClicked.connect(self.PlotMPP1)
        self.ui.pushButton_SaveMpp.clicked.connect(self.GraphMPPsave_as)
        # self.ui.listWidget_MppSamples.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.ui.spinBox_MppFontsize.valueChanged.connect(self.PlotMPP)
        
        self.ui.listWidget_HistoGroups.itemClicked.connect(self.UpdateHistGraph)
        self.ui.comboBox_HistoParam.currentTextChanged.connect(self.UpdateHistGraph)
        self.ui.comboBox_HistoScanDirect.currentTextChanged.connect(self.UpdateHistGraph)
        self.ui.comboBox_HistoType.currentTextChanged.connect(self.UpdateHistGraph)
        self.ui.checkBox_Histxscale.toggled.connect(self.UpdateHistGraph)
        self.ui.spinBox_HistoBins.valueChanged.connect(self.UpdateHistGraph)
        self.ui.spinBox_HistxscaleMin.valueChanged.connect(self.UpdateHistGraph)
        self.ui.spinBox_HistxscaleMax.valueChanged.connect(self.UpdateHistGraph)
        self.ui.pushButton_SaveHistoGraph.clicked.connect(self.GraphHistsave_as)
        # self.ui.listWidget_HistoGroups.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)

        # # self.ui.listWidget_BoxPlotGroup.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        # self.ui.listWidget_BoxPlotGroup.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        # self.ui.listWidget_BoxPlotGroup.itemClicked.connect(self.UpdateBoxGraph)
        # self.ui.comboBox_BoxPlotParam.currentTextChanged.connect(self.UpdateBoxGraph)
        # self.ui.checkBox_BoxPlotAftermpp.toggled.connect(self.UpdateBoxGraph)
        # self.ui.checkBox_BoxPlotBkg.toggled.connect(self.UpdateBoxGraph)
        # self.ui.spinBox_BoxPlotRotation.valueChanged.connect(self.UpdateBoxGraph)
        # self.ui.spinBox_BoxPlotFontsize.valueChanged.connect(self.UpdateBoxGraph)
        # self.ui.spinBox_markerSize.valueChanged.connect(self.UpdateBoxGraph)
        # self.ui.checkBox_BoxPlotBoxPlot.toggled.connect(self.UpdateBoxGraph)
        # self.ui.checkBox_BoxPlotRevForw.toggled.connect(self.UpdateBoxGraph)
        # self.ui.pushButton_SaveBoxPlot.clicked.connect(self.GraphBoxsave_as)
        # self.ui.pushButton_groupanova.clicked.connect(self.ANOVAboxplot)
        
        self.ui.listWidget_BoxPlotGroup2_2.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.ui.listWidget_BoxPlotGroup2_2.itemClicked.connect(self.UpdateBoxGraph2)
        self.ui.comboBox_BoxPlotParam_2.currentTextChanged.connect(self.UpdateBoxGraph2)
        self.ui.checkBox_BoxPlotAftermpp_2.toggled.connect(self.UpdateBoxGraph2)
        self.ui.checkBox_BoxPlotBkg_2.toggled.connect(self.UpdateBoxGraph2)
        self.ui.spinBox_BoxPlotRotation_2.valueChanged.connect(self.UpdateBoxGraph2)
        self.ui.spinBox_BoxPlotFontsize_2.valueChanged.connect(self.UpdateBoxGraph2)
        self.ui.spinBox_markerSize_2.valueChanged.connect(self.UpdateBoxGraph2)
        self.ui.checkBox_BoxPlotBoxPlot_2.toggled.connect(self.UpdateBoxGraph2)
        self.ui.checkBox_BoxPlotRevForw_2.toggled.connect(self.UpdateBoxGraph2)
        self.ui.pushButton_SaveBoxPlot_2.clicked.connect(self.GraphBoxsave_as2)
        self.ui.pushButton_groupanova_2.clicked.connect(self.ANOVAboxplot2)
        self.ui.checkBox_onlygroup1.toggled.connect(self.UpdateBoxGraph2)
        
        
        
        self.ui.pushButton_SavePVPTime.clicked.connect(self.GraphTimesave_as)
        self.ui.pushButton_PVPTimePlot.clicked.connect(self.plottingTimefromTable)
        self.ui.checkBox_PVPTimeBig4.toggled.connect(self.UpdateTimeGraph)
        self.ui.checkBox_PVPTimeLine.toggled.connect(self.UpdateTimeGraph)
        self.ui.checkBox_PVPTimeRelativeTime.toggled.connect(self.UpdateTimeGraph)
        self.ui.checkBox_PVPTimeNormal.toggled.connect(self.UpdateTimeGraph)
        self.ui.checkBox_bestofRevFor.toggled.connect(self.UpdateTimeGraph)
        self.ui.checkBox_BestEffPixDay.toggled.connect(self.UpdateTimeGraph)
        self.ui.spinBox_PVPTimeNormalPoint.valueChanged.connect(self.UpdateTimeGraph)
        self.ui.comboBox_PVPTimeParam.currentTextChanged.connect(self.UpdateTimeGraph)
        
        self.ui.pushButton_SavePVPGraph.clicked.connect(self.GraphCompsave_as)
        self.ui.comboBox_PVPx.currentTextChanged.connect(self.UpdateCompGraph)
        self.ui.comboBox_PVPy.currentTextChanged.connect(self.UpdateCompGraph)
        # self.ui.listWidget_ParamComp.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.ui.listWidget_ParamComp.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.ui.listWidget_ParamComp.itemClicked.connect(self.UpdateCompGraph)
        
        self.ui.pushButton_AutoAnalysis.clicked.connect(self.ExportAutoAnalysis)
        
        self.ui.tabWidget_JV.currentChanged.connect(self.onclicklegendtabJV)
        self.ui.tabWidget_MPP.currentChanged.connect(self.onclicklegendtabMPP)
        
        self.ui.comboBox_FFloss.currentTextChanged.connect(self.FFlosses)
        self.ui.pushButton_exportFFloss.clicked.connect(self.exportFFloss)

        self.ui.pushButton_SGFilter.clicked.connect(self.SavitzkyGolayFiltering)
        self.ui.pushButton_goback.clicked.connect(self.backtoOriginal)
        self.ui.pushButton_recalcJV.clicked.connect(self.RecalcJV)
        self.ui.pushButton_bumpsRem.clicked.connect(self.BumpsRemoval)
        
    
    def addmpl(self, fig, whereLayout, whereWidget):
        self.canvas = FixFigureCanvas(fig)
        whereLayout.addWidget(self.canvas)
        self.canvas.draw()
        self.toolbar = NavigationToolbar(self.canvas, 
                whereWidget, coordinates=True)
        whereLayout.addWidget(self.toolbar)
        
    def closeEvent(self, event):
        
        """ what happens when close the program"""
        
        close = QMessageBox.question(self,
                                     "QUIT",
                                     "Are you sure?",
                                      QMessageBox.Yes | QMessageBox.No)
        if close == QMessageBox.Yes:
            self.Reinitializetheapp()
            event.accept()
            try:
                app.quit()
            except:
                pass
                # self.parent.MainWindowAllApps()
        else:
            event.ignore()
    
    def Reinitializetheapp(self):
        global file_path,testdata,DATA,DATAJVforexport,DATAJVtabforexport,DATAmppforexport,DATAgroupforexport,DATAHistforexport,DATAcompforexport,DATAtimeevolforexport
        global takenforplot,takenforplotmpp,takenforplotTime,DATAMPP,DATAdark,DATAFV,DATAFFloss,numbLightfiles,numbDarkfiles,DBisconnected,instructionsofsearch
        global IVlegendMod,colorstylelist,colormapname,MPPlegendMod,MPPlinestyle,titIV,titmpp,titStat,samplesgroups,groupstoplotcomp,groupcomment
        global listofanswer,listoflinestyle,listofcolorstyle,listoflinewidthstyle,listofanswermpp,listoflinestylempp,listofcolorstylempp,listoflinewidthstylempp
        
        file_path=[]
        testdata = []
        DATA = {} #{sampleID: {"SampleName":, "CellNumber": , "MeasDayTime":, "CellSurface":, "Voc":, "Jsc":, "FF":, "Eff":, "Pmpp":, "Vmpp":, "Jmpp":, "Roc":, "Rsc":, "VocFF":, "RscJsc":, "NbPoints":, "Delay":, "IntegTime":, "Vstart":, "Vend":, "Illumination":, "ScanDirection":, "ImaxComp":, "Isenserange":,"AreaJV":, "Operator":, "MeasComment":, "IVData":}]
        DATAJVforexport=[]
        DATAJVtabforexport=[]
        DATAmppforexport=[]
        DATAgroupforexport=[]
        DATAHistforexport=[]
        DATAcompforexport=[]
        DATAtimeevolforexport={}#key: [[realtimeF, relativetimeF, valueF, normalizedvaluetot0F, realtimeR, relativetimeR, valueR, normalizedvaluetot0R]]
        takenforplot=[]
        takenforplotmpp=[]
        takenforplotTime=[]
        
        DATAMPP = {}
        DATAdark = []
        DATAFV=[]
        DATAFFloss={}
        
        numbLightfiles=0
        numbDarkfiles=0
        
        DBisconnected=0
        instructionsofsearch=''
        self.ui.lineEdit_DBpath.setText('')
        try:
            self.theCursor.close()
            self.db_conn.close()
            self.ui.pushButton_chooseDB.setStyleSheet("background-color: red")
            DBisconnected=0
        except:
            pass
        
        IVlegendMod = []
        colorstylelist = ['black', 'red', 'blue', 'brown', 'green','cyan','magenta','olive','navy','orange','gray','aliceblue','antiquewhite','aqua','aquamarine','azure','beige','bisque','blanchedalmond','blue','blueviolet','brown','burlywood','cadetblue','chartreuse','chocolate','coral','cornflowerblue','cornsilk','crimson','darkblue','darkcyan','darkgoldenrod','darkgray','darkgreen','darkkhaki','darkmagenta','darkolivegreen','darkorange','darkorchid','darkred','darksalmon','darkseagreen','darkslateblue','darkslategray','darkturquoise','darkviolet','deeppink','deepskyblue','dimgray','dodgerblue','firebrick','floralwhite','forestgreen','fuchsia','gainsboro','ghostwhite','gold','goldenrod','greenyellow','honeydew','hotpink','indianred','indigo','ivory','khaki','lavender','lavenderblush','lawngreen','lemonchiffon','lightblue','lightcoral','lightcyan','lightgoldenrodyellow','lightgreen','lightgray','lightpink','lightsalmon','lightseagreen','lightskyblue','lightslategray','lightsteelblue','lightyellow','lime','limegreen','linen','magenta','maroon','mediumaquamarine','mediumblue','mediumorchid','mediumpurple','mediumseagreen','mediumslateblue','mediumspringgreen','mediumturquoise','mediumvioletred','midnightblue','mintcream','mistyrose','moccasin','navajowhite','navy','oldlace','olive','olivedrab','orange','orangered','orchid','palegoldenrod','palegreen','paleturquoise','palevioletred','papayawhip','peachpuff','peru','pink','plum','powderblue','purple','red','rosybrown','royalblue','saddlebrown','salmon','sandybrown','seagreen','seashell','sienna','silver','skyblue','slateblue','slategray','snow','springgreen','steelblue','tan','teal','thistle','tomato','turquoise','violet','wheat','white','whitesmoke','yellow','yellowgreen']
        colormapname="jet"
        
        MPPlegendMod = []
        MPPlinestyle = []
        
        titIV =0
        titmpp=0
        titStat=0
        samplesgroups=["Default group"]
        groupstoplotcomp=["Default group"]
        groupcomment={}
        
        listofanswer=[]
        listoflinestyle=[]
        listofcolorstyle=[]
        listoflinewidthstyle=[]
        
        listofanswermpp=[]
        listoflinestylempp=[]
        listofcolorstylempp=[]
        listoflinewidthstylempp=[]
        
        #clear table
        self.updateTable(DATA)
        self.Confirmgroupnamechanges()
        self.Confirmgroup2namechanges()
        self.ui.listWidget_MppSamples.clear()
        self.ui.comboBox_FFloss.clear()
        self.PlotIV()
        self.PlotMPP()
        self.UpdateHistGraph()
        self.UpdateBoxGraph2()
        self.fig31.clear()
        self.UpdateTimeGraph()

        self.TimeEvolgraph.clear()
        self.UpdateCompGraph()
        #initialize the progress bar
        self.setProgressVal(0)
        
    def startimporting(self):
        global DATA 
        global DATAFV
        global DATAMPP,resultsdirectory
        global DATAdark, numbLightfiles, numbDarkfiles, file_path
        
        
        finished=0
        j=0
        
        if self.ui.checkBox_GBsolarsim.isChecked():
            j=2
            # print("GBpklab-SERIS files")
            self.thread = Thread_getdatalistsfromGBpklabSERIS(file_path)
            self.thread.change_value.connect(self.setProgressVal)
            self.thread.finished.connect(self.ImportFinished)
            self.thread.start()
            finished=1
        
        while j<1:
            file_pathnew=[]
            file_path = QFileDialog.getOpenFileNames(caption = 'Please select the JV files')[0]
            if file_path!='':
                filetypes=[os.path.splitext(item)[1] for item in file_path]
#                print(list(set(filetypes)))
                if len(list(set(filetypes)))==1 or (''in list(set(filetypes)) and '.txt'in list(set(filetypes))) or (".itx" in list(set(filetypes)) and '.txt'in list(set(filetypes))):
                    directory = os.path.join(str(Path(file_path[0]).parent.parent),'resultFilesIV')
                    resultsdirectory=directory
                    if not os.path.exists(directory):
                        os.makedirs(directory)
                        os.chdir(directory)
                    else :
                        os.chdir(directory)
                    filetype=list(set(filetypes))[0]
                    if filetype==".txt" or filetype=='':
                        filetoread = open(file_path[0],"r", encoding='ISO-8859-1')
                        filerawdata = filetoread.readlines()
                        # print(filerawdata[0])
                        if '#Singapore Solar Simulator, Python'in filerawdata[0]:
                            print("SERISpython files")
                            self.thread = Thread_getdatalistsfromCUBpyfiles(file_path)
                            self.thread.change_value.connect(self.setProgressVal)
                            self.thread.finished.connect(self.ImportFinished)
                            self.thread.start()
                            finished=1
                        break
                    elif filetype==".xlsx":
                        print('xlsx')
                        celltest=[]
                        for k in range(len(file_path)):
                            wb = xlrd.open_workbook(file_path[k])
                            xlsheet = wb.sheet_by_index(0)
                            celltest.append(str(xlsheet.cell(10,1).value))
                        if len(list(set(celltest)))==1:
                            print('SERISsinusGUI files')
                            self.thread = Thread_getdatalistsfromSERISsinusGUI(file_path)
                            self.thread.change_value.connect(self.setProgressVal)
                            self.thread.finished.connect(self.ImportFinished)
                            self.thread.start()
                            finished=1
                            break
                        else:
                            QMessageBox.information(self,'Information', "Multiple types of files... please choose one!")
                            j+=1
                            
                    else:
                        QMessageBox.information(self,'Information', "neither .iv or .xls IV files... try again")
                        j+=1
                else:
                    QMessageBox.information(self,'Information', "Multiple types of files... please choose one!")
                    j+=1
            else:
                QMessageBox.information(self,'Information', "Please select IV files")
                j+=1

    def ImportFinished(self):
        global file_path
        global DATA, DATAdark, DATAFFloss
        global DATAMPP, numbLightfiles, numbDarkfiles
        
        print(len(DATA))
        print(len(DATAMPP.keys()))
        
        if DATAMPP!={}:
            self.ui.listWidget_MppSamples.clear()
            self.ui.listWidget_MppSamples.addItems(DATAMPP.keys())

        self.updateTable(DATA)
        
        self.ui.comboBox_FFloss.addItems(DATAFFloss.keys())
        
        # self.updategrouptoplotdropbutton()
        # self.updateCompgrouptoplotdropbutton()
        # self.updateHistgrouptoplotdropbutton()
        # self.UpdateGroupGraph(1)
        # self.UpdateCompGraph(1)
        
    def setProgressVal(self, val):
        global DATA
        
        self.ui.progressBar_ImportData.setValue(val)
        
        # self.updateTable(DATA) #fills the table live as data comes. problem is that if the user messes around with the table while it's loading the data gets messed up. 

        
    def updateTable(self, dictdata):
        try:
            self.ui.tableWidget.setRowCount(0)
            self.ui.tableWidget.setSortingEnabled(False)
            self.ui.tableWidget.setRowCount(len(dictdata.keys()))
            self.ui.tableWidget.setColumnCount(19)
            self.ui.tableWidget.setHorizontalHeaderLabels(
                ['Group','Group2','SampleName','DateTime','Eff. (%)','FF (%)', 'Voc (mV)', 'Jsc (mA/cm2)','Isc (A)', 
                 'Roc (ohm.cm2)', 'Rsc (ohm.cm2)', 'Pmpp (W/m2)', 'Vmpp (mV)', 'Jmpp (mA/cm2)','Area','ScanDirect.','Illum.','SunInten.', 'Comment'])
            i=0
            
            
            for key in dictdata.keys():
                self.ui.tableWidget.setItem(i,0,QTableWidgetItem(dictdata[key]['Group']))
                try:
                    self.ui.tableWidget.setItem(i,1,QTableWidgetItem(dictdata[key]['Group2']))
                except:
                    dictdata[key]['Group2']=''
                    self.ui.tableWidget.setItem(i,1,QTableWidgetItem(dictdata[key]['Group2']))
                    
                self.ui.tableWidget.setItem(i,2,QTableWidgetItem(dictdata[key]['SampleName']))
                self.ui.tableWidget.setItem(i,3,QTableWidgetItem(str(dictdata[key]['MeasDayTime2'])))
                item3=QtWidgets.QTableWidgetItem()
                item3.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Eff']))
                self.ui.tableWidget.setItem(i,4,item3)
                item4=QtWidgets.QTableWidgetItem()
                item4.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['FF']))
                self.ui.tableWidget.setItem(i,5,item4)
                item5=QtWidgets.QTableWidgetItem()
                item5.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Voc']))
                self.ui.tableWidget.setItem(i,6,item5)
                item6=QtWidgets.QTableWidgetItem()
                item6.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Jsc']))
                self.ui.tableWidget.setItem(i,7,item6)
                item7=QtWidgets.QTableWidgetItem()
                item7.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Isc']))
                self.ui.tableWidget.setItem(i,8,item7)
                item8=QtWidgets.QTableWidgetItem()
                item8.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Roc']))
                self.ui.tableWidget.setItem(i,9,item8)
                item9=QtWidgets.QTableWidgetItem()
                item9.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Rsc']))
                self.ui.tableWidget.setItem(i,10,item9)
                item10=QtWidgets.QTableWidgetItem()
                item10.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Pmpp']))
                self.ui.tableWidget.setItem(i,11,item10)
                item11=QtWidgets.QTableWidgetItem()
                item11.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Vmpp']))
                self.ui.tableWidget.setItem(i,12,item11)
                item12=QtWidgets.QTableWidgetItem()
                item12.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Jmpp']))
                self.ui.tableWidget.setItem(i,13,item12)
                item13=QtWidgets.QTableWidgetItem()
                item13.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['CellSurface']))
                self.ui.tableWidget.setItem(i,14,item13)
                self.ui.tableWidget.setItem(i,15,QTableWidgetItem(dictdata[key]['ScanDirection']))
                self.ui.tableWidget.setItem(i,16,QTableWidgetItem(dictdata[key]['Illumination']))
                item14=QtWidgets.QTableWidgetItem()
                item14.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['sunintensity']))
                self.ui.tableWidget.setItem(i,17,item14)
                self.ui.tableWidget.setItem(i,18,QTableWidgetItem(dictdata[key]['MeasComment']))
                i+=1
            self.ui.tableWidget.setSortingEnabled(True)
        except RuntimeError:
            pass
        
            
    def ClearTable(self):
        self.ui.tableWidget.setRowCount(0)
    
    def DeleteMPP(self):
        global DATAMPP
        
        items = self.ui.listWidget_MppSamples.selectedItems()
        sampleselected = []
        for i in range(len(items)):
            sampleselected.append(str(self.ui.listWidget_MppSamples.selectedItems()[i].text()))
        for item in sampleselected:
            del(DATAMPP[item])
        # if DATAMPP!={}:
        self.ui.listWidget_MppSamples.clear()
        self.ui.listWidget_MppSamples.addItems(DATAMPP.keys())
        
        # print(DATAMPP.keys())
        
    def DeleteRows(self):
        global DATA, DATAFFloss
        # print("before: Data has "+str(len(DATA.keys()))+' elements and table has '+str(self.ui.tableWidget.rowCount()))
        
        rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
        sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
        
        for item in sampleselected:
            del(DATA[item])
        
        self.updateTable(DATA)
        
        self.ui.comboBox_FFloss.clear()
        self.ui.comboBox_FFloss.addItems(DATAFFloss.keys())

    def ConfirmMPPAreachanges(self):
        global DATAMPP
        
        items = self.ui.listWidget_MppSamples.selectedItems()
        selectedmpptraces = []
        for i in range(len(items)):
            selectedmpptraces.append(str(self.ui.listWidget_MppSamples.selectedItems()[i].text()))
        
        for i in selectedmpptraces:
            oldarea=float(DATAMPP[i]['CellSurface'])
            newarea=float(self.ui.doubleSpinBox_mppChangeArea.value())
            DATAMPP[i]['CellSurface']=newarea
            DATAMPP[i]['MppData'][1]=[x*oldarea/newarea for x in DATAMPP[i]['MppData'][1]]
            DATAMPP[i]['MppData'][3]=[x*oldarea/newarea for x in DATAMPP[i]['MppData'][3]]
        self.PlotMPP()
        
    def ConfirmAreachanges(self):
        global DATA
        print('areachange')
        if self.ui.checkBox_applyArea.isChecked():
            print('checked')
            rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
            sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
            for i in sampleselected:
                oldarea=float(DATA[i]['CellSurface'])
                newarea=float(self.ui.doubleSpinBox_newarea.value())
                DATA[i]['CellSurface']=newarea
                # Jsc, Jmpp, eff, Pmpp, roc, rsc, 
                DATA[i]['Jsc']=DATA[i]['Jsc']*oldarea/newarea
                DATA[i]['Jmpp']=DATA[i]['Jmpp']*oldarea/newarea
                DATA[i]['Eff']=DATA[i]['Eff']*oldarea/newarea
                DATA[i]['Pmpp']=DATA[i]['Pmpp']*oldarea/newarea
                DATA[i]['Roc']=DATA[i]['Roc']*newarea/oldarea
                DATA[i]['Rsc']=DATA[i]['Rsc']*newarea/oldarea
                
                #JV data
                DATA[i]['IVData'][1]=[float(x)*oldarea/newarea for x in DATA[i]['IVData'][1]]

        else:
            for i in range(self.ui.tableWidget.rowCount()):
                sn=self.ui.tableWidget.item(i,2).text()+'_'+str(self.ui.tableWidget.item(i,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(i,8).text()))+'_'+str(float(self.ui.tableWidget.item(i,5).text()))
                oldarea=DATA[sn]['CellSurface']
                newarea=float(self.ui.tableWidget.item(i,14).text())
                DATA[sn]['CellSurface']=newarea
                # Jsc, Jmpp, eff, Pmpp, roc, rsc, 
                DATA[sn]['Jsc']=DATA[sn]['Jsc']*oldarea / newarea
                DATA[sn]['Jmpp']=DATA[sn]['Jmpp']*oldarea/newarea
                DATA[sn]['Eff']=DATA[sn]['Eff']*oldarea/newarea
                DATA[sn]['Pmpp']=DATA[sn]['Pmpp']*oldarea/newarea
                DATA[sn]['Roc']=DATA[sn]['Roc']*newarea/oldarea
                DATA[sn]['Rsc']=DATA[sn]['Rsc']*newarea/oldarea
                #JV data
                DATA[sn]['IVData'][1]=[x*oldarea/newarea for x in DATA[sn]['IVData'][1]]
                
        self.updateTable(DATA)
    
    def ConfirmSunIntenchanges(self):
        global DATA
        print('SunIntenchange')
        if self.ui.checkBox_applySunInten.isChecked():
            print('checked')
            rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
            sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
            for i in sampleselected:
                DATA[i]['sunintensity']=float(self.ui.doubleSpinBox_newSunInten.value())

        else:
            for i in range(self.ui.tableWidget.rowCount()):
                sn=self.ui.tableWidget.item(i,2).text()+'_'+str(self.ui.tableWidget.item(i,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(i,8).text()))+'_'+str(float(self.ui.tableWidget.item(i,5).text()))
                DATA[i]['sunintensity']=float(self.ui.tableWidget.item(i,16).text())
                
        self.updateTable(DATA)
        
    def switchD2L(self):
        global DATA, DATAdark
        
        #change in DATA
        rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
        sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
        for i in sampleselected:
            if DATA[i]['Illumination']=="Dark":
                DATA[i]['Illumination']="Light"
            else:
                DATA[i]['Illumination']="Dark"
        self.updateTable(DATA)
        #update table and graphs

    def ConfirmSampleNameChanges(self):
        global DATA
        if self.ui.checkBox_allNameChange.isChecked():
            rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
            sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
            for i in sampleselected:
                DATA[i]['SampleName']=self.ui.lineEdit_NameChange.text()+"_"+DATA[i]["Cellletter"]+"_"+i.split('_')[3]
                DATA[i]["SampleNameID"]=DATA[i]["SampleName"]+'_'+str(DATA[i]["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(DATA[i]["Isc"]))+'_'+str(float(DATA[i]["FF"]))
                DATA[i]["IVlinestyle"][0]=DATA[i]['SampleName']
                DATA[i]["batchname"]=DATA[i]['SampleName'].split('_')[0]
                DATA[i]["DepID"]=self.ui.lineEdit_NameChange.text()
                copydata=copy.deepcopy(DATA[i])
                DATA.pop(i, None)
                DATA[copydata['SampleNameID']]=copydata
                # print(copydata['SampleNameID'])
            self.updateTable(DATA)
        else:
            for i in range(self.ui.tableWidget.rowCount()):
                sn=self.ui.tableWidget.item(i,2).text()+'_'+str(self.ui.tableWidget.item(i,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(i,8).text()))+'_'+str(float(self.ui.tableWidget.item(i,5).text()))
                DATA[sn]['SampleName']=self.ui.tableWidget.item(i,2).text()+"_"+DATA[sn]["Cellletter"]+"_"+sn.split('_')[3]
                DATA[sn]["SampleNameID"]=DATA[sn]["SampleName"]+'_'+str(DATA[sn]["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(DATA[sn]["Isc"]))+'_'+str(float(DATA[sn]["FF"]))
                copydata=copy.deepcopy(DATA[sn])
                DATA.pop(sn, None)
                DATA[copydata['SampleNameID']]=copydata
            self.updateTable(DATA)
            
    def ConfirmCommentChanges(self):
        global DATA
        if self.ui.checkBox_allcomment.isChecked():
            rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
            sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
            for i in sampleselected:
                DATA[i]['MeasComment']=self.ui.lineEdit_applynewcomment.text()
            self.updateTable(DATA)
        else:
            for i in range(self.ui.tableWidget.rowCount()):
                sn=self.ui.tableWidget.item(i,2).text()+'_'+str(self.ui.tableWidget.item(i,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(i,8).text()))+'_'+str(float(self.ui.tableWidget.item(i,5).text()))
                DATA[sn]['MeasComment']=self.ui.tableWidget.item(i,18).text()
            self.updateTable(DATA)
                
    def Confirmgroup2namechanges(self):
        global DATA, samplesgroups,samplesgroups12,groupSubgroups
        
        # self.ui.listWidget_BoxPlotGroup.clear()
        # self.ui.listWidget_BoxPlotGroup2.clear()
        self.ui.listWidget_HistoGroups.clear()
        self.ui.listWidget_ParamComp.clear()
        
        # self.ui.listWidget_BoxPlotGroup_2.clear()
        self.ui.listWidget_BoxPlotGroup2_2.clear()

        if self.ui.checkBox_applygroupname.isChecked():
            rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
            sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
            for i in sampleselected:
                DATA[i]['Group2']=self.ui.lineEdit_applygroup.text()
            self.updateTable(DATA)
        else:
            for i in range(self.ui.tableWidget.rowCount()):
                sn=self.ui.tableWidget.item(i,2).text()+'_'+str(self.ui.tableWidget.item(i,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(i,8).text()))+'_'+str(float(self.ui.tableWidget.item(i,5).text()))
                DATA[sn]['Group2']=self.ui.tableWidget.item(i,1).text()
        
        groupnames=[self.ui.tableWidget.item(i,0).text() for i in range(self.ui.tableWidget.rowCount())]
        group1and2names=[self.ui.tableWidget.item(i,0).text()+' * '+self.ui.tableWidget.item(i,1).text() for i in range(self.ui.tableWidget.rowCount())]
        
        groupnames=sorted(list(set(groupnames)))
        samplesgroups=groupnames
        group1and2names=sorted(list(set(group1and2names)))
        samplesgroups12=group1and2names
        
        # self.ui.listWidget_BoxPlotGroup.addItems(groupnames)
        # self.ui.listWidget_BoxPlotGroup2.addItems(group1and2names)
        self.ui.listWidget_HistoGroups.addItems(group1and2names)
        self.ui.listWidget_ParamComp.addItems(group1and2names)
        
        # self.ui.listWidget_BoxPlotGroup_2.addItems(groupnames)
        self.ui.listWidget_BoxPlotGroup2_2.addItems(group1and2names)
        
    def Confirmgroupnamechanges(self):
        global DATA, groupcomment, samplesgroups, samplesgroups12,groupSubgroups
        # self.ui.listWidget_BoxPlotGroup.clear()
        # self.ui.listWidget_BoxPlotGroup2.clear()
        # self.ui.listWidget_BoxPlotGroup_2.clear()
        self.ui.listWidget_BoxPlotGroup2_2.clear()
        self.ui.listWidget_HistoGroups.clear()
        self.ui.listWidget_ParamComp.clear()
        
        if self.ui.checkBox_applygroupname.isChecked():
            rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
            sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
            for i in sampleselected:
                DATA[i]['Group']=self.ui.lineEdit_applygroup.text()
            self.updateTable(DATA)
        else:
            for i in range(self.ui.tableWidget.rowCount()):
                sn=self.ui.tableWidget.item(i,2).text()+'_'+str(self.ui.tableWidget.item(i,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(i,8).text()))+'_'+str(float(self.ui.tableWidget.item(i,5).text()))
                DATA[sn]['Group']=self.ui.tableWidget.item(i,0).text()
        
        groupnames=[self.ui.tableWidget.item(i,0).text() for i in range(self.ui.tableWidget.rowCount())]
        group1and2names=[self.ui.tableWidget.item(i,0).text()+' * '+self.ui.tableWidget.item(i,1).text() for i in range(self.ui.tableWidget.rowCount())]
        
        groupnames=sorted(list(set(groupnames)))
        samplesgroups=groupnames
        group1and2names=sorted(list(set(group1and2names)))
        samplesgroups12=group1and2names
        
        # self.ui.listWidget_BoxPlotGroup.addItems(groupnames)
        # self.ui.listWidget_BoxPlotGroup2.addItems(group1and2names)
        # self.ui.listWidget_BoxPlotGroup_2.addItems(groupnames)
        self.ui.listWidget_BoxPlotGroup2_2.addItems(group1and2names)
        self.ui.listWidget_HistoGroups.addItems(group1and2names)
        self.ui.listWidget_ParamComp.addItems(group1and2names)
        
        # print(type(groupcomment))
        # print(groupcomment)
        # print(groupnames)
        for groupname in groupnames:
            if groupname not in groupcomment.keys():
                groupcomment[groupname]=""
        
        self.clearLayout(self.ui.gridLayout_23)
        self.ui.scrollArea_DBgroupcomment = QtWidgets.QScrollArea(self.ui.frame_3)
        self.ui.scrollArea_DBgroupcomment.setWidgetResizable(True)
        self.ui.scrollArea_DBgroupcomment.setObjectName("scrollArea_DBgroupcomment")
        self.ui.scrollAreaWidgetContents_3 = QtWidgets.QWidget()
        self.ui.scrollAreaWidgetContents_3.setGeometry(QtCore.QRect(0, 0, 500, 400))
        self.ui.scrollAreaWidgetContents_3.setObjectName("scrollAreaWidgetContents_3")
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ui.scrollAreaWidgetContents_3.sizePolicy().hasHeightForWidth())
        self.ui.scrollAreaWidgetContents_3.setSizePolicy(sizePolicy)
        self.ui.verticalLayout_DBgroupcomment = QtWidgets.QVBoxLayout(self.ui.scrollAreaWidgetContents_3)
        self.ui.verticalLayout_DBgroupcomment.setObjectName("verticalLayout_DBgroupcomment")
        
        item1=0
        for itemm in groupnames:
            # print(item1)
            self.frame = QtWidgets.QFrame(self.ui.scrollAreaWidgetContents_3)
            self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
            self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
            self.frame.setObjectName("frame")
            self.horizontalLayout = QtWidgets.QHBoxLayout(self.frame)
            self.horizontalLayout.setObjectName("horizontalLayout")
            
            label = QtWidgets.QLabel(self.frame)
            label.setText(itemm)
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(label.sizePolicy().hasHeightForWidth())
            label.setSizePolicy(sizePolicy)
            self.horizontalLayout.addWidget(label)
            
            comment=QtWidgets.QLineEdit(self.frame)
            comment.setText(groupcomment[itemm])
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(comment.sizePolicy().hasHeightForWidth())
            comment.setSizePolicy(sizePolicy)
            self.horizontalLayout.addWidget(comment)
            comment.textChanged.connect(partial(self.UpdateGroupcomment,itemm))
            # comment.textChanged.connect(self.UpdateGroupcomment)
            item1+=1
            self.ui.verticalLayout_DBgroupcomment.addWidget(self.frame)
        self.ui.scrollArea_DBgroupcomment.setWidget(self.ui.scrollAreaWidgetContents_3)
        self.ui.gridLayout_23.addWidget(self.ui.scrollArea_DBgroupcomment, 0, 0, 1, 1)
    
    def UpdateGroupcomment(self,groupname,newcomment):
        global groupcomment
        # print('updategroupcomment')
        # print(type(groupcomment))
        groupcomment[groupname]=newcomment
        # print(groupcomment)
        
        
    def SavitzkyGolayFiltering(self):
        global DATA
        
        if self.ui.spinBox_SG1.value()>self.ui.spinBox_SG2.value() and self.ui.spinBox_SG1.value()%2==1:
            #get samples selected
            rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
            sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
            
            for item in sampleselected:
                y=DATA[item]['IVData'][1]
                y=array(y)
                # print(y)
                # print(self.ui.spinBox_SG1.value())
                DATA[item]['IVData'][1] = list(savitzky_golay(y, window_size=self.ui.spinBox_SG1.value(), order=self.ui.spinBox_SG2.value()))
                
            self.PlotIV2(sampleselected)
                
            #loop on data to apply SGfilter
            #save in Data mod lists
            #recalculate the JV parameter with external function
            #update IV graph
            #update Table
            
                        
                                
                # self.UpdateEQEGraph()
        else:
            QMessageBox.information(self,'Information', "the SG window-size must be larger than the SG order, positive and odd.")

    def backtoOriginal(self):
        global DATA
        # samplestakenforplot = self.ui.listWidget.selectedItems()
        # samplestakenforplot=[x.text() for x in samplestakenforplot]
        # if samplestakenforplot!=[]:
        #     for item in samplestakenforplot:

        #         DATA[item]['DATAmod'][2]=DATA[item]['DATAorig'][2]

                #recalcualte JV parameters
        
        # self.UpdateEQEGraph()
        
#%%#############FFlosses
        
    def FFlosses(self):
        global DATA, DATAFFloss
        
        try:
            currentsample= self.ui.comboBox_FFloss.currentText()

            self.ui.doubleSpinBox_FFloss_Eff.setValue(DATAFFloss[currentsample]['Eff'])
            self.ui.doubleSpinBox_FFloss_Voc.setValue(DATAFFloss[currentsample]['Voc'])
            self.ui.doubleSpinBox_FFloss_Jsc.setValue(DATAFFloss[currentsample]['Jsc'])
            self.ui.doubleSpinBox_FFloss_Vmpp.setValue(DATAFFloss[currentsample]['Vmpp'])
            self.ui.doubleSpinBox_FFloss_Jmpp.setValue(DATAFFloss[currentsample]['Jmpp'])
            self.ui.doubleSpinBox_FFloss_FF.setValue(DATAFFloss[currentsample]['FF'])
            self.ui.doubleSpinBox_FFloss_Rs.setValue(DATAFFloss[currentsample]['Roc'])
            self.ui.doubleSpinBox_FFloss_Rsh.setValue(DATAFFloss[currentsample]['Rsh'])
            self.ui.doubleSpinBox_FFloss_pFF.setValue(DATAFFloss[currentsample]['pFF'])
            self.ui.doubleSpinBox_FFloss_pVoc.setValue(DATAFFloss[currentsample]['pVoc'])
            self.ui.doubleSpinBox_FFloss_pJsc.setValue(DATAFFloss[currentsample]['pJsc'])
            self.ui.doubleSpinBox_FFloss_pVmpp.setValue(DATAFFloss[currentsample]['pVmpp'])
            self.ui.doubleSpinBox_FFloss_pJmpp.setValue(DATAFFloss[currentsample]['pJmpp'])
            
            
            # xideal=[x*DATAFFloss[currentsample]['Voc']/1000 for x in range(1000)]
            # yideal=[DATAFFloss[currentsample]['Jsc']-(DATAFFloss[currentsample]['Jsc']/math.exp(DATAFFloss[currentsample]['Voc']/25.7))*(math.exp(item/25.7)-1) for item in xideal]
            # powerideal=[xideal[i]*yideal[i]/1000 for i in range(1000)]
            self.ui.doubleSpinBox_FFloss_iFF.setValue(DATAFFloss[currentsample]['idFF'])
            # self.ui.doubleSpinBox_FFloss_iFF.setValue(100*max(powerideal)/(DATAFFloss[currentsample]['Voc']*DATAFFloss[currentsample]['Jsc']/1000))
            self.ui.doubleSpinBox_FFloss_iVmpp.setValue(DATAFFloss[currentsample]['idVmpp'])
            # self.ui.doubleSpinBox_FFloss_iVmpp.setValue(xideal[powerideal.index(max(powerideal))])
            self.ui.doubleSpinBox_FFloss_iJmpp.setValue(DATAFFloss[currentsample]['idJmpp'])
            # self.ui.doubleSpinBox_FFloss_iJmpp.setValue(yideal[powerideal.index(max(powerideal))])
            self.ui.doubleSpinBox_FFloss_VmppJmppRs.setValue(DATAFFloss[currentsample]['VmppPlusJmppRs'])
            # VmppJmppRs=DATAFFloss[currentsample]['Vmpp']+DATAFFloss[currentsample]['Jmpp']*DATAFFloss[currentsample]['Roc']
            # self.ui.doubleSpinBox_FFloss_VmppJmppRs.setValue(VmppJmppRs)
            self.ui.doubleSpinBox_FFloss_AbsRsloss.setValue(DATAFFloss[currentsample]['SerisResistanceAbs'])
            # self.ui.doubleSpinBox_FFloss_AbsRsloss.setValue(100*(DATAFFloss[currentsample]['Jmpp']*DATAFFloss[currentsample]['Jmpp']*DATAFFloss[currentsample]['Roc'])/(DATAFFloss[currentsample]['Voc']*DATAFFloss[currentsample]['Jsc']))
            self.ui.doubleSpinBox_FFloss_AbsRshloss.setValue(DATAFFloss[currentsample]['Shuntterm'])
            self.ui.doubleSpinBox_FFloss_lossseries.setValue(DATAFFloss[currentsample]['FFlossSeries'])
            self.ui.doubleSpinBox_FFloss_lossshunt.setValue(DATAFFloss[currentsample]['FFlossShunt'])
            self.ui.doubleSpinBox_FFloss_lossJ02.setValue(DATAFFloss[currentsample]['FFlossJo2'])
            # self.ui.doubleSpinBox_FFloss_AbsRshloss.setValue(100*VmppJmppRs*VmppJmppRs/(DATAFFloss[currentsample]['Voc']*DATAFFloss[currentsample]['Jsc']*DATAFFloss[currentsample]['Rsh']))
            # if self.ui.doubleSpinBox_FFloss_pFF.value()<self.ui.doubleSpinBox_FFloss_FF.value():
            #     self.ui.doubleSpinBox_FFloss_lossseries.setValue(self.ui.doubleSpinBox_FFloss_AbsRsloss.value())
            # else:
            #     self.ui.doubleSpinBox_FFloss_lossseries.setValue(self.ui.doubleSpinBox_FFloss_pFF.value()-self.ui.doubleSpinBox_FFloss_FF.value())
            # self.ui.doubleSpinBox_FFloss_lossshunt.setValue(self.ui.doubleSpinBox_FFloss_AbsRshloss.value())
            # self.ui.doubleSpinBox_FFloss_lossJ02.setValue(self.ui.doubleSpinBox_FFloss_iFF.value()-self.ui.doubleSpinBox_FFloss_lossseries.value()-self.ui.doubleSpinBox_FFloss_lossshunt.value()-DATAFFloss[currentsample]['FF'])
            
            #plot in stacked histogram the FF losses and measured FF
            
            # self.FFlossgraph.hist([[],[],], 1, density=True, histtype='bar', stacked=True)
            # sizes=[DATAFFloss[currentsample]['FF'],DATAFFloss[currentsample]['FFlossSeries'],DATAFFloss[currentsample]['FFlossJo2'],DATAFFloss[currentsample]['FFlossShunt']]
            # labels=['meas. FF','FFlossSeries','FFlossJo2','FFlossShunt']
            # colors=['black','red','blue','green']
            # self.FFlossgraph.squarify.plot(sizes=sizes, label=labels, color=colors, alpha=0.6 )
            # self.FFlossgraph.axis('off')
            p4=self.FFlossgraph.bar(0,DATAFFloss[currentsample]['FF']+DATAFFloss[currentsample]['FFlossSeries']+DATAFFloss[currentsample]['FFlossJo2']+DATAFFloss[currentsample]['FFlossShunt'],0.3,yerr=0,color='green')
            p3=self.FFlossgraph.bar(0,DATAFFloss[currentsample]['FF']+DATAFFloss[currentsample]['FFlossSeries']+DATAFFloss[currentsample]['FFlossJo2'],0.3,yerr=0,color='blue')
            p2=self.FFlossgraph.bar(0,DATAFFloss[currentsample]['FF']+DATAFFloss[currentsample]['FFlossSeries'],0.3,yerr=0,color='red')
            p1=self.FFlossgraph.bar(0,DATAFFloss[currentsample]['FF'],0.3,yerr=0,color='black')
            p5=self.FFlossgraph.bar(1,DATAFFloss[currentsample]['idFF'],0.3,yerr=0,color='grey')
            self.FFlossgraph.legend((p1[0], p2[0], p3[0], p4[0], p5[0]), ('FF,'+" %.1f" % DATAFFloss[currentsample]['FF']+'%', 'FFlossSeries,'+" %.1f" % DATAFFloss[currentsample]['FFlossSeries']+'%','FFlossJo2,'+" %.1f" % DATAFFloss[currentsample]['FFlossJo2']+'%','FFlossShunt,'+" %.1f" % DATAFFloss[currentsample]['FFlossShunt']+'%','Ideal FF,'+" %.1f" % DATAFFloss[currentsample]['idFF']+'%'))
            
            self.FFlossgraph.get_xaxis().set_visible(False)
            self.fig8.canvas.draw_idle()
        except KeyError:
            pass

    def exportFFloss(self):
        global DATAFFloss
        path = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]

        self.fig8.savefig(path, dpi=300)
        currentsample= self.ui.comboBox_FFloss.currentText()
        
        DATAFFlossforexport='Eff'+'\t'+str(DATAFFloss[currentsample]['Eff'])+'\n'+\
        'Voc'+'\t'+str(DATAFFloss[currentsample]['Voc'])+'\n'+\
        'Jsc'+'\t'+str(DATAFFloss[currentsample]['Jsc'])+'\n'+\
        'Vmpp'+'\t'+str(DATAFFloss[currentsample]['Vmpp'])+'\n'+\
        'Jmpp'+'\t'+str(DATAFFloss[currentsample]['Jmpp'])+'\n'+\
        'FF'+'\t'+str(DATAFFloss[currentsample]['FF'])+'\n'+\
        'Rs'+'\t'+str(DATAFFloss[currentsample]['Roc'])+'\n'+\
        'Rsh'+'\t'+str(DATAFFloss[currentsample]['Rsh'])+'\n'+\
        'pFF'+'\t'+str(DATAFFloss[currentsample]['pFF'])+'\n'+\
        'pVoc'+'\t'+str(DATAFFloss[currentsample]['pVoc'])+'\n'+\
        'pJsc'+'\t'+str(DATAFFloss[currentsample]['pJsc'])+'\n'+\
        'pVmpp'+'\t'+str(DATAFFloss[currentsample]['pVmpp'])+'\n'+\
        'pJmpp'+'\t'+str(DATAFFloss[currentsample]['pJmpp'])+'\n'+\
        'iFF'+'\t'+str(DATAFFloss[currentsample]['idFF'])+'\n'+\
        'iVmpp'+'\t'+str(DATAFFloss[currentsample]['idVmpp'])+'\n'+\
        'iJmpp'+'\t'+str(DATAFFloss[currentsample]['idJmpp'])+'\n'+\
        'VmppJmppRs'+'\t'+str(DATAFFloss[currentsample]['VmppPlusJmppRs'])+'\n'+\
        'AbsRsloss'+'\t'+str(DATAFFloss[currentsample]['SerisResistanceAbs'])+'\n'+\
        'AbsRshloss'+'\t'+str(DATAFFloss[currentsample]['Shuntterm'])+'\n'+\
        'lossseries'+'\t'+str(DATAFFloss[currentsample]['FFlossSeries'])+'\n'+\
        'lossshunt'+'\t'+str(DATAFFloss[currentsample]['FFlossShunt'])+'\n'+\
        'lossJ02'+'\t'+str(DATAFFloss[currentsample]['FFlossJo2'])+'\n'
            
        file = open(str(path[:-4]+"_dat.txt"),'w', encoding='ISO-8859-1')
        file.writelines("%s" % item for item in DATAFFlossforexport)
        file.close()
        
#%%#############Database
    def chooseandconnectDB(self):
        global DBisconnected
        if self.ui.lineEdit_DBpath.text()=='':
            f = QFileDialog.getSaveFileName(self, 'Select database')[0]
            if '.db' not in f:
                self.ui.lineEdit_DBpath.setText(f+'.db')
            else:
                self.ui.lineEdit_DBpath.setText(f)
            
        self.db_conn=sqlite3.connect(self.ui.lineEdit_DBpath.text())
        self.theCursor=self.db_conn.cursor()
        CreateAllTables(self.db_conn)
        DBisconnected=1
        self.ui.pushButton_chooseDB.setStyleSheet("background-color: green")
        
    def loadtoDB(self):
        global DATA, DATAMPP,samplesgroups,groupcomment,DBisconnected
        
        if DBisconnected:
            self.Confirmgroupnamechanges()
            self.Confirmgroup2namechanges()
            
            devicetype=self.ui.comboBox_DBdevicetype.currentText()
            generalcomment=self.ui.lineEdit_DBcomments.text()
            
            if self.ui.comboBox_DBJVorMPP.currentText()=='add JV data to DB':
                # print('JV')
                DATAx=DATA
            else:
                # print('MPP')
                DATAx=DATAMPP
                
            for sample in DATAx.keys():
                # print(sample)
                #username in Table users        
                self.theCursor.execute("SELECT id FROM users WHERE user=?",(DATAx[sample]['Operator'],))
                users_id_exists = self.theCursor.fetchone()
                if users_id_exists==None:
                    self.theCursor.execute("INSERT INTO users (user) VALUES (?)",
                                    (DATAx[sample]['Operator'],))
                    users_id_exists=self.theCursor.lastrowid
                else:
                    users_id_exists=users_id_exists[0]
                
                # #batchname in Table batch, with users_id
                self.theCursor.execute("SELECT id FROM batch WHERE batchname =? AND users_id =?",(DATAx[sample]['batchname'],users_id_exists,))
                batch_id_exists = self.theCursor.fetchone()
                if batch_id_exists==None:
                    self.theCursor.execute("INSERT INTO batch (batchname,users_id) VALUES (?,?)",
                                    (DATAx[sample]['batchname'],users_id_exists,))
                    batch_id_exists=self.theCursor.lastrowid
                else:
                    batch_id_exists=batch_id_exists[0]
                    
                #Table Groups
                self.theCursor.execute("SELECT id FROM Groups WHERE GroupName =?",(DATAx[sample]['Group'],))
                group_id_exists = self.theCursor.fetchone()
                if group_id_exists==None:
                    self.theCursor.execute("INSERT INTO Groups (GroupName,LayerStack) VALUES (?,?)",
                                    (DATAx[sample]['Group'],groupcomment[DATAx[sample]['Group']]))
                    group_id_exists=self.theCursor.lastrowid
                else:
                    group_id_exists=group_id_exists[0]
                    
                #samplename in table samples, with batch_id
                self.theCursor.execute("SELECT id FROM samples WHERE samplename =? AND batch_id =?",(DATAx[sample]['SampleName'],batch_id_exists,))
                samples_id_exists = self.theCursor.fetchone()
                if samples_id_exists==None:
                    self.theCursor.execute("INSERT INTO samples (samplename,batch_id,DeviceType) VALUES (?,?,?)",
                                    (DATAx[sample]['SampleName'],batch_id_exists,devicetype,))
                    samples_id_exists=self.theCursor.lastrowid
                else:
                    samples_id_exists=samples_id_exists[0]
                    
                #sample pixel area
                self.theCursor.execute("SELECT id FROM pixelarea WHERE pixel_area =?",(DATAx[sample]['CellSurface'],))
                pixelarea_id_exists = self.theCursor.fetchone()
                if pixelarea_id_exists==None:
                    self.theCursor.execute("INSERT INTO pixelarea (pixel_area) VALUES (?)",
                                    (DATAx[sample]['CellSurface'],))
                    pixelarea_id_exists=self.theCursor.lastrowid
                else:
                    pixelarea_id_exists=pixelarea_id_exists[0]
                    
                #table cells
                self.theCursor.execute("SELECT id FROM cells WHERE cellname =? AND samples_id =? AND batch_id =?",(DATAx[sample]['Cellletter'],samples_id_exists,batch_id_exists,))
                cells_id_exists = self.theCursor.fetchone()
                if cells_id_exists==None:
                    self.theCursor.execute("INSERT INTO cells (cellname,pixelarea_id,samples_id,batch_id) VALUES (?,?,?,?)",
                                    (DATAx[sample]['Cellletter'],pixelarea_id_exists,samples_id_exists,batch_id_exists,))
                    cells_id_exists=self.theCursor.lastrowid
                else:
                    cells_id_exists=cells_id_exists[0]
                    
                if self.ui.comboBox_DBJVorMPP.currentText()=='add JV data to DB':
                    if '_SunsVoc' not in DATAx[sample]['SampleName']:
                        #tables JVmeas
                        self.theCursor.execute("SELECT id FROM JVmeas WHERE MeasurementLongName =? AND cells_id =? AND samples_id =? AND batch_id =?",(DATAx[sample]['SampleNameID'],cells_id_exists,samples_id_exists,batch_id_exists,))
                        JVmeas_id_exists = self.theCursor.fetchone()
                        if JVmeas_id_exists==None:
                            self.theCursor.execute("INSERT INTO JVmeas (DateTimeJV, Eff, Voc,Jsc,Isc, FF, Vmpp, Jmpp,Pmpp,Roc,Rsc,ScanDirect,Delay,IntegTime,Vmin,Vmax,MeasSetup,NbPoints,LightDark,IlluminationIntensity,commentJV,MeasurementLongName,SerialNumber,linktorawdata,samples_id,batch_id,cells_id,Groups_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                            (DATAx[sample]['MeasDayTime2'],
                                             DATAx[sample]['Eff'],DATAx[sample]['Voc'],DATAx[sample]['Jsc'],
                                             DATAx[sample]['Isc'],DATAx[sample]['FF'],DATAx[sample]['Vmpp'],DATAx[sample]['Jmpp'],
                                             DATAx[sample]['Pmpp'],DATAx[sample]['Roc'],DATAx[sample]['Rsc'],DATAx[sample]['ScanDirection'],
                                             DATAx[sample]['Delay'],DATAx[sample]['IntegTime'],DATAx[sample]['Vstart'],DATAx[sample]['Vend'],
                                             DATAx[sample]['Setup'],DATAx[sample]['NbPoints'],
                                             DATAx[sample]['Illumination'],DATAx[sample]['sunintensity'],generalcomment,
                                             DATAx[sample]['SampleNameID'],DATAx[sample]['SerialNumber'],DATAx[sample]['filepath'],
                                             samples_id_exists,batch_id_exists,cells_id_exists,group_id_exists,))
                            JVmeas_id_exists=self.theCursor.lastrowid
                        else:
                            JVmeas_id_exists=JVmeas_id_exists[0]
                    else:
                        #tables SunsVoc
                        self.theCursor.execute("SELECT id FROM SunsVoc WHERE MeasurementLongName =? AND cells_id =? AND samples_id =? AND batch_id =?",(DATAx[sample]['SampleNameID'],cells_id_exists,samples_id_exists,batch_id_exists,))
                        SunsVoc_id_exists = self.theCursor.fetchone()
                        if SunsVoc_id_exists==None:
                            self.theCursor.execute("INSERT INTO SunsVoc (DateTimeSunV,pIsc,pJsc,pVoc,pFF,pPmpp,pETA,temperature,commentSV,MeasurementLongName,SerialNumber,linktorawdata,samples_id,batch_id,cells_id,group_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                            (DATAx[sample]['MeasDayTime2'],DATAx[sample]['Isc'],DATAx[sample]['Jsc'],DATAx[sample]['Voc'],DATAx[sample]['FF'],DATAx[sample]['Pmpp'],DATAx[sample]['Eff'],
                                             DATAx[sample]['ChuckTemp'],generalcomment,
                                             DATAx[sample]['SampleNameID'],DATAx[sample]['SerialNumber'],DATAx[sample]['filepath'],
                                             samples_id_exists,batch_id_exists,cells_id_exists,group_id_exists,))
                            SunsVoc_id_exists=self.theCursor.lastrowid
                        else:
                            SunsVoc_id_exists=SunsVoc_id_exists[0]
    
                else:
                    #tables Mppmeas
                    self.theCursor.execute("SELECT id FROM MPPmeas WHERE MeasurementLongName =? AND cells_id =? AND samples_id =? AND batch_id =?",(DATAx[sample]['SampleNameID'],cells_id_exists,samples_id_exists,batch_id_exists,))
                    MPPmeas_id_exists = self.theCursor.fetchone()
                    if MPPmeas_id_exists==None:
                        self.theCursor.execute("INSERT INTO MPPmeas (DateTimeMPP,TrackingAlgo,MeasSetup,TrackingDuration,Vstart,Vstep,Delay,PowerEnd,commentmpp,LightDark,IlluminationIntensity,MeasurementLongName,SerialNumber,linktorawdata,samples_id,batch_id, cells_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                        (DATAx[sample]['MeasDayTime2'],'perturbe&observe',DATAx[sample]['Setup'],
                                         DATAx[sample]['trackingduration'],DATAx[sample]['Vstart'],DATAx[sample]['Vstep'],DATAx[sample]['Delay'],
                                         DATAx[sample]['PowerEnd'],generalcomment,DATAx[sample]['Illumination'],
                                         DATAx[sample]['sunintensity'],DATAx[sample]['SampleNameID'],DATAx[sample]['SerialNumber'],DATAx[sample]['filepath'],
                                         samples_id_exists,batch_id_exists,cells_id_exists,))
                        MPPmeas_id_exists=self.theCursor.lastrowid
                    else:
                        MPPmeas_id_exists=MPPmeas_id_exists[0]
                self.db_conn.commit()
                
            self.theCursor.close()
            self.db_conn.close()
            self.ui.pushButton_chooseDB.setStyleSheet("background-color: red")
            DBisconnected=0
            QMessageBox.information(self,'Information', "Loading finished")
        else:
            QMessageBox.information(self,'Information', "A .db file should be create and connected to first.")
            
    def connectToDBreading(self):
        global DBisconnected
        
        if DBisconnected:
            self.theCursor.close()
            self.db_conn.close()
            DBisconnected=0
            
        if self.ui.lineEdit_dbpathRead.text()=='':
            f = QFileDialog.getOpenFileName(self, 'Select database')[0]
            if '.db' in f:
                self.ui.lineEdit_dbpathRead.setText(f)
                self.db_conn=sqlite3.connect(self.ui.lineEdit_dbpathRead.text())
                self.theCursor=self.db_conn.cursor()
                DBisconnected=1
                self.onconnecting()
            else:
                QMessageBox.information(self,'Information', "not a .db files?")
        else:
            self.db_conn=sqlite3.connect(self.ui.lineEdit_dbpathRead.text())
            self.theCursor=self.db_conn.cursor()
            DBisconnected=1
            self.onconnecting()
        
    def onconnecting(self):
        self.ui.comboBox_DBTime.clear()
        timecriterialist=timecriteria
        for i in range(len(timecriterialist)):
            self.ui.comboBox_DBTime.addItem(timecriterialist[i])
                
    def on_combobox_DBTime_changed(self):

        self.ui.comboBox_DBTimeYaxis.clear()
        
        yaxisoptionlist=fromtocriteria
        for i in range(len(yaxisoptionlist)):
            if 'JV' in self.ui.comboBox_DBTime.currentText():
                if 'MPP' not in yaxisoptionlist[i] and 'SunsVoc' not in yaxisoptionlist[i]:
                    self.ui.comboBox_DBTimeYaxis.addItem(yaxisoptionlist[i])
            elif 'MPP' in self.ui.comboBox_DBTime.currentText():
                if 'JV' not in yaxisoptionlist[i] and 'SunsVoc' not in yaxisoptionlist[i]:
                    self.ui.comboBox_DBTimeYaxis.addItem(yaxisoptionlist[i])
            elif 'SunsVoc' in self.ui.comboBox_DBTime.currentText():
                if 'JV' not in yaxisoptionlist[i] and 'MPP' not in yaxisoptionlist[i]:
                    self.ui.comboBox_DBTimeYaxis.addItem(yaxisoptionlist[i])
        
        self.ui.comboBox_DBTime_restrictions.clear()
        
        restrictionslist=fromtocriteria + dropdowncriteria
            
        for i in range(len(restrictionslist)):
            if 'JV' in self.ui.comboBox_DBTime.currentText():
                if 'MPP' not in restrictionslist[i] and 'SunsVoc' not in restrictionslist[i]:
                    self.ui.comboBox_DBTime_restrictions.addItem(restrictionslist[i])
            elif 'MPP' in self.ui.comboBox_DBTime.currentText():
                if 'JV' not in restrictionslist[i] and 'SunsVoc' not in restrictionslist[i]:
                    self.ui.comboBox_DBTime_restrictions.addItem(restrictionslist[i])
            elif 'SunsVoc' in self.ui.comboBox_DBTime.currentText():
                if 'JV' not in restrictionslist[i] and 'MPP' not in restrictionslist[i]:
                    self.ui.comboBox_DBTime_restrictions.addItem(restrictionslist[i])

    def on_combobox_DBTimeRestrictions_changed(self):
        criteria=self.ui.comboBox_DBTime_restrictions.currentText()
        if criteria in fromtocriteria:
            # print(criteria)
            try:
                self.theCursor.execute("SELECT "+criteria.split('.')[1]+" FROM "+criteria.split('.')[0])
                listoptions=list(set([x[0] for x in self.theCursor.fetchall()]))
                minimum=min(listoptions)
                maximum=max(listoptions)
                self.ui.lineEdit_DBTime_From.setText(str(minimum))
                self.ui.lineEdit_DBTime_To.setText(str(maximum))
            except ValueError:
                pass
        elif criteria in dropdowncriteria:
            # print('here')
            self.ui.listWidget_DBTime_Restrictions.clear()
            self.theCursor.execute("SELECT "+criteria.split('.')[1]+" FROM "+criteria.split('.')[0])
            listoptions=list(set([x[0] for x in self.theCursor.fetchall()]))
            for item in listoptions:
                self.ui.listWidget_DBTime_Restrictions.addItem(str(item))
    
    def Addrestriction(self):
        
        currentrestrictionparam = self.ui.comboBox_DBTime_restrictions.currentText()
        if currentrestrictionparam in fromtocriteria:
            self.ui.listWidget_DBTime_chosenRestrictions.addItem(currentrestrictionparam+'__'+self.ui.lineEdit_DBTime_From.text()+'__'+self.ui.lineEdit_DBTime_To.text())
        elif currentrestrictionparam in dropdowncriteria:
            selectedrestparam=self.ui.listWidget_DBTime_Restrictions.selectedItems()
            for i in selectedrestparam:
                self.ui.listWidget_DBTime_chosenRestrictions.addItem(currentrestrictionparam+'__'+i.text())
        self.SearchAndPlot()
                
    def Removerestriction(self):
        self.ui.listWidget_DBTime_chosenRestrictions.takeItem(self.ui.listWidget_DBTime_chosenRestrictions.currentRow())
        self.SearchAndPlot()
        
    def AutoGraphExport(self):
        global instructionsofsearch
        
        
        #add red lines and text as in report
        
        current_path = os.getcwd()
        fname = QFileDialog.getSaveFileName(self, 'Export graph', current_path, ".png", "graph file (*.png);; All Files (*)")[0]
        listofimages=[]
        criteria='JVmeas.Eff'
        timecrit='JVmeas.DateTimeJV'
        
        items=['samples.DeviceType__Single jct Perovskite', 'JVmeas.FF__15__88', 'JVmeas.LightDark__Light']
        self.DBplot(items,criteria,timecrit,1)
        
        self.DBgraph.set_ylim([0,20.5])
        # self.fig7.autofmt_xdate()
        self.DBgraph.axhline(y=20, color='r')
        self.DBgraph.set_xlim([datetime.datetime.now() - datetime.timedelta(days=6*30), datetime.datetime.now() + datetime.timedelta(days=3)])
        self.fig7.canvas.draw()
        self.fig7.canvas.flush_events()
        
        self.fig7.savefig(str(fname[:-4]+'_'+datetime.datetime.now().strftime("%Y%m%d")+"_SG.png"), dpi=300)
        listofimages.append(str(fname[:-4]+'_'+datetime.datetime.now().strftime("%Y%m%d")+"_SG.png"))
        
        items=['samples.DeviceType__Perovskite/Silicon 2TT', 'JVmeas.FF__15__88', 'JVmeas.LightDark__Light']
        self.DBplot(items,criteria,timecrit,1)
        
        self.DBgraph.set_ylim([0,30.5])
        # self.fig7.autofmt_xdate()
        self.DBgraph.axhline(y=30, color='r')
        self.DBgraph.axhline(y=25, color='r',linestyle='dashed')
        self.DBgraph.set_xlim([datetime.datetime.now() - datetime.timedelta(days=6*30), datetime.datetime.now() + datetime.timedelta(days=3)])
        self.fig7.canvas.draw()
        self.fig7.canvas.flush_events()
        
        self.fig7.savefig(str(fname[:-4]+'_'+datetime.datetime.now().strftime("%Y%m%d")+"_2tt.png"), dpi=300)
        listofimages.append(str(fname[:-4]+'_'+datetime.datetime.now().strftime("%Y%m%d")+"_2tt.png"))
        
        images = list(map(ImageTk.open, listofimages))
        widths, heights = zip(*(i.size for i in images))
        total_width=max(widths)
        height=sum(heights)
        new_im = ImageTk.new('RGB', (total_width, height), (255, 255, 255))
        new_im.paste(im=images[0],box=(0,0))
        os.remove(listofimages[0])
        new_im.paste(im=images[1],box=(0,heights[1]))
        os.remove(listofimages[1])
        new_im.save(str(fname+'_'+datetime.datetime.now().strftime("%Y%m%d")+".png"))
        
        
    def SearchAndPlot(self):
        global instructionsofsearch
        
        criteria=self.ui.comboBox_DBTimeYaxis.currentText()
        timecrit=self.ui.comboBox_DBTime.currentText()
        
        items = []
        for index in range(self.ui.listWidget_DBTime_chosenRestrictions.count()):
             items.append(self.ui.listWidget_DBTime_chosenRestrictions.item(index).text()) 
        
        # print(items)
        self.DBplot(items,criteria,timecrit)
        

        
        
    def DBplot(self,items=[],criteria='',timecrit='',auto=0):
        global instructionsofsearch
        criteriaListdetailled=[]
        criteriaListdetailled2=[]
        for item in items:
            if len(item.split('__'))==2:
                criteriaListdetailled.append([item.split('__')[0],[item.split('__')[1]]])
            elif len(item.split('__'))==3:
                criteriaListdetailled2.append([item.split('__')[0],item.split('__')[1],item.split('__')[2]])
        
             
        #get the from
             
        #get where: list of id matches
        #get where: from listbox, blabla = blabli or ...
        #get where: from fromto, (blabla between ll and ll) and ...
        parameterList=[timecrit,criteria]
        parametertables=list(set([x.split('.')[0] for x in parameterList]))
        parametertables=sorted(parametertables, key=lambda s: s.casefold())
        criteriaList=[x.split('__')[0] for x in items]
        tablenames=list(set(["batch","samples"]+[timecrit.split('.')[0],criteria.split('.')[0]]+[x.split('.')[0] for x in criteriaList]))
        # print(items)
        # print(criteriaList)
        # print(tablenames)
        
        if criteria.split('.')[0]==timecrit.split('.')[0]:
            SelectInstructions="SELECT "+timecrit +', ' + criteria +" FROM "#+criteria.split('.')[0] + ',samples, batch, cells WHERE '
            for item in tablenames:
                SelectInstructions+=item+', '
            SelectInstructions=SelectInstructions[:-2]+" WHERE "
            wherelist=[criteria.split('.')[0]]
        else:
            SelectInstructions="SELECT "+timecrit +', ' + criteria +" FROM "#+criteria.split('.')[0]+', ' + timecrit.split('.')[0] + ', samples, batch, cells WHERE '
            for item in tablenames:
                SelectInstructions+=item+', '
            SelectInstructions=SelectInstructions[:-2]+" WHERE "
            wherelist=[criteria.split('.')[0]]
             
        items+=[criteria.split('.')[0],timecrit.split('.')[0]]
        # print(items)
        # criteriaList=[self.listWidget2.item(i).text() for i in range(self.listWidget2.count())]
        
        # tablenames=list(set(["batch","samples","cells"]+[x.split('.')[0] for x in items]))
        # tablenames=list(set([x.split('.')[0] for x in items]))
        # print(tablenames)
        wherelist=["samples.batch_id = batch.id AND "]
        # wherelist=[]
        try:
            for item in tablenames:
                self.theCursor.execute("SELECT * FROM "+item)
                headcol=[x[0] for x in self.theCursor.description]  
                headcol=[x[:-3] for x in headcol if '_id' in x] 
                # print(headcol)
                for item2 in headcol:
                    if item2 in tablenames:
                        wherelist.append(item+'.'+item2+'_id = '+item2+'.id AND ')

            wherelist=list(set(wherelist))
            # print(wherelist)
            for item in wherelist:
                SelectInstructions+=item
            for item in criteriaListdetailled:
                SelectInstructions+='('
                for item2 in item[1]:
                    SelectInstructions+= item[0]+' = '+"'"+str(item2)+"' OR "
                SelectInstructions=SelectInstructions[:-4]+') AND '
            for item in criteriaListdetailled2:
                SelectInstructions+= '('+item[0] + ' BETWEEN ' + item[1] + ' AND ' + item[2] +') AND '
            
            # print(SelectInstructions[:-4])
            self.theCursor.execute(SelectInstructions[:-4])
            instructionsofsearch=SelectInstructions[:-4]
            data=self.theCursor.fetchall()
            
            date=[datetime.datetime.strptime(x[0],'%Y-%m-%d %H:%M:%S') for x in data]
            ydata=[x[1] for x in data]
            
            maxdatx=[date[ydata.index(max(ydata))]]
            maxdaty=[max(ydata)]
            
            self.DBgraph.clear()
            self.DBgraph.plot(date,ydata,'o')
            if auto:
                self.DBgraph.plot(maxdatx,maxdaty,'ro')
            self.DBgraph.set_xlabel('Date/Time')
            self.DBgraph.set_ylabel(criteria.split('.')[1])
            self.DBgraph.yaxis.set_minor_locator(AutoMinorLocator())
            for tick in self.DBgraph.get_xticklabels():
                tick.set_rotation(5)
            self.fig7.canvas.draw()
            self.fig7.canvas.flush_events()
        except:
            pass
        
    def ExportDBsearchresults(self):
        global instructionsofsearch
        # print(self.instructionsofsearch.split('WHERE')[1])
        
        # self.theCursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        # alltablesnames=self.theCursor.fetchall()
        # alltablesnames=[item[0] for item in alltablesnames]
        # # print(alltablesnames)
        # allcolumns=[]
        # for tablename in alltablesnames:
        #     self.theCursor.execute("SELECT * FROM "+tablename+";")
        #     allcolumns+=[tablename+"."+member[0] for member in self.theCursor.description]
        # # print(allcolumns)
        # instructionsstring='SELECT '
        # for item in allcolumns:
        #     if 'sqlite'not in item and 'id' not in item:
        #         instructionsstring+=item+', '
        # instructionsstring=instructionsstring[:-2]+' FROM batch '
        # # for item in alltablesnames:
        # #     if 'sqlite'not in item:
        # #         instructionsstring+=item+', '
        # # instructionsstring=instructionsstring[:-2]
        
        # #add the INNER JOIN
        # wherelist=[" INNER JOIN samples ON batch.id = samples.batch_id "]
        # for item in alltablesnames:
        #     self.theCursor.execute("SELECT * FROM "+item)
        #     headcol=[x[0] for x in self.theCursor.description]  
        #     headcol=[x[:-3] for x in headcol if '_id' in x]  
        #     for item2 in headcol:
        #         if item2 in alltablesnames:
        #             wherelist.append(' INNER JOIN '+item+' ON '+item2+'.id = '+item+'.'+item2+'_id ')
        # wherelist=list(set(wherelist))
        # # print(wherelist)
        # for item in wherelist:
        #     instructionsstring+=item
        # #+' WHERE '+self.instructionsofsearch.split('WHERE')[1]#+" ORDER BY JVmeas.MeasurementLongName ASC"

        
        self.ui.comboBox_DBTime.currentText()
        if 'JVmeas' in self.ui.comboBox_DBTime.currentText():
            # print(instructionsofsearch)
            meas='JVmeas'
        elif 'MPPmeas' in self.ui.comboBox_DBTime.currentText():
            # print('mpp')
            meas='MPPmeas'
        elif 'Suns' in self.ui.comboBox_DBTime.currentText():
            # print('sunsvoc')
            meas='SunsVoc'
        
        instructionsstring= """SELECT * FROM """+meas+""" 
                        LEFT OUTER JOIN batch ON """+meas+""".batch_id = batch.id 
                        LEFT OUTER JOIN users ON batch.users_id = users.id 
                        LEFT OUTER JOIN samples ON """+meas+""".samples_id=samples.id
                        LEFT OUTER JOIN cells ON """+meas+""".cells_id=cells.id
                        LEFT OUTER JOIN pixelarea ON cells.pixelarea_id=pixelarea.id
                        LEFT OUTER JOIN Groups ON """+meas+""".Groups_id=Groups.id
                        """
        tablenames=[meas,'batch','users','samples','cells','pixelarea','Groups']
        allcolumns=[]
        for tablename in tablenames:
            self.theCursor.execute("SELECT * FROM "+tablename+";")
            allcolumns+=[tablename+"."+member[0] for member in self.theCursor.description]
        
        
        listinstruc=instructionsofsearch.split('WHERE ')[1].split(' AND ')
        
        addwheres=[x for x in listinstruc if 'id' not in x]
        # print(addwheres)
        if addwheres!=[]:
            instructionsstring+=' WHERE '
            for item in addwheres:
                instructionsstring+=item+' AND '
            instructionsstring=instructionsstring[:-4]
        # print(instructionsstring)
        
        path = QFileDialog.getSaveFileName(caption = 'Select where to save the extracted data')
        if '.xlsx' not in path:
            path=str(path[0])+'.xlsx'
        else:
            path =str(path[0])
        workbook = xlsxwriter.Workbook(path, {'nan_inf_to_errors': True})
        self.theCursor.execute(instructionsstring)
        data=list(set(self.theCursor.fetchall()))
        # print(data)
        if data!=[]:
            data=[tuple(allcolumns)]+data
            worksheetx = workbook.add_worksheet('DATA')
            for item in range(len(data)):
                for item0 in range(len(data[item])):
                    worksheetx.write(item,item0,data[item][item0])
        workbook.close()




#%%#############
    def clearLayout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clearLayout(item.layout())
    
    def onclicklegendtabJV(self,indexoftab):
        # print(indexoftab)   
        if indexoftab==1 and self.ui.checkBox_JVLegend.isChecked():
            self.populateJV()
        elif indexoftab==2 and self.ui.checkBox_JVLegend.isChecked():
            # print('changed1')
            self.populateJVreorder()
    def onclicklegendtabMPP(self,indexoftab):
        # print(indexoftab)   
        if indexoftab==1 and self.ui.checkBox_MppLegend.isChecked():
            self.populateMPP()
    
    def populateJVreorder(self):
        global DATA
        global takenforplot
        
        self.ui.listWidget_reorderJV.clear()
        self.ui.listWidget_reorderJV.addItems(takenforplot)
        self.ui.listWidget_reorderJV.currentItemChanged.connect(self.orderchanged)
        
    def orderchanged(self):
        global DATA
        global takenforplot
        # print("changed")
        # print(takenforplot)
        takenforplot = []
        for x in range(self.ui.listWidget_reorderJV.count()):
            takenforplot.append(self.ui.listWidget_reorderJV.item(x).text())
        # print(takenforplot)
        self.populateJV()
        self.UpdateJVLegMod()
        
        
    def populateJV(self):
        global DATA
        global takenforplot
        global IVlegendMod
        global IVlinestyle
        global colorstylelist
        global listofanswer
        global listoflinestyle
        global listofcolorstyle, listoflinewidthstyle
        
        sampleselected=takenforplot
        listofanswer=[]
        listoflinestyle=[]
        listofcolorstyle=[]
        listoflinewidthstyle=[]
        
        for item in sampleselected:
            listofanswer.append(DATA[item]["IVlinestyle"][0])
            listoflinestyle.append(DATA[item]["IVlinestyle"][1])
            listofcolorstyle.append(DATA[item]["IVlinestyle"][2])
            listoflinewidthstyle.append(str(DATA[item]["IVlinestyle"][3]))
        
        self.clearLayout(self.ui.gridLayout_18)
        self.ui.scrollArea_JVlegend = QtWidgets.QScrollArea(self.ui.EditLegendJV)
        self.ui.scrollArea_JVlegend.setWidgetResizable(True)
        self.ui.scrollArea_JVlegend.setObjectName("scrollArea_JVlegend")
        self.ui.scrollAreaWidgetContents_2 = QtWidgets.QWidget()
        self.ui.scrollAreaWidgetContents_2.setGeometry(QtCore.QRect(0, 0, 500, 400))
        self.ui.scrollAreaWidgetContents_2.setObjectName("scrollAreaWidgetContents")
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ui.scrollAreaWidgetContents_2.sizePolicy().hasHeightForWidth())
        self.ui.scrollAreaWidgetContents_2.setSizePolicy(sizePolicy)
        self.ui.verticalLayout_JVlegend = QtWidgets.QVBoxLayout(self.ui.scrollAreaWidgetContents_2)
        self.ui.verticalLayout_JVlegend.setObjectName("verticalLayout_JVlegend")
        
        item1=0
        for itemm in sampleselected:
            # print(item1)
            self.frame = QtWidgets.QFrame(self.ui.scrollAreaWidgetContents_2)
            self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
            self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
            self.frame.setObjectName("frame")
            self.horizontalLayout = QtWidgets.QHBoxLayout(self.frame)
            self.horizontalLayout.setObjectName("horizontalLayout")
            
            label = QtWidgets.QLabel(self.frame)
            label.setText(DATA[itemm]['SampleName'])
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(label.sizePolicy().hasHeightForWidth())
            label.setSizePolicy(sizePolicy)
            self.horizontalLayout.addWidget(label)
            
            listofanswer[item1]=QtWidgets.QLineEdit(self.frame)
            listofanswer[item1].setText(DATA[itemm]["IVlinestyle"][0])
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(listofanswer[item1].sizePolicy().hasHeightForWidth())
            listofanswer[item1].setSizePolicy(sizePolicy)
            self.horizontalLayout.addWidget(listofanswer[item1])
            listofanswer[item1].textChanged.connect(self.UpdateJVLegMod)
            
            listoflinestyle[item1] = QtWidgets.QComboBox(self.frame)
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(listoflinestyle[item1].sizePolicy().hasHeightForWidth())
            listoflinestyle[item1].setObjectName("comboBox_matname"+str(item1))
            listoflinestyle[item1].addItems(["-","--","-.",":"])
            listoflinestyle[item1].setCurrentText(DATA[itemm]["IVlinestyle"][1])
            self.horizontalLayout.addWidget(listoflinestyle[item1])
            listoflinestyle[item1].currentTextChanged.connect(self.UpdateJVLegMod)
            
            listofcolorstyle[item1] = QtWidgets.QPushButton('Select Color', self.frame)
            listofcolorstyle[item1].setObjectName("button_"+str(item1))
            listofcolorstyle[item1].setStyleSheet("color:"+str(DATA[itemm]["IVlinestyle"][2])+";")
            self.horizontalLayout.addWidget(listofcolorstyle[item1])
            listofcolorstyle[item1].clicked.connect(partial(self.getColor,itemm))
            
            listoflinewidthstyle[item1] = QtWidgets.QSpinBox(self.frame)
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(listoflinewidthstyle[item1].sizePolicy().hasHeightForWidth())
            listoflinewidthstyle[item1].setSizePolicy(sizePolicy)
            listoflinewidthstyle[item1].setMaximum(9999999)
            listoflinewidthstyle[item1].setObjectName("spinBox_"+str(item1))
            listoflinewidthstyle[item1].setValue(DATA[itemm]["IVlinestyle"][3])
            self.horizontalLayout.addWidget(listoflinewidthstyle[item1])
            listoflinewidthstyle[item1].valueChanged.connect(self.UpdateJVLegMod)
            item1+=1
            self.ui.verticalLayout_JVlegend.addWidget(self.frame)
                
        
        self.ui.scrollArea_JVlegend.setWidget(self.ui.scrollAreaWidgetContents_2)
        self.ui.gridLayout_18.addWidget(self.ui.scrollArea_JVlegend, 0, 0, 1, 1)
        
    def getColor(self,rowitem):
        global DATA, listofcolorstyle
        color = QColorDialog.getColor()
        DATA[rowitem]["IVlinestyle"][2]=color.name()
        
        self.UpdateJVLegMod()
        self.populateJV()
        
    def UpdateJVLegMod(self):
        global DATA, takenforplot
        global listofanswer
        global listoflinestyle
        global listofcolorstyle,listoflinewidthstyle
        sampleselected=takenforplot
        leglist=[]
        for e in listofanswer:
            if type(e)!=str:
                leglist.append(e.text())
            else:
                leglist.append(e)
        for item in range(len(sampleselected)):
            DATA[sampleselected[item]]["IVlinestyle"][0]=leglist[item]
        
        # for item in range(len(sampleselected)):
        #     DATA[sampleselected[item]]["IVlinestyle"][2]=listofcolorstyle[item]
            
        leglist=[]
        for e in listoflinestyle:
            if type(e)!=str:
                leglist.append(e.currentText())
            else:
                leglist.append(e)
        for item in range(len(sampleselected)):
            DATA[sampleselected[item]]["IVlinestyle"][1]=leglist[item]
            
        leglist=[]
        for e in listoflinewidthstyle:
            if type(e)!=str:
                leglist.append(e.value())
            else:
                leglist.append(e) 
        for item in range(len(sampleselected)):
            DATA[sampleselected[item]]["IVlinestyle"][3]=int(leglist[item])
                
        # self.populateJV()
        self.PlotIV2(takenforplot)
        
    def populateMPP(self):
        global DATAMPP
        global takenforplotmpp
        global MPPlegendMod
        global MPPlinestyle
        global colorstylelist
        global listofanswermpp
        global listoflinestylempp
        global listofcolorstylempp, listoflinewidthstylempp
        items = self.ui.listWidget_MppSamples.selectedItems()
        sampleselected = []
        for i in range(len(items)):
            sampleselected.append(str(self.ui.listWidget_MppSamples.selectedItems()[i].text()))
        
        listofanswermpp=[]
        listoflinestylempp=[]
        listofcolorstylempp=[]
        listoflinewidthstylempp=[]
        
        for item in sampleselected:
            listofanswermpp.append(DATAMPP[item]["MPPlinestyle"][0])
        
        for item in sampleselected:
            listoflinestylempp.append(DATAMPP[item]["MPPlinestyle"][1])
            listofcolorstylempp.append(DATAMPP[item]["MPPlinestyle"][2])
            listoflinewidthstylempp.append(str(DATAMPP[item]["MPPlinestyle"][3]))
        
        self.clearLayout(self.ui.gridLayout_3)
        self.ui.scrollArea_MPPTlegend = QtWidgets.QScrollArea(self.ui.EditLegendMPP)
        self.ui.scrollArea_MPPTlegend.setWidgetResizable(True)
        self.ui.scrollArea_MPPTlegend.setObjectName("scrollArea_MPPTlegend")
        self.ui.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.ui.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 500, 400))
        self.ui.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ui.scrollAreaWidgetContents.sizePolicy().hasHeightForWidth())
        self.ui.scrollAreaWidgetContents.setSizePolicy(sizePolicy)
        self.ui.verticalLayout_3 = QtWidgets.QVBoxLayout(self.ui.scrollAreaWidgetContents)
        self.ui.verticalLayout_3.setObjectName("verticalLayout_3")
        
        item1=0
        for itemm in sampleselected:
            # print(item1)
            self.frame = QtWidgets.QFrame(self.ui.scrollAreaWidgetContents)
            self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
            self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
            self.frame.setObjectName("frame")
            self.horizontalLayout = QtWidgets.QHBoxLayout(self.frame)
            self.horizontalLayout.setObjectName("horizontalLayout")
            
            label = QtWidgets.QLabel(self.frame)
            label.setText(DATAMPP[itemm]['SampleName'])
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(label.sizePolicy().hasHeightForWidth())
            label.setSizePolicy(sizePolicy)
            self.horizontalLayout.addWidget(label)
            
            listofanswermpp[item1]=QtWidgets.QLineEdit(self.frame)
            listofanswermpp[item1].setText(DATAMPP[itemm]["MPPlinestyle"][0])
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(listofanswermpp[item1].sizePolicy().hasHeightForWidth())
            listofanswermpp[item1].setSizePolicy(sizePolicy)
            self.horizontalLayout.addWidget(listofanswermpp[item1])
            listofanswermpp[item1].textChanged.connect(self.UpdateMPPLegMod)
            
            listoflinestylempp[item1] = QtWidgets.QComboBox(self.frame)
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(listoflinestylempp[item1].sizePolicy().hasHeightForWidth())
            listoflinestylempp[item1].setObjectName("comboBox_matname"+str(item1))
            listoflinestylempp[item1].addItems(["-","--","-.",":"])
            listoflinestylempp[item1].setCurrentText(DATAMPP[itemm]["MPPlinestyle"][1])
            self.horizontalLayout.addWidget(listoflinestylempp[item1])
            listoflinestylempp[item1].currentTextChanged.connect(self.UpdateMPPLegMod)
            
            listofcolorstylempp[item1] = QtWidgets.QPushButton('Select Color', self.frame)
            listofcolorstylempp[item1].setObjectName("button_"+str(item1))
            listofcolorstylempp[item1].setStyleSheet("color:"+str(DATAMPP[itemm]["MPPlinestyle"][2])+";")
            self.horizontalLayout.addWidget(listofcolorstylempp[item1])
            listofcolorstylempp[item1].clicked.connect(partial(self.getColormpp,itemm))
            
            listoflinewidthstylempp[item1] = QtWidgets.QSpinBox(self.frame)
            sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
            sizePolicy.setHorizontalStretch(0)
            sizePolicy.setVerticalStretch(0)
            sizePolicy.setHeightForWidth(listoflinewidthstylempp[item1].sizePolicy().hasHeightForWidth())
            listoflinewidthstylempp[item1].setSizePolicy(sizePolicy)
            listoflinewidthstylempp[item1].setMaximum(9999999)
            listoflinewidthstylempp[item1].setObjectName("spinBox_"+str(item1))
            listoflinewidthstylempp[item1].setValue(DATAMPP[itemm]["MPPlinestyle"][3])
            self.horizontalLayout.addWidget(listoflinewidthstylempp[item1])
            listoflinewidthstylempp[item1].valueChanged.connect(self.UpdateMPPLegMod)
            item1+=1
            self.ui.verticalLayout_3.addWidget(self.frame)
                
        self.ui.scrollArea_MPPTlegend.setWidget(self.ui.scrollAreaWidgetContents)
        self.ui.gridLayout_3.addWidget(self.ui.scrollArea_MPPTlegend, 0, 0, 1, 1)
        
    def getColormpp(self,rowitem):
        global DATAMPP, listofcolorstyle
        color = QColorDialog.getColor()
        DATAMPP[rowitem]["MPPlinestyle"][2]=color.name()
        
        self.UpdateMPPLegMod()
        self.populateMPP()
        
    def UpdateMPPLegMod(self):
        global DATAMPP
        global listofanswermpp
        global listoflinestylempp
        global listofcolorstylempp,listoflinewidthstylempp
        # print('updatelegmod')
        items = self.ui.listWidget_MppSamples.selectedItems()
        sampleselected = []
        for i in range(len(items)):
            sampleselected.append(str(self.ui.listWidget_MppSamples.selectedItems()[i].text()))
            
        leglist=[]
        for e in listofanswermpp:
            if type(e)!=str:
                leglist.append(e.text())
            else:
                leglist.append(e)
        for item in range(len(sampleselected)):
            DATAMPP[sampleselected[item]]["MPPlinestyle"][0]=leglist[item]
        
        # for item in range(len(sampleselected)):
        #     DATA[sampleselected[item]]["IVlinestyle"][2]=listofcolorstyle[item]
            
        leglist=[]
        for e in listoflinestylempp:
            if type(e)!=str:
                leglist.append(e.currentText())
            else:
                leglist.append(e)
        for item in range(len(sampleselected)):
            DATAMPP[sampleselected[item]]["MPPlinestyle"][1]=leglist[item]
            
        leglist=[]
        for e in listoflinewidthstylempp:
            if type(e)!=str:
                leglist.append(e.value())
            else:
                leglist.append(e) 
        for item in range(len(sampleselected)):
            DATAMPP[sampleselected[item]]["MPPlinestyle"][3]=int(leglist[item])
                
        # self.populateJV()
        self.PlotMPP()
#%%#############
    def PlotMPP1(self):
        if self.ui.checkBox_MppLegend.isChecked():
            self.populateMPP()
            self.PlotMPP()
        else:
            self.PlotMPP()
    def PlotMPP(self):
        global DATAMPP, MPPlegendMod, MPPlinestyle
        global DATAmppforexport
        DATAmppforexport=[]
        items = self.ui.listWidget_MppSamples.selectedItems()
        selectedmpptraces = []
        for i in range(len(items)):
            selectedmpptraces.append(str(self.ui.listWidget_MppSamples.selectedItems()[i].text()))
        
        self.MPPgraph.clear()
        if selectedmpptraces!=[]:
            alltimes=[]
            for item in selectedmpptraces:
                alltimes.append(DATAMPP[item]["MeasDayTime2"])
            mintime=min(alltimes)
            # print(alltimes)
            # print(mintime)
            TimeUnit=1
            if self.ui.comboBox_MPPT_TimeUnit.currentText()=='Seconds':
                TimeUnit=1
                self.MPPgraph.set_xlabel('Time (s)')
                startcolx=["Time","s",""]
            elif self.ui.comboBox_MPPT_TimeUnit.currentText()=='Minutes':
                TimeUnit=60
                self.MPPgraph.set_xlabel('Time (min)')
                startcolx=["Time","min",""]
            elif self.ui.comboBox_MPPT_TimeUnit.currentText()=='Hours':
                TimeUnit=60*60
                self.MPPgraph.set_xlabel('Time (hrs)')
                startcolx=["Time","hrs",""]
            elif self.ui.comboBox_MPPT_TimeUnit.currentText()=='Days':
                TimeUnit=60*60*24
                self.MPPgraph.set_xlabel('Time (days)')
                startcolx=["Time","days",""]
                
            for item in selectedmpptraces:
                
                if self.ui.comboBox_MPPT_TrueOrRelative.currentText()=='RelativeTime, each':
                    x = DATAMPP[item]["MppData"][2]
                    x = [time/TimeUnit for time in x]
                elif self.ui.comboBox_MPPT_TrueOrRelative.currentText()=='RelativeTime, all':
                    x=[(DATAMPP[item]["MeasDayTime2"]-mintime).total_seconds() +time for time in DATAMPP[item]["MppData"][2]]
                    x = [time/TimeUnit for time in x]
                elif self.ui.comboBox_MPPT_TrueOrRelative.currentText()=='True Date/Time':
                    x = [DATAMPP[item]["MeasDayTime2"]+datetime.timedelta(seconds=time) for time in DATAMPP[item]["MppData"][2]]
                    self.MPPgraph.set_xlabel('Time')
                    startcolx=["Time","-",""]
                    
                    
                y = DATAMPP[item]["MppData"][3]
                
                colx=startcolx+x
                coly=["Power","mW/cm2",DATAMPP[item]["SampleName"]]+y
                DATAmppforexport.append(colx)
                DATAmppforexport.append(coly)
                
                if self.ui.checkBox_MppLegend.isChecked():
                    self.MPPgraph.plot(x,y,label=DATAMPP[item]["MPPlinestyle"][0],linestyle=DATAMPP[item]["MPPlinestyle"][1],color=DATAMPP[item]["MPPlinestyle"][2],linewidth=DATAMPP[item]["MPPlinestyle"][3])
                else:
                    self.MPPgraph.plot(x,y,linestyle=DATAMPP[item]["MPPlinestyle"][1],color=DATAMPP[item]["MPPlinestyle"][2],linewidth=DATAMPP[item]["MPPlinestyle"][3])
                    
            self.MPPgraph.set_ylabel('Power (mW/cm'+'\xb2'+')')
            # self.MPPgraph.set_xlabel('Time (s)')
            
            for item in ([self.MPPgraph.title, self.MPPgraph.xaxis.label, self.MPPgraph.yaxis.label] +
                                 self.MPPgraph.get_xticklabels() + self.MPPgraph.get_yticklabels()):
                item.set_fontsize(self.ui.spinBox_MppFontsize.value())
            
            if self.ui.checkBox_MppLegend.isChecked():
                if self.ui.radioButton_MppTopleft.isChecked():
                    self.leg=self.MPPgraph.legend(loc=2, fontsize = self.ui.spinBox_MppFontsize.value())
                elif self.ui.radioButton_MppTopright.isChecked():
                    self.leg=self.MPPgraph.legend(loc=1, fontsize = self.ui.spinBox_MppFontsize.value())
                elif self.ui.radioButton_MppBottomleft.isChecked():
                    self.leg=self.MPPgraph.legend(loc=3, fontsize = self.ui.spinBox_MppFontsize.value())
                elif self.ui.radioButton_MppBottomright.isChecked():
                    self.leg=self.MPPgraph.legend(loc=4, fontsize = self.ui.spinBox_MppFontsize.value())
                elif self.ui.radioButton_MppOutside.isChecked():
                    self.leg=self.MPPgraph.legend(bbox_to_anchor=(1, 1), loc='upper left', ncol=1, fontsize = self.ui.spinBox_MppFontsize.value())
                elif self.ui.radioButton_MppBest.isChecked():
                    self.leg=self.MPPgraph.legend(loc=0, fontsize = self.ui.spinBox_MppFontsize.value())
            
            DATAmppforexport=map(list, six.moves.zip_longest(*DATAmppforexport, fillvalue=' '))

        self.fig2.canvas.draw_idle()
        self.fig2.canvas.flush_events()        
        
    def GraphMPPsave_as(self):
        global DATAmppforexport
        path = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]

        if self.ui.checkBox_MppLegend.isChecked():
            self.fig2.savefig(path, dpi=300, bbox_extra_artists=(self.leg,))#, transparent=True)
        else:
            self.fig2.savefig(path, dpi=300)#, transparent=True)
            
        DATAmppforexport1=[]            
        for item in DATAmppforexport:
            line=""
            for item1 in item:
                line=line+str(item1)+"\t"
            line=line[:-1]+"\n"
            DATAmppforexport1.append(line)
            
        file = open(str(path[:-4]+"_dat.txt"),'w', encoding='ISO-8859-1')
        file.writelines("%s" % item for item in DATAmppforexport1)
        file.close()

#%%#############
        
    def PlotIV(self):
        global takenforplot
        rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
        sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
        # print(sampleselected)
        takenforplot=sampleselected
        self.PlotIV2(sampleselected)
        
    def PlotIV2(self,sampleselected):
        global DATA, DATAJVforexport, DATAJVtabforexport
        
        DATAJVforexport=[]
        DATAJVtabforexport=[]
        
        self.JVgraph.clear()
        if sampleselected!=[]:
            for item in sampleselected:
                x = DATA[item]["IVData"][0]
                y = DATA[item]["IVData"][1]
                
                colx=["Voltage","mV",""]+x
                coly=["Current density","ma/cm2",DATA[item]["SampleName"]]+y
                DATAJVforexport.append(colx)
                DATAJVforexport.append(coly)
                DATAJVtabforexport.append([DATA[item]["SampleName"],str(DATA[item]["MeasDayTime2"]),'%.f' % float(DATA[item]["Voc"]),'%.2f' % float(DATA[item]["Jsc"]),'%.2f' % float(DATA[item]["FF"]),'%.2f' % float(DATA[item]["Eff"]),'%.2f' % float(DATA[item]["Roc"]),'%.2f' % float(DATA[item]["Rsc"]),'%.2f' % float(DATA[item]["Vstart"]),'%.2f' % float(DATA[item]["Vend"]),'%.2f' % float(DATA[item]["CellSurface"])])
    
                if self.ui.checkBox_JVLegend.isChecked():
                    # print(DATA[item]["IVlinestyle"][0])
                    self.JVgraph.plot(x,y,label=DATA[item]["IVlinestyle"][0],linestyle=DATA[item]["IVlinestyle"][1],color=DATA[item]["IVlinestyle"][2],linewidth=DATA[item]["IVlinestyle"][3])
                else:
                    self.JVgraph.plot(x,y,linestyle=DATA[item]["IVlinestyle"][1],color=DATA[item]["IVlinestyle"][2],linewidth=DATA[item]["IVlinestyle"][3])
            self.JVgraph.set_xlabel('Voltage (V)')#,**csfont)
            self.JVgraph.set_ylabel('Current density (mA/cm'+'\xb2'+')')#,**csfont)
            self.JVgraph.axhline(y=0, color='k')
            self.JVgraph.axvline(x=0, color='k')
            for item in ([self.JVgraph.title, self.JVgraph.xaxis.label, self.JVgraph.yaxis.label] +
                                 self.JVgraph.get_xticklabels() + self.JVgraph.get_yticklabels()):
                item.set_fontsize(self.ui.spinBox_JVfontsize.value())
            
            if self.ui.checkBox_MinorTicksJV.isChecked():
                self.JVgraph.xaxis.set_minor_locator(AutoMinorLocator())
                self.JVgraph.yaxis.set_minor_locator(AutoMinorLocator())
            
            DATAJVforexport=map(list, six.moves.zip_longest(*DATAJVforexport, fillvalue=' '))
            DATAJVtabforexport.insert(0,[" ","DateTime","Voc", "Jsc", "FF","Eff","Roc","Rsc","Vstart","Vend","Cellsurface"])
            
            if self.ui.checkBox_JVLegend.isChecked():
                if self.ui.radioButton_JVtopleft.isChecked():
                    self.leg=self.JVgraph.legend(loc=2, fontsize = self.ui.spinBox_JVfontsize.value())
                elif self.ui.radioButton_JVtopright.isChecked():
                    self.leg=self.JVgraph.legend(loc=1, fontsize = self.ui.spinBox_JVfontsize.value())
                elif self.ui.radioButton_JVbottomleft.isChecked():
                    self.leg=self.JVgraph.legend(loc=3, fontsize = self.ui.spinBox_JVfontsize.value())
                elif self.ui.radioButton_JVbottomright.isChecked():
                    self.leg=self.JVgraph.legend(loc=4, fontsize = self.ui.spinBox_JVfontsize.value())
                elif self.ui.radioButton_JVoutside.isChecked():
                    self.leg=self.JVgraph.legend(bbox_to_anchor=(1, 1), loc='upper left', ncol=1, fontsize = self.ui.spinBox_JVfontsize.value())
                elif self.ui.radioButton_JVBest.isChecked():
                    self.leg=self.JVgraph.legend(loc=0, fontsize = self.ui.spinBox_JVfontsize.value())

        self.fig1.canvas.draw_idle()
        self.fig1.canvas.flush_events()

    def GraphJVsave_as(self):
        global DATA, DATAJVforexport, DATAJVtabforexport
        
        f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]

        if self.ui.checkBox_JVLegend.isChecked():
            self.fig1.savefig(f, dpi=300, bbox_extra_artists=(self.leg,))#, transparent=True)
        else:
            self.fig1.savefig(f, dpi=300)#, transparent=True)
        
        DATAJVforexport1=[]
        for item in DATAJVforexport:
            line=""
            for item1 in item:
                line=line+str(item1)+"\t"
            line=line[:-1]+"\n"
            DATAJVforexport1.append(line)
            
        file = open(str(f[:-4]+"_dat.txt"),'w', encoding='ISO-8859-1')
        file.writelines("%s" % item for item in DATAJVforexport1)
        file.close()   
        
        DATAJVforexport1=[]
        for item in DATAJVtabforexport:
            line=""
            for item1 in item:
                line=line+str(item1)+"\t"
            line=line[:-1]+"\n"
            DATAJVforexport1.append(line)
        file = open(str(f[:-4]+"_tab.txt"),'w', encoding='ISO-8859-1')
        file.writelines("%s" % item for item in DATAJVforexport1)
        file.close()
        
        
#%%#############
        
    def UpdateHistGraph(self):
        global DATA
        global DATAHistforexport
        
        DATAHistforexport=[]
        numbbins=int(self.ui.spinBox_HistoBins.value())
        DATAx=copy.deepcopy(DATA)
        
        samplesgroups = self.ui.listWidget_HistoGroups.selectedItems()
        samplesgroups=[item.text() for item in samplesgroups]
        # print(samplesgroups)
        groupnames=[]
        #sorting data
        if samplesgroups==[]:
            self.Histgraph.clear()
        else:
            grouplistdict=[]
            if self.ui.comboBox_HistoScanDirect.currentText()=="Allmeas":    #select all data points
                for item in range(len(samplesgroups)):
                    listdata=[]
                    for item1 in DATAx.keys():
                        if DATAx[item1]["Group"]+' * '+DATAx[item1]["Group2"]==samplesgroups[item] and DATAx[item1]["Illumination"]=='Light':
                            listdata.append(DATAx[item1][self.ui.comboBox_HistoParam.currentText()])
                    groupnames.append(samplesgroups[item])
                    grouplistdict.append(listdata)
            elif self.ui.comboBox_HistoScanDirect.currentText()=="OnlyRev":
                for item in range(len(samplesgroups)):
                    listdata=[]
                    for item1 in DATAx.keys():
                        if DATAx[item1]["Group"]+' * '+DATAx[item1]["Group2"]==samplesgroups[item] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["ScanDirection"]=="Reverse":
                            listdata.append(DATAx[item1][self.ui.comboBox_HistoParam.currentText()])
                    groupnames.append(samplesgroups[item])
                    grouplistdict.append(listdata)
                    
            elif self.ui.comboBox_HistoScanDirect.currentText()=="OnlyForw":
                for item in range(len(samplesgroups)):
                    listdata=[]
                    for item1 in DATAx.keys():
                        if DATAx[item1]["Group"]+' * '+DATAx[item1]["Group2"]==samplesgroups[item] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["ScanDirection"]=="Forward":
                            listdata.append(DATAx[item1][self.ui.comboBox_HistoParam.currentText()])
                    groupnames.append(samplesgroups[item])
                    grouplistdict.append(listdata)
            elif self.ui.comboBox_HistoScanDirect.currentText()=="Bestof/pix":  
                for item in range(len(samplesgroups)):
                    listofthegroup=[]
                    for item1 in DATAx.keys():
                        if DATAx[item1]["Group"]+' * '+DATAx[item1]["Group2"]==samplesgroups[item] and DATAx[item1]["Illumination"]=='Light':
                            listofthegroup.append(DATAx[item1])
                    if len(listofthegroup)!=0:
                        
                        grouper = itemgetter("DepID", "Cellletter")
                        result = []
                        keylist=[]
                        for key, grp in groupby(sorted(listofthegroup, key = grouper), grouper):
                            result.append(list(grp))
                            keylist.append(key)
                        # print(result)
                        # print(keylist)
                        listdata=[]
                        for item1 in range(len(keylist)):
                            listdata1=[]
                            for item2 in range(len(result[item1])):
                                listdata1.append(result[item1][item2][self.ui.comboBox_HistoParam.currentText()])
                            listdata.append(max(listdata1))
                            
                        groupnames.append(samplesgroups[item])        
                        grouplistdict.append(listdata)
            elif self.ui.comboBox_HistoScanDirect.currentText()=="Bestof/subst":  
                for item in range(len(samplesgroups)):
                    listofthegroup=[]
                    for item1 in DATAx.keys():
                        if DATAx[item1]["Group"]+' * '+DATAx[item1]["Group2"]==samplesgroups[item] and DATAx[item1]["Illumination"]=='Light':
                            listofthegroup.append(DATAx[item1])
                    if len(listofthegroup)!=0:
                        grouper = itemgetter("DepID")
                        result = []
                        keylist=[]
                        for key, grp in groupby(sorted(listofthegroup, key = grouper), grouper):
                            result.append(list(grp))
#                            print(len(result))
                            keylist.append(key)
                        
                        listdata=[]
                        for item1 in range(len(keylist)):
                            listdata1=[]
                            for item2 in range(len(result[item1])):
                                listdata1.append(result[item1][item2][self.ui.comboBox_HistoParam.currentText()])
                            listdata.append(max(listdata1))
                            
                        groupnames.append(samplesgroups[item])        
                        grouplistdict.append(listdata)


            self.Histgraph.clear()
            if self.ui.checkBox_Histxscale.isChecked():
                self.Histgraph.hist(grouplistdict,bins=numbbins,range=[self.ui.spinBox_HistxscaleMin.value(), self.ui.spinBox_HistxscaleMax.value()],histtype= self.ui.comboBox_HistoType.currentText(), density=False, cumulative=False, alpha=0.6, edgecolor='black', linewidth=1.2, label=groupnames)
            else:
                self.Histgraph.hist(grouplistdict,bins=numbbins,histtype= self.ui.comboBox_HistoType.currentText(), density=False, cumulative=False, alpha=0.6, edgecolor='black', linewidth=1.2, label=groupnames)
                
            self.Histgraph.set_xlabel(self.ui.comboBox_HistoParam.currentText())
            self.Histgraph.set_ylabel('counts')
            self.Histgraph.legend()
        
            DATAHistforexport=list(map(list, six.moves.zip_longest(*grouplistdict, fillvalue=' ')))
            DATAHistforexport=[groupnames]+DATAHistforexport

        
        self.fig6.canvas.draw_idle()
        self.fig6.canvas.flush_events()

    def GraphHistsave_as(self):
        global DATA, DATAHistforexport
        
        f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]
        self.fig6.savefig(f, dpi=300)#, transparent=True)
                       
        DATAHistforexport1=[]
        for item in DATAHistforexport:
            line=""
            for item1 in item:
                line=line+str(item1)+"\t"
            line=line[:-1]+"\n"
            DATAHistforexport1.append(line)
            
        file = open(str(f[:-4]+"_dat.txt"),'w', encoding='ISO-8859-1')
        file.writelines("%s" % item for item in DATAHistforexport1)
        file.close()
        
        
    def RecalcJV(self):
        global DATA
    def BumpsRemoval(self):
        global DATA
        
        rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
        sampleselected=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
        newsampleslist=[]
        for item in sampleselected:
            y=DATA[item]['IVData'][1]
            x=DATA[item]['IVData'][0]
            
            for item1 in range(len(y)-2):
                pt0=y[item1]
                pt1=y[item1+1]
                pt2=y[item1+2]
                
                if abs(pt1-pt2)>0.1 and abs(pt1-pt0)>0.1 and (math.copysign(1,pt0-pt1) != math.copysign(1,pt1-pt2)):
                    # print(x[item1+1])
                    DATA[item]['IVData'][1][item1+1]=(DATA[item]['IVData'][1][item1]+DATA[item]['IVData'][1][item1+2])/2
                    params=extract_jv_params(DATA[item]['IVData'])
                    DATA[item]["Voc"]= float(params['Voc'])*1000 #mV
                    DATA[item]["Jsc"]= float(params['Jsc'])#mA/cm2
                    DATA[item]["Isc"]=float(params['Jsc']*DATA[item]["CellSurface"])
                    DATA[item]["FF"]=float(params['FF']) #%
                    DATA[item]["Eff"]=float(params['Pmax'])#%
                    DATA[item]["Pmpp"]=float(DATA[item]["Eff"])*10#W/cm2
                    DATA[item]["VocFF"]=DATA[item]["Voc"]*DATA[item]["FF"]
                    DATA[item]["Roc"]=float(params['Roc'])
                    DATA[item]["Rsc"]=float(params['Rsc'])
                    DATA[item]["RscJsc"]=DATA[item]["Rsc"]*DATA[item]["Jsc"]
                    DATA[item]["Vmpp"]=float(params['Vmpp'])
                    DATA[item]["Jmpp"]=float(params['Jmpp'])
                    newkeyname=DATA[item]["SampleName"]+'_'+str(DATA[item]["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(DATA[item]["Isc"]))+'_'+str(float(DATA[item]["FF"]))
                    DATA[newkeyname]=DATA[item]
                    del(DATA[item])
                    newsampleslist.append(newkeyname)
                else:
                    newsampleslist.append(item)
        
        self.PlotIV2(newsampleslist)
        self.updateTable(DATA)
        
        
#%%#############
    def ANOVAboxplot2(self):
        global DATA, DATAgroupforexport
        
        #anova for subgroups, between group1, inside group1 subgroups, between all group1-group2
        
        f = QFileDialog.getSaveFileName(self, 'Save ANOVA', ".xlsx", "All Files (*)")[0]+'.xlsx'
        # shutil.copy(r"C:\Users\serjw\Documents\nBox\PythonScripts\SERIS-pythonscripts\Python-Solar-Data-Analyser-SERIS\apps\template_Anova_JV.xlsx", f)
        shutil.copy(os.path.join(exedirectory,'template_Anova_JV.xlsx'),f)
        
        
    # def ANOVAboxplot(self):
    #     global DATA, DATAgroupforexport
        
    #     f = QFileDialog.getSaveFileName(self, 'Save ANOVA', ".xlsx", "All Files (*)")[0]+'.xlsx'
    #     # shutil.copy(r"C:\Users\serjw\Documents\nBox\PythonScripts\SERIS-pythonscripts\Python-Solar-Data-Analyser-SERIS\apps\template_Anova_JV.xlsx", f)
    #     shutil.copy(os.path.join(exedirectory,'template_Anova_JV.xlsx'),f)
        
    #     #open existing wb
    #     wb = Opxl.load_workbook(f)
    #     #create the worksheets with param
    #     parameters=['Eff','Voc','Jsc','FF']
    #     target = wb['Eff']
    #     for param in range(len(parameters)-1):
    #         wb.copy_worksheet(target)
    #     DATAx=[DATA[key] for key in DATA.keys()]
    #     namessheet=wb.sheetnames
    #     grouplistdict=[]
    #     for item in range(len(samplesgroups)):
    #         groupdict={}
    #         groupdict["Group"]=samplesgroups[item]
    #         listofthegroup=[]
    #         listofthegroupNames=[]
    #         for item1 in range(len(DATAx)):
    #             if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=="Light":
    #                 listofthegroup.append(DATAx[item1])
    #                 listofthegroupNames.append(DATAx[item1]['DepID']+'_'+DATAx[item1]['Cellletter'])
    #         groupdict["numbCell"]=len(list(set(listofthegroupNames)))
    #         listofthegroupRev=[]
    #         listofthegroupFor=[]
    #         for item1 in range(len(listofthegroup)):
    #             if listofthegroup[item1]["ScanDirection"]=="Reverse":
    #                 listofthegroupRev.append(listofthegroup[item1])
    #             else:
    #                 listofthegroupFor.append(listofthegroup[item1])
               
    #         groupdict["RevVoc"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
    #         groupdict["ForVoc"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
    #         groupdict["RevJsc"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
    #         groupdict["ForJsc"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
    #         groupdict["RevFF"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
    #         groupdict["ForFF"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
    #         groupdict["RevEff"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
    #         groupdict["ForEff"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
            
    #         grouplistdict.append(groupdict)
        
    #     for sheetnb in range(len(namessheet)):
    #         wb_sheet = wb[namessheet[sheetnb]]
    #         wb_sheet.title = parameters[sheetnb]
            
    #         summary=[]
    #         listofaverages=[]
    #         listofaveragesFRtogether=[]
    #         listofallmeas=[]
    #         ttest1=[{},{}]
    #         ttest2=[{},{}]
    #         ttest3=[{},{}]
            
    #         ssbetweengroups=[{'group':[],'dat':[],'counts':[],'mean':[]},{'group':[],'dat':[],'counts':[],'mean':[]}]#0:reverse,1:forward
            
    #         for item in range(len(samplesgroups)):
    #             ncell=1
    #             partlistofaveragesFRtogether=[]
    #             if grouplistdict[item]['For'+parameters[sheetnb]]!=[]:
    #                 dat=grouplistdict[item]['For'+parameters[sheetnb]]
    #                 lengthofgroup=len(dat)
    #                 #group, numbcell, scan direct, numbmeas, average, sum, std dev, std error, variance, sum of squares
    #                 summary.append([grouplistdict[item]["Group"],grouplistdict[item]["numbCell"],"Forward",lengthofgroup,sum(dat,0.0)/lengthofgroup,sum(dat,0.0),np.std(dat),np.std(dat)/math.sqrt(lengthofgroup),np.var(dat),sumofsquaredev(dat)])
    #                 ncell=0
    #                 listofaverages.append(sum(dat,0.0)/lengthofgroup)
    #                 listofallmeas+=dat
    #                 partlistofaveragesFRtogether+=dat
    #                 ssbetweengroups[1]['group'].append(grouplistdict[item]["Group"])
    #                 ssbetweengroups[1]['dat'].append(dat)
    #                 ssbetweengroups[1]['counts'].append(lengthofgroup)
    #                 ssbetweengroups[1]['mean'].append(sum(dat,0.0)/lengthofgroup)
    #                 ttest1[1][grouplistdict[item]["Group"]]=((np.std(dat))**2)/lengthofgroup
    #                 ttest2[1][grouplistdict[item]["Group"]]=sum(dat,0.0)/lengthofgroup
    #                 ttest3[1][grouplistdict[item]["Group"]]=lengthofgroup
                    
    #             if grouplistdict[item]['Rev'+parameters[sheetnb]]!=[]:
    #                 dat=grouplistdict[item]['Rev'+parameters[sheetnb]]
    #                 listofaverages.append(sum(dat,0.0)/lengthofgroup)
    #                 listofallmeas+=dat
    #                 partlistofaveragesFRtogether+=dat
    #                 ssbetweengroups[0]['group'].append(grouplistdict[item]["Group"])
    #                 ssbetweengroups[0]['dat'].append(dat)
    #                 ssbetweengroups[0]['counts'].append(lengthofgroup)
    #                 ssbetweengroups[0]['mean'].append(sum(dat,0.0)/lengthofgroup)
    #                 ttest1[0][grouplistdict[item]["Group"]]=((np.std(dat))**2)/lengthofgroup
    #                 ttest2[0][grouplistdict[item]["Group"]]=sum(dat,0.0)/lengthofgroup
    #                 ttest3[0][grouplistdict[item]["Group"]]=lengthofgroup
    #                 if ncell==0:
    #                     lengthofgroup=len(dat)
    #                     summary.append([grouplistdict[item]["Group"]," ","Reverse",lengthofgroup,sum(dat,0.0)/lengthofgroup,sum(dat,0.0),np.std(dat),np.std(dat)/math.sqrt(lengthofgroup),np.var(dat),sumofsquaredev(dat)])
    #                 else:
    #                     lengthofgroup=len(dat)
    #                     summary.append([grouplistdict[item]["Group"],grouplistdict[item]["numbCell"],"Reverse",lengthofgroup,sum(dat,0.0)/lengthofgroup,sum(dat,0.0),np.std(dat),np.std(dat)/math.sqrt(lengthofgroup),np.var(dat),sumofsquaredev(dat)])
    #             listofaveragesFRtogether.append(sum(partlistofaveragesFRtogether)/len(partlistofaveragesFRtogether))
            
            
    #         for item in range(len(summary)):
    #             for item0 in range(len(summary[item])):
    #                 wb_sheet.cell(item+5,item0+2).value = summary[item][item0]
            
    #         wb_sheet.cell(5,14).value = len(listofallmeas)
    #         wb_sheet.cell(5,15).value = sum(listofaverages)/len(listofaverages)
    #         wb_sheet.cell(5,16).value = sum(listofaveragesFRtogether)/len(listofaveragesFRtogether)
    #         wb_sheet.cell(5,17).value = sum(listofallmeas)/len(listofallmeas)
            
    #         sumofsquares=[]
    #         for item in range(len(ssbetweengroups[0]['group'])):
    #             wb_sheet.cell(item+5,20).value = ssbetweengroups[0]['group'][item]
    #             wb_sheet.cell(item+5,21).value = ssbetweengroups[0]['counts'][item]
    #             wb_sheet.cell(item+5,22).value = ssbetweengroups[0]['mean'][item]
    #             wb_sheet.cell(item+5,23).value = sumofsquaredev(ssbetweengroups[0]['dat'][item])
    #             ss=(ssbetweengroups[0]['mean'][item]-sum(ssbetweengroups[0]['mean'])/len(ssbetweengroups[0]['mean']))**2
    #             sumofsquares.append(ss)
    #             wb_sheet.cell(item+5,24).value = ss
    #         wb_sheet.cell(3,22).value = sum(ssbetweengroups[0]['mean'])/len(ssbetweengroups[0]['mean'])
    #         wb_sheet.cell(3,24).value = sum(sumofsquares)
    #         wb_sheet.cell(7,27).value =sumofsquaredev(listofallmeas)
    #         wb_sheet.cell(7,28).value = len(listofallmeas)-1
            
    #         comb = list(combinations(ssbetweengroups[0]['group'], 2))
    #         for item in range(len(comb)):
    #             wb_sheet.cell(5+item,34).value=comb[item][0]+' & '+comb[item][1]
    #             wb_sheet.cell(5+item,35).value=ttest2[0][comb[item][0]]-ttest2[0][comb[item][1]]
    #             wb_sheet.cell(5+item,36).value=math.sqrt(ttest1[0][comb[item][0]]+ttest1[0][comb[item][1]])
    #             wb_sheet.cell(5+item,38).value=ttest3[0][comb[item][0]]+ttest3[0][comb[item][1]]-2
            
    #         sumofsquares=[]
    #         for item in range(len(ssbetweengroups[1]['group'])):
    #             wb_sheet.cell(item+5,43).value = ssbetweengroups[1]['group'][item]
    #             wb_sheet.cell(item+5,44).value = ssbetweengroups[1]['counts'][item]
    #             wb_sheet.cell(item+5,45).value = ssbetweengroups[1]['mean'][item]
    #             wb_sheet.cell(item+5,46).value = sumofsquaredev(ssbetweengroups[1]['dat'][item])
    #             ss=(ssbetweengroups[1]['mean'][item]-sum(ssbetweengroups[1]['mean'])/len(ssbetweengroups[1]['mean']))**2
    #             sumofsquares.append(ss)
    #             wb_sheet.cell(item+5,47).value = ss
                
    #         wb_sheet.cell(3,45).value = sum(ssbetweengroups[1]['mean'])/len(ssbetweengroups[1]['mean'])
    #         wb_sheet.cell(3,47).value = sum(sumofsquares)
            
    #         wb_sheet.cell(7,50).value =sumofsquaredev(listofallmeas)
    #         wb_sheet.cell(7,51).value = len(listofallmeas)-1
            
    #         comb = list(combinations(ssbetweengroups[1]['group'], 2))
    #         for item in range(len(comb)):
    #             wb_sheet.cell(5+item,57).value=comb[item][0]+' & '+comb[item][1]
    #             wb_sheet.cell(5+item,58).value=ttest2[1][comb[item][0]]-ttest2[1][comb[item][1]]
    #             wb_sheet.cell(5+item,59).value=math.sqrt(ttest1[1][comb[item][0]]+ttest1[1][comb[item][1]])
    #             wb_sheet.cell(5+item,61).value=ttest3[1][comb[item][0]]+ttest3[1][comb[item][1]]-2
                
            
    #     wb.save(f)
        
        
        
    def UpdateBoxGraph2(self):
        global DATA
        global DATAgroupforexport,groupSubgroups
        
        DATAgroupforexport=[]
        fontsizegroup=self.ui.spinBox_BoxPlotFontsize_2.value()
        DATAx=copy.deepcopy(DATA)
        samplesgroups=[]
        
        samplesgroups2 = self.ui.listWidget_BoxPlotGroup2_2.selectedItems()
        samplesgroups2=[item.text() for item in samplesgroups2]
        
        markersize=self.ui.spinBox_markerSize_2.value()
        
        print('\n')
        if self.ui.checkBox_onlygroup1.isChecked():
            for i in DATAx.keys():
                DATAx[i]['Group2']=''
            samplesgroups2=[i.split(' * ')[0]+' * ' for i in samplesgroups2]
        
        print(len(samplesgroups2))
        if len(samplesgroups2)>0:
            for subsample in samplesgroups2:
                newsamplesgroups=subsample.split(' * ')[0]
                if newsamplesgroups not in samplesgroups:
                    samplesgroups.append(newsamplesgroups)
            
            groupSubgroups={}
            for sample in samplesgroups:
                print(sample)
                groupSubgroups[sample]={}
                for subsample in samplesgroups2:
                    if sample in subsample:
                        groupSubgroups[sample][subsample.split(' * ')[1]]={}
                        groupdict=groupSubgroups[sample][subsample.split(' * ')[1]]
                        groupdict["Group"]=subsample
                        listofthegroup=[]
                        listofthegroup2=[]
                        for item1 in DATAx.keys():
                            if not self.ui.checkBox_BoxPlotAftermpp_2.isChecked():
                                if DATAx[item1]["Group"]==sample and DATAx[item1]["Group2"]==subsample.split(' * ')[1] and DATAx[item1]["Illumination"]=='Light':
                                    listofthegroup.append(DATAx[item1])
                            else:
                                if DATAx[item1]["Group"]==sample and DATAx[item1]["Group2"]==subsample.split(' * ')[1] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["aftermpp"]==0:
                                    listofthegroup.append(DATAx[item1])
                                elif DATAx[item1]["Group"]==sample and DATAx[item1]["Group2"]==subsample.split(' * ')[1] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["aftermpp"]==1:
                                    listofthegroup2.append(DATAx[item1])
                        print(len(listofthegroup))
                        if len(listofthegroup)!=0:
                            if not self.ui.checkBox_BoxPlotRevForw_2.isChecked():
                                listofthegroupRev=[]
                                listofthegroupFor=[]
                                for item1 in range(len(listofthegroup)):
                                    if listofthegroup[item1]["ScanDirection"]=="Reverse":
                                        listofthegroupRev.append(listofthegroup[item1])
                                    else:
                                        listofthegroupFor.append(listofthegroup[item1])
                            else:
                                grouper = itemgetter("DepID", "Cellletter",'ScanDirection')
                                result = []
                                for key, grp in groupby(sorted(listofthegroup, key = grouper), grouper):
                                    result.append(list(grp))
                                result1=[]
                                for item in result:
                                    result1.append(sorted(item,key=itemgetter('Eff'),reverse=True)[0])
                                grouper = itemgetter('ScanDirection')
                                result2 = []
                                for key, grp in groupby(sorted(result1, key = grouper), grouper):
                                    result2.append(list(grp))
                                
                                listofthegroupRev=[]
                                listofthegroupFor=[]
                                
                                if result2[0][0]['ScanDirection']=='Forward':
                                    listofthegroupFor=result2[0]
                                    if len(result2)>1:
                                        listofthegroupRev=result2[1]
                                else:
                                    listofthegroupRev=result2[0]
                                    if len(result2)>1:
                                        listofthegroupFor=result2[1]
                
                            groupdict["RevVoc"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
                            groupdict["ForVoc"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
                            groupdict["RevJsc"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
                            groupdict["ForJsc"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
                            groupdict["RevFF"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
                            groupdict["ForFF"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
                            groupdict["RevEff"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
                            groupdict["ForEff"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
                            groupdict["RevRoc"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
                            groupdict["ForRoc"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
                            groupdict["RevRsc"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
                            groupdict["ForRsc"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
                            groupdict["RevVmpp"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
                            groupdict["ForVmpp"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
                            groupdict["RevJmpp"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
                            groupdict["ForJmpp"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
                            
                        if len(listofthegroup2)!=0:
                            if not self.ui.checkBox_BoxPlotRevForw_2.isChecked():
                                listofthegroupRev=[]
                                listofthegroupFor=[]
                                for item1 in range(len(listofthegroup2)):
                                    if listofthegroup2[item1]["ScanDirection"]=="Reverse":
                                        listofthegroupRev.append(listofthegroup2[item1])
                                    else:
                                        listofthegroupFor.append(listofthegroup2[item1])
                            else:
                                grouper = itemgetter("DepID", "Cellletter",'ScanDirection')
                                result = []
                                for key, grp in groupby(sorted(listofthegroup2, key = grouper), grouper):
                                    result.append(list(grp))
                                
                                result1=[]
                                
                                for item in result:
                                    result1.append(sorted(item,key=itemgetter('Eff'),reverse=True)[0])
                                
                                grouper = itemgetter('ScanDirection')
                                result2 = []
                                for key, grp in groupby(sorted(result1, key = grouper), grouper):
                                    result2.append(list(grp))
                                
                                listofthegroupRev=[]
                                listofthegroupFor=[]
                                
                                if result2[0][0]['ScanDirection']=='Forward':
                                    listofthegroupFor=result2[0]
                                    if len(result2)>1:
                                        listofthegroupRev=result2[1]
                                else:
                                    listofthegroupRev=result2[0]
                                    if len(result2)>1:
                                        listofthegroupFor=result2[1]
                            groupdict["RevVocAMPP"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
                            groupdict["ForVocAMPP"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
                            groupdict["RevJscAMPP"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
                            groupdict["ForJscAMPP"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
                            groupdict["RevFFAMPP"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
                            groupdict["ForFFAMPP"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
                            groupdict["RevEffAMPP"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
                            groupdict["ForEffAMPP"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
                            groupdict["RevRocAMPP"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
                            groupdict["ForRocAMPP"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
                            groupdict["RevRscAMPP"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
                            groupdict["ForRscAMPP"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
                            groupdict["RevVmppAMPP"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
                            groupdict["ForVmppAMPP"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
                            groupdict["RevJmppAMPP"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
                            groupdict["ForJmppAMPP"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
                        else:
                            groupdict["RevVocAMPP"]=[]
                            groupdict["ForVocAMPP"]=[]
                            groupdict["RevJscAMPP"]=[]
                            groupdict["ForJscAMPP"]=[]
                            groupdict["RevFFAMPP"]=[]
                            groupdict["ForFFAMPP"]=[]
                            groupdict["RevEffAMPP"]=[]
                            groupdict["ForEffAMPP"]=[]
                            groupdict["RevRocAMPP"]=[]
                            groupdict["ForRocAMPP"]=[]
                            groupdict["RevRscAMPP"]=[]
                            groupdict["ForRscAMPP"]=[]
                            groupdict["RevVmppAMPP"]=[]
                            groupdict["ForVmppAMPP"]=[]
                            groupdict["RevJmppAMPP"]=[]
                            groupdict["ForJmppAMPP"]=[]
            
            self.fig31.clear()
            ncols=len(groupSubgroups.keys())
            group1names=samplesgroups#list(groupSubgroups.keys())
            markersize=self.ui.spinBox_markerSize_2.value()
            groupchoice=self.ui.comboBox_BoxPlotParam_2.currentText()
            if ncols>0:
                ax = self.fig31.add_subplot(1,ncols,1)
                if self.ui.checkBox_BoxPlotBoxPlot_2.isChecked():
                    if not self.ui.checkBox_BoxPlotAftermpp_2.isChecked():
                        ax.boxplot([groupSubgroups[group1names[0]][item]["Rev"+groupchoice]+groupSubgroups[group1names[0]][item]["For"+groupchoice] for item in groupSubgroups[group1names[0]].keys()], showmeans=False, meanline=False,notch=False,boxprops = dict(linestyle='-', linewidth=1, color='black'),showfliers=False)
                    else:
                        ax.boxplot([groupSubgroups[group1names[0]][item]["Rev"+groupchoice]+groupSubgroups[group1names[0]][item]["For"+groupchoice]+groupSubgroups[group1names[0]][item]["Rev"+groupchoice+"AMPP"]+groupSubgroups[group1names[0]][item]["For"+groupchoice+"AMPP"] for item in groupSubgroups[group1names[0]].keys()], showmeans=False, meanline=False,notch=False,boxprops = dict(linestyle='-', linewidth=1, color='black'),showfliers=False)
                        
                subkeys=list(groupSubgroups[group1names[0]].keys())
                for item in range(len(subkeys)):
                    if not self.ui.checkBox_BoxPlotAftermpp_2.isChecked():
                        y=groupSubgroups[group1names[0]][subkeys[item]]["Rev"+groupchoice]
                        x=np.random.normal(item+1,0.04,size=len(y))
                        ax.scatter(x,y,s=markersize,color='red', alpha=0.5)
                        DATAgroupforexport.append([group1names[0]+'-'+subkeys[item],"Rev"+groupchoice]+y)
                        y=groupSubgroups[group1names[0]][subkeys[item]]["For"+groupchoice]
                        x=np.random.normal(item+1,0.04,size=len(y))
                        ax.scatter(x,y,s=markersize,color='blue', alpha=0.5)
                        DATAgroupforexport.append([group1names[0]+'-'+subkeys[item],"For"+groupchoice]+y)
                    else:
                        y=groupSubgroups[group1names[0]][subkeys[item]]["Rev"+groupchoice]
                        x=np.random.normal(item+0.9,0.04,size=len(y))
                        ax.scatter(x,y,s=markersize,color='red', alpha=0.5)
                        DATAgroupforexport.append([group1names[0]+'-'+subkeys[item],"Rev"+groupchoice]+y)
                        y=groupSubgroups[group1names[0]][subkeys[item]]["For"+groupchoice]
                        x=np.random.normal(item+0.9,0.04,size=len(y))
                        ax.scatter(x,y,s=markersize,color='blue', alpha=0.5)
                        DATAgroupforexport.append([group1names[0]+'-'+subkeys[item],"For"+groupchoice]+y)
                        
                        y=groupSubgroups[group1names[0]][subkeys[item]]["Rev"+groupchoice+"AMPP"]
                        x=np.random.normal(item+1.1,0.04,size=len(y))
                        ax.scatter(x,y,s=markersize,color='orange', alpha=0.5)
                        DATAgroupforexport.append([group1names[0]+'-'+subkeys[item],"Rev"+groupchoice+"AMPP"]+y)
                        y=groupSubgroups[group1names[0]][subkeys[item]]["For"+groupchoice+"AMPP"]
                        x=np.random.normal(item+1.1,0.04,size=len(y))
                        ax.scatter(x,y,s=markersize,color='lightblue', alpha=0.5)
                        DATAgroupforexport.append([group1names[0]+'-'+subkeys[item],"For"+groupchoice+"AMPP"]+y)
                ax.set_xlabel(group1names[0])
                ax.set_ylabel(groupchoice)
                ax.set_xticklabels(groupSubgroups[group1names[0]].keys())
                if self.ui.checkBox_minorticksBoxPlot.isChecked():
                    ax.yaxis.set_minor_locator(AutoMinorLocator())
                if not self.ui.checkBox_BoxPlotBoxPlot_2.isChecked():
                    namelist=list(groupSubgroups[group1names[0]].keys())
                    if namelist !=[]:
                        span=range(1,len(namelist)+1)
                        ax.set_xticks(span)
                        ax.set_xticklabels(namelist)
                        ax.set_xlim([0.5,span[-1]+0.5])
                for item in ([ax.title, ax.xaxis.label, ax.yaxis.label] +
                              ax.get_xticklabels() + ax.get_yticklabels()):
                    item.set_fontsize(fontsizegroup)
                for tick in ax.get_xticklabels():
                    tick.set_rotation(self.ui.spinBox_BoxPlotRotation_2.value())
            if ncols>1:
                for colnumb in range(1,ncols):
                    ax2 = self.fig31.add_subplot(1,ncols,colnumb+1,sharey=ax)
                    subkeys=list(groupSubgroups[group1names[colnumb]].keys())
                    for item in range(len(subkeys)):
                        if not self.ui.checkBox_BoxPlotAftermpp_2.isChecked():
                            y=groupSubgroups[group1names[colnumb]][subkeys[item]]["Rev"+groupchoice]
                            x=np.random.normal(item+1,0.04,size=len(y))
                            ax2.scatter(x,y,s=markersize,color='red', alpha=0.5)
                            DATAgroupforexport.append([group1names[colnumb]+'-'+subkeys[item],"Rev"+groupchoice]+y)
                            y=groupSubgroups[group1names[colnumb]][subkeys[item]]["For"+groupchoice]
                            x=np.random.normal(item+1,0.04,size=len(y))
                            ax2.scatter(x,y,s=markersize,color='blue', alpha=0.5)
                            DATAgroupforexport.append([group1names[colnumb]+'-'+subkeys[item],"For"+groupchoice]+y)
                        else:
                            y=groupSubgroups[group1names[colnumb]][subkeys[item]]["Rev"+groupchoice]
                            x=np.random.normal(item+0.9,0.04,size=len(y))
                            ax2.scatter(x,y,s=markersize,color='red', alpha=0.5)
                            DATAgroupforexport.append([group1names[colnumb]+'-'+subkeys[item],"Rev"+groupchoice]+y)
                            y=groupSubgroups[group1names[colnumb]][subkeys[item]]["For"+groupchoice]
                            x=np.random.normal(item+0.9,0.04,size=len(y))
                            ax2.scatter(x,y,s=markersize,color='blue', alpha=0.5)
                            DATAgroupforexport.append([group1names[colnumb]+'-'+subkeys[item],"For"+groupchoice]+y)
                            
                            y=groupSubgroups[group1names[colnumb]][subkeys[item]]["Rev"+groupchoice+"AMPP"]
                            x=np.random.normal(item+1.1,0.04,size=len(y))
                            ax2.scatter(x,y,s=markersize,color='orange', alpha=0.5)
                            DATAgroupforexport.append([group1names[colnumb]+'-'+subkeys[item],"Rev"+groupchoice+"AMPP"]+y)
                            y=groupSubgroups[group1names[colnumb]][subkeys[item]]["For"+groupchoice+"AMPP"]
                            x=np.random.normal(item+1.1,0.04,size=len(y))
                            ax2.scatter(x,y,s=markersize,color='lightblue', alpha=0.5)
                            DATAgroupforexport.append([group1names[colnumb]+'-'+subkeys[item],"For"+groupchoice+"AMPP"]+y)

                    if self.ui.checkBox_BoxPlotBoxPlot_2.isChecked():
                        if not self.ui.checkBox_BoxPlotAftermpp_2.isChecked():
                            ax2.boxplot([groupSubgroups[group1names[colnumb]][item]["Rev"+groupchoice]+groupSubgroups[group1names[colnumb]][item]["For"+groupchoice] for item in groupSubgroups[group1names[colnumb]].keys()], showmeans=False, meanline=False,notch=False,boxprops = dict(linestyle='-', linewidth=1, color='black'),showfliers=False)
                        else:
                            ax2.boxplot([groupSubgroups[group1names[colnumb]][item]["Rev"+groupchoice]+groupSubgroups[group1names[colnumb]][item]["For"+groupchoice]+groupSubgroups[group1names[colnumb]][item]["Rev"+groupchoice+"AMPP"]+groupSubgroups[group1names[colnumb]][item]["For"+groupchoice+"AMPP"] for item in groupSubgroups[group1names[colnumb]].keys()], showmeans=False, meanline=False,notch=False,boxprops = dict(linestyle='-', linewidth=1, color='black'),showfliers=False)
                    
                    ax2.yaxis.set_visible(False)
                    ax2.set_xlabel(group1names[colnumb])
                    ax2.set_xticklabels(groupSubgroups[group1names[colnumb]].keys())
                    if self.ui.checkBox_minorticksBoxPlot.isChecked():
                        ax2.yaxis.set_minor_locator(AutoMinorLocator())
                    if not self.ui.checkBox_BoxPlotBoxPlot_2.isChecked():
                        namelist=list(groupSubgroups[group1names[colnumb]].keys())
                        if namelist !=[]:
                            span=range(1,len(namelist)+1)
                            ax2.set_xticks(span)
                            ax2.set_xticklabels(namelist)
                            ax2.set_xlim([0.5,span[-1]+0.5])
                    for item in ([ax2.title, ax2.xaxis.label, ax2.yaxis.label] +
                                  ax2.get_xticklabels() + ax2.get_yticklabels()):
                        item.set_fontsize(fontsizegroup)
                    for tick in ax2.get_xticklabels():
                        tick.set_rotation(self.ui.spinBox_BoxPlotRotation_2.value())
        
        DATAgroupforexport=list(map(list, six.moves.zip_longest(*DATAgroupforexport, fillvalue=' ')))
        
        self.fig31.canvas.draw_idle()

        
        
#     def UpdateBoxGraph(self):
#         global DATA
#         global DATAgroupforexport
        
#         DATAgroupforexport=[]
#         fontsizegroup=self.ui.spinBox_BoxPlotFontsize.value()
#         DATAx=copy.deepcopy(DATA)
        
#         # samplesgroups = self.ui.listWidget_BoxPlotGroup.selectedItems()
#         samplesgroups=[item.text() for item in samplesgroups]
        
#         samplesgroups2 = self.ui.listWidget_BoxPlotGroup2.selectedItems()
#         samplesgroups2=[item.text() for item in samplesgroups2]
        
#         # print(samplesgroups)
        
#         markersize=self.ui.spinBox_markerSize.value()
        
#         if len(samplesgroups)>0:    #if user defined group names different than "Default group"        
#             grouplistdict=[]
#             if not self.ui.checkBox_BoxPlotRevForw.isChecked():    #select all data points
#                 if not self.ui.checkBox_BoxPlotAftermpp.isChecked():#all points without separation
#                     for sample in samplesgroups:
# #                        print(samplesgroups[item])
#                         groupdict={}
#                         groupdict["Group"]=sample
#                         listofthegroup=[]
#                         for item1 in DATAx.keys():
#                             if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=='Light':
#                                 listofthegroup.append(DATAx[item1])
#                         if len(listofthegroup)!=0:
#                             listofthegroupRev=[]
#                             listofthegroupFor=[]
#                             for item1 in range(len(listofthegroup)):
#                                 if listofthegroup[item1]["ScanDirection"]=="Reverse":
#                                     listofthegroupRev.append(listofthegroup[item1])
#                                 else:
#                                     listofthegroupFor.append(listofthegroup[item1])
                            
#                             groupdict["RevVoc"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
#                             groupdict["ForVoc"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
#                             groupdict["RevJsc"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
#                             groupdict["ForJsc"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
#                             groupdict["RevFF"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
#                             groupdict["ForFF"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
#                             groupdict["RevEff"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
#                             groupdict["ForEff"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
#                             groupdict["RevRoc"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
#                             groupdict["ForRoc"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
#                             groupdict["RevRsc"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
#                             groupdict["ForRsc"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
#                             groupdict["RevVmpp"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
#                             groupdict["ForVmpp"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
#                             groupdict["RevJmpp"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
#                             groupdict["ForJmpp"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
                            
#                             grouplistdict.append(groupdict)
                
#                 else:#for separation before/after mpp
#                     for sample in samplesgroups:
#                         groupdict={}
#                         groupdict["Group"]=sample
#                         listofthegroup=[]
#                         for item1 in DATAx.keys():
#                             if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["aftermpp"]==0:
#                                 listofthegroup.append(DATAx[item1])
#                         if len(listofthegroup)!=0:
#                             listofthegroupRev=[]
#                             listofthegroupFor=[]
#                             for item1 in range(len(listofthegroup)):
#                                 if listofthegroup[item1]["ScanDirection"]=="Reverse":
#                                     listofthegroupRev.append(listofthegroup[item1])
#                                 else:
#                                     listofthegroupFor.append(listofthegroup[item1])
                            
#                             groupdict["RevVoc"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
#                             groupdict["ForVoc"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
#                             groupdict["RevJsc"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
#                             groupdict["ForJsc"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
#                             groupdict["RevFF"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
#                             groupdict["ForFF"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
#                             groupdict["RevEff"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
#                             groupdict["ForEff"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
#                             groupdict["RevRoc"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
#                             groupdict["ForRoc"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
#                             groupdict["RevRsc"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
#                             groupdict["ForRsc"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
#                             groupdict["RevVmpp"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
#                             groupdict["ForVmpp"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
#                             groupdict["RevJmpp"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
#                             groupdict["ForJmpp"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
                            
# #                            grouplistdict.append(groupdict)
#                         listofthegroup2=[]
#                         for item1 in DATAx.keys():
#                             if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["aftermpp"]==1:
#                                 listofthegroup2.append(DATAx[item1])
#                         if len(listofthegroup2)!=0:
#                             listofthegroupRev=[]
#                             listofthegroupFor=[]
#                             for item1 in range(len(listofthegroup2)):
#                                 if listofthegroup2[item1]["ScanDirection"]=="Reverse":
#                                     listofthegroupRev.append(listofthegroup2[item1])
#                                 else:
#                                     listofthegroupFor.append(listofthegroup2[item1])
                            
#                             groupdict["RevVocAMPP"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
#                             groupdict["ForVocAMPP"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
#                             groupdict["RevJscAMPP"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
#                             groupdict["ForJscAMPP"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
#                             groupdict["RevFFAMPP"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
#                             groupdict["ForFFAMPP"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
#                             groupdict["RevEffAMPP"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
#                             groupdict["ForEffAMPP"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
#                             groupdict["RevRocAMPP"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
#                             groupdict["ForRocAMPP"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
#                             groupdict["RevRscAMPP"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
#                             groupdict["ForRscAMPP"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
#                             groupdict["RevVmppAMPP"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
#                             groupdict["ForVmppAMPP"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
#                             groupdict["RevJmppAMPP"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
#                             groupdict["ForJmppAMPP"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
#                         else:
#                             groupdict["RevVocAMPP"]=[]
#                             groupdict["ForVocAMPP"]=[]
#                             groupdict["RevJscAMPP"]=[]
#                             groupdict["ForJscAMPP"]=[]
#                             groupdict["RevFFAMPP"]=[]
#                             groupdict["ForFFAMPP"]=[]
#                             groupdict["RevEffAMPP"]=[]
#                             groupdict["ForEffAMPP"]=[]
#                             groupdict["RevRocAMPP"]=[]
#                             groupdict["ForRocAMPP"]=[]
#                             groupdict["RevRscAMPP"]=[]
#                             groupdict["ForRscAMPP"]=[]
#                             groupdict["RevVmppAMPP"]=[]
#                             groupdict["ForVmppAMPP"]=[]
#                             groupdict["RevJmppAMPP"]=[]
#                             groupdict["ForJmppAMPP"]=[]                            
#                         grouplistdict.append(groupdict)
                    
                    
#             elif self.ui.checkBox_BoxPlotRevForw.isChecked():      #select only the best RevFor of each cell
#                 if not self.ui.checkBox_BoxPlotAftermpp.isChecked():
#                     for sample in samplesgroups:
#                         groupdict={}
#                         groupdict["Group"]=sample
#                         listofthegroup=[]
#                         for item1 in DATAx.keys():
#                             if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=='Light':
#                                 listofthegroup.append(DATAx[item1])
#                         if len(listofthegroup)!=0:
#                             grouper = itemgetter("DepID", "Cellletter",'ScanDirection')
#                             result = []
#                             for key, grp in groupby(sorted(listofthegroup, key = grouper), grouper):
#                                 result.append(list(grp))
                            
#                             result1=[]
                            
#                             for item in result:
#                                 result1.append(sorted(item,key=itemgetter('Eff'),reverse=True)[0])
                            
#                             grouper = itemgetter('ScanDirection')
#                             result2 = []
#                             for key, grp in groupby(sorted(result1, key = grouper), grouper):
#                                 result2.append(list(grp))
                            
#                             listofthegroupRev=[]
#                             listofthegroupFor=[]
                            
#                             if result2[0][0]['ScanDirection']=='Forward':
#                                 listofthegroupFor=result2[0]
#                                 if len(result2)>1:
#                                     listofthegroupRev=result2[1]
#                             else:
#                                 listofthegroupRev=result2[0]
#                                 if len(result2)>1:
#                                     listofthegroupFor=result2[1]
        
#                             groupdict["RevVoc"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
#                             groupdict["ForVoc"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
#                             groupdict["RevJsc"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
#                             groupdict["ForJsc"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
#                             groupdict["RevFF"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
#                             groupdict["ForFF"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
#                             groupdict["RevEff"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
#                             groupdict["ForEff"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
#                             groupdict["RevRoc"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
#                             groupdict["ForRoc"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
#                             groupdict["RevRsc"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
#                             groupdict["ForRsc"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
#                             groupdict["RevVmpp"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
#                             groupdict["ForVmpp"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
#                             groupdict["RevJmpp"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
#                             groupdict["ForJmpp"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
                            
#                             grouplistdict.append(groupdict)
#                 else: #if aftermppchecked
# #                    print("aftermpp is checked")
# #                    print(samplesgroups)
#                     for sample in samplesgroups:
#                         groupdict={}
#                         groupdict["Group"]=sample
#                         listofthegroup=[]
#                         for item1 in DATAx.keys():
#                             if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["aftermpp"]==0:
#                                 listofthegroup.append(DATAx[item1])
#                         if len(listofthegroup)!=0:
#                             grouper = itemgetter("DepID", "Cellletter",'ScanDirection')
#                             result = []
#                             for key, grp in groupby(sorted(listofthegroup, key = grouper), grouper):
#                                 result.append(list(grp))
                            
#                             result1=[]
                            
#                             for item in result:
#                                 result1.append(sorted(item,key=itemgetter('Eff'),reverse=True)[0])
                            
#                             grouper = itemgetter('ScanDirection')
#                             result2 = []
#                             for key, grp in groupby(sorted(result1, key = grouper), grouper):
#                                 result2.append(list(grp))
                            
#                             listofthegroupRev=[]
#                             listofthegroupFor=[]
                            
#                             if result2[0][0]['ScanDirection']=='Forward':
#                                 listofthegroupFor=result2[0]
#                                 if len(result2)>1:
#                                     listofthegroupRev=result2[1]
#                             else:
#                                 listofthegroupRev=result2[0]
#                                 if len(result2)>1:
#                                     listofthegroupFor=result2[1]
        
#                             groupdict["RevVoc"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
#                             groupdict["ForVoc"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
#                             groupdict["RevJsc"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
#                             groupdict["ForJsc"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
#                             groupdict["RevFF"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
#                             groupdict["ForFF"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
#                             groupdict["RevEff"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
#                             groupdict["ForEff"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
#                             groupdict["RevRoc"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
#                             groupdict["ForRoc"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
#                             groupdict["RevRsc"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
#                             groupdict["ForRsc"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
#                             groupdict["RevVmpp"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
#                             groupdict["ForVmpp"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
#                             groupdict["RevJmpp"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
#                             groupdict["ForJmpp"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
                            
# #                            grouplistdict.append(groupdict)
#                         listofthegroup2=[]
#                         for item1 in DATAx.keys():
#                             if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=='Light' and DATAx[item1]["aftermpp"]==1:
#                                 listofthegroup2.append(DATAx[item1])
#                         if len(listofthegroup2)!=0:
# #                            print("listofthegroup2nonzero")
#                             grouper = itemgetter("DepID", "Cellletter",'ScanDirection')
#                             result = []
#                             for key, grp in groupby(sorted(listofthegroup2, key = grouper), grouper):
#                                 result.append(list(grp))
                            
#                             result1=[]
                            
#                             for item in result:
#                                 result1.append(sorted(item,key=itemgetter('Eff'),reverse=True)[0])
                            
#                             grouper = itemgetter('ScanDirection')
#                             result2 = []
#                             for key, grp in groupby(sorted(result1, key = grouper), grouper):
#                                 result2.append(list(grp))
                            
#                             listofthegroupRev=[]
#                             listofthegroupFor=[]
                            
#                             if result2[0][0]['ScanDirection']=='Forward':
#                                 listofthegroupFor=result2[0]
#                                 if len(result2)>1:
#                                     listofthegroupRev=result2[1]
#                             else:
#                                 listofthegroupRev=result2[0]
#                                 if len(result2)>1:
#                                     listofthegroupFor=result2[1]
        
#                             groupdict["RevVocAMPP"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
#                             groupdict["ForVocAMPP"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
#                             groupdict["RevJscAMPP"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
#                             groupdict["ForJscAMPP"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
#                             groupdict["RevFFAMPP"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
#                             groupdict["ForFFAMPP"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
#                             groupdict["RevEffAMPP"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
#                             groupdict["ForEffAMPP"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
#                             groupdict["RevRocAMPP"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
#                             groupdict["ForRocAMPP"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
#                             groupdict["RevRscAMPP"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
#                             groupdict["ForRscAMPP"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
#                             groupdict["RevVmppAMPP"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
#                             groupdict["ForVmppAMPP"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
#                             groupdict["RevJmppAMPP"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
#                             groupdict["ForJmppAMPP"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
#                         else:
#                             groupdict["RevVocAMPP"]=[]
#                             groupdict["ForVocAMPP"]=[]
#                             groupdict["RevJscAMPP"]=[]
#                             groupdict["ForJscAMPP"]=[]
#                             groupdict["RevFFAMPP"]=[]
#                             groupdict["ForFFAMPP"]=[]
#                             groupdict["RevEffAMPP"]=[]
#                             groupdict["ForEffAMPP"]=[]
#                             groupdict["RevRocAMPP"]=[]
#                             groupdict["ForRocAMPP"]=[]
#                             groupdict["RevRscAMPP"]=[]
#                             groupdict["ForRscAMPP"]=[]
#                             groupdict["RevVmppAMPP"]=[]
#                             groupdict["ForVmppAMPP"]=[]
#                             groupdict["RevJmppAMPP"]=[]
#                             groupdict["ForJmppAMPP"]=[]    
#                         grouplistdict.append(groupdict)
                            
#             self.Boxgraph.clear()
#             names=samplesgroups
#             #                print("aftermpp1")
#             # if self.GroupChoice.get()=="Eff":
#             groupchoice=self.ui.comboBox_BoxPlotParam.currentText()
#             # print(groupchoice)

#             if not self.ui.checkBox_BoxPlotAftermpp.isChecked():#aftermpp checkbox is not checked
#                 Effsubfig = self.Boxgraph 
#                 #names=samplesgroups
#                 valsRev=[]
#                 for item in names:
#                     valsRev.append([i["Rev"+groupchoice] for i in grouplistdict if i["Group"]==item and "Rev"+groupchoice in i])
#                 valsFor=[]
#                 for item in names:
#                     valsFor.append([i["For"+groupchoice] for i in grouplistdict if i["Group"]==item and "For"+groupchoice in i])
#                 valstot=[]
                
#                 for item in names:
#                     d=[item,"Rev"+groupchoice]
#                     for i in grouplistdict: 
#                         if i["Group"]==item and "Rev"+groupchoice in i:
#                             d+=i["Rev"+groupchoice]
#                     if d!=[]:
#                         DATAgroupforexport.append(d)
#                     d=[item,"For"+groupchoice]
#                     for i in grouplistdict: 
#                         if i["Group"]==item and "For"+groupchoice in i:
#                             d+=i["For"+groupchoice]
#                     if d!=[]:
#                         DATAgroupforexport.append(d)
#                 # print(DATAgroupforexport)
#                 DATAgroupforexport=list(map(list, six.moves.zip_longest(*DATAgroupforexport, fillvalue=' ')))
#                 # print(DATAgroupforexport)
#                 Rev=[]
#                 Forw=[]
#                 namelist=[]
# #                        print(names)
#                 for i in range(len(names)):
#                      if valsRev!=[]:
#                          if valsRev[i]!=[]:
#                              if valsRev[i][0]!=[]:
#                                  Rev.append(valsRev[i][0])
#                              else:
#                                  Rev.append([])
#                      if valsFor!=[]:
#                          if valsFor[i]!=[]:
#                              if valsFor[i][0]!=[]:
#                                  Forw.append(valsFor[i][0])
#                              else:
#                                  Forw.append([])
#                      if valsRev!=[] or valsFor!=[]: 
#                          if valsRev[i]!=[] or valsFor[i]!=[]: 
#                              if valsRev[i][0]!=[] or valsFor[i][0]!=[]:
#                                  valstot.append(valsRev[i][0]+valsFor[i][0])
#                                  namelist.append(names[i])
# #                        print(namelist)  
                
#                 if self.ui.checkBox_BoxPlotBoxPlot.isChecked():
#                     # print('box')
#                     Effsubfig.boxplot(valstot,0,'',labels=namelist)
#                 # print(valstot)
#                 # print(Rev)
#                 # print(Forw)
#                 for i in range(len(namelist)):
#                     y=Rev[i]
#                     if len(y)>0:
#                         x=np.random.normal(i+1,0.04,size=len(y))
#                         Effsubfig.scatter(x,y,s=markersize,color='red', alpha=0.5)
#                     y=Forw[i]
#                     if len(y)>0:
#                         x=np.random.normal(i+1,0.04,size=len(y))
#                         Effsubfig.scatter(x,y,s=markersize,color='blue', alpha=0.5) 
                    
#             else:
# #                        print("aftermpp")
#                 Effsubfig = self.Boxgraph 
#                 #names=samplesgroups
#                 valsRev=[]
#                 for item in names:
#                     valsRev.append([i["Rev"+groupchoice] for i in grouplistdict if i["Group"]==item and "Rev"+groupchoice in i])
#                 valsFor=[]
#                 for item in names:
#                     valsFor.append([i["For"+groupchoice] for i in grouplistdict if i["Group"]==item and "For"+groupchoice in i])
#                 valsRevAMPP=[]
#                 for item in names:
# #                            v=[i["RevEffAMPP"] for i in grouplistdict if i["Group"]==item and "RevEffAMPP" in i]
#                     valsRevAMPP.append([i["Rev"+groupchoice+"AMPP"] for i in grouplistdict if i["Group"]==item and "Rev"+groupchoice+"AMPP" in i])
# #                        print(len(valsRevAMPP))
#                 valsForAMPP=[]
#                 for item in names:
#                     valsForAMPP.append([i["For"+groupchoice+"AMPP"] for i in grouplistdict if i["Group"]==item and "For"+groupchoice+"AMPP" in i])
# #                        print(len(valsForAMPP))
                
                
                
#                 for item in names:
#                     try:
#                         DATAgroupforexport.append([item,"Rev"+groupchoice]+[i["Rev"+groupchoice] for i in grouplistdict if i["Group"]==item and "Rev"+groupchoice in i][0])
#                     except IndexError:
# #                                print("indexError1")
#                         DATAgroupforexport.append([item,"Rev"+groupchoice]+[])
#                     try:
#                         DATAgroupforexport.append([item,"For"+groupchoice]+[i["For"+groupchoice] for i in grouplistdict if i["Group"]==item and "For"+groupchoice in i][0])
#                     except IndexError:
# #                                print("indexError1")
#                         DATAgroupforexport.append([item,"For"+groupchoice]+[])
#                     try:
#                         DATAgroupforexport.append([item,"Rev"+groupchoice+"AMPP"]+[i["Rev"+groupchoice+"AMPP"] for i in grouplistdict if i["Group"]==item and "Rev"+groupchoice+"AMPP" in i][0])
#                     except IndexError:
# #                                print("indexError1")
#                         DATAgroupforexport.append([item,"Rev"+groupchoice+"AMPP"]+[])
#                     try:
#                         DATAgroupforexport.append([item,"For"+groupchoice+"AMPP"]+[i["For"+groupchoice+"AMPP"] for i in grouplistdict if i["Group"]==item and "For"+groupchoice+"AMPP" in i][0])
#                     except IndexError:
# #                                print("indexError2")
#                         DATAgroupforexport.append([item,"For"+groupchoice+"AMPP"]+[])
#                 # print(DATAgroupforexport)
#                 DATAgroupforexport=map(list, six.moves.zip_longest(*DATAgroupforexport, fillvalue=' '))
                
#                 valstot=[]
#                 Rev=[]
#                 Forw=[]
#                 RevAMPP=[]
#                 ForwAMPP=[]
#                 namelist=[]
#                 for i in range(len(names)):
#                     if valsRev[i]!=[]:
#                         if valsRev[i][0]!=[]:
#                              Rev.append(valsRev[i][0])
#                         else:
#                              Rev.append([])
#                     else:
#                         Rev.append([])
#                     if valsFor[i]!=[]:
#                         if valsFor[i][0]!=[]:
#                              Forw.append(valsFor[i][0])
#                         else:
#                              Forw.append([])
#                     else:
#                         Forw.append([])
#                     if valsRevAMPP[i]!=[]:
#                         if valsRevAMPP[i][0]!=[]:
#                              RevAMPP.append(valsRevAMPP[i][0])
#                         else:
#                              RevAMPP.append([])
#                     else:
#                         RevAMPP.append([])
#                     if valsForAMPP[i]!=[]:    
#                         if valsForAMPP[i][0]!=[]:
#                              ForwAMPP.append(valsForAMPP[i][0])
#                         else:
#                              ForwAMPP.append([])  
#                     else:
#                         ForwAMPP.append([])
#                     try:    
#                         if valsRev[i][0]!=[] or valsFor[i][0]!=[] or valsRevAMPP[i][0]!=[] or valsForAMPP[i][0]!=[]:
#                              valstot.append(valsRev[i][0]+valsFor[i][0]+valsRevAMPP[i][0]+valsForAMPP[i][0])
#                              namelist.append(names[i])
#                     except IndexError:
#                         toaddtovalstot=[]
#                         try:
#                             toaddtovalstot+=valsRev[i][0]
#                         except:
#                             pass
#                         try:
#                             toaddtovalstot+=valsFor[i][0]
#                         except:
#                             pass
#                         try:
#                             toaddtovalstot+=valsRevAMPP[i][0]
#                         except:
#                             pass
#                         try:
#                             toaddtovalstot+=valsForAMPP[i][0]
#                         except:
#                             pass
                            
#                 if namelist!=[]:            
#                     if self.ui.checkBox_BoxPlotBoxPlot.isChecked():
#                         Effsubfig.boxplot(valstot,0,'',labels=namelist)
                
#                     for i in range(len(namelist)):
# #                            if len(listofthegroup)!=0:
#                         y=Rev[i]
#                         if len(y)>0:
#                             x=np.random.normal(i+0.9,0.04,size=len(y))
#                             Effsubfig.scatter(x,y,s=markersize,color='red', alpha=0.5)
#                         y=Forw[i]
#                         if len(y)>0:
#                             x=np.random.normal(i+0.9,0.04,size=len(y))
#                             Effsubfig.scatter(x,y,s=markersize,color='blue', alpha=0.5) 
#                         y=RevAMPP[i]
#                         if len(y)>0:
#                             x=np.random.normal(i+1.1,0.04,size=len(y))
#                             Effsubfig.scatter(x,y,s=markersize,color='orange', alpha=0.5)
#                         y=ForwAMPP[i]
#                         if len(y)>0:
#                             x=np.random.normal(i+1.1,0.04,size=len(y))
#                             Effsubfig.scatter(x,y,s=markersize,color='lightblue', alpha=0.5) 
                        
#             if not self.ui.checkBox_BoxPlotBoxPlot.isChecked():
#                 if namelist!=[]:
#                     span=range(1,len(namelist)+1)
# #                            print(namelist)
# #                            print(span)
# #                        plt.xticks(span,namelist)
#                     Effsubfig.set_xticks(span)
#                     Effsubfig.set_xticklabels(namelist)
#                     Effsubfig.set_xlim([0.5,span[-1]+0.5])
            
#             # if self.minmaxgroupgraphcheck.get()==1:
#             #     Effsubfig.set_ylim([self.minYgroupgraph.get(),self.maxYgroupgraph.get()])
                
#             Effsubfig.set_ylabel(groupchoice)
#             for item in ([Effsubfig.title, Effsubfig.xaxis.label, Effsubfig.yaxis.label] +
#                          Effsubfig.get_xticklabels() + Effsubfig.get_yticklabels()):
#                 item.set_fontsize(fontsizegroup)
            
#             for tick in Effsubfig.get_xticklabels():
#                 tick.set_rotation(self.ui.spinBox_BoxPlotRotation.value())
        
#         self.fig3.canvas.draw_idle()
        
    def GraphBoxsave_as2(self):
        global DATA
        global DATAgroupforexport
        
        try:
            if not self.ui.checkBox_BoxPlotBig4_2.isChecked():
                f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]
                if self.ui.checkBox_BoxPlotBkg_2.isChecked():
                    self.fig31.savefig(f, dpi=300, transparent=True, bbox_inches='tight')
                else:
                    self.fig31.savefig(f,dpi=300, transparent=False, bbox_inches='tight')
                # print(DATAgroupforexport)
                DATAgroupforexport1=[]
                for item in DATAgroupforexport:
                    line=""
                    for item1 in item:
                        line=line+str(item1)+"\t"
                    line=line[:-1]+"\n"
                    DATAgroupforexport1.append(line)
                
                file = open(str(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam_2.currentText()+"dat.txt"),'w', encoding='ISO-8859-1')
                file.writelines("%s" % item for item in DATAgroupforexport1)
                file.close()
            elif self.ui.checkBox_BoxPlotBig4_2.isChecked():
                
                f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]
                listofimages=[]
                for param in ["Eff","FF","Voc","Jsc"]:
                    self.ui.comboBox_BoxPlotParam_2.setCurrentIndex(self.ui.comboBox_BoxPlotParam_2.findText(param, QtCore.Qt.MatchFixedString))
                    self.UpdateBoxGraph2()
                    # self.fig31.tight_layout()
                    if self.ui.checkBox_BoxPlotBkg_2.isChecked():
                        self.fig31.savefig(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam_2.currentText()+f[-4:], dpi=300, transparent=True, bbox_inches='tight')
                    else:
                        self.fig31.savefig(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam_2.currentText()+f[-4:],dpi=300, transparent=False, bbox_inches='tight')
                    listofimages.append(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam_2.currentText()+f[-4:])
                    
                    DATAgroupforexport1=[]
                    for item in DATAgroupforexport:
                        line=""
                        for item1 in item:
                            line=line+str(item1)+"\t"
                        line=line[:-1]+"\n"
                        DATAgroupforexport1.append(line)
                    
                    file = open(str(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam_2.currentText()+"dat.txt"),'w', encoding='ISO-8859-1')
                    file.writelines("%s" % item for item in DATAgroupforexport1)
                    file.close()
                self.ui.comboBox_BoxPlotParam_2.setCurrentIndex(self.ui.comboBox_BoxPlotParam_2.findText("Eff", QtCore.Qt.MatchFixedString))
                
                images = list(map(ImageTk.open, listofimages))
                widths, heights = zip(*(i.size for i in images))
                total_width = max([widths[0],widths[1]])+max([widths[2],widths[3]])
                max_height = max([heights[0],heights[2]])+max([heights[1],heights[3]])
                new_im = ImageTk.new('RGB', (total_width, max_height), (255, 255, 255))
                new_im.paste(im=images[0],box=(0,0))
                new_im.paste(im=images[1],box=(0,max([heights[0],heights[2]])))
                new_im.paste(im=images[2],box=(max([widths[0],widths[1]]),0))
                new_im.paste(im=images[3],box=(max([widths[0],widths[1]]),max([heights[0],heights[2]])))
                new_im.save(f[:-4]+"_Big4"+f[-4:])
                
                
        except:
            print("there is an exception with save groupboxgraph")
        
    # def GraphBoxsave_as(self):
    #     global DATA
    #     global DATAgroupforexport
        
    #     try:
    #         if not self.ui.checkBox_BoxPlotBig4.isChecked():
    #             f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]
    #             if self.ui.checkBox_BoxPlotBkg.isChecked():
    #                 self.fig3.savefig(f, dpi=300, transparent=True)
    #             else:
    #                 self.fig3.savefig(f,dpi=300, transparent=False)
    #             print(DATAgroupforexport)
    #             DATAgroupforexport1=[]            
    #             for item in DATAgroupforexport:
    #                 line=""
    #                 for item1 in item:
    #                     line=line+str(item1)+"\t"
    #                 line=line[:-1]+"\n"
    #                 DATAgroupforexport1.append(line)
                
    #             file = open(str(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam.currentText()+"dat.txt"),'w', encoding='ISO-8859-1')
    #             file.writelines("%s" % item for item in DATAgroupforexport1)
    #             file.close()
    #         elif self.ui.checkBox_BoxPlotBig4.isChecked():
                
    #             f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]
                
    #             for param in ["Eff","FF","Voc","Jsc"]:
    #                 self.ui.comboBox_BoxPlotParam.setCurrentIndex(self.ui.comboBox_BoxPlotParam.findText(param, QtCore.Qt.MatchFixedString))
    #                 self.UpdateBoxGraph()
    #                 if self.ui.checkBox_BoxPlotBkg.isChecked():
    #                     self.fig3.savefig(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam.currentText()+f[-4:], dpi=300, transparent=True)
    #                 else:
    #                     self.fig3.savefig(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam.currentText()+f[-4:],dpi=300, transparent=False)
                        
    #                 DATAgroupforexport1=[]
    #                 for item in DATAgroupforexport:
    #                     line=""
    #                     for item1 in item:
    #                         line=line+str(item1)+"\t"
    #                     line=line[:-1]+"\n"
    #                     DATAgroupforexport1.append(line)
                    
    #                 file = open(str(f[:-4]+"_"+self.ui.comboBox_BoxPlotParam.currentText()+"dat.txt"),'w', encoding='ISO-8859-1')
    #                 file.writelines("%s" % item for item in DATAgroupforexport1)
    #                 file.close()
    #             self.ui.comboBox_BoxPlotParam.setCurrentIndex(self.ui.comboBox_BoxPlotParam.findText("Eff", QtCore.Qt.MatchFixedString))
    #     except:
    #         print("there is an exception with save groupboxgraph")

#%%#############
    def plottingTimefromTable(self):
        global takenforplotTime
        rows=list(set(index.row() for index in self.ui.tableWidget.selectedIndexes()))
        takenforplotTime=[self.ui.tableWidget.item(row,2).text()+'_'+str(self.ui.tableWidget.item(row,3).text()).replace(' ','_').replace(':','-')+'_'+str(float(self.ui.tableWidget.item(row,8).text()))+'_'+str(float(self.ui.tableWidget.item(row,5).text())) for row in rows]
        self.UpdateTimeGraph()

    def UpdateTimeGraph(self):
        global DATA, takenforplotTime, colorstylelist, DATAtimeevolforexport
#        print("")
#        print(takenforplotTime)
        #"MeasDayTime2"
        DATAtimeevolforexport={}
        # print(takenforplotTime)
        if takenforplotTime!=[]:
            if self.ui.checkBox_BestEffPixDay.isChecked()==False and self.ui.checkBox_bestofRevFor.isChecked()==False:
                TimeDatDict={}
                self.TimeEvolgraph.clear()
                for item in takenforplotTime:
                    newkey=item.split('_')[0]+'_'+item.split('_')[1]+'_'+item.split('_')[2]
                    if newkey not in TimeDatDict.keys():
                        TimeDatDict[newkey]={'reverse':{'Voc':[[],[]],'Jsc':[[],[]],'FF':[[],[]],'Eff':[[],[]]},'forward':{'Voc':[[],[]],'Jsc':[[],[]],'FF':[[],[]],'Eff':[[],[]]}}
                    for item11 in DATA.keys():
                        item1=DATA[item11]
                        if item1["SampleNameID"]==item:
                            if item1["ScanDirection"]=="Reverse" and item1["Illumination"]=="Light":
                                TimeDatDict[newkey]['reverse']['Voc'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['reverse']['Voc'][1].append(item1["Voc"])
                                TimeDatDict[newkey]['reverse']['Jsc'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['reverse']['Jsc'][1].append(item1["Jsc"])
                                TimeDatDict[newkey]['reverse']['FF'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['reverse']['FF'][1].append(item1["FF"])
                                TimeDatDict[newkey]['reverse']['Eff'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['reverse']['Eff'][1].append(item1["Eff"])
                            elif item1["ScanDirection"]=="Forward" and item1["Illumination"]=="Light":
                                TimeDatDict[newkey]['forward']['Voc'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['forward']['Voc'][1].append(item1["Voc"])
                                TimeDatDict[newkey]['forward']['Jsc'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['forward']['Jsc'][1].append(item1["Jsc"])
                                TimeDatDict[newkey]['forward']['FF'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['forward']['FF'][1].append(item1["FF"])
                                TimeDatDict[newkey]['forward']['Eff'][0].append(item1["MeasDayTime2"])
                                TimeDatDict[newkey]['forward']['Eff'][1].append(item1["Eff"])
        #        num_plots = len(TimeDatDict.keys())          
        #        plt.gca().set_prop_cycle(plt.cycler('color', plt.cm.Spectral(np.linspace(0, 1, num_plots))))
                   
        #        print(list(TimeDatDict.keys())) 
                # minx=min(TimeDatDict[newkey]['forward'][self.TimeChoice.get()][0]+TimeDatDict[newkey]['reverse'][self.TimeChoice.get()][0])
                # maxx=max(TimeDatDict[newkey]['forward'][self.TimeChoice.get()][0]+TimeDatDict[newkey]['reverse'][self.TimeChoice.get()][0])
                # print(TimeDatDict)
                for key in list(TimeDatDict.keys()):
                    partdatatime=[[],[],[],[],[],[],[],[]]
                    # if minx>min(TimeDatDict[key]['forward'][self.TimeChoice.get()][0]+TimeDatDict[key]['reverse'][self.TimeChoice.get()][0]):
                    #     minx=min(TimeDatDict[key]['forward'][self.TimeChoice.get()][0]+TimeDatDict[key]['reverse'][self.TimeChoice.get()][0])
                    # if maxx<max(TimeDatDict[key]['forward'][self.TimeChoice.get()][0]+TimeDatDict[key]['reverse'][self.TimeChoice.get()][0]):
                    #     maxx=max(TimeDatDict[key]['forward'][self.TimeChoice.get()][0]+TimeDatDict[key]['reverse'][self.TimeChoice.get()][0])
                    try:
                        xfor, yfor=zip(*sorted(zip(TimeDatDict[key]['forward'][self.ui.comboBox_PVPTimeParam.currentText()][0],TimeDatDict[key]['forward'][self.ui.comboBox_PVPTimeParam.currentText()][1]), key = lambda x: x[1]))
                        xfor=list(xfor)
                        yfor=list(yfor)
                        yfor.sort(key=dict(zip(yfor, xfor)).get)
                        xfor=sorted(xfor)
                        partdatatime[0]=xfor
                        partdatatime[1]=[(m-xfor[0]).total_seconds()/3600 for m in xfor]
                        partdatatime[2]=yfor
                        if self.ui.spinBox_PVPTimeNormalPoint.value()==-1:
                            try: 
                                partdatatime[3]=[(m)/(yfor[0]) for m in yfor]
                            except ZeroDivisionError:
                                pass
                        else:
                            foundnormalsetpt=0
                            for item in range(len(partdatatime[1])):
                                if partdatatime[1][item]>=self.ui.spinBox_PVPTimeNormalPoint.value():
                                    partdatatime[3]=[(m)/(partdatatime[2][item]) for m in yfor]
                                    foundnormalsetpt=1
                                    break
                            if foundnormalsetpt==0:
                                partdatatime[3]=[(m)/(yfor[0]) for m in yfor]
                    except ValueError:
                        pass
                    try:
                        xrev, yrev=zip(*sorted(zip(TimeDatDict[key]['reverse'][self.ui.comboBox_PVPTimeParam.currentText()][0],TimeDatDict[key]['reverse'][self.ui.comboBox_PVPTimeParam.currentText()][1]), key = lambda x: x[1]))                
                        xrev=list(xrev)
                        yrev=list(yrev)
                        yrev.sort(key=dict(zip(yrev, xrev)).get)
                        xrev=sorted(xrev)
                        partdatatime[4]=xrev
                        partdatatime[5]=[(m-xrev[0]).total_seconds()/3600 for m in xrev]
                        partdatatime[6]=yrev
                        
                        if self.ui.spinBox_PVPTimeNormalPoint.value()==-1:
                            try:
                                partdatatime[7]=[(m)/(yrev[0]) for m in yrev]
                            except ZeroDivisionError:
                                print('there is a zero devision error. we pass...')
                        else:
                            foundnormalsetpt=0
                            for item in range(len(partdatatime[5])):
                                if partdatatime[5][item]>=self.ui.spinBox_PVPTimeNormalPoint.value():
                                    partdatatime[7]=[(m)/(partdatatime[6][item]) for m in yrev]
                                    foundnormalsetpt=1
                                    break
                            if foundnormalsetpt==0:
                                partdatatime[7]=[(m)/(yrev[0]) for m in yrev]
                        
                    except ValueError:
                        pass
                    DATAtimeevolforexport[key]=partdatatime
                 
                color1=0 
                for key in list(DATAtimeevolforexport.keys()):
                    xfor=DATAtimeevolforexport[key][0]
                    yfor=DATAtimeevolforexport[key][2]
                    xrev=DATAtimeevolforexport[key][4]
                    yrev=DATAtimeevolforexport[key][6]
                    if self.ui.checkBox_PVPTimeRelativeTime.isChecked():
                        xfor=DATAtimeevolforexport[key][1]
                        xrev=DATAtimeevolforexport[key][5]
                    if self.ui.checkBox_PVPTimeNormal.isChecked():
                        yfor=DATAtimeevolforexport[key][3]
                        yrev=DATAtimeevolforexport[key][7]
                
                    if self.ui.checkBox_PVPTimeLine.isChecked():
                        self.TimeEvolgraph.plot(xfor, yfor, linestyle='--', marker='o',color=colorstylelist[color1],label=key+'_For')
                        self.TimeEvolgraph.plot(xrev, yrev, linestyle='-', marker='o', color=colorstylelist[color1], alpha=0.5,label=key+'_Rev')
                    else:
                        self.TimeEvolgraph.plot(xfor, yfor, linestyle='', marker='o',color=colorstylelist[color1],label=key+'_For')
                        self.TimeEvolgraph.plot(xrev, yrev, linestyle='', marker='o', color=colorstylelist[color1], alpha=0.5,label=key+'_Rev')  
                    color1=color1+1
                    
                
                if self.ui.checkBox_PVPTimeRelativeTime.isChecked():
                    self.TimeEvolgraph.set_xlabel('Time (hours)')
                else:
                    # self.TimeEvolfig.set_xlim(minx-0.05*(maxx-minx),maxx+0.05*(maxx-minx))    
                    self.TimeEvolgraph.set_xlabel('Time')
                
                # if self.minmaxtimegraphcheck.get():
                #     self.TimeEvolfig.set_ylim(self.minYtimegraph.get(),self.maxYtimegraph.get())
                
                self.TimeEvolgraph.set_ylabel(self.ui.comboBox_PVPTimeParam.currentText())
                for tick in self.TimeEvolgraph.get_xticklabels():
                    tick.set_rotation(20)
                self.TimeEvolgraphleg=self.TimeEvolgraph.legend(loc='lower left', bbox_to_anchor=(1, 0))
                self.fig4.canvas.draw_idle()
                
            elif self.ui.checkBox_bestofRevFor.isChecked() and not self.ui.checkBox_BestEffPixDay.isChecked():
                print("bestrevfor")
                
                
            elif not self.ui.checkBox_bestofRevFor.isChecked() and self.ui.checkBox_BestEffPixDay.isChecked():   
#                print("bestoftheday")
                TimeDatDict={}
                self.TimeEvolgraph.clear()
                for item in takenforplotTime:
                    newkey=item.split('_')[0]+'_'+item.split('_')[1]#per substrate e.g. 41_10
                    if newkey not in TimeDatDict.keys():
                        TimeDatDict[newkey]={}
                    for item11 in DATA.keys():
                        item1=DATA[item11]
                        if item1["SampleNameID"]==item:
                            newdatekey=str(item1["MeasDayTime2"].date())
                            if newdatekey not in TimeDatDict[newkey].keys():
                                TimeDatDict[newkey][newdatekey]={'Voc':[[],[]],'Jsc':[[],[]],'FF':[[],[]],'Eff':[[],[]]}
                            
                            TimeDatDict[newkey][newdatekey]['Voc'][0].append(item1["MeasDayTime2"])
                            TimeDatDict[newkey][newdatekey]['Voc'][1].append(item1["Voc"])
                            TimeDatDict[newkey][newdatekey]['Jsc'][0].append(item1["MeasDayTime2"])
                            TimeDatDict[newkey][newdatekey]['Jsc'][1].append(item1["Jsc"])
                            TimeDatDict[newkey][newdatekey]['FF'][0].append(item1["MeasDayTime2"])
                            TimeDatDict[newkey][newdatekey]['FF'][1].append(item1["FF"])
                            TimeDatDict[newkey][newdatekey]['Eff'][0].append(item1["MeasDayTime2"])
                            TimeDatDict[newkey][newdatekey]['Eff'][1].append(item1["Eff"])
                
                for key0 in list(TimeDatDict.keys()):
#                    print(key0)
                    TimeDatDict[key0]['bestEffofday']={'Voc':[[],[]],'Jsc':[[],[]],'FF':[[],[]],'Eff':[[],[]]}
                    for key in list(TimeDatDict[key0].keys()):
#                        print(key)
#                        print(max(TimeDatDict[key0][key]['Eff'][1]))
                        
                        ind=TimeDatDict[key0][key]['Eff'][1].index(max(TimeDatDict[key0][key]['Eff'][1]))
                        
                        TimeDatDict[key0]['bestEffofday']['Voc'][0].append(TimeDatDict[key0][key]['Voc'][0][ind])
                        TimeDatDict[key0]['bestEffofday']['Voc'][1].append(TimeDatDict[key0][key]['Voc'][1][ind])
                        TimeDatDict[key0]['bestEffofday']['Jsc'][0].append(TimeDatDict[key0][key]['Jsc'][0][ind])
                        TimeDatDict[key0]['bestEffofday']['Jsc'][1].append(TimeDatDict[key0][key]['Jsc'][1][ind])
                        TimeDatDict[key0]['bestEffofday']['FF'][0].append(TimeDatDict[key0][key]['FF'][0][ind])
                        TimeDatDict[key0]['bestEffofday']['FF'][1].append(TimeDatDict[key0][key]['FF'][1][ind])
                        TimeDatDict[key0]['bestEffofday']['Eff'][0].append(TimeDatDict[key0][key]['Eff'][0][ind])
                        TimeDatDict[key0]['bestEffofday']['Eff'][1].append(TimeDatDict[key0][key]['Eff'][1][ind])
                
                # minx=min(TimeDatDict[newkey]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][0])
                # maxx=max(TimeDatDict[newkey]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][0])
                
                for key in list(TimeDatDict.keys()):
                    partdatatime=[[],[],[],[]]
                    # if minx>min(TimeDatDict[key]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][0]):
                    #     minx=min(TimeDatDict[key]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][0])
                    # if maxx<max(TimeDatDict[key]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][0]):
                    #     maxx=max(TimeDatDict[key]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][0])
                    try:
                        xfor, yfor=zip(*sorted(zip(TimeDatDict[key]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][0],TimeDatDict[key]['bestEffofday'][self.ui.comboBox_PVPTimeParam.currentText()][1]), key = lambda x: x[1]))
                        xfor=list(xfor)
                        yfor=list(yfor)
                        yfor.sort(key=dict(zip(yfor, xfor)).get)
                        xfor=sorted(xfor)
                        partdatatime[0]=xfor
                        partdatatime[1]=[(m-xfor[0]).total_seconds()/3600 for m in xfor]
                        partdatatime[2]=yfor
                        if self.ui.spinBox_PVPTimeNormalPoint.value()==-1:
                            partdatatime[3]=[(m)/(yfor[0]) for m in yfor]
                        else:
                            foundnormalsetpt=0
                            for item in range(len(partdatatime[1])):
                                if partdatatime[1][item]>=self.ui.spinBox_PVPTimeNormalPoint.value():
                                    partdatatime[3]=[(m)/(partdatatime[2][item]) for m in yfor]
                                    foundnormalsetpt=1
                                    break
                            if foundnormalsetpt==0:
                                partdatatime[3]=[(m)/(yfor[0]) for m in yfor]
                    except ValueError:
                        pass
                    
                    DATAtimeevolforexport[key]=partdatatime
                 
                color1=0 
                for key in list(DATAtimeevolforexport.keys()):
                    xfor=DATAtimeevolforexport[key][0]
                    yfor=DATAtimeevolforexport[key][2]
                    
                    if self.ui.checkBox_PVPTimeRelativeTime.isChecked():
                        xfor=DATAtimeevolforexport[key][1]
                    if self.ui.checkBox_PVPTimeNormal.isChecked():
                        yfor=DATAtimeevolforexport[key][3]
                
                    if self.ui.checkBox_PVPTimeLine.isChecked():
                        self.TimeEvolgraph.plot(xfor, yfor, linestyle='-', marker='o',color=colorstylelist[color1],label=key+'_Best')
                    else:
                        self.TimeEvolgraph.plot(xfor, yfor, linestyle='', marker='o',color=colorstylelist[color1],label=key+'_Best')
                    color1=color1+1
                    
                
                if self.ui.checkBox_PVPTimeRelativeTime.isChecked():
                    self.TimeEvolgraph.set_xlabel('Time (hours)')
                else:
                    # self.TimeEvolgraph.set_xlim(minx-0.05*(maxx-minx),maxx+0.05*(maxx-minx))    
                    self.TimeEvolgraph.set_xlabel('Time')
                
                # if self.minmaxtimegraphcheck.get():
                #     self.TimeEvolgraph.set_ylim(self.minYtimegraph.get(),self.maxYtimegraph.get())
                
                self.TimeEvolgraph.set_ylabel(self.ui.comboBox_PVPTimeParam.currentText())
                for tick in self.TimeEvolgraph.get_xticklabels():
                    tick.set_rotation(20)
                self.TimeEvolgraphleg=self.TimeEvolgraph.legend(loc='lower left', bbox_to_anchor=(1, 0))
                
        self.fig4.canvas.draw_idle()
        
    def GraphTimesave_as(self):
        global DATAtimeevolforexport
        if not self.ui.checkBox_PVPTimeBig4.isChecked():
            f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]
            
            self.fig4.savefig(f, dpi=300, transparent=False)
            
                            
            for key in list(DATAtimeevolforexport.keys()):
                DATAgroupforexport1=["realtimeF\trelativetimeF\tvalueF\tnormalizedvaluetot0F\trealtimeR\trelativetimeR\tvalueR\tnormalizedvaluetot0R\n"] 
                templist=map(list, six.moves.zip_longest(*DATAtimeevolforexport[key], fillvalue=' '))
                for item in templist:
                    line=""
                    for item1 in item:
                        line=line+str(item1)+"\t"
                    line=line[:-1]+"\n"
                    DATAgroupforexport1.append(line)
                file = open(str(f[:-4]+"_"+self.ui.comboBox_PVPTimeParam.currentText()+'_'+str(key)+"_dat.txt"),'w', encoding='ISO-8859-1')
                file.writelines("%s" % item for item in DATAgroupforexport1)
                file.close()
            
        elif self.ui.checkBox_PVPTimeBig4.isChecked():
            
            f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]                
            for param in ["Eff","Voc","Jsc","FF"]:
                self.ui.comboBox_PVPTimeParam.setCurrentIndex(self.ui.comboBox_PVPTimeParam.findText(param, QtCore.Qt.MatchFixedString))
                self.UpdateTimeGraph()
                self.fig4.savefig(f[:-4]+"_"+self.ui.comboBox_PVPTimeParam.currentText()+f[-4:], dpi=300, transparent=False)
                
                for key in list(DATAtimeevolforexport.keys()):
                    DATAgroupforexport1=["realtimeF\trelativetimeF\tvalueF\tnormalizedvaluetot0F\trealtimeR\trelativetimeR\tvalueR\tnormalizedvaluetot0R\n"] 
                    templist=map(list, six.moves.zip_longest(*DATAtimeevolforexport[key], fillvalue=' '))
                    for item in templist:
                        line=""
                        for item1 in item:
                            line=line+str(item1)+"\t"
                        line=line[:-1]+"\n"
                        DATAgroupforexport1.append(line)
                    file = open(str(f[:-4]+"_"+self.ui.comboBox_PVPTimeParam.currentText()+'_'+str(key)+"_dat.txt"),'w', encoding='ISO-8859-1')
                    file.writelines("%s" % item for item in DATAgroupforexport1)
                    file.close()

#%%#############
    def UpdateCompGraph(self):
        global DATA
        global DATAcompforexport
        
        DATAcompforexport=[]
        DATAx=copy.deepcopy(DATA)
        
        samplesgroups = self.ui.listWidget_ParamComp.selectedItems()
        samplesgroups=[item.text() for item in samplesgroups]
        
#        print(samplesgroups)
        
        if samplesgroups==[]:
            self.ParamParamgraph.clear()
        else:
            grouplistdict={}
            for item in range(len(samplesgroups)):
                groupdict={}
                groupdict["Group"]=samplesgroups[item]
                listofthegroup=[]
                for item1 in DATAx.keys():
                    if DATAx[item1]["Group"]+' * '+DATAx[item1]["Group2"]==groupdict["Group"] and DATAx[item1]["Illumination"]=='Light':
                        listofthegroup.append(DATAx[item1])
               
                if len(listofthegroup)!=0:
                    listofthegroupRev=[]
                    listofthegroupFor=[]
                    for item1 in range(len(listofthegroup)):
                        if listofthegroup[item1]["ScanDirection"]=="Reverse":
                            listofthegroupRev.append(listofthegroup[item1])
                        else:
                            listofthegroupFor.append(listofthegroup[item1])
                    
                    groupdict["Voc"]={}
                    groupdict["Jsc"]={}
                    groupdict["FF"]={}
                    groupdict["Eff"]={}
                    groupdict["Roc"]={}
                    groupdict["Rsc"]={}
                    groupdict["Vmpp"]={}
                    groupdict["Jmpp"]={}
                    
                    
                    groupdict["Voc"]["Rev"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
                    groupdict["Voc"]["For"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
                    groupdict["Jsc"]["Rev"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
                    groupdict["Jsc"]["For"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
                    groupdict["FF"]["Rev"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
                    groupdict["FF"]["For"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
                    groupdict["Eff"]["Rev"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
                    groupdict["Eff"]["For"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
                    groupdict["Roc"]["Rev"]=[x['Roc'] for x in listofthegroupRev if 'Roc' in x]
                    groupdict["Roc"]["For"]=[x['Roc'] for x in listofthegroupFor if 'Roc' in x]
                    groupdict["Rsc"]["Rev"]=[x['Rsc'] for x in listofthegroupRev if 'Rsc' in x]
                    groupdict["Rsc"]["For"]=[x['Rsc'] for x in listofthegroupFor if 'Rsc' in x]
                    groupdict["Vmpp"]["Rev"]=[x['Vmpp'] for x in listofthegroupRev if 'Vmpp' in x]
                    groupdict["Vmpp"]["For"]=[x['Vmpp'] for x in listofthegroupFor if 'Vmpp' in x]
                    groupdict["Jmpp"]["Rev"]=[x['Jmpp'] for x in listofthegroupRev if 'Jmpp' in x]
                    groupdict["Jmpp"]["For"]=[x['Jmpp'] for x in listofthegroupFor if 'Jmpp' in x]
                    
#                    grouplistdict.append(groupdict)
                    grouplistdict[samplesgroups[item]]=groupdict
            colormapname="jet"
            cmap = plt.get_cmap(colormapname)
            colors = cmap(np.linspace(0, 1.0, len(list(grouplistdict.keys()))))
            colors=[tuple(item) for item in colors]  
             
            self.ParamParamgraph.clear()
            indexcolor=0
            for group in list(grouplistdict.keys()):
                DATAcompforexport.append([self.ui.comboBox_PVPx.currentText(),'']+grouplistdict[group][self.ui.comboBox_PVPx.currentText()]['Rev'])
                DATAcompforexport.append([self.ui.comboBox_PVPy.currentText(),group+'_Rev']+grouplistdict[group][self.ui.comboBox_PVPy.currentText()]['Rev'])
                self.ParamParamgraph.scatter(grouplistdict[group][self.ui.comboBox_PVPx.currentText()]['Rev'],grouplistdict[group][self.ui.comboBox_PVPy.currentText()]['Rev']
                                            ,label=group+'_Rev',color=colors[indexcolor],marker="o")
                DATAcompforexport.append([self.ui.comboBox_PVPx.currentText(),'']+grouplistdict[group][self.ui.comboBox_PVPx.currentText()]['For'])
                DATAcompforexport.append([self.ui.comboBox_PVPy.currentText(),group+'For']+grouplistdict[group][self.ui.comboBox_PVPy.currentText()]['For'])
                self.ParamParamgraph.scatter(grouplistdict[group][self.ui.comboBox_PVPx.currentText()]['For'],grouplistdict[group][self.ui.comboBox_PVPy.currentText()]['For']
                                            ,label=group+'_For',color=colors[indexcolor],marker="s")
                indexcolor+=1
            
            DATAcompforexport=map(list, six.moves.zip_longest(*DATAcompforexport, fillvalue=' '))
            
            self.ParamParamgraph.set_ylabel(self.ui.comboBox_PVPy.currentText())    
            self.ParamParamgraph.set_xlabel(self.ui.comboBox_PVPx.currentText()) 
#            self.CompParamGroupfig.legend()
            self.leg=self.ParamParamgraph.legend(loc='lower left', bbox_to_anchor=(1, 0))
            
        self.fig5.canvas.draw_idle()
    
    def GraphCompsave_as(self):
        global DATAcompforexport
        
        try:
            f = QFileDialog.getSaveFileName(self, 'Save graph and data', ".png", "graph file (*.png);; All Files (*)")[0]
            self.fig5.savefig(f, dpi=300, bbox_extra_artists=(self.leg,))#, transparent=True)
                           
            DATAcompforexport1=[]            
            for item in DATAcompforexport:
                line=""
                for item1 in item:
                    line=line+str(item1)+"\t"
                line=line[:-1]+"\n"
                DATAcompforexport1.append(line)
                
            file = open(str(f[:-4]+"_dat.txt"),'w', encoding='ISO-8859-1')
            file.writelines("%s" % item for item in DATAcompforexport1)
            file.close()
        
        except:
            print("there is an exception") 

#%%#############
    def SaveSession(self):
        global file_path,testdata,DATA,DATAJVforexport,DATAJVtabforexport,DATAmppforexport,DATAgroupforexport
        global DATAHistforexport,DATAcompforexport,DATAtimeevolforexport,takenforplot,takenforplotmpp,takenforplotTime
        global DATAMPP,DATAdark,DATAFV, DATAFFloss,numbLightfiles,numbDarkfiles, instructionsofsearch,IVlegendMod,colorstylelist,colormapname
        global MPPlegendMod,MPPlinestyle,titIV,titmpp,titStat,samplesgroups,groupstoplotcomp,grouptoplot
        global groupcomment,listofanswer,listoflinestyle,listofcolorstyle,listoflinewidthstyle,listofanswermpp
        global listoflinestylempp,listofcolorstylempp,listoflinewidthstylempp
        
        global resultsdirectory
        
        directory = os.path.join(resultsdirectory,'savedsession')
        if not os.path.exists(directory):
            os.makedirs(directory)
            os.chdir(directory)
        else :
            os.chdir(directory)
            
        # current_path = os.getcwd()
        directory=QFileDialog.getExistingDirectory(self, 'Select directory')
        os.chdir(directory)
        
        listofglobalvariables= [file_path,testdata,DATA,DATAJVforexport,DATAJVtabforexport,DATAmppforexport,DATAgroupforexport,
            DATAHistforexport,DATAcompforexport,DATAtimeevolforexport,takenforplot,takenforplotmpp,takenforplotTime,
            DATAMPP,DATAdark,DATAFV,DATAFFloss,numbLightfiles,numbDarkfiles,instructionsofsearch,IVlegendMod,colorstylelist,colormapname,
            MPPlegendMod,MPPlinestyle,titIV,titmpp,titStat,samplesgroups,groupstoplot,groupstoplotcomp,
            groupcomment]
            # listofanswer,listoflinestyle,listofcolorstyle,listoflinewidthstyle,listofanswermpp,
            # listoflinestylempp,listofcolorstylempp,listoflinewidthstylempp]
        
        for item in range(len(listofglobalvariables)):
            # print(item)
            pickle.dump(listofglobalvariables[item],open(str(item)+'.pkl','wb'), protocol=pickle.HIGHEST_PROTOCOL)
        
        self.saveconfigsgui(directory)
        
        print("dumped")
        """
        try:
            self.dumpfilename = filedialog.asksaveasfilename(defaultextension=".pkl")
            dill.dump_session(self.dumpfilename)
        except:
            print("there is an exception")
        """
        
    def LoadSession(self):
        global file_path,testdata,DATA,DATAJVforexport,DATAJVtabforexport,DATAmppforexport,DATAgroupforexport
        global DATAHistforexport,DATAcompforexport,DATAtimeevolforexport,takenforplot,takenforplotmpp,takenforplotTime
        global DATAMPP,DATAdark,DATAFV,DATAFFloss,numbLightfiles,numbDarkfiles,instructionsofsearch,IVlegendMod,colorstylelist,colormapname
        global MPPlegendMod,MPPlinestyle,titIV,titmpp,titStat,samplesgroups,groupstoplotcomp,grouptoplot
        global groupcomment,listofanswer,listoflinestyle,listofcolorstyle,listoflinewidthstyle,listofanswermpp
        global listoflinestylempp,listofcolorstylempp,listoflinewidthstylempp
        current_path = os.getcwd()
        path=QFileDialog.getExistingDirectory(self, 'Select directory')
        os.chdir(path)
        
        listofglobalvariables= ["file_path","testdata","DATA","DATAJVforexport","DATAJVtabforexport","DATAmppforexport","DATAgroupforexport",
            "DATAHistforexport","DATAcompforexport","DATAtimeevolforexport","takenforplot","takenforplotmpp","takenforplotTime",
            "DATAMPP","DATAdark","DATAFV", "DATAFFloss", "numbLightfiles","numbDarkfiles","instructionsofsearch","IVlegendMod","colorstylelist","colormapname",
            "MPPlegendMod","MPPlinestyle","titIV","titmpp","titStat","samplesgroups","groupstoplot","groupstoplotcomp",
            "groupcomment"]
        # ,"listofanswer","listoflinestyle","listofcolorstyle","listoflinewidthstyle","listofanswermpp",
        #     "listoflinestylempp","listofcolorstylempp","listoflinewidthstylempp"]
        
        for item in range(len(listofglobalvariables)):
            globals()[listofglobalvariables[item]] = pickle.load(open(str(item)+'.pkl','rb'))

        self.loadconfigsgui(path)
        """
        try:
            self.dumpfilename = filedialog.asksaveasfilename(defaultextension=".pkl")
            dill.load_session(self.dumpfilename)
        except:
            print("there is an exception")
        """
        print("loaded")
            
        if DATAMPP!={}:
            self.ui.listWidget_MppSamples.clear()
            self.ui.listWidget_MppSamples.addItems(DATAMPP.keys())
            
        if DATA!={}:
            titIV =0
            titmpp=0
            titStat=0
            self.updateTable(DATA)
            self.Confirmgroupnamechanges()
            self.Confirmgroup2namechanges()
            
    def loadconfigsgui(self,directory):
        
        if directory=='':
            current_path = os.getcwd()
            fname = QFileDialog.getSaveFileName(self, 'load file', current_path,"Text files (*.txt)")[0]
        else:
            fname=os.path.join(directory,'guiconfigs.txt')
        
        with open(fname,'r') as file:
            for line in file:
                if 'checkBox_JVLegend' in line:
                    self.ui.checkBox_JVLegend.setChecked(eval(line.split('\t')[1]))
                elif 'spinBox_JVfontsize' in line:
                    self.ui.spinBox_JVfontsize.setValue(int(line.split('\t')[1]))
                elif 'checkBox_MppLegend' in line:
                    self.ui.checkBox_MppLegend.setChecked(eval(line.split('\t')[1]))
                elif 'spinBox_MppFontsize' in line:
                    self.ui.spinBox_MppFontsize.setValue(int(line.split('\t')[1]))
                elif 'checkBox_BoxPlotBig4' in line:
                    self.ui.checkBox_BoxPlotBig4_2.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_BoxPlotRevForw' in line:
                    self.ui.checkBox_BoxPlotRevForw_2.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_BoxPlotBoxPlot' in line:
                    self.ui.checkBox_BoxPlotBoxPlot_2.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_BoxPlotAftermpp' in line:
                    self.ui.checkBox_BoxPlotAftermpp_2.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_BoxPlotBkg' in line:
                    self.ui.checkBox_BoxPlotBkg_2.setChecked(eval(line.split('\t')[1]))
                elif 'spinBox_markerSize' in line:
                    self.ui.spinBox_markerSize_2.setValue(int(line.split('\t')[1]))
                elif 'spinBox_BoxPlotFontsize' in line:
                    self.ui.spinBox_BoxPlotFontsize_2.setValue(int(line.split('\t')[1]))
                elif 'spinBox_BoxPlotRotation' in line:
                    self.ui.spinBox_BoxPlotRotation_2.setValue(int(line.split('\t')[1]))
                elif 'comboBox_BoxPlotParam' in line:
                    index = self.ui.comboBox_BoxPlotParam_2.findText(line.split('\t')[1][:-1], QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.ui.comboBox_BoxPlotParam_2.setCurrentIndex(index)
                elif 'checkBox_PVPTimeBig4' in line:
                    self.ui.checkBox_PVPTimeBig4.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_PVPTimeLine' in line:
                    self.ui.checkBox_PVPTimeLine.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_PVPTimeRelativeTime' in line:
                    self.ui.checkBox_PVPTimeRelativeTime.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_PVPTimeNormal' in line:
                    self.ui.checkBox_PVPTimeNormal.setChecked(eval(line.split('\t')[1]))
                elif 'spinBox_PVPTimeNormalPoint' in line:
                    self.ui.spinBox_PVPTimeNormalPoint.setValue(float(line.split('\t')[1]))
                elif 'comboBox_PVPTimeParam' in line:
                    index = self.ui.comboBox_PVPTimeParam.findText(line.split('\t')[1][:-1], QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.ui.comboBox_PVPTimeParam.setCurrentIndex(index)
                elif 'checkBox_bestofRevFor' in line:
                    self.ui.checkBox_bestofRevFor.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_BestEffPixDay' in line:
                    self.ui.checkBox_BestEffPixDay.setChecked(eval(line.split('\t')[1]))
                elif 'comboBox_PVPx' in line:
                    index = self.ui.comboBox_PVPx.findText(line.split('\t')[1][:-1], QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.ui.comboBox_PVPx.setCurrentIndex(index)
                elif 'comboBox_PVPy' in line:
                    index = self.ui.comboBox_PVPy.findText(line.split('\t')[1][:-1], QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.ui.comboBox_PVPy.setCurrentIndex(index)
                elif 'checkBox_Histxscale' in line:
                    self.ui.checkBox_Histxscale.setChecked(eval(line.split('\t')[1]))
                elif 'spinBox_HistoBins' in line:
                    self.ui.spinBox_HistoBins.setValue(int(line.split('\t')[1]))
                elif 'spinBox_HistxscaleMin' in line:
                    self.ui.spinBox_HistxscaleMin.setValue(int(line.split('\t')[1]))
                elif 'spinBox_HistxscaleMax' in line:
                    self.ui.spinBox_HistxscaleMax.setValue(int(line.split('\t')[1]))
                elif 'comboBox_HistoType' in line:
                    index = self.ui.comboBox_HistoType.findText(line.split('\t')[1][:-1], QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.ui.comboBox_HistoType.setCurrentIndex(index)
                elif 'comboBox_HistoScanDirect' in line:
                    index = self.ui.comboBox_HistoScanDirect.findText(line.split('\t')[1][:-1], QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.ui.comboBox_HistoScanDirect.setCurrentIndex(index)
                elif 'comboBox_HistoParam' in line:
                    index = self.ui.comboBox_HistoParam.findText(line.split('\t')[1][:-1], QtCore.Qt.MatchFixedString)
                    if index >= 0:
                         self.ui.comboBox_HistoParam.setCurrentIndex(index)
                elif 'checkBox_AAxlsxsummary' in line:
                    self.ui.checkBox_AAxlsxsummary.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_AAstatgraphs' in line:
                    self.ui.checkBox_AAstatgraphs.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_AAivgraphs' in line:
                    self.ui.checkBox_AAivgraphs.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_AAmppgraphs' in line:
                    self.ui.checkBox_AAmppgraphs.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_AAtxtfiles' in line:
                    self.ui.checkBox_AAtxtfiles.setChecked(eval(line.split('\t')[1]))
                elif 'checkBox_applygroupname' in line:
                    self.ui.checkBox_applygroupname.setChecked(eval(line.split('\t')[1]))
        
    def saveconfigsgui(self,directory):
        
        if directory=='':
            current_path = os.getcwd()
            fname = QFileDialog.getSaveFileName(self, 'Save file', current_path,"Text files (*.txt)")[0]
        else:
            fname=os.path.join(directory,'guiconfigs.txt')
            # print(fname)
        
        with open(fname,'w') as file:
            text='checkBox_JVLegend\t'+ str(self.ui.checkBox_JVLegend.isChecked())+'\n'+\
                'spinBox_JVfontsize\t'+ str(self.ui.spinBox_JVfontsize.value())+'\n'+\
                'checkBox_MppLegend\t'+ str(self.ui.checkBox_MppLegend.isChecked())+'\n'+\
                'spinBox_MppFontsize\t'+ str(self.ui.spinBox_MppFontsize.value())+'\n'+\
                'checkBox_BoxPlotBig4\t'+ str(self.ui.checkBox_BoxPlotBig4_2.isChecked())+'\n'+\
                'checkBox_BoxPlotRevForw\t'+ str(self.ui.checkBox_BoxPlotRevForw_2.isChecked())+'\n'+\
                'checkBox_BoxPlotBoxPlot\t'+ str(self.ui.checkBox_BoxPlotBoxPlot_2.isChecked())+'\n'+\
                'checkBox_BoxPlotAftermpp\t'+ str(self.ui.checkBox_BoxPlotAftermpp_2.isChecked())+'\n'+\
                'checkBox_BoxPlotBkg\t'+ str(self.ui.checkBox_BoxPlotBkg_2.isChecked())+'\n'+\
                'spinBox_markerSize\t'+ str(self.ui.spinBox_markerSize_2.value())+'\n'+\
                'spinBox_BoxPlotFontsize\t'+ str(self.ui.spinBox_BoxPlotFontsize_2.value())+'\n'+\
                'spinBox_BoxPlotRotation\t'+ str(self.ui.spinBox_BoxPlotRotation_2.value())+'\n'+\
                'comboBox_BoxPlotParam\t'+ str(self.ui.comboBox_BoxPlotParam_2.currentText())+'\n'+\
                'checkBox_PVPTimeBig4\t'+ str(self.ui.checkBox_PVPTimeBig4.isChecked())+'\n'+\
                'checkBox_PVPTimeLine\t'+ str(self.ui.checkBox_PVPTimeLine.isChecked())+'\n'+\
                'checkBox_PVPTimeRelativeTime\t'+ str(self.ui.checkBox_PVPTimeRelativeTime.isChecked())+'\n'+\
                'checkBox_PVPTimeNormal\t'+ str(self.ui.checkBox_PVPTimeNormal.isChecked())+'\n'+\
                'spinBox_PVPTimeNormalPoint\t'+ str(self.ui.spinBox_PVPTimeNormalPoint.value())+'\n'+\
                'comboBox_PVPTimeParam\t'+ str(self.ui.comboBox_PVPTimeParam.currentText())+'\n'+\
                'checkBox_bestofRevFor\t'+ str(self.ui.checkBox_bestofRevFor.isChecked())+'\n'+\
                'checkBox_BestEffPixDay\t'+ str(self.ui.checkBox_BestEffPixDay.isChecked())+'\n'+\
                'comboBox_PVPx\t'+ str(self.ui.comboBox_PVPx.currentText())+'\n'+\
                'comboBox_PVPy\t'+ str(self.ui.comboBox_PVPy.currentText())+'\n'+\
                'checkBox_Histxscale\t'+ str(self.ui.checkBox_Histxscale.isChecked())+'\n'+\
                'spinBox_HistoBins\t'+ str(self.ui.spinBox_HistoBins.value())+'\n'+\
                'spinBox_HistxscaleMin\t'+ str(self.ui.spinBox_HistxscaleMin.value())+'\n'+\
                'spinBox_HistxscaleMax\t'+ str(self.ui.spinBox_HistxscaleMax.value())+'\n'+\
                'comboBox_HistoType\t'+ str(self.ui.comboBox_HistoType.currentText())+'\n'+\
                'comboBox_HistoScanDirect\t'+ str(self.ui.comboBox_HistoScanDirect.currentText())+'\n'+\
                'comboBox_HistoParam\t'+ str(self.ui.comboBox_HistoParam.currentText())+'\n'+\
                'checkBox_AAxlsxsummary\t'+ str(self.ui.checkBox_AAxlsxsummary.isChecked())+'\n'+\
                'checkBox_AAstatgraphs\t'+ str(self.ui.checkBox_AAstatgraphs.isChecked())+'\n'+\
                'checkBox_AAivgraphs\t'+ str(self.ui.checkBox_AAivgraphs.isChecked())+'\n'+\
                'checkBox_AAmppgraphs\t'+ str(self.ui.checkBox_AAmppgraphs.isChecked())+'\n'+\
                'checkBox_AAtxtfiles\t'+ str(self.ui.checkBox_AAtxtfiles.isChecked())+'\n'+\
                'checkBox_applygroupname\t'+ str(self.ui.checkBox_applygroupname.isChecked())
                
            file.write(text)
            
#%%#############
            
    def ExportAutoAnalysis(self):
        global DATA
        global DATAdark
        global DATAFV
        global DATAMPP
        global samplesgroups

        current_path = os.getcwd()
        folderName=QFileDialog.getExistingDirectory(self, 'Select directory')
        # folderName = filedialog.askdirectory(title = "choose a folder to export the auto-analysis results", initialdir=os.path.dirname(current_path))
        os.chdir(folderName)
        
        DATAx=[DATA[key] for key in DATA.keys()]
        DATAMPPx=[DATAMPP[key] for key in DATAMPP.keys()]
        # DATAdarkx=[DATAdark[key] for key in DATAdark.keys()]
        sorted_datajv = sorted(DATAx, key=itemgetter('DepID')) 
        sorted_datampp = sorted(DATAMPPx, key=itemgetter('DepID')) 
        sorted_datadark = sorted(DATAdark, key=itemgetter('DepID'))
        
        # print('depID: ',sorted_datajv[0]['DepID'])
        # print('SampleName: ',sorted_datajv[0]['SampleName'])
        # print(sorted_datajv[0]['SampleNameID'])

        QMessageBox.information(self, 'Information',
                                AA(DATAx,DATAMPPx,DATAdark,sorted_datajv,sorted_datampp,sorted_datadark,self))


# class Thread_AA(QThread):
    
#     finished = pyqtSignal()
#     change_value = pyqtSignal(int)
#     def __init__(self, DATAx,DATAMPPx,DATAdarkx,sorted_datajv,sorted_datampp,sorted_datadark, parent=None):
#         QThread.__init__(self, parent)
#         self.DATAx=DATAx
#         self.DATAMPPx=DATAMPPx
#         self.DATAdarkx=DATAdarkx
#         self.sorted_datajv=sorted_datajv
#         self.sorted_datampp=sorted_datampp
#         self.sorted_datadark=sorted_datadark
        
#     def run(self):
#         # global DATA, DATAdark, colorstylelist
#         # global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
#         print('threadstart')
#         DATAx=self.DATAx
#         DATAMPP=self.DATAMPPx
#         DATAdark=self.DATAdarkx
#         sorted_datajv=self.sorted_datajv
#         sorted_datampp=self.sorted_datampp
#         sorted_datadark=self.sorted_datadark
#         DATAbysubstrate=[] 
#         DATAmppbysubstrate=[]
#         DATAdarkbysubstrate=[] 
#         bestEff=[]
#         bestvocff =[]

def AA(DATAx,DATAMPPx,DATAdarkx,sorted_datajv,sorted_datampp,sorted_datadark,window):
    
    DATAx=copy.deepcopy(DATAx)
    DATAMPP=DATAMPPx
    DATAdark=DATAdarkx
    sorted_datajv=sorted_datajv
    sorted_datampp=sorted_datampp
    sorted_datadark=sorted_datadark
    DATAbysubstrate=[] 
    DATAmppbysubstrate=[]
    DATAdarkbysubstrate=[] 
    bestEff=[]
    bestvocff =[]
    
    batchname=DATAx[0]["batchname"]
    
    # plt.show(block=False)
    
    for key, group in groupby(sorted_datadark, key=lambda x:x['DepID']):
        substratpartdat=[key,list(group)]
        DATAdarkbysubstrate.append(copy.deepcopy(substratpartdat))
        try:
            if window.ui.checkBox_AAtxtfiles.isChecked():
                contenttxtfile=["","",""]
                for item in range(len(substratpartdat[1])):
                    contenttxtfile[0] += "Voltage\t" + "Current density\t" 
                    contenttxtfile[1] += "mV\t" + "mA/cm2\t"
                    contenttxtfile[2] += " \t" + substratpartdat[1][item]["SampleName"] + '\t'
                contenttxtfile[0]=contenttxtfile[0][:-1]+'\n'
                contenttxtfile[1]=contenttxtfile[1][:-1]+'\n'
                contenttxtfile[2]=contenttxtfile[2][:-1]+'\n'
                #find max length of subjv lists => fill the others with blancks
                lengthmax=max([len(substratpartdat[1][x]["IVData"][0]) for x in range(len(substratpartdat[1]))])
                for item in range(len(substratpartdat[1])):
                    while (len(substratpartdat[1][item]["IVData"][0])<lengthmax):
                        substratpartdat[1][item]["IVData"][0].append('')   
                        substratpartdat[1][item]["IVData"][1].append('') 
                #append them all in the content list
                for item0 in range(lengthmax):
                    ligne=""
                    for item in range(len(substratpartdat[1])):
                        ligne += str(substratpartdat[1][item]["IVData"][0][item0]) +'\t' + str(substratpartdat[1][item]["IVData"][1][item0]) +'\t'   
                    ligne=ligne[:-1]+'\n'    
                    contenttxtfile.append(ligne)
                #export it to txt files
                file = open(str(substratpartdat[0]) + '_lowIllum.txt','w', encoding='ISO-8859-1')
                file.writelines("%s" % item for item in contenttxtfile)
                file.close()
        except:
            print("there's an issue with creating JVdark txt files")
        try:
            if window.ui.checkBox_AAivgraphs.isChecked():
                plt.clf()
                plt.close("all")
                fig, axs =plt.subplots(1,2)
                x1=min(DATAdarkbysubstrate[-1][1][0]["IVData"][0])
                x2=max(DATAdarkbysubstrate[-1][1][0]["IVData"][0])
                y1=1.1*max(DATAdarkbysubstrate[-1][1][0]["IVData"][1])
                if min(DATAdarkbysubstrate[-1][1][0]["IVData"][1])>-10:
                    y2=min(DATAdarkbysubstrate[-1][1][0]["IVData"][1])
                else:
                    y2=-10
                for item in range(len(substratpartdat[1])):
                    axs[0].plot(DATAdarkbysubstrate[-1][1][item]["IVData"][0],DATAdarkbysubstrate[-1][1][item]["IVData"][1], label=DATAdarkbysubstrate[-1][1][item]["SampleName"])
                    if min(DATAdarkbysubstrate[-1][1][item]["IVData"][0])<x1:
                        x1=copy.deepcopy(min(DATAdarkbysubstrate[-1][1][item]["IVData"][0]))
                    if max(DATAdarkbysubstrate[-1][1][item]["IVData"][0])>x2:
                        x2=copy.deepcopy(max(DATAdarkbysubstrate[-1][1][item]["IVData"][0]))
                    if max(DATAdarkbysubstrate[-1][1][item]["IVData"][1])>y1:
                        y1=copy.deepcopy(max(DATAdarkbysubstrate[-1][1][item]["IVData"][1]))
                    if min(DATAdarkbysubstrate[-1][1][item]["IVData"][1])<y2 and min(DATAdarkbysubstrate[-1][1][item]["IVData"][1])>-10:
                        y2=copy.deepcopy(max(DATAdarkbysubstrate[-1][1][item]["IVData"][1]))
                    slist=DATAdarkbysubstrate[-1][1][item]
                axs[0].set_title("Low Illumination: "+str(substratpartdat[0]))
                axs[0].set_xlabel('Voltage (mV)')
                axs[0].set_ylabel('Current density (mA/cm'+'\xb2'+')')
                axs[0].axhline(y=0, color='k')
                axs[0].axvline(x=0, color='k')
                axs[0].axis([x1,x2,y2,y1])
                for item in range(len(substratpartdat[1])):
                    axs[1].semilogy(DATAdarkbysubstrate[-1][1][item]["IVData"][0],list(map(abs, DATAdarkbysubstrate[-1][1][item]["IVData"][1])), label=DATAdarkbysubstrate[-1][1][item]["SampleName"])
                axs[1].set_title("logscale")
                axs[1].set_xlabel('Voltage (mV)')
                axs[1].axhline(y=0, color='k')
                axs[1].axvline(x=0, color='k')
                box = axs[0].get_position()
                axs[0].set_position([box.x0, box.y0, box.width, box.height])
                box = axs[1].get_position()
                axs[1].set_position([box.x0, box.y0, box.width, box.height])
                lgd=axs[1].legend(loc='center left', bbox_to_anchor=(1, 0.5), ncol=1)
                #axs[1].legend(bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.2)
                plt.savefig(str(substratpartdat[0])+'_lowIllum.png',dpi=300, bbox_extra_artists=(lgd,),bbox_inches="tight")
                plt.close(fig) 
                plt.close('all')
                plt.clf()
        except:
            print("there's an issue with creating JV lowillum graphs")
    plt.clf()
    try:
        for key, group in groupby(sorted_datampp, key=lambda x:x['DepID']):
            substratpartdat=[key,list(group)]
            DATAmppbysubstrate.append(copy.deepcopy(substratpartdat))
            for item0 in range(len(substratpartdat[1])):
                if window.ui.checkBox_AAtxtfiles.isChecked():
                    contenttxtfile=["Voltage\tCurrent density\tTime\tPmpp\tTime\tVstep\n","V\tmA/cm2\ts\tW/m2\ts\tV\n"]
                    for item in range(len(substratpartdat[1][item0]["MppData"][0])):
                        contenttxtfile.append(str(substratpartdat[1][item0]["MppData"][0][item])+"\t"+str(substratpartdat[1][item0]["MppData"][1][item])+"\t"+str(substratpartdat[1][item0]["MppData"][2][item])+"\t"+str(substratpartdat[1][item0]["MppData"][3][item])+"\t"+str(substratpartdat[1][item0]["MppData"][2][item])+"\t"+str(substratpartdat[1][item0]["MppData"][4][item])+"\n")
                    #export to txt files
                    file = open(str(substratpartdat[1][item0]["SampleName"]) + '_Pmpp.txt','w', encoding='ISO-8859-1')
                    file.writelines("%s" % item for item in contenttxtfile)
                    file.close()
                #export figures
                if window.ui.checkBox_AAmppgraphs.isChecked():
                    plt.plot(substratpartdat[1][item0]["MppData"][2],substratpartdat[1][item0]["MppData"][3])
                    plt.xlabel('Time (s)')
                    plt.ylabel('Power (mW/cm'+'\xb2'+')')        
                    plt.savefig(str(substratpartdat[1][item0]["SampleName"]) + '_Pmpp.png',dpi=300)
                plt.close('all')
                plt.clf()
    except:
        print("there's an issue with creating pmpp txt files")
    
    # print('start substrate graph')
    for key, group in groupby(sorted_datajv, key=lambda x:x['DepID']):
        substratpartdat=[key,list(group)]
        DATAbysubstrate.append(copy.deepcopy(substratpartdat))
        if window.ui.checkBox_AAtxtfiles.isChecked():
            contenttxtfile=["","",""]
            for item in range(len(substratpartdat[1])):
                contenttxtfile[0] += "Voltage\t" + "Current density\t" 
                contenttxtfile[1] += "mV\t" + "mA/cm2\t"
                contenttxtfile[2] += " \t" + substratpartdat[1][item]["SampleName"] + '\t'
            contenttxtfile[0]=contenttxtfile[0][:-1]+'\n'
            contenttxtfile[1]=contenttxtfile[1][:-1]+'\n'
            contenttxtfile[2]=contenttxtfile[2][:-1]+'\n'
            #print(contenttxtfile)  
            #find max length of subjv lists => fill the others with blancks
            lengthmax=max([len(substratpartdat[1][x]["IVData"][0]) for x in range(len(substratpartdat[1]))])
            for item in range(len(substratpartdat[1])):
                while (len(substratpartdat[1][item]["IVData"][0])<lengthmax):
                    substratpartdat[1][item]["IVData"][0].append('')   
                    substratpartdat[1][item]["IVData"][1].append('') 
            #append them all in the content list
            for item0 in range(lengthmax):
                ligne=""
                for item in range(len(substratpartdat[1])):
                    ligne += str(substratpartdat[1][item]["IVData"][0][item0]) +'\t' + str(substratpartdat[1][item]["IVData"][1][item0]) +'\t'   
                ligne=ligne[:-1]+'\n'    
                contenttxtfile.append(ligne)
            #export it to txt files
            file = open(str(substratpartdat[0]) + '.txt','w', encoding='ISO-8859-1')
            file.writelines("%s" % item for item in contenttxtfile)
            file.close()
        #graphs by substrate with JV table, separate graph and table production, then reassemble to export...
        plt.clf()
        if window.ui.checkBox_AAivgraphs.isChecked() or window.ui.checkBox_AAstatgraphs.isChecked() or window.ui.checkBox_AAmppgraphs.isChecked():
            collabel=("Voc","Jsc","FF","Eff","Roc","Rsc","Vstart","Vend","CellSurface")
            nrows, ncols = len(substratpartdat[1])+1, len(collabel)
            hcell, wcell = 0.3, 1.
            hpad, wpad = 0, 0 
            fig2=plt.figure(figsize=(ncols*wcell+wpad, nrows*hcell+hpad))
            ax2 = fig2.add_subplot(111)
            
            fig1=plt.figure()
            ax3 = fig1.add_subplot(111)
            
            item=0
            while item < len(DATAbysubstrate[-1][1]):
                try:
                    x1=min(DATAbysubstrate[-1][1][item]["IVData"][0])
                    # print(x1)
                    x2=max(DATAbysubstrate[-1][1][item]["IVData"][0])
                    y1=1.1*max(DATAbysubstrate[-1][1][item]["IVData"][1])
                    if min(DATAbysubstrate[-1][1][item]["IVData"][1])>-10:
                        y2=min(DATAbysubstrate[-1][1][item]["IVData"][1])
                    else:
                        y2=-10
                    break
                except TypeError:
                    item+=1
                    
            tabledata=[]
            rowlabel=[]
            for item in range(len(substratpartdat[1])):
                # print(item)
                # print(DATAbysubstrate[-1][1][item]["IVData"][0])
                try:
                    if min(DATAbysubstrate[-1][1][item]["IVData"][0])<x1:
                        # x1=copy.deepcopy(min(DATAbysubstrate[-1][1][item]["IVData"][0]))
                        x1=min(DATAbysubstrate[-1][1][item]["IVData"][0])
                    if max(DATAbysubstrate[-1][1][item]["IVData"][0])>x2:
                        x2=copy.deepcopy(max(DATAbysubstrate[-1][1][item]["IVData"][0]))
                    if max(DATAbysubstrate[-1][1][item]["IVData"][1])>y1:
                        y1=copy.deepcopy(1.1*max(DATAbysubstrate[-1][1][item]["IVData"][1]))
                    if min(DATAbysubstrate[-1][1][item]["IVData"][1])<y2 and min(DATAbysubstrate[-1][1][item]["IVData"][1])>-10:
                        y2=copy.deepcopy(min(DATAbysubstrate[-1][1][item]["IVData"][1]))
                    ax3.plot(DATAbysubstrate[-1][1][item]["IVData"][0],DATAbysubstrate[-1][1][item]["IVData"][1], label=DATAbysubstrate[-1][1][item]["SampleName"])
                    slist=DATAbysubstrate[-1][1][item]
                    rowlabel.append(slist["SampleName"])
                    tabledata.append(['%.f' % float(slist["Voc"]),'%.2f' % float(slist["Jsc"]),'%.2f' % float(slist["FF"]),'%.2f' % float(slist["Eff"]),'%.2f' % float(slist["Roc"]),'%.2f' % float(slist["Rsc"]),'%.2f' % float(slist["Vstart"]),'%.2f' % float(slist["Vend"]),'%.2f' % float(slist["CellSurface"])])
                except:
                    print('exception <, ',DATAbysubstrate[-1][1][item]["SampleName"])
                    pass
                
            ax3.set_title(str(substratpartdat[0]))
            ax3.set_xlabel('Voltage (mV)')
            ax3.set_ylabel('Current density (mA/cm'+'\xb2'+')')
            ax3.axhline(y=0, color='k')
            ax3.axvline(x=0, color='k')
            ax3.axis([x1,x2,y2,y1])
            
            if window.ui.checkBox_AAivgraphs.isChecked():
                rowlabel=tuple(rowlabel)
                the_table = ax2.table(cellText=tabledata,colLabels=collabel,rowLabels=rowlabel,loc='center')
                the_table.set_fontsize(18)
                ax2.axis('off')
                fig2.savefig(str(substratpartdat[0])+'_table.png',dpi=200,bbox_inches="tight")
                handles, labels = ax3.get_legend_handles_labels()
                lgd = ax3.legend(handles, labels, loc='center left', bbox_to_anchor=(1, 0.5))
                fig1.savefig(str(substratpartdat[0])+'.png',dpi=200, bbox_extra_artists=(lgd,),bbox_inches="tight")
            
                images = list(map(ImageTk.open, [str(substratpartdat[0])+'.png',str(substratpartdat[0])+'_table.png']))
                widths, heights = zip(*(i.size for i in images))
                total_width = max(widths)
                max_height = sum(heights)
                new_im = ImageTk.new('RGB', (total_width, max_height), (255, 255, 255))
                new_im.paste(im=images[0],box=(0,0))
                new_im.paste(im=images[1],box=(0,heights[0]))
                new_im.save(str(substratpartdat[0])+'.png')
                
                os.remove(str(substratpartdat[0])+'_table.png')
            plt.close(fig2)
            plt.close(fig1)
            plt.close('all')
            plt.clf()
            
            # print('starting best FR')
            #graph best FR of this substrate
            best = sorted(DATAbysubstrate[-1][1], key=itemgetter('VocFF'), reverse=True)
            item=0
            while item<len(best):
                if float(best[item]["FF"])>10 and float(best[item]["Jsc"])<40 and 'SunsVoc' not in best[item]["SampleName"]:
                    bestvocff.append(copy.deepcopy(best[item]))
                    break
                else:
                    item+=1
            best = sorted(DATAbysubstrate[-1][1], key=itemgetter('Eff'), reverse=True)
            item=0
            while item<len(best):
                if float(best[item]["FF"])>10 and float(best[item]["Jsc"])<40 and 'SunsVoc' not in best[item]["SampleName"]:
                    fig=plt.figure()
                    ax=fig.add_subplot(111)
                    bestEff.append(copy.deepcopy(best[item]))
                    if best[item]["ScanDirection"]=="Reverse":
                        ax.plot(best[item]["IVData"][0],best[item]["IVData"][1],"r", label=best[item]["SampleName"])
                        text = best[item]["ScanDirection"]+"; "+"Rsc: " + '%.f' % float(best[item]["Rsc"]) +" Ohm*cm2; "+"; "+"Roc: " + '%.f' % float(best[item]["Roc"]) +" Ohm*cm2; "+"; "+"\nVoc: " + '%.f' % float(best[item]["Voc"]) +" mV; " + "Jsc: " + '%.2f' % float(best[item]["Jsc"]) +" mA/cm2; " +"FF: " + '%.2f' % float(best[item]["FF"]) +" %; " +"Eff: " + '%.2f' % float(best[item]["Eff"]) +" %"
                        ax.set_title('Best:'+ best[item]["SampleName"]+"\n"+text, fontsize = 10, color="r")
                    elif best[item]["ScanDirection"]=="Forward":
                        ax.plot(best[item]["IVData"][0],best[item]["IVData"][1],"k", label=best[item]["SampleName"]) 
                        text = best[item]["ScanDirection"]+"; "+"Rsc: " + '%.f' % float(best[item]["Rsc"]) +" Ohm*cm2; "+"; "+"Roc: " + '%.f' % float(best[item]["Roc"]) +" Ohm*cm2; "+"; "+"\nVoc: " + '%.f' % float(best[item]["Voc"]) +" mV; " + "Jsc: " + '%.2f' % float(best[item]["Jsc"]) +" mA/cm2; " +"FF: " + '%.2f' % float(best[item]["FF"]) +" %; " +"Eff: " + '%.2f' % float(best[item]["Eff"]) +" %"
                        ax.set_title('Best:'+ best[item]["SampleName"]+"\n"+text, fontsize = 10, color="k")
                    pos=0
                    if best[item]["ScanDirection"]=="Reverse":
                        for item0 in range(item+1,len(best),1):
                            if best[item0]["ScanDirection"]=="Forward" and best[item]["Cellletter"]==best[item0]["Cellletter"]:
                                #other=best[item0]
                                pos=item0
                                ax.plot(best[pos]["IVData"][0],best[pos]["IVData"][1],"k", label=best[pos]["SampleName"])
                                ax.set_title('Best:'+ best[item]["SampleName"]+"-"+best[pos]["SampleName"]+"\n"+text, fontsize = 10, color="r")
                                break
                        
                    elif best[item]["ScanDirection"]=="Forward":
                        for item0 in range(item+1,len(best),1):
                            if best[item0]["ScanDirection"]=="Reverse" and best[item]["Cellletter"]==best[item0]["Cellletter"]:
                                #other=best[item0]
                                pos=item0
                                ax.plot(best[pos]["IVData"][0],best[pos]["IVData"][1],"r", label=best[pos]["SampleName"])
                                ax.set_title('Best:'+ best[item]["SampleName"]+"-"+best[pos]["SampleName"]+"\n"+text, fontsize = 10, color="k")
                                break
                    for item0 in range(len(DATAx)):
                        if DATAx[item0]["DepID"]==best[item]["DepID"] and DATAx[item0]["Cellletter"]==best[item]["Cellletter"] and DATAx[item0]["Illumination"]=="Dark":
                            ax.plot(DATAx[item0]["IVData"][0],DATAx[item0]["IVData"][1],color='gray',linestyle='dashed', label=DATAx[item0]["SampleName"])
                            break
                    
                    ax.set_xlabel('Voltage (mV)')
                    ax.set_ylabel('Current density (mA/cm'+'\xb2'+')')
                    ax.axhline(y=0, color='k')
                    ax.axvline(x=0, color='k')
                    
                    x1=min(best[item]["IVData"][0][0],best[pos]["IVData"][0][0])
                    x2=max(float(best[item]["IVData"][0][-1]),float(best[pos]["IVData"][0][-1]))
                    y1=1.1*max(best[item]["IVData"][1]+best[pos]["IVData"][1])
                    if min(best[item]["IVData"][1]+best[pos]["IVData"][1])<-10:
                        y2=-10
                    else:
                        y2=min(best[item]["IVData"][1]+best[pos]["IVData"][1])
                    ax.axis([x1,x2,y2,y1])
                    if window.ui.checkBox_AAivgraphs.isChecked():
                        handles, labels = ax.get_legend_handles_labels()
                        lgd = ax.legend(handles, labels, loc='center left', bbox_to_anchor=(1, 0.5))
                        fig.savefig(str(substratpartdat[0])+'_BestRevForw.png',dpi=300, bbox_extra_artists=(lgd,),bbox_inches="tight")
                    plt.close('all')
                    plt.clf()
                    break
                else:
                    item+=1 
            plt.close('all')
            plt.clf()
            #specific power graph
            if window.ui.checkBox_AAmppgraphs.isChecked(): 
                for item in range(len(DATAmppbysubstrate)):
                    if substratpartdat[0]==DATAmppbysubstrate[item][0]:
                        for item0 in range(len(DATAmppbysubstrate[item][1])):
                            fig=plt.figure()
                            ax=fig.add_subplot(111)
                            ax.plot([],[],label="Initial scans",color="White") 
                            checkaftermpp=1
                            for item1 in range(len(DATAbysubstrate[-1][1])):
                                if DATAmppbysubstrate[item][1][item0]["Cellletter"]==DATAbysubstrate[-1][1][item1]["Cellletter"] and DATAbysubstrate[-1][1][item1]["Illumination"]=="Light":
                                    if DATAbysubstrate[-1][1][item1]["aftermpp"] and checkaftermpp:
                                        ax.plot([],[],label="After mpp",color="White") 
                                        checkaftermpp=0
                                        ax.plot(DATAbysubstrate[-1][1][item1]["IVData"][0],[-a*b for a,b in zip(DATAbysubstrate[-1][1][item1]["IVData"][0],DATAbysubstrate[-1][1][item1]["IVData"][1])],label=DATAbysubstrate[-1][1][item1]["SampleName"])   
                                    else:
                                        ax.plot(DATAbysubstrate[-1][1][item1]["IVData"][0],[-a*b for a,b in zip(DATAbysubstrate[-1][1][item1]["IVData"][0],DATAbysubstrate[-1][1][item1]["IVData"][1])],label=DATAbysubstrate[-1][1][item1]["SampleName"])
                            ax.plot([abs(a) for a in DATAmppbysubstrate[item][1][item0]["MppData"][0]],DATAmppbysubstrate[item][1][item0]["MppData"][3])
                            ax.set_xlabel('Voltage (mV)')
                            ax.set_ylabel('Specific power (mW/cm$^2$)')
                            ax.axhline(y=0, color='k')
                            ax.axvline(x=0, color='k')
                            ax.set_xlim(left=0)
                            ax.set_ylim(bottom=0)
                            ax.legend()
                            fig.savefig(DATAmppbysubstrate[item][1][item0]["SampleName"]+'_specpower.png',dpi=300,bbox_inches="tight")
                            plt.close("all")
                            plt.clf()
                        break
    plt.close("all")
    plt.clf()
    print('start besteff stat graph')
    if DATAx[0]["Setup"]=="TFIV" or DATAx[0]["Setup"]=="SSIgorC215":
#            try:        
        if window.ui.checkBox_AAstatgraphs.isChecked():
            #graph with all best efficient cells from all substrates
            fig=plt.figure()
            ax=fig.add_subplot(111)
            bestEff2=[item for item in bestEff if item["Illumination"]=="Light"]
            bestEffsorted = sorted(bestEff2, key=itemgetter('Eff'), reverse=True) 
            tabledata=[]
            rowlabel=[]
            minJscfind=[]
            maxcurrentfind=[]
            minVfind=[]
            maxVfind=[]
            for item in range(len(bestEffsorted)):
                ax.plot(bestEffsorted[item]["IVData"][0],bestEffsorted[item]["IVData"][1], label=bestEffsorted[item]["SampleName"]) 
                rowlabel.append(bestEffsorted[item]["SampleName"])
                tabledata.append(['%.f' % float(bestEffsorted[item]["Voc"]),'%.2f' % float(bestEffsorted[item]["Jsc"]),'%.2f' % float(bestEffsorted[item]["FF"]),'%.2f' % float(bestEffsorted[item]["Eff"]),'%.2f' % float(bestEffsorted[item]["Roc"]),'%.2f' % float(bestEffsorted[item]["Rsc"]),'%.2f' % float(bestEffsorted[item]["Vstart"]),'%.2f' % float(bestEffsorted[item]["Vend"]),'%.2f' % float(bestEffsorted[item]["CellSurface"])])
                minJscfind.append(min(bestEffsorted[item]["IVData"][1]))
                maxcurrentfind.append(max(bestEffsorted[item]["IVData"][1]))
                minVfind.append(min(bestEffsorted[item]["IVData"][0]))
                maxVfind.append(max(bestEffsorted[item]["IVData"][0]))
            ax.set_xlabel('Voltage (mV)')
            ax.set_ylabel('Current density (mA/cm'+'\xb2'+')')
            ax.axhline(y=0, color='k')
            ax.axvline(x=0, color='k')
            x1=min(minVfind)
            x2=max(maxVfind)
            y1=1.1*max(minJscfind)
            if min(maxcurrentfind)<-10:
                y2=-10
            else:
                y2=min(maxcurrentfind)
            ax.axis([x1,x2,y2,y1])
            handles, labels = ax.get_legend_handles_labels()
            lgd = ax.legend(handles, labels, loc='center left', bbox_to_anchor=(1, 0.5))
            fig.savefig(batchname+'_MostEfficientCells.png',dpi=300, bbox_extra_artists=(lgd,),bbox_inches="tight")
            plt.close()
            collabel=("Voc","Jsc","FF","Eff","Roc","Rsc","Vstart","Vend","CellSurface")
            nrows, ncols = len(bestEffsorted)+1, len(collabel)
            hcell, wcell = 0.3, 1.
            hpad, wpad = 0, 0 
            fig=plt.figure(figsize=(ncols*wcell+wpad, nrows*hcell+hpad))
            ax = fig.add_subplot(111)
            rowlabel=tuple(rowlabel)
            the_table = ax.table(cellText=tabledata,colLabels=collabel,rowLabels=rowlabel,loc='center')
            the_table.set_fontsize(18)
            ax.axis('off')
            fig.savefig('MostEfficientCellstable.png',dpi=300,bbox_inches="tight")
            plt.close("all")
            images = list(map(ImageTk.open, [batchname+'_MostEfficientCells.png','MostEfficientCellstable.png']))
            widths, heights = zip(*(i.size for i in images))
            total_width = max(widths)
            max_height = sum(heights)
            new_im = ImageTk.new('RGB', (total_width, max_height), (255, 255, 255))
            new_im.paste(im=images[0],box=(0,0))
            new_im.paste(im=images[1],box=(0,heights[0]))
            new_im.save(batchname+'_MostEfficientCells.png')
            os.remove('MostEfficientCellstable.png')
            plt.close()
            plt.clf()
            #graph with all best voc*FF cells from all substrates  
            fig=plt.figure()
            ax=fig.add_subplot(111)
            bestvocff2=[item for item in bestvocff if item["Illumination"]=="Light"]
            bestvocffsorted = sorted(bestvocff2, key=itemgetter('VocFF'), reverse=True) 
            tabledata=[]
            rowlabel=[]
            minJscfind=[]
            maxcurrentfind=[]
            minVfind=[]
            maxVfind=[]
            for item in range(len(bestvocffsorted)):
                x=bestvocffsorted[item]["IVData"][0]
                y=bestvocffsorted[item]["IVData"][1]
                ax.plot(x,y, label=bestvocffsorted[item]["SampleName"]) 
                rowlabel.append(bestvocffsorted[item]["SampleName"])
                tabledata.append(['%.f' % float(bestvocffsorted[item]["Voc"]),'%.2f' % float(bestvocffsorted[item]["Jsc"]),'%.2f' % float(bestvocffsorted[item]["FF"]),'%.2f' % float(bestvocffsorted[item]["Eff"]),'%.2f' % float(bestvocffsorted[item]["Roc"]),'%.2f' % float(bestvocffsorted[item]["Rsc"]),'%.2f' % float(bestvocffsorted[item]["Vstart"]),'%.2f' % float(bestvocffsorted[item]["Vend"]),'%.2f' % float(bestvocffsorted[item]["CellSurface"])])
                minJscfind.append(min(bestvocffsorted[item]["IVData"][1]))
                maxcurrentfind.append(max(bestvocffsorted[item]["IVData"][1]))
                minVfind.append(min(bestvocffsorted[item]["IVData"][0]))
                maxVfind.append(max(bestvocffsorted[item]["IVData"][0]))
            ax.set_xlabel('Voltage (mV)')
            ax.set_ylabel('Current density (mA/cm'+'\xb2'+')')
            ax.axhline(y=0, color='k')
            ax.axvline(x=0, color='k')
            x1=min(minVfind)
            x2=max(maxVfind)
            y1=1.1*max(minJscfind)
            if min(maxcurrentfind)<-10:
                y2=-10
            else:
                y2=min(maxcurrentfind)
            ax.axis([x1,x2,y2,y1])
            handles, labels = ax.get_legend_handles_labels()
            lgd = ax.legend(handles, labels, loc='center left', bbox_to_anchor=(1, 0.5))
            fig.savefig(batchname+'_bestvocff.png',dpi=300, bbox_extra_artists=(lgd,),bbox_inches="tight")
            plt.close(fig)
            plt.clf()
            collabel=("Voc","Jsc","FF","Eff","Roc","Rsc","Vstart","Vend","CellSurface")
            nrows, ncols = len(bestvocffsorted)+1, len(collabel)
            hcell, wcell = 0.3, 1.
            hpad, wpad = 0, 0 
            fig=plt.figure(figsize=(ncols*wcell+wpad, nrows*hcell+hpad))
            ax = fig.add_subplot(111)
            rowlabel=tuple(rowlabel)
            the_table = ax.table(cellText=tabledata,colLabels=collabel,rowLabels=rowlabel,loc='center')
            the_table.set_fontsize(18)
            ax.axis('off')
            fig.savefig('bestvocfftable.png',dpi=300,bbox_inches="tight")
            plt.close(fig)
            plt.clf()
            images = list(map(ImageTk.open, [batchname+'_bestvocff.png','bestvocfftable.png']))
            widths, heights = zip(*(i.size for i in images))
            total_width = max(widths)
            max_height = sum(heights)
            new_im = ImageTk.new('RGB', (total_width, max_height), (255, 255, 255))
            new_im.paste(im=images[0],box=(0,0))
            new_im.paste(im=images[1],box=(0,heights[0]))
            new_im.save(batchname+'_bestvocff.png')
            os.remove('bestvocfftable.png')
            plt.close("all")
            plt.clf()
#            except:
#                print("there's an issue with creating Bestof graphs")

    plt.clf()    
    plt.close("all")    
    if len(samplesgroups)>1:            
        grouplistdict=[]
        for item in range(len(samplesgroups)):
            groupdict={}
            groupdict["Group"]=samplesgroups[item]
            listofthegroup=[]
            listofthegroupNames=[]
            for item1 in range(len(DATAx)):
                if DATAx[item1]["Group"]==groupdict["Group"] and DATAx[item1]["Illumination"]=="Light":
                    listofthegroup.append(DATAx[item1])
                    listofthegroupNames.append(DATAx[item1]['DepID']+'_'+DATAx[item1]['Cellletter'])
            groupdict["numbCell"]=len(list(set(listofthegroupNames)))
            listofthegroupRev=[]
            listofthegroupFor=[]
            for item1 in range(len(listofthegroup)):
                if listofthegroup[item1]["ScanDirection"]=="Reverse":
                    listofthegroupRev.append(listofthegroup[item1])
                else:
                    listofthegroupFor.append(listofthegroup[item1])
           
            groupdict["RevVoc"]=[x['Voc'] for x in listofthegroupRev if 'Voc' in x]
            groupdict["ForVoc"]=[x['Voc'] for x in listofthegroupFor if 'Voc' in x]
            groupdict["RevJsc"]=[x['Jsc'] for x in listofthegroupRev if 'Jsc' in x]
            groupdict["ForJsc"]=[x['Jsc'] for x in listofthegroupFor if 'Jsc' in x]
            groupdict["RevFF"]=[x['FF'] for x in listofthegroupRev if 'FF' in x]
            groupdict["ForFF"]=[x['FF'] for x in listofthegroupFor if 'FF' in x]
            groupdict["RevEff"]=[x['Eff'] for x in listofthegroupRev if 'Eff' in x]
            groupdict["ForEff"]=[x['Eff'] for x in listofthegroupFor if 'Eff' in x]
            
            grouplistdict.append(groupdict)
    plt.close("all")  
    plt.clf()
    if window.ui.checkBox_AAxlsxsummary.isChecked():   
        workbook = xlsxwriter.Workbook(batchname+'_Summary.xlsx')
        
        LandD=DATAx + DATAdark
        timeLandD =sorted(LandD, key=itemgetter('MeasDayTime')) 
        
        if len(samplesgroups)>1:
#                try:
            worksheet = workbook.add_worksheet("GroupStat")
            summary=[]
            for item in range(len(samplesgroups)):
                ncell=1
                if grouplistdict[item]["ForVoc"]!=[]:
                    lengthofgroup=len(grouplistdict[item]["ForVoc"])
                    summary.append([grouplistdict[item]["Group"],grouplistdict[item]["numbCell"],"Forward",lengthofgroup,sum(grouplistdict[item]["ForVoc"],0.0)/lengthofgroup,np.std(grouplistdict[item]["ForVoc"]),sum(grouplistdict[item]["ForJsc"],0.0)/lengthofgroup,np.std(grouplistdict[item]["ForJsc"]),sum(grouplistdict[item]["ForFF"],0.0)/lengthofgroup,np.std(grouplistdict[item]["ForFF"]),sum(grouplistdict[item]["ForEff"],0.0)/lengthofgroup,np.std(grouplistdict[item]["ForEff"])])
                    ncell=0
                if grouplistdict[item]["RevVoc"]!=[]:  
                    if ncell==0:
                        lengthofgroup=len(grouplistdict[item]["RevVoc"])
                        summary.append([grouplistdict[item]["Group"]," ","Reverse",lengthofgroup,sum(grouplistdict[item]["RevVoc"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevVoc"]),sum(grouplistdict[item]["RevJsc"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevJsc"]),sum(grouplistdict[item]["RevFF"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevFF"]),sum(grouplistdict[item]["RevEff"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevEff"])])
                    else:
                        lengthofgroup=len(grouplistdict[item]["RevVoc"])
                        summary.append([grouplistdict[item]["Group"],grouplistdict[item]["numbCell"],"Reverse",lengthofgroup,sum(grouplistdict[item]["RevVoc"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevVoc"]),sum(grouplistdict[item]["RevJsc"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevJsc"]),sum(grouplistdict[item]["RevFF"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevFF"]),sum(grouplistdict[item]["RevEff"],0.0)/lengthofgroup,np.std(grouplistdict[item]["RevEff"])])

            summary.insert(0, [" ", " "," ", "-", "mV","-","mA/cm2","-","%","-","%","-"])
            summary.insert(0, ["Group","#Cells","Scan Dir.","#ofmeas", "Voc"," ","Jsc"," ","FF"," ","Eff"," "])
            summary.insert(0, [" "," "," "," ", "Avg","StdDev","Avg","StdDev","Avg","StdDev","Avg","StdDev"])
            for item in range(len(summary)):
                for item0 in range(len(summary[item])):
                    worksheet.write(item,item0, str(summary[item][item0]))
            # except:
            #     print("exception: excel summary - groupstat")
    
        if timeLandD!=[]:
            try:
                worksheet = workbook.add_worksheet("AllJVrawdata")
                summary=[]
                for item in range(len(timeLandD)):
                    summary.append([timeLandD[item]["Group"],timeLandD[item]["SampleName"],timeLandD[item]["Cellletter"],timeLandD[item]["MeasDayTime"],timeLandD[item]["CellSurface"],str(timeLandD[item]["Voc"]),str(timeLandD[item]["Jsc"]),str(timeLandD[item]["FF"]),str(timeLandD[item]["Eff"]),str(timeLandD[item]["Pmpp"]),str(timeLandD[item]["Vmpp"]),str(timeLandD[item]["Jmpp"]),str(timeLandD[item]["Roc"]),str(timeLandD[item]["Rsc"]),str(timeLandD[item]["VocFF"]),str(timeLandD[item]["RscJsc"]),str(timeLandD[item]["NbPoints"]),timeLandD[item]["Delay"],timeLandD[item]["IntegTime"],timeLandD[item]["Vstart"],timeLandD[item]["Vend"],timeLandD[item]["Illumination"],timeLandD[item]["ScanDirection"],str('%.2f' % float(timeLandD[item]["ImaxComp"])),timeLandD[item]["Isenserange"],str(timeLandD[item]["AreaJV"]),timeLandD[item]["Operator"],timeLandD[item]["MeasComment"],timeLandD[item]["RefNomCurr"],timeLandD[item]["RefMeasCurr"],str(timeLandD[item]["AirTemp"]),str(timeLandD[item]["ChuckTemp"])])
                summary.insert(0, ["-", "-", "-","-","cm2","mV","mA/cm2","%","%","W/cm2","mV","mA/cm2","Ohm*cm2","Ohm*cm2","-","-","-","s","s","mV","mV","-","-","A","A","-","-","-","mA","mA","DegC","DegC"])
                summary.insert(0, ["Group","Sample Name", "Cell","MeasDayTime","Cell Surface","Voc","Jsc","FF","Eff","Pmpp","Vmpp","Jmpp","Roc","Rsc","VocFF","RscJsc","NbPoints","Delay","IntegTime","Vstart","Vend","Illumination","ScanDirection","ImaxComp","Isenserange","AreaJV","Operator","MeasComment","RefNomCurr","RefMeasCurr","AirTemp","ChuckTemp"])
                for item in range(len(summary)):
                    for item0 in range(len(summary[item])):
                        worksheet.write(item,item0,summary[item][item0])
            except:
                print("exception: excel summary - AllJVrawdata")
        
        if DATAx!=[]:
            try:
                worksheet = workbook.add_worksheet("rawdataLight")
                summary=[]
                for item in range(len(DATAx)):
                    if DATAx[item]["Illumination"]=='Light':
                        summary.append([DATAx[item]["Group"],DATAx[item]["SampleName"],DATAx[item]["Cellletter"],DATAx[item]["MeasDayTime"],DATAx[item]["CellSurface"],str(DATAx[item]["Voc"]),str(DATAx[item]["Jsc"]),str(DATAx[item]["FF"]),str(DATAx[item]["Eff"]),str(DATAx[item]["Pmpp"]),str(DATAx[item]["Vmpp"]),str(DATAx[item]["Jmpp"]),str(DATAx[item]["Roc"]),str(DATAx[item]["Rsc"]),str(DATAx[item]["VocFF"]),str(DATAx[item]["RscJsc"]),str(DATAx[item]["NbPoints"]),str(DATAx[item]["Delay"]),str(DATAx[item]["IntegTime"]),str(DATAx[item]["Vstart"]),str(DATAx[item]["Vend"]),str(DATAx[item]["Illumination"]),str(DATAx[item]["ScanDirection"]),str('%.2f' % float(DATAx[item]["ImaxComp"])),str(DATAx[item]["Isenserange"]),str(DATAx[item]["AreaJV"]),DATAx[item]["Operator"],DATAx[item]["MeasComment"],timeLandD[item]["RefNomCurr"],timeLandD[item]["RefMeasCurr"],str(timeLandD[item]["AirTemp"]),str(timeLandD[item]["ChuckTemp"])])
                summary.insert(0, ["-", "-", "-","-","cm2","mV","mA/cm2","%","%","W/cm2","mV","mA/cm2","Ohm*cm2","Ohm*cm2","-","-","-","s","s","mV","mV","-","-","A","A","-","-","-","mA","mA","DegC","DegC"])
                summary.insert(0, ["Group","Sample Name", "Cell","MeasDayTime","Cell Surface","Voc","Jsc","FF","Eff","Pmpp","Vmpp","Jmpp","Roc","Rsc","VocFF","RscJsc","NbPoints","Delay","IntegTime","Vstart","Vend","Illumination","ScanDirection","ImaxComp","Isenserange","AreaJV","Operator","MeasComment","RefNomCurr","RefMeasCurr","AirTemp","ChuckTemp"])
                for item in range(len(summary)):
                    for item0 in range(len(summary[item])):
                        worksheet.write(item,item0,summary[item][item0])
                worksheet = workbook.add_worksheet("rawdatadark")
                summary=[]
                for item in range(len(DATAx)):
                    if DATAx[item]["Illumination"]=='Dark':
                        summary.append([DATAx[item]["Group"],DATAx[item]["SampleName"],DATAx[item]["Cellletter"],DATAx[item]["MeasDayTime"],DATAx[item]["CellSurface"],str(DATAx[item]["Voc"]),str(DATAx[item]["Jsc"]),str(DATAx[item]["FF"]),str(DATAx[item]["Eff"]),str(DATAx[item]["Pmpp"]),str(DATAx[item]["Vmpp"]),str(DATAx[item]["Jmpp"]),str(DATAx[item]["Roc"]),str(DATAx[item]["Rsc"]),str(DATAx[item]["VocFF"]),str(DATAx[item]["RscJsc"]),str(DATAx[item]["NbPoints"]),str(DATAx[item]["Delay"]),str(DATAx[item]["IntegTime"]),str(DATAx[item]["Vstart"]),str(DATAx[item]["Vend"]),str(DATAx[item]["Illumination"]),str(DATAx[item]["ScanDirection"]),str('%.2f' % float(DATAx[item]["ImaxComp"])),str(DATAx[item]["Isenserange"]),str(DATAx[item]["AreaJV"]),DATAx[item]["Operator"],DATAx[item]["MeasComment"],timeLandD[item]["RefNomCurr"],timeLandD[item]["RefMeasCurr"],str(timeLandD[item]["AirTemp"]),str(timeLandD[item]["ChuckTemp"])])
                summary.insert(0, ["-", "-", "-","-","cm2","mV","mA/cm2","%","%","W/cm2","mV","mA/cm2","Ohm*cm2","Ohm*cm2","-","-","-","s","s","mV","mV","-","-","A","A","-","-","-","mA","mA","DegC","DegC"])
                summary.insert(0, ["Group","Sample Name", "Cell","MeasDayTime","Cell Surface","Voc","Jsc","FF","Eff","Pmpp","Vmpp","Jmpp","Roc","Rsc","VocFF","RscJsc","NbPoints","Delay","IntegTime","Vstart","Vend","Illumination","ScanDirection","ImaxComp","Isenserange","AreaJV","Operator","MeasComment","RefNomCurr","RefMeasCurr","AirTemp","ChuckTemp"])
                for item in range(len(summary)):
                    for item0 in range(len(summary[item])):
                        worksheet.write(item,item0,summary[item][item0])
            except:
                print("exception: excel summary - rawdataLight")
#                if DATAdark!=[]:
#                    worksheet = workbook.add_worksheet("rawdatadark")
#                    summary=[]
#                    for item in range(len(DATAdark)):
#                        summary.append([DATAdark[item]["Group"],DATAdark[item]["SampleName"],DATAdark[item]["Cellletter"],DATAdark[item]["MeasDayTime"],DATAdark[item]["CellSurface"],str(DATAdark[item]["Voc"]),str(DATAdark[item]["Jsc"]),str(DATAdark[item]["FF"]),str(DATAdark[item]["Eff"]),str(DATAdark[item]["Pmpp"]),str(DATAdark[item]["Vmpp"]),str(DATAdark[item]["Jmpp"]),str(DATAdark[item]["Roc"]),str(DATAdark[item]["Rsc"]),str(DATAdark[item]["VocFF"]),str(DATAdark[item]["RscJsc"]),str(DATAdark[item]["NbPoints"]),DATAdark[item]["Delay"],DATAdark[item]["IntegTime"],DATAdark[item]["Vstart"],DATAdark[item]["Vend"],DATAdark[item]["Illumination"],DATAdark[item]["ScanDirection"],str('%.2f' % float(DATAdark[item]["ImaxComp"])),DATAdark[item]["Isenserange"],str(DATAdark[item]["AreaJV"]),DATAdark[item]["Operator"],DATAdark[item]["MeasComment"],timeLandD[item]["RefNomCurr"],timeLandD[item]["RefMeasCurr"],str(timeLandD[item]["AirTemp"]),str(timeLandD[item]["ChuckTemp"])])
#                    summary.insert(0, ["-", "-", "-","cm2","mV","mA/cm2","%","%","W/cm2","mV","mA/cm2","Ohm*cm2","Ohm*cm2","-","-","-","s","s","mV","mV","-","-","A","A","-","-","-","mA","mA","DegC","DegC"])
#                    summary.insert(0, ["Sample Name", "Cell","MeasDayTime","Cell Surface","Voc","Jsc","FF","Eff","Pmpp","Vmpp","Jmpp","Roc","Rsc","VocFF","RscJsc","NbPoints","Delay","IntegTime","Vstart","Vend","Illumination","ScanDirection","ImaxComp","Isenserange","AreaJV","Operator","MeasComment","RefNomCurr","RefMeasCurr","AirTemp","ChuckTemp"])
#                    for item in range(len(summary)):
#                        for item0 in range(len(summary[item])):
#                            worksheet.write(item,item0,summary[item][item0])
                    
        sorted_bestEff= sorted(bestEff, key=itemgetter('Eff'), reverse=True) 
        if sorted_bestEff!=[]:  
            try:
                worksheet = workbook.add_worksheet("besteff")
                summary=[]
                for item in range(len(sorted_bestEff)):
                    summary.append([sorted_bestEff[item]["Group"],sorted_bestEff[item]["SampleName"],sorted_bestEff[item]["Cellletter"],sorted_bestEff[item]["MeasDayTime"],sorted_bestEff[item]["CellSurface"],str(sorted_bestEff[item]["Voc"]),str(sorted_bestEff[item]["Jsc"]),str(sorted_bestEff[item]["FF"]),str(sorted_bestEff[item]["Eff"]),str(sorted_bestEff[item]["Pmpp"]),str(sorted_bestEff[item]["Vmpp"]),str(sorted_bestEff[item]["Jmpp"]),str(sorted_bestEff[item]["Roc"]),str(sorted_bestEff[item]["Rsc"]),str(sorted_bestEff[item]["VocFF"]),str(sorted_bestEff[item]["RscJsc"]),str(sorted_bestEff[item]["NbPoints"]),sorted_bestEff[item]["Delay"],sorted_bestEff[item]["IntegTime"],sorted_bestEff[item]["Vstart"],sorted_bestEff[item]["Vend"],sorted_bestEff[item]["Illumination"],sorted_bestEff[item]["ScanDirection"],str('%.2f' % float(sorted_bestEff[item]["ImaxComp"])),sorted_bestEff[item]["Isenserange"],str(sorted_bestEff[item]["AreaJV"]),sorted_bestEff[item]["Operator"],sorted_bestEff[item]["MeasComment"],timeLandD[item]["RefNomCurr"],timeLandD[item]["RefMeasCurr"],str(timeLandD[item]["AirTemp"]),str(timeLandD[item]["ChuckTemp"])])
                summary.insert(0, ["-", "-", "-", "-","cm2","mV","mA/cm2","%","%","W/cm2","mV","mA/cm2","Ohm*cm2","Ohm*cm2","-","-","-","s","s","mV","mV","-","-","A","A","-","-","-","mA","mA","DegC","DegC"])
                summary.insert(0, ["Group","Sample Name", "Cell","MeasDayTime","Cell Surface","Voc","Jsc","FF","Eff","Pmpp","Vmpp","Jmpp","Roc","Rsc","VocFF","RscJsc","NbPoints","Delay","IntegTime","Vstart","Vend","Illumination","ScanDirection","ImaxComp","Isenserange","AreaJV","Operator","MeasComment","RefNomCurr","RefMeasCurr","AirTemp","ChuckTemp"])
                for item in range(len(summary)):
                    for item0 in range(len(summary[item])):
                        worksheet.write(item,item0,summary[item][item0])
            except:
                print("exception: excel summary - besteff")
        sorted_bestvocff= sorted(bestvocff, key=itemgetter('VocFF'), reverse=True) 
        if sorted_bestvocff!=[]: 
            try:
                worksheet = workbook.add_worksheet("bestvocff")
                summary=[]
                for item in range(len(sorted_bestvocff)):
                    summary.append([sorted_bestvocff[item]["Group"], sorted_bestvocff[item]["SampleName"],sorted_bestvocff[item]["Cellletter"],sorted_bestvocff[item]["MeasDayTime"],sorted_bestvocff[item]["CellSurface"],str(sorted_bestvocff[item]["Voc"]),str(sorted_bestvocff[item]["Jsc"]),str(sorted_bestvocff[item]["FF"]),str(sorted_bestvocff[item]["Eff"]),str(sorted_bestvocff[item]["Pmpp"]),str(sorted_bestvocff[item]["Vmpp"]),str(sorted_bestvocff[item]["Jmpp"]),str(sorted_bestvocff[item]["Roc"]),str(sorted_bestvocff[item]["Rsc"]),str(sorted_bestvocff[item]["VocFF"]),str(sorted_bestvocff[item]["RscJsc"]),str(sorted_bestvocff[item]["NbPoints"]),sorted_bestvocff[item]["Delay"],sorted_bestvocff[item]["IntegTime"],sorted_bestvocff[item]["Vstart"],sorted_bestvocff[item]["Vend"],sorted_bestvocff[item]["Illumination"],sorted_bestvocff[item]["ScanDirection"],str('%.2f' % float(sorted_bestvocff[item]["ImaxComp"])),sorted_bestvocff[item]["Isenserange"],str(sorted_bestvocff[item]["AreaJV"]),sorted_bestvocff[item]["Operator"],sorted_bestvocff[item]["MeasComment"],timeLandD[item]["RefNomCurr"],timeLandD[item]["RefMeasCurr"],str(timeLandD[item]["AirTemp"]),str(timeLandD[item]["ChuckTemp"])])
                summary.insert(0, ["-", "-", "-","-","cm2","mV","mA/cm2","%","%","W/cm2","mV","mA/cm2","Ohm*cm2","Ohm*cm2","-","-","-","s","s","mV","mV","-","-","A","A","-","-","-","mA","mA","DegC","DegC"])
                summary.insert(0, ["Group","Sample Name", "Cell","MeasDayTime","Cell Surface","Voc","Jsc","FF","Eff","Pmpp","Vmpp","Jmpp","Roc","Rsc","VocFF","RscJsc","NbPoints","Delay","IntegTime","Vstart","Vend","Illumination","ScanDirection","ImaxComp","Isenserange","AreaJV","Operator","MeasComment","RefNomCurr","RefMeasCurr","AirTemp","ChuckTemp"])
                for item in range(len(summary)):
                    for item0 in range(len(summary[item])):
                        worksheet.write(item,item0,summary[item][item0])
            except:
                print("exception: excel summary - bestvocff")
        
        if DATAMPP!=[]: 
            try:
                worksheet = workbook.add_worksheet("Pmpp")
                summary=[]
                for item in range(len(DATAMPP)):
                    summary.append([DATAMPP[item]["Group"],DATAMPP[item]["SampleName"],DATAMPP[item]["Cellletter"],DATAMPP[item]["MeasDayTime"],float('%.2f' % float(DATAMPP[item]["CellSurface"])),DATAMPP[item]["Delay"],DATAMPP[item]["IntegTime"],float(DATAMPP[item]["Vstep"]),float(DATAMPP[item]["Vstart"]),float('%.1f' % float(DATAMPP[item]["MppData"][2][-1])),DATAMPP[item]["Operator"],DATAMPP[item]["MeasComment"]])
                summary.insert(0, ["Group","Sample Name", "Cell","MeasDayTime","Cell Surface","Delay","IntegTime","Vstep","Vstart","ExecTime","Operator","MeasComment"])
                for item in range(len(summary)):
                    for item0 in range(len(summary[item])):
                        worksheet.write(item,item0,summary[item][item0])
            except:
                print("exception: excel summary - Pmpp")
        
        if DATAFV!=[]: 
            try:
                worksheet = workbook.add_worksheet("fixedvoltage")
                summary=[]
                for item in range(len(DATAFV)):
                    summary.append([DATAFV[item]["Group"],DATAFV[item]["SampleName"],DATAFV[item]["Cellletter"],DATAFV[item]["MeasDayTime"],float('%.2f' % float(DATAFV[item]["CellSurface"])),DATAFV[item]["Delay"],DATAFV[item]["IntegTime"],DATAFV[item]["NbCycle"],float(DATAFV[item]["Vstep"]),float(DATAFV[item]["ExecTime"]),float(DATAFV[item]["TimeatZero"]),DATAFV[item]["Operator"],DATAFV[item]["MeasComment"]])
                summary.insert(0, ["Group", "Sample Name", "Cell","MeasDayTime","Cell Surface","Delay","IntegTime","NbCycle","Initial voltage step", "Time at voltage bias", "Time at zero", "Operator","MeasComment"])
                for item in range(len(summary)):
                    for item0 in range(len(summary[item])):
                        worksheet.write(item,item0,summary[item][item0])
            except:
                print("exception: excel summary - fixedvoltage")
                
        if LandD!=[]:   
            try:
                sorted_dataall = sorted(LandD, key=itemgetter('DepID')) 
                for key, group in groupby(sorted_dataall, key=lambda x:x['DepID']):
                    partdat=list(group)
                    worksheet = workbook.add_worksheet(key)
                    summary=[]
                    for item in range(len(partdat)):
                        summary.append([partdat[item]["Group"],partdat[item]["SampleName"],partdat[item]["Cellletter"],partdat[item]["MeasDayTime"],partdat[item]["CellSurface"],str(partdat[item]["Voc"]),str(partdat[item]["Jsc"]),str(partdat[item]["FF"]),str(partdat[item]["Eff"]),str(partdat[item]["Pmpp"]),str(partdat[item]["Vmpp"]),str(partdat[item]["Jmpp"]),str(partdat[item]["Roc"]),str(partdat[item]["Rsc"]),str(partdat[item]["VocFF"]),str(partdat[item]["RscJsc"]),str(partdat[item]["NbPoints"]),partdat[item]["Delay"],partdat[item]["IntegTime"],partdat[item]["Vstart"],partdat[item]["Vend"],partdat[item]["Illumination"],partdat[item]["ScanDirection"],str('%.2f' % float(partdat[item]["ImaxComp"])),partdat[item]["Isenserange"],str(partdat[item]["AreaJV"]),partdat[item]["Operator"],partdat[item]["MeasComment"],timeLandD[item]["RefNomCurr"],timeLandD[item]["RefMeasCurr"],str(timeLandD[item]["AirTemp"]),str(timeLandD[item]["ChuckTemp"])])
                    summary.insert(0, ["-", "-", "-","-","cm2","mV","mA/cm2","%","%","W/cm2","mV","mA/cm2","Ohm*cm2","Ohm*cm2","-","-","-","s","s","mV","mV","-","-","A","A","-","-","-","mA","mA","DegC","DegC"])
                    summary.insert(0, ["Group", "Sample Name", "Cell","MeasDayTime","Cell Surface","Voc","Jsc","FF","Eff","Pmpp","Vmpp","Jmpp","Roc","Rsc","VocFF","RscJsc","NbPoints","Delay","IntegTime","Vstart","Vend","Illumination","ScanDirection","ImaxComp","Isenserange","AreaJV","Operator","MeasComment","RefNomCurr","RefMeasCurr","AirTemp","ChuckTemp"])
                    for item in range(len(summary)):
                        for item0 in range(len(summary[item])):
                            worksheet.write(item,item0,summary[item][item0])
            except:
                print("exception: excel summary - LandD")
                        
        workbook.close()
    
    plt.close("all")
    plt.clf()
            
    if window.ui.checkBox_AAstatgraphs.isChecked():
        fig = plt.figure()
        Effsubfig = fig.add_subplot(224)
        Vocsubfig = fig.add_subplot(221)
        Jscsubfig = fig.add_subplot(222)
        FFsubfig = fig.add_subplot(223)
        
        listofparam={'Eff':[Effsubfig,'Efficiency (%)'],'Voc':[Vocsubfig,'Voc (mV)'],'Jsc':[Jscsubfig,'Jsc (mA/cm'+'\xb2'+')'],'FF':[FFsubfig,'FF (%)']}
        usednamesempty=0
        for param in listofparam.keys():
            usednames=[]
            eff={"A":[],"B":[],"C":[],"D":[],"E":[],"F":[],"S":[]}
            for item in DATAx:
                if item["Illumination"]=='Light':
                    if item["Cellletter"] not in eff.keys():
                        eff[item["Cellletter"]]=[]
                    eff[item["Cellletter"]].append(item[param])
                    usednames.append(item["Cellletter"])
            if usednames==[]:
                usednamesempty=1
                break
            else:
                usednames=list(set(usednames))
                usednames.sort()
                limitlist=[]
                for i in range(len(usednames)):
                    y=eff[usednames[i]]
                    limitlist.append(eff[usednames[i]])
                    if len(y)>0:
                        x=np.random.normal(i+1,0.04,size=len(y))
                        listofparam[param][0].scatter(x,y,s=15,color='red', alpha=0.5)
                flatlimitlist = []
                for sublist in limitlist:
                    for item in sublist:
                        flatlimitlist.append(item)
                span=range(1,len(usednames)+1)
                # print(span)
                listofparam[param][0].set_xticks(span)
                listofparam[param][0].set_xticklabels(usednames)
                listofparam[param][0].set_xlim([0.5,span[-1]+0.5])
                try:
                    listofparam[param][0].set_ylim([min(flatlimitlist)-1,max(flatlimitlist)+1])
                except ValueError:
                    pass
                listofparam[param][0].set_ylabel(listofparam[param][1])
        if usednamesempty==0:
            fig.subplots_adjust(wspace=.25)
            fig.savefig(batchname+'_StatCells.png',dpi=300,bbox_inches="tight")
            
            plt.close("all")
            plt.clf()
        
    #stat graphs
    if window.ui.checkBox_AAstatgraphs.isChecked():
        #group
#            try:
        plt.close("all")
        plt.clf()
        # print(len(samplesgroups))
        # print(samplesgroups)
        if len(samplesgroups)>1:
            fig = plt.figure()
            Effsubfig = fig.add_subplot(224) 
            names=samplesgroups
            valsRev=[]
            for item in names:
                valsRev.append([i["RevEff"] for i in grouplistdict if i["Group"]==item and "RevEff" in i])
            valsFor=[]
            for item in names:
                valsFor.append([i["ForEff"] for i in grouplistdict if i["Group"]==item and "ForEff" in i])
                
            valstot=[]
            Rev=[]
            Forw=[]
            namelist=[]
            for i in range(len(names)):
                if valsRev[i][0]==[] and valsFor[i][0]==[]:
                    print(" ")
                else:
                    Rev.append(valsRev[i][0])
                    Forw.append(valsFor[i][0])
                    valstot.append(valsRev[i][0]+valsFor[i][0])
                    namelist.append(names[i])
            
            Effsubfig.boxplot(valstot,0,'',labels=namelist)
            
            for i in range(len(namelist)):
                y=Rev[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    Effsubfig.scatter(x,y,s=5,color='red', alpha=0.5)
                y=Forw[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    Effsubfig.scatter(x,y,s=5,color='blue', alpha=0.5)  
            
            #Effsubfig.set_xlabel('Red=reverse; Blue=forward')
            
            Effsubfig.set_ylabel('Efficiency (%)')
            for item in ([Effsubfig.title, Effsubfig.xaxis.label, Effsubfig.yaxis.label] +
                          Effsubfig.get_xticklabels() + Effsubfig.get_yticklabels()):
                item.set_fontsize(4)
            
            Vocsubfig = fig.add_subplot(221) 
            names=samplesgroups
            valsRev=[]
            for item in names:
                valsRev.append([i["RevVoc"] for i in grouplistdict if i["Group"]==item and "RevVoc" in i])
            valsFor=[]
            for item in names:
                valsFor.append([i["ForVoc"] for i in grouplistdict if i["Group"]==item and "ForVoc" in i])
                
            valstot=[]
            Rev=[]
            Forw=[]
            namelist=[]
            for i in range(len(names)):
                if valsRev[i][0]==[] and valsFor[i][0]==[]:
                    print(" ")
                else:
                    Rev.append(valsRev[i][0])
                    Forw.append(valsFor[i][0])
                    valstot.append(valsRev[i][0]+valsFor[i][0])
                    namelist.append(names[i])
            
            Vocsubfig.boxplot(valstot,0,'',labels=namelist)
            
            for i in range(len(namelist)):
                y=Rev[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    Vocsubfig.scatter(x,y,s=5,color='red', alpha=0.5)
                y=Forw[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    Vocsubfig.scatter(x,y,s=5,color='blue', alpha=0.5)  
            
            #Vocsubfig.set_xlabel('Red=reverse; Blue=forward')
            
            Vocsubfig.set_ylabel('Voc (mV)')
            for item in ([Vocsubfig.title, Vocsubfig.xaxis.label, Vocsubfig.yaxis.label] +
                          Vocsubfig.get_xticklabels() + Vocsubfig.get_yticklabels()):
                item.set_fontsize(4)
                
            Jscsubfig = fig.add_subplot(222) 
            names=samplesgroups
            valsRev=[]
            for item in names:
                valsRev.append([i["RevJsc"] for i in grouplistdict if i["Group"]==item and "RevJsc" in i])
            valsFor=[]
            for item in names:
                valsFor.append([i["ForJsc"] for i in grouplistdict if i["Group"]==item and "ForJsc" in i])
                
            valstot=[]
            Rev=[]
            Forw=[]
            namelist=[]
            for i in range(len(names)):
                if valsRev[i][0]==[] and valsFor[i][0]==[]:
                    print(" ")
                else:
                    Rev.append(valsRev[i][0])
                    Forw.append(valsFor[i][0])
                    valstot.append(valsRev[i][0]+valsFor[i][0])
                    namelist.append(names[i])
            
            Jscsubfig.boxplot(valstot,0,'',labels=namelist)
            
            for i in range(len(namelist)):
                y=Rev[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    Jscsubfig.scatter(x,y,s=5,color='red', alpha=0.5)
                y=Forw[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    Jscsubfig.scatter(x,y,s=5,color='blue', alpha=0.5)  
            
            #Jscsubfig.set_xlabel('Red=reverse; Blue=forward')
            
            Jscsubfig.set_ylabel('Jsc (mA/cm'+'\xb2'+')')
            for item in ([Jscsubfig.title, Jscsubfig.xaxis.label, Jscsubfig.yaxis.label] +
                          Jscsubfig.get_xticklabels() + Jscsubfig.get_yticklabels()):
                item.set_fontsize(4)
            
            FFsubfig = fig.add_subplot(223) 
            names=samplesgroups
            valsRev=[]
            for item in names:
                valsRev.append([i["RevFF"] for i in grouplistdict if i["Group"]==item and "RevFF" in i])
            valsFor=[]
            for item in names:
                valsFor.append([i["ForFF"] for i in grouplistdict if i["Group"]==item and "ForFF" in i])
                
            valstot=[]
            Rev=[]
            Forw=[]
            namelist=[]
            for i in range(len(names)):
                if valsRev[i][0]==[] and valsFor[i][0]==[]:
                    print(" ")
                else:
                    Rev.append(valsRev[i][0])
                    Forw.append(valsFor[i][0])
                    valstot.append(valsRev[i][0]+valsFor[i][0])
                    namelist.append(names[i])
            
            FFsubfig.boxplot(valstot,0,'',labels=namelist)
            
            for i in range(len(namelist)):
                y=Rev[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    FFsubfig.scatter(x,y,s=5,color='red', alpha=0.5)
                y=Forw[i]
                if len(y)>0:
                    x=np.random.normal(i+1,0.04,size=len(y))
                    FFsubfig.scatter(x,y,s=5,color='blue', alpha=0.5)  
            
            #FFsubfig.set_xlabel('Red=reverse; Blue=forward')
            
            FFsubfig.set_ylabel('FF (%)')
            for item in ([FFsubfig.title, FFsubfig.xaxis.label, FFsubfig.yaxis.label] +
                          FFsubfig.get_xticklabels() + FFsubfig.get_yticklabels()):
                item.set_fontsize(4)
                
            FFsubfig.annotate('Red=reverse; Blue=forward', xy=(1.3,-0.2), xycoords='axes fraction', fontsize=4,
                        horizontalalignment='right', verticalalignment='bottom')
            annotation="#ofCells: "
            for item in range(len(samplesgroups)):
                if samplesgroups[item] in namelist:
                    annotation+=samplesgroups[item]+"=>"+str(grouplistdict[item]["numbCell"])+"; "
            FFsubfig.annotate(annotation, xy=(0,-0.3), xycoords='axes fraction', fontsize=4,
                        horizontalalignment='left', verticalalignment='bottom')
            
            fig.subplots_adjust(wspace=.25)
            fig.savefig(batchname+'_StatGroupgraph.png',dpi=300,bbox_inches="tight")
            
            
            
            plt.close("all")
        plt.clf()
#            except:
#                print("Exception: statgraphs - group")
        
        #time
        # if DATAx[0]["Setup"]=="TFIV" or DATAx[0]["Setup"]=='SSIgorC215':
        # try: 
        time=[]
        for i in DATAx:
            if i["Illumination"]=="Light":
                time.append(i["MeasDayTime2"])
        # for i in DATAx:
        #     if i["Illumination"]=="Light":
        #         if i["MeasDayTime"].split(' ')[-1]=='PM' and float(i["MeasDayTime"].split(' ')[-2].split(':')[0])!=12: 
        #             time.append(float(i["MeasDayTime"].split(' ')[-2].split(':')[0])+12+ float(i["MeasDayTime"].split(' ')[-2].split(':')[1])/60 + float(i["MeasDayTime"].split(' ')[-2].split(':')[2])/3600)
        #         else:
        #             time.append(float(i["MeasDayTime"].split(' ')[-2].split(':')[0])+ float(i["MeasDayTime"].split(' ')[-2].split(':')[1])/60 + float(i["MeasDayTime"].split(' ')[-2].split(':')[2])/3600)
                               
        if time!=[]:
            Voct=[i["Voc"] for i in DATAx if i["Illumination"]=="Light"]
            Jsct=[i["Jsc"] for i in DATAx if i["Illumination"]=="Light"]
            FFt=[i["FF"] for i in DATAx if i["Illumination"]=="Light"]
            Efft=[i["Eff"] for i in DATAx if i["Illumination"]=="Light"]
            
            fig = plt.figure()
            Vocsubfig = fig.add_subplot(221) 
            Vocsubfig.scatter(time, Voct, s=5, c='k', alpha=0.5)
            Vocsubfig.set_ylabel('Voc (mV)')
            for item in ([Vocsubfig.title, Vocsubfig.xaxis.label, Vocsubfig.yaxis.label] +
                          Vocsubfig.get_xticklabels() + Vocsubfig.get_yticklabels()):
                item.set_fontsize(8)
            
            
            Jscsubfig = fig.add_subplot(222) 
            Jscsubfig.scatter(time, Jsct, s=5, c='k', alpha=0.5)
            Jscsubfig.set_ylabel('Jsc (mA/cm'+'\xb2'+')')
            for item in ([Jscsubfig.title, Jscsubfig.xaxis.label, Jscsubfig.yaxis.label] +
                          Jscsubfig.get_xticklabels() + Jscsubfig.get_yticklabels()):
                item.set_fontsize(8)
            
            
            FFsubfig = fig.add_subplot(223) 
            FFsubfig.scatter(time, FFt, s=5, c='k', alpha=0.5)
            FFsubfig.set_xlabel('Time')
            FFsubfig.set_ylabel('FF (%)')
            for item in ([FFsubfig.title, FFsubfig.xaxis.label, FFsubfig.yaxis.label] +
                          FFsubfig.get_xticklabels() + FFsubfig.get_yticklabels()):
                item.set_fontsize(8)
            
            
            Effsubfig = fig.add_subplot(224) 
            Effsubfig.scatter(time, Efft, s=5, c='k', alpha=0.5)
            Effsubfig.set_xlabel('Time')
            Effsubfig.set_ylabel('Eff (%)')
            for item in ([Effsubfig.title, Effsubfig.xaxis.label, Effsubfig.yaxis.label] +
                          Effsubfig.get_xticklabels() + Effsubfig.get_yticklabels()):
                item.set_fontsize(8)
            
            difftime=max(time)-min(time)
            Vocsubfig.set_xlim([min(time)-0.1*difftime,max(time)+0.1*difftime])
            Jscsubfig.set_xlim([min(time)-0.1*difftime,max(time)+0.1*difftime])
            FFsubfig.set_xlim([min(time)-0.1*difftime,max(time)+0.1*difftime])
            Effsubfig.set_xlim([min(time)-0.1*difftime,max(time)+0.1*difftime])
            
            # Vocsubfig.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d-%H:%M"))
            # Jscsubfig.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d-%H:%M"))
            # FFsubfig.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d-%H:%M"))
            # Effsubfig.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d-%H:%M"))
            
            # Vocsubfig.xaxis.set_major_locator(mdates.DayLocator())
            # Jscsubfig.xaxis.set_major_locator(mdates.DayLocator())
            # FFsubfig.xaxis.set_major_locator(mdates.DayLocator())
            # Effsubfig.xaxis.set_major_locator(mdates.DayLocator())
            
            fig.autofmt_xdate()
            
            fig.subplots_adjust(wspace=.25)
            fig.savefig(batchname+'_StatTimegraph.png',dpi=300,bbox_inches="tight")
            plt.close("all")
            # except:
            #     print("Exception: statgraph - time")
            plt.clf()
        
        #Resistances
        try:
            Rsclist=[float(i["Rsc"]) for i in DATAx]
            Roclist=[float(i["Roc"]) for i in DATAx]
            Voct=[i["Voc"] for i in DATAx]
            Jsct=[i["Jsc"] for i in DATAx]
            FFt=[i["FF"] for i in DATAx]
            Efft=[i["Eff"] for i in DATAx]
            
            
            fig = plt.figure()
            Vocsubfig = fig.add_subplot(221) 
            Vocsubfig.scatter(Rsclist, Voct, s=5, c='k', alpha=0.5)
            Vocsubfig.set_ylabel('Voc (mV)')
            for item in ([Vocsubfig.title, Vocsubfig.xaxis.label, Vocsubfig.yaxis.label] +
                          Vocsubfig.get_xticklabels() + Vocsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            Vocsubfig.set_xlim(left=0)
            Vocsubfig.set_ylim(bottom=0)
            Vocsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            
            
            Jscsubfig = fig.add_subplot(222) 
            Jscsubfig.scatter(Rsclist, Jsct, s=5, c='k', alpha=0.5)
            Jscsubfig.set_ylabel('Jsc (mA/cm'+'\xb2'+')')
            for item in ([Jscsubfig.title, Jscsubfig.xaxis.label, Jscsubfig.yaxis.label] +
                          Jscsubfig.get_xticklabels() + Jscsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            Jscsubfig.set_xlim(left=0)
            Jscsubfig.set_ylim(bottom=0)
            Jscsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            
            FFsubfig = fig.add_subplot(223) 
            FFsubfig.scatter(Rsclist, FFt, s=5, c='k', alpha=0.5)
            FFsubfig.set_xlabel('Rsc')
            FFsubfig.set_ylabel('FF (%)')
            for item in ([FFsubfig.title, FFsubfig.xaxis.label, FFsubfig.yaxis.label] +
                          FFsubfig.get_xticklabels() + FFsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            FFsubfig.set_xlim(left=0)
            FFsubfig.set_ylim(bottom=0)
            FFsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            
            Effsubfig = fig.add_subplot(224) 
            Effsubfig.scatter(Rsclist, Efft, s=5, c='k', alpha=0.5)
            Effsubfig.set_xlabel('Rsc')
            Effsubfig.set_ylabel('Eff (%)')
            for item in ([Effsubfig.title, Effsubfig.xaxis.label, Effsubfig.yaxis.label] +
                          Effsubfig.get_xticklabels() + Effsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            Effsubfig.set_xlim(left=0)
            Effsubfig.set_ylim(bottom=0)
            Effsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            fig.autofmt_xdate()
            fig.subplots_adjust(wspace=.3)
            fig.savefig(batchname+'_StatRscgraph.png',dpi=300,bbox_inches="tight")
            plt.close("all")
            
            
            fig = plt.figure()
            Vocsubfig = fig.add_subplot(221) 
            Vocsubfig.scatter(Roclist, Voct, s=5, c='k', alpha=0.5)
            Vocsubfig.set_ylabel('Voc (mV)')
            for item in ([Vocsubfig.title, Vocsubfig.xaxis.label, Vocsubfig.yaxis.label] +
                          Vocsubfig.get_xticklabels() + Vocsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            Vocsubfig.set_xlim(left=0)
            Vocsubfig.set_ylim(bottom=0)
            Vocsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            
            Jscsubfig = fig.add_subplot(222) 
            Jscsubfig.scatter(Roclist, Jsct, s=5, c='k', alpha=0.5)
            Jscsubfig.set_ylabel('Jsc (mA/cm'+'\xb2'+')')
            for item in ([Jscsubfig.title, Jscsubfig.xaxis.label, Jscsubfig.yaxis.label] +
                          Jscsubfig.get_xticklabels() + Jscsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            Jscsubfig.set_xlim(left=0)
            Jscsubfig.set_ylim(bottom=0)
            Jscsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            
            FFsubfig = fig.add_subplot(223) 
            FFsubfig.scatter(Roclist, FFt, s=5, c='k', alpha=0.5)
            FFsubfig.set_xlabel('Roc')
            FFsubfig.set_ylabel('FF (%)')
            for item in ([FFsubfig.title, FFsubfig.xaxis.label, FFsubfig.yaxis.label] +
                          FFsubfig.get_xticklabels() + FFsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            FFsubfig.set_xlim(left=0)
            FFsubfig.set_ylim(bottom=0)
            FFsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            
            Effsubfig = fig.add_subplot(224) 
            Effsubfig.scatter(Roclist, Efft, s=5, c='k', alpha=0.5)
            Effsubfig.set_xlabel('Roc')
            Effsubfig.set_ylabel('Eff (%)')
            for item in ([Effsubfig.title, Effsubfig.xaxis.label, Effsubfig.yaxis.label] +
                          Effsubfig.get_xticklabels() + Effsubfig.get_yticklabels()):
                item.set_fontsize(8)
            #plt.xticks(np.arange(min(time), max(time)+1, 1.0))
            Effsubfig.set_xlim(left=0)
            Effsubfig.set_ylim(bottom=0)
            Effsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
            
            fig.subplots_adjust(wspace=.3)
            fig.savefig(batchname+'_StatRocgraph.png',dpi=300,bbox_inches="tight")
            plt.close("all")
            plt.clf()
        except:
            print("Exception: statgraph - resistance")
        
        #stat graph with diff colors for ABC and Forw Rev, by substrate
        #get substrate number without run number
        # if DATAx[0]["Setup"]=="TFIV" or DATAx[0]["Setup"]=='SSIgorC215':
            # try:
        fig = plt.figure()
        
        VocAFy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        VocAFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        VocBFy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        VocBFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        VocCFy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        VocCFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        VocDFy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        VocDFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        VocEFy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        VocEFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        VocFFy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        VocFFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]

        VocSFy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        VocSFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        VocARy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        VocARx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        VocBRy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        VocBRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        VocCRy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        VocCRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        VocDRy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        VocDRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        VocERy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        VocERx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        VocFRy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        VocFRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        VocSRy=[float(i["Voc"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        VocSRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        Vocsubfig = fig.add_subplot(221) 
        Vocsubfig.scatter(VocAFx, VocAFy, s=5, facecolors='none', edgecolors='r', lw=0.5)
        Vocsubfig.scatter(VocBFx, VocBFy, s=5, facecolors='none', edgecolors='g', lw=0.5)
        Vocsubfig.scatter(VocCFx, VocCFy, s=5, facecolors='none', edgecolors='b', lw=0.5)
        Vocsubfig.scatter(VocARx, VocARy, s=5, facecolors='r', edgecolors='r', lw=0.5)
        Vocsubfig.scatter(VocBRx, VocBRy, s=5, facecolors='g', edgecolors='g', lw=0.5)
        Vocsubfig.scatter(VocCRx, VocCRy, s=5, facecolors='b', edgecolors='b', lw=0.5)
        Vocsubfig.scatter(VocDFx, VocDFy, s=5, facecolors='none', edgecolors='c', lw=0.5)
        Vocsubfig.scatter(VocEFx, VocEFy, s=5, facecolors='none', edgecolors='m', lw=0.5)
        Vocsubfig.scatter(VocFFx, VocFFy, s=5, facecolors='none', edgecolors='k', lw=0.5)
        Vocsubfig.scatter(VocDRx, VocDRy, s=5, facecolors='c', edgecolors='c', lw=0.5)
        Vocsubfig.scatter(VocERx, VocERy, s=5, facecolors='m', edgecolors='m', lw=0.5)
        Vocsubfig.scatter(VocFRx, VocFRy, s=5, facecolors='k', edgecolors='k', lw=0.5)
        Vocsubfig.scatter(VocSFx, VocSFy, s=5, facecolors='none', edgecolors='y', lw=0.5)
        Vocsubfig.scatter(VocSRx, VocSRy, s=5, facecolors='y', edgecolors='y', lw=0.5)
#                    Vocsubfig.scatter(VocSFx, VocSFy, s=5, facecolors='none', edgecolors='k', lw=0.5)
#                    Vocsubfig.scatter(VocSRx, VocSRy, s=5, edgecolors='k', lw=0.5)
        Vocsubfig.set_ylabel('Voc (mV)')
        Vocsubfig.set_xlabel("Sample #")
        for item in ([Vocsubfig.title, Vocsubfig.xaxis.label, Vocsubfig.yaxis.label] + Vocsubfig.get_xticklabels() + Vocsubfig.get_yticklabels()):
            item.set_fontsize(4)
        Vocsubfig.set_ylim(bottom=0)
        Vocsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
        try:
            Vocsubfig.set_xticks(np.arange(float(min(VocAFx+VocBFx+VocCFx+VocARx+VocBRx+VocCRx+VocDFx+VocEFx+VocFFx+VocDRx+VocERx+VocFRx+VocSFx+VocSRx))-0.5,float(max(VocAFx+VocBFx+VocCFx+VocARx+VocBRx+VocCRx+VocDFx+VocEFx+VocFFx+VocDRx+VocERx+VocFRx+VocSFx+VocSRx))+0.5,1), minor=True)
        except:
            pass
        #Vocsubfig.set_xticks(np.arange(float(min(VocAFx))-0.5,float(max(VocAFx))+0.5,1), minor=True)
        Vocsubfig.xaxis.grid(False, which='major')
        Vocsubfig.xaxis.grid(True, which='minor', color='k', linestyle='-', alpha=0.2)
        
        try:
            Vocsubfig.axis([float(min(VocAFx+VocBFx+VocCFx+VocARx+VocBRx+VocCRx+VocDFx+VocEFx+VocFFx+VocDRx+VocERx+VocFRx+VocSFx+VocSRx))-0.5,float(max(VocAFx+VocBFx+VocCFx+VocARx+VocBRx+VocCRx+VocDFx+VocEFx+VocFFx+VocDRx+VocERx+VocFRx+VocSFx+VocSRx))+0.5,0.5*float(min(VocAFy+VocBFy+VocCFy+VocARy+VocBRy+VocCRy+VocDFy+VocEFy+VocFFy+VocDRy+VocERy+VocFRy+VocSFy+VocSRy)),1.1*float(max(VocAFy+VocBFy+VocCFy+VocARy+VocBRy+VocCRy+VocDFy+VocEFy+VocFFy+VocDRy+VocERy+VocFRy+VocSFy+VocSRy))])
        except ValueError:
            pass
        for axis in ['top','bottom','left','right']:
            Vocsubfig.spines[axis].set_linewidth(0.5)
        Vocsubfig.tick_params(axis='x', which='both',bottom='off', top='off')
        
        
        
        JscAFy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        JscAFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        JscBFy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        JscBFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        JscCFy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        JscCFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        JscARy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        JscARx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        JscBRy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        JscBRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        JscCRy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        JscCRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]

        JscDFy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        JscDFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        JscEFy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        JscEFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        JscFFy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        JscFFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        JscSFy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        JscSFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        JscDRy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        JscDRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        JscERy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        JscERx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        JscFRy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        JscFRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]                    
        
        JscSRy=[float(i["Jsc"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        JscSRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]                    
        
        Jscsubfig = fig.add_subplot(222) 
        Jscsubfig.scatter(JscAFx, JscAFy, s=5, facecolors='none', edgecolors='r', lw=0.5)
        Jscsubfig.scatter(JscBFx, JscBFy, s=5, facecolors='none', edgecolors='g', lw=0.5)
        Jscsubfig.scatter(JscCFx, JscCFy, s=5, facecolors='none', edgecolors='b', lw=0.5)
        Jscsubfig.scatter(JscARx, JscARy, s=5, facecolors='r', edgecolors='r', lw=0.5)
        Jscsubfig.scatter(JscBRx, JscBRy, s=5, facecolors='g', edgecolors='g', lw=0.5)
        Jscsubfig.scatter(JscCRx, JscCRy, s=5, facecolors='b', edgecolors='b', lw=0.5)
        Jscsubfig.scatter(JscDFx, JscDFy, s=5, facecolors='none', edgecolors='c', lw=0.5)
        Jscsubfig.scatter(JscEFx, JscEFy, s=5, facecolors='none', edgecolors='m', lw=0.5)
        Jscsubfig.scatter(JscFFx, JscFFy, s=5, facecolors='none', edgecolors='k', lw=0.5)
        Jscsubfig.scatter(JscDRx, JscDRy, s=5, facecolors='c', edgecolors='c', lw=0.5)
        Jscsubfig.scatter(JscERx, JscERy, s=5, facecolors='m', edgecolors='m', lw=0.5)
        Jscsubfig.scatter(JscFRx, JscFRy, s=5, facecolors='k', edgecolors='k', lw=0.5)
        Jscsubfig.scatter(JscSFx, JscSFy, s=5, facecolors='none', edgecolors='y', lw=0.5)
        Jscsubfig.scatter(JscSRx, JscSRy, s=5, facecolors='y', edgecolors='y', lw=0.5)
        
        Jscsubfig.set_ylabel('Jsc (mA/cm'+'\xb2'+')')
        Jscsubfig.set_xlabel("Sample #")
        for item in ([Jscsubfig.title, Jscsubfig.xaxis.label, Jscsubfig.yaxis.label] +
                      Jscsubfig.get_xticklabels() + Jscsubfig.get_yticklabels()):
            item.set_fontsize(4)
        Jscsubfig.set_ylim(bottom=0)
        Jscsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
        
        try:
            Jscsubfig.set_xticks(np.arange(float(min(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx+JscSRx+JscSFx))-0.5,float(max(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx+JscSRx+JscSFx))+0.5,1), minor=True)
        except:
            pass
        #Jscsubfig.set_xticks(np.arange(float(min(JscAFx))-0.5,float(max(JscAFx))+0.5,1), minor=True)
        Jscsubfig.xaxis.grid(False, which='major')
        Jscsubfig.xaxis.grid(True, which='minor', color='k', linestyle='-', alpha=0.2)
        
        try:
            Jscsubfig.axis([float(min(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx+JscSRx+JscSFx))-0.5,float(max(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx+JscSRx+JscSFx))+0.5,0.5*float(min(JscAFy+JscBFy+JscCFy+JscARy+JscBRy+JscCRy+JscDFy+JscEFy+JscFFy+JscDRy+JscERy+JscFRy+JscSRy+JscSFy)),1.1*float(max(JscAFy+JscBFy+JscCFy+JscARy+JscBRy+JscCRy+JscDFy+JscEFy+JscFFy+JscDRy+JscERy+JscFRy+JscSRy+JscSFy))])
        except ValueError:
            pass
#                    print([float(min(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx))-0.5,float(max(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx))+0.5,0.5*float(min(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx)),1.1*float(max(JscAFx+JscBFx+JscCFx+JscARx+JscBRx+JscCRx+JscDFx+JscEFx+JscFFx+JscDRx+JscERx+JscFRx))])
        for axis in ['top','bottom','left','right']:
          Jscsubfig.spines[axis].set_linewidth(0.5)
        Jscsubfig.tick_params(axis='x', which='both',bottom='off', top='off')
        
        
        FFAFy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        FFAFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        FFBFy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        FFBFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        FFCFy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        FFCFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        FFARy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        FFARx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        FFBRy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        FFBRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        FFCRy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        FFCRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]

        FFDFy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        FFDFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        FFEFy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        FFEFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        FFFFy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        FFFFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        FFSFy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        FFSFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        FFDRy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        FFDRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        FFERy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        FFERx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        FFFRy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        FFFRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]                    
        
        FFSRy=[float(i["FF"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        FFSRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]                    
        
        FFsubfig = fig.add_subplot(223) 
        FFsubfig.scatter(FFAFx, FFAFy, s=5, facecolors='none', edgecolors='r', lw=0.5)
        FFsubfig.scatter(FFBFx, FFBFy, s=5, facecolors='none', edgecolors='g', lw=0.5)
        FFsubfig.scatter(FFCFx, FFCFy, s=5, facecolors='none', edgecolors='b', lw=0.5)
        FFsubfig.scatter(FFARx, FFARy, s=5, facecolors='r', edgecolors='r', lw=0.5)
        FFsubfig.scatter(FFBRx, FFBRy, s=5, facecolors='g', edgecolors='g', lw=0.5)
        FFsubfig.scatter(FFCRx, FFCRy, s=5, facecolors='b', edgecolors='b', lw=0.5)
        FFsubfig.scatter(FFDFx, FFDFy, s=5, facecolors='none', edgecolors='c', lw=0.5)
        FFsubfig.scatter(FFEFx, FFEFy, s=5, facecolors='none', edgecolors='m', lw=0.5)
        FFsubfig.scatter(FFFFx, FFFFy, s=5, facecolors='none', edgecolors='k', lw=0.5)
        FFsubfig.scatter(FFDRx, FFDRy, s=5, facecolors='c', edgecolors='c', lw=0.5)
        FFsubfig.scatter(FFERx, FFERy, s=5, facecolors='m', edgecolors='m', lw=0.5)
        FFsubfig.scatter(FFFRx, FFFRy, s=5, facecolors='k', edgecolors='k', lw=0.5)
        FFsubfig.scatter(FFSFx, FFSFy, s=5, facecolors='none', edgecolors='y', lw=0.5)
        FFsubfig.scatter(FFSRx, FFSRy, s=5, facecolors='y', edgecolors='y', lw=0.5)
        
        FFsubfig.set_ylabel('FF (%)')
        FFsubfig.set_xlabel("Sample #")
        for item in ([FFsubfig.title, FFsubfig.xaxis.label, FFsubfig.yaxis.label] +
                      FFsubfig.get_xticklabels() + FFsubfig.get_yticklabels()):
            item.set_fontsize(4)
        FFsubfig.set_ylim(bottom=0)
        FFsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
        try:
            FFsubfig.set_xticks(np.arange(float(min(FFAFx+FFBFx+FFCFx+FFARx+FFBRx+FFCRx+FFDFx+FFEFx+FFFFx+FFDRx+FFERx+FFFRx+FFSRx+FFSFx))-0.5,float(max(FFAFx+FFBFx+FFCFx+FFARx+FFBRx+FFCRx+FFDFx+FFEFx+FFFFx+FFDRx+FFERx+FFFRx+FFSRx+FFSFx))+0.5,1), minor=True)
        except:
            pass
        #FFsubfig.set_xticks(np.arange(float(min(FFAFx))-0.5,float(max(FFAFx))+0.5,1), minor=True)
        FFsubfig.xaxis.grid(False, which='major')
        FFsubfig.xaxis.grid(True, which='minor', color='k', linestyle='-', alpha=0.2)
        
        try:
            FFsubfig.axis([float(min(FFAFx+FFBFx+FFCFx+FFARx+FFBRx+FFCRx+FFDFx+FFEFx+FFFFx+FFDRx+FFERx+FFFRx+FFSRx+FFSFx))-0.5,float(max(FFAFx+FFBFx+FFCFx+FFARx+FFBRx+FFCRx+FFDFx+FFEFx+FFFFx+FFDRx+FFERx+FFFRx+FFSRx+FFSFx))+0.5,0.5*float(min(FFAFy+FFBFy+FFCFy+FFARy+FFBRy+FFCRy+FFDFy+FFEFy+FFFFy+FFDRy+FFERy+FFFRy+FFSRy+FFSFy)),1.1*float(max(FFAFy+FFBFy+FFCFy+FFARy+FFBRy+FFCRy+FFDFy+FFEFy+FFFFy+FFDRy+FFERy+FFFRy+FFSRy+FFSFy))])
        except ValueError:
            pass
        for axis in ['top','bottom','left','right']:
          FFsubfig.spines[axis].set_linewidth(0.5)
        FFsubfig.tick_params(axis='x', which='both',bottom='off', top='off')
        
        
        EffAFy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        EffAFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        EffBFy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        EffBFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        EffCFy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        EffCFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        EffARy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        EffARx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='A' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        EffBRy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        EffBRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='B' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        EffCRy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        EffCRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='C' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        EffDFy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        EffDFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        EffEFy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        EffEFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        EffFFy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        EffFFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        EffSFy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        EffSFx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Forward" and i["Illumination"]=="Light"]
        
        EffDRy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        EffDRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='D' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        EffERy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        EffERx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='E' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        EffFRy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        EffFRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='F' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        EffSRy=[float(i["Eff"]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        EffSRx=[int(i["DepID"].split('_')[1]) for i in DATAx if i["Cellletter"]=='S' and i["ScanDirection"]=="Reverse" and i["Illumination"]=="Light"]
        
        Effsubfig = fig.add_subplot(224) 
        Effsubfig.scatter(EffAFx, EffAFy, s=5, facecolors='none', edgecolors='r', lw=0.5)
        Effsubfig.scatter(EffBFx, EffBFy, s=5, facecolors='none', edgecolors='g', lw=0.5)
        Effsubfig.scatter(EffCFx, EffCFy, s=5, facecolors='none', edgecolors='b', lw=0.5)
        Effsubfig.scatter(EffARx, EffARy, s=5, facecolors='r', edgecolors='r', lw=0.5)
        Effsubfig.scatter(EffBRx, EffBRy, s=5, facecolors='g', edgecolors='g', lw=0.5)
        Effsubfig.scatter(EffCRx, EffCRy, s=5, facecolors='b', edgecolors='b', lw=0.5)
        Effsubfig.scatter(EffDFx, EffDFy, s=5, facecolors='none', edgecolors='c', lw=0.5)
        Effsubfig.scatter(EffEFx, EffEFy, s=5, facecolors='none', edgecolors='m', lw=0.5)
        Effsubfig.scatter(EffFFx, EffFFy, s=5, facecolors='none', edgecolors='k', lw=0.5)
        Effsubfig.scatter(EffDRx, EffDRy, s=5, facecolors='c', edgecolors='c', lw=0.5)
        Effsubfig.scatter(EffERx, EffERy, s=5, facecolors='m', edgecolors='m', lw=0.5)
        Effsubfig.scatter(EffFRx, EffFRy, s=5, facecolors='k', edgecolors='k', lw=0.5)
        Effsubfig.scatter(EffSFx, EffSFy, s=5, facecolors='none', edgecolors='y', lw=0.5)
        Effsubfig.scatter(EffSRx, EffSRy, s=5, facecolors='y', edgecolors='y', lw=0.5)
        
        Effsubfig.set_ylabel('Eff (%)')
        Effsubfig.set_xlabel("Sample #")
        for item in ([Effsubfig.title, Effsubfig.xaxis.label, Effsubfig.yaxis.label] +
                      Effsubfig.get_xticklabels() + Effsubfig.get_yticklabels()):
            item.set_fontsize(4)
        Effsubfig.set_ylim(bottom=0)
        Effsubfig.xaxis.set_major_locator(MaxNLocator(integer=True))
        try:
            Effsubfig.set_xticks(np.arange(float(min(EffAFx+EffBFx+EffCFx+EffARx+EffBRx+EffCRx+EffDFx+EffEFx+EffFFx+EffDRx+EffERx+EffFRx+EffSRx+EffSFx))-0.5,float(max(EffAFx+EffBFx+EffCFx+EffARx+EffBRx+EffCRx+EffDFx+EffEFx+EffFFx+EffDRx+EffERx+EffFRx+EffSRx+EffSFx))+0.5,1), minor=True)
        except:
            pass
        Effsubfig.xaxis.grid(False, which='major')
        Effsubfig.xaxis.grid(True, which='minor', color='k', linestyle='-', alpha=0.2)
        
        try:
            Effsubfig.axis([float(min(EffAFx+EffBFx+EffCFx+EffARx+EffBRx+EffCRx+EffDFx+EffEFx+EffFFx+EffDRx+EffERx+EffFRx+EffSRx+EffSFx))-0.5,float(max(EffAFx+EffBFx+EffCFx+EffARx+EffBRx+EffCRx+EffDFx+EffEFx+EffFFx+EffDRx+EffERx+EffFRx+EffSRx+EffSFx))+0.5,0.5*float(min(EffAFy+EffBFy+EffCFy+EffARy+EffBRy+EffCRy+EffDFy+EffEFy+EffFFy+EffDRy+EffERy+EffFRy+EffSRy+EffSFy)),1.1*float(max(EffAFy+EffBFy+EffCFy+EffARy+EffBRy+EffCRy+EffDFy+EffEFy+EffFFy+EffDRy+EffERy+EffFRy+EffSRy+EffSFy))])
        except ValueError:
            pass
        for axis in ['top','bottom','left','right']:
          Effsubfig.spines[axis].set_linewidth(0.5)
        Effsubfig.tick_params(axis='x', which='both',bottom='off', top='off')
        
        
        FFsubfig.annotate('Red=A; Green=B; Blue=C; Cyan=D; Magenta=E; Black=F; Yellow=S; empty=Forward; full=Reverse;', xy=(1.55,-0.3), xycoords='axes fraction', fontsize=4,
                        horizontalalignment='right', verticalalignment='bottom')
        
        fig.savefig(batchname+'_StatJVgraph.png',dpi=300,bbox_inches="tight")
        plt.close("all")
        plt.clf()
            # except:
            #     print("Exception: statgraph - bysubstrate")
                
    plt.close("all")
    plt.close(fig)
    plt.close(fig1) 
    
    if window.ui.checkBox_AAstatgraphs.isChecked():
        try:
            images = list(map(ImageTk.open, [batchname+'_StatCells.png',batchname+'_StatTimegraph.png',batchname+'_StatJVgraph.png',batchname+'_StatGroupgraph.png']))
            widths, heights = zip(*(i.size for i in images))
            total_width = max(widths[0]+widths[2],widths[1]+widths[3])
            max_height = max(heights[0]+heights[1],heights[2]+heights[3])
            new_im = ImageTk.new('RGB', (total_width, max_height), (255, 255, 255))
            new_im.paste(im=images[0],box=(0,0))
            new_im.paste(im=images[1],box=(0,max(heights[0],heights[2])))
            new_im.paste(im=images[2],box=(max(widths[0],widths[1]),0))
            new_im.paste(im=images[3],box=(max(widths[0],widths[1]),max(heights[0],heights[2])))
            new_im.save(batchname+'_controls.png')
        except:
            try:
                images = list(map(ImageTk.open, [batchname+'_StatCells.png',batchname+'_StatTimegraph.png',batchname+'_StatJVgraph.png']))
                widths, heights = zip(*(i.size for i in images))
                total_width = max(widths[0]+widths[2],2*widths[1])
                max_height = max(heights[0]+heights[1],2*heights[2])
                new_im = ImageTk.new('RGB', (total_width, max_height), (255, 255, 255))
                new_im.paste(im=images[0],box=(0,0))
                new_im.paste(im=images[1],box=(0,max(heights[0],heights[2])))
                new_im.paste(im=images[2],box=(max(widths[0],widths[1]),0))
                new_im.save(batchname+'_controls.png')
            except:#if gets here, probably there is not light files, only dark or mppt, so graphs were not generated
                print('no graphs?')
                pass
                
    return 'autoanalysis is finished'

#%%###########################################################################################
##############################################################################
##############################################################################
    


class Thread_getdatalistsfromGBpklabSERIS(QThread):
    finished = pyqtSignal()
    change_value = pyqtSignal(int)
    def __init__(self, file_path, parent=None):
        QThread.__init__(self, parent)
        self.file_path=file_path
        
    def run(self):
        global DATA, DATAdark
        global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
        
        numbnewMPPfiles=0
        numbnewIVfiles=0
        
        path_to_project = QFileDialog.getExistingDirectory(None, "Open Directory", "C:/", QFileDialog.ShowDirsOnly|QFileDialog.DontResolveSymlinks)        
        batchname=os.path.basename(os.path.normpath(path_to_project))
        print(batchname)
        
        from fnmatch import fnmatch

        root = path_to_project
        pattern = "IVRawData*.csv"
        filespathsMeasParam=[]
        for path, subdirs, files in os.walk(root):
            for name in files:
                if fnmatch(name, pattern):
                    filespathsMeasParam.append(os.path.join(path, name))

        pattern = "IVRawData*.csv"
        subfolders=[]
        filespaths=[]
        filesnames=[]
        for path, subdirs, files in os.walk(root):
            subfolders.append(subdirs)
            for name in files:
                if fnmatch(name, pattern):
                    filespaths.append(os.path.join(path, name))
                    filesnames.append(name)
                    numbnewIVfiles+=1
        # print(subfolders[0])
        
        num_plotsIV=len(DATA.keys())+2*numbnewIVfiles
        cmap = plt.get_cmap(colormapname)
        colorsIV = cmap(np.linspace(0, 1.0, num_plotsIV))
        colorsIV=[matplotlib.colors.to_hex(item) for item in colorsIV]
        
        #read files
        for i in range(len(filespaths)):
            # print(filespaths[i])
            filetype = 0
            partdict = {}
            partdict["filepath"]=filespaths[i].replace('\\','/')
            filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
            print(filename)
            samplenumber=filespaths[i].split('/')[-1].split('\\')[1]
            partdict["Cellletter"]='A'
            partdict["batchname"]=batchname
            partdict["DepID"]=batchname+'_'+samplenumber
            partdict["SampleName"]=partdict["DepID"]+'_'+filename.replace('-','')
            
            partdict["Illumination"]="Light"
            #for reverse
            
            partdict["MeasDayTime2"]=datetime.datetime.strptime(modification_date(partdict["filepath"]), "%Y-%m-%d %H:%M:%S")
            partdict["MeasDayTime"]=str(partdict["MeasDayTime2"])
            # print(type(partdict["MeasDayTime2"]))
            # print(type(partdict["MeasDayTime"]))
            # break
            partdict["sunintensity"]=1
            partdict["MeasComment"]="-"
            partdict["aftermpp"]=0
            partdict["Vstart"]=1.2
            partdict["Vend"]=-0.2
            partdict["NbPoints"]=71
            partdict["Delay"]=999
            partdict["IntegTime"]=999
            partdict["ImaxComp"]=-1
            partdict["Isenserange"]=-1
            partdict["Operator"]=''
            partdict["AreaJV"] =""
            partdict["Group"]="Default group"
            partdict["Setup"]="GBpklabSERIS"
            partdict["RefNomCurr"]=999
            partdict["RefMeasCurr"]=999
            partdict["AirTemp"]=999
            partdict["ChuckTemp"]=999
            partdict["Celltype"]=''#M2-244.32
            partdict["SerialNumber"]=''#forSERIS
            
            
            #read file rawdata
            ivpartdatrev = [[],[]]#[voltage,current]
            ivpartdatforw = [[],[]]#[voltage,current]
            print(filespaths[i])
            import csv
            with open(filespaths[i], newline='') as csvfile:
                datareader = list(csv.reader(csvfile, delimiter=','))
                datareader=datareader[1:]
                for row in datareader:
                    # print(row)
                    # print(row[0] +', '+spamreader[0][0])
                    # print(float(row[3]))
                    # print(float(row[4]))
                    # print('')

                    if row[0]==datareader[0][0]:
                        # print(row)
                        ivpartdatrev[0].append(float(row[3]))
                        ivpartdatforw[0].append(float(row[3]))
                        ivpartdatrev[1].append(float(row[4]))
                    else:
                        # print(row)
                        ivpartdatforw[1].append(float(row[3]))
            ivpartdatforw[0].reverse()
            
            with open(filespaths[i].replace('IVRawData','IVMeasurement'), newline='') as csvfile:
                datareader = list(csv.reader(csvfile, delimiter=','))
                Isc=float(datareader[1][1])
                Jsc=float(datareader[1][7])
                area=Isc/Jsc 
            partdict["CellSurface"]=area
            #change I to J in JVdata
            ivpartdatrev[1]=[x/area for x in ivpartdatrev[1]]
            ivpartdatforw[1]=[x/area for x in ivpartdatforw[1]]
            #specific for Reverse
            partdict["ScanDirection"]="Reverse"
            partdict["IVData"]=ivpartdatrev
            partdict["SampleName"]=partdict["DepID"]+'_'+filename.replace('-','').replace('IVRawData','')+'_Rev'
            partdict["IVlinestyle"]=[partdict["SampleName"],"-",colorsIV[len(DATA.keys())],2]
            
            params=extract_jv_params(ivpartdatrev)
            # print(params)
            partdict["Voc"]= float(params['Voc'])*1000 #mV
            partdict["Jsc"]= float(params['Jsc'])#mA/cm2
            partdict["Isc"]=float(params['Jsc']*partdict["CellSurface"])
            partdict["FF"]=float(params['FF']) #%
            partdict["Eff"]=float(params['Pmax'])#%
            partdict["Pmpp"]=float(partdict["Eff"])*10#W/cm2
            partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
            partdict["Roc"]=float(params['Roc'])
            partdict["Rsc"]=float(params['Rsc'])
            partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
            partdict["Vmpp"]=float(params['Vmpp'])
            partdict["Jmpp"]=float(params['Jmpp'])
            
            
            partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(partdict["Isc"]))+'_'+str(float(partdict["FF"]))
            DATA[partdict["SampleNameID"]]=partdict
            numbLightfiles+=1
            
            #specific for Forward
            partdict2=copy.deepcopy(partdict)
            partdict2["ScanDirection"]="Forward"
            partdict2["IVData"]=ivpartdatforw
            partdict2["SampleName"]=partdict2["DepID"]+'_'+filename.replace('-','').replace('IVRawData','')+'_For'
            partdict2["IVlinestyle"]=[partdict2["SampleName"],"-",colorsIV[len(DATA.keys())],2]
            
            params=extract_jv_params(ivpartdatforw)
            
            partdict2["Voc"]= float(params['Voc'])*1000 #mV
            partdict2["Jsc"]= float(params['Jsc'])#mA/cm2
            partdict2["Isc"]=float(params['Jsc']*partdict2["CellSurface"])
            partdict2["FF"]=float(params['FF']) #%
            partdict2["Eff"]=float(params['Pmax'])#%
            partdict2["Pmpp"]=float(partdict2["Eff"])*10#W/cm2
            partdict2["VocFF"]=partdict2["Voc"]*partdict2["FF"]
            partdict2["Roc"]=float(params['Roc'])
            partdict2["Rsc"]=float(params['Rsc'])
            partdict2["RscJsc"]=partdict2["Rsc"]*partdict2["Jsc"]
            partdict2["Vmpp"]=float(params['Vmpp'])
            partdict2["Jmpp"]=float(params['Jmpp'])
            
            
            partdict2["SampleNameID"]=partdict2["SampleName"]+'_'+str(partdict2["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(partdict2["Isc"]))+'_'+str(float(partdict2["FF"]))
            DATA[partdict2["SampleNameID"]]=partdict2
            numbLightfiles+=1
            
            self.change_value.emit(100*(i+1)/len(filespaths))
        self.finished.emit()


def xldate_to_datetime(xldatetime): #something like 43705.6158241088

      tempDate = datetime.datetime(1899, 12, 31)
      (days, portion) = math.modf(xldatetime)

      deltaDays = datetime.timedelta(days=days)
      #changing the variable name in the edit
      secs = int(24 * 60 * 60 * portion)
      detlaSeconds = datetime.timedelta(seconds=secs)
      TheTime = (tempDate + deltaDays + detlaSeconds )
      return TheTime.strftime("%Y-%m-%d %H:%M:%S")
  
class Thread_getdatalistsfromCUBpyfiles(QThread):
    finished = pyqtSignal()
    change_value = pyqtSignal(int)
    def __init__(self, file_path, parent=None):
        QThread.__init__(self, parent)
        self.file_path=file_path
        
    def run(self):
        global DATA, DATAdark
        global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
        print('threadstart')
        
        numbnewMPPfiles=0
        numbnewIVfiles=0
        for i in range(len(self.file_path)):
            filename=os.path.splitext(os.path.basename(self.file_path[i]))[0]
            if "MPPT" in filename or "FixedCurrent" in filename or "FixedVoltage" in filename:
                numbnewMPPfiles+=1
            else:
                numbnewIVfiles+=1
        num_plotsMPP=len(DATAMPP.keys())+numbnewMPPfiles
        cmap = plt.get_cmap(colormapname)
        colorsMPP = cmap(np.linspace(0, 1.0, num_plotsMPP))
        colorsMPP=[matplotlib.colors.to_hex(item) for item in colorsMPP]
        num_plotsIV=len(DATA.keys())+numbnewIVfiles
        colorsIV = cmap(np.linspace(0, 1.0, num_plotsIV))
        colorsIV=[matplotlib.colors.to_hex(item) for item in colorsIV]
        
        for i in range(len(self.file_path)):
            filetoread = open(self.file_path[i],"r", encoding='ISO-8859-1')
            filerawdata = filetoread.readlines()
            # print(i)
            filetype = 0
            partdict = {}
            partdict["filepath"]=self.file_path[i]
            filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
            
            if "MPPT" in filename or "FixedCurrent" in filename or "FixedVoltage" in filename:
                filetype = 2
                num_plots=num_plotsMPP
                colors = colorsMPP
            else:
                filetype = 1 
                num_plots=num_plotsIV
                colors = colorsIV
            
            if filetype ==1 : #J-V files from CUBoulder python software
                partdict["DepID"]=filename[:filename.index('pX')-1]
                aftername=filename[filename.index('pX'):]
                # print(filename)
                partdict["Cellletter"]=aftername.split('_')[0][2:]
                partdict["batchname"]=partdict["DepID"].split('_')[0]
                partdict["SampleName"]=partdict["DepID"]+"_"+partdict["Cellletter"]+"_"+aftername.split('_')[4]
                # print(partdict["SampleName"])
                if "_lt_" in aftername:
                    partdict["Illumination"]="Light"
                else:
                    partdict["Illumination"]="Dark"
                    
                if "_rev_" in aftername:
                    partdict["ScanDirection"]="Reverse"
                else:
                    partdict["ScanDirection"]="Forward" 
                
                for item in range(len(filerawdata)):
                    if "DateTime:" in filerawdata[item]:
                        partdict["MeasDayTime2"]=parser.parse(filerawdata[item][10:-1])
                        partdict["MeasDayTime"]=filerawdata[item][10:-1]
                        # print(type(partdict["MeasDayTime2"]))
                        # print(type(partdict["MeasDayTime"]))
                        # print(partdict["MeasDayTime2"])
                        # print(partdict["MeasDayTime"])
                        break
                for item in range(len(filerawdata)):
                    if "#sun:" in filerawdata[item]:
                        partdict["sunintensity"]=float(filerawdata[item][6:-1])
                        break
                partdict["MeasComment"]="-"
                for item in range(len(filerawdata)):
                    if "Comment" in filerawdata[item]:
                        # print(filerawdata[item][9:-1])
                        partdict["MeasComment"]=filerawdata[item][9:-1]
                        break
                if "aftermpp" in partdict["MeasComment"]:
                    partdict["aftermpp"]=1
                else:
                    partdict["aftermpp"]=0
                for item in range(len(filerawdata)):
                    if "minvoltage:" in filerawdata[item]:
                        partdict["Vstart"]=float(filerawdata[item][12:-1])
                        break
                for item in range(len(filerawdata)):
                    if "maxvoltage:" in filerawdata[item]:
                        partdict["Vend"]=float(filerawdata[item][12:-1])
                        break
                if partdict["ScanDirection"]=="Reverse":
                    if partdict["Vstart"]<partdict["Vend"]:
                        vend=partdict["Vend"]
                        partdict["Vend"]=partdict["Vstart"]
                        partdict["Vstart"]=vend
                else:
                    if partdict["Vstart"]>partdict["Vend"]:
                        vend=partdict["Vend"]
                        partdict["Vend"]=partdict["Vstart"]
                        partdict["Vstart"]=vend 
                for item in range(len(filerawdata)):
                    if "JVstepsize:" in filerawdata[item]:
                        partdict["NbPoints"]=abs(partdict["Vend"]-partdict["Vstart"])/float(filerawdata[item][12:-1])
                        break    
                for item in range(len(filerawdata)):
                    if "PixArea:" in filerawdata[item]:
                        partdict["CellSurface"]=float(filerawdata[item][9:-1])
                        #print(partdict["CellSurface"])
                        break
                for item in range(len(filerawdata)):
                    if "delaypoints:" in filerawdata[item]:
                        partdict["Delay"]=float(filerawdata[item][13:-1])
                        break
                for item in range(len(filerawdata)):
                    if "integtime:" in filerawdata[item]:
                        partdict["IntegTime"]=float(filerawdata[item][11:-1])
                        break
                for item in range(len(filerawdata)):
                    if "#IV data" in filerawdata[item]:
                            pos=item+2
                            break
                        
                ivpartdat = [[],[]]#[voltage,current]
                for item in range(pos,len(filerawdata),1):
                    try:
                        ivpartdat[0].append(float(filerawdata[item].split("\t")[0]))
                        ivpartdat[1].append(-float(filerawdata[item].split("\t")[1]))
                    except:
                        break
                partdict["IVData"]=ivpartdat
                
                y=ivpartdat[1]
            
                for item1 in range(len(y)-2):
                    pt0=y[item1]
                    pt1=y[item1+1]
                    pt2=y[item1+2]
                    if abs(pt0-pt1)>0.2 and abs(pt1-pt2)>0.2 and (math.copysign(1,pt0-pt1) != math.copysign(1,pt1-pt2)):
                        partdict["IVData"][1][item1+1]=(partdict["IVData"][1][item1]+partdict["IVData"][1][item1+2])/2
                
                params=extract_jv_params(ivpartdat)
                partdict["Voc"]= float(params['Voc'])*1000 #mV
                partdict["Jsc"]= float(params['Jsc'])#mA/cm2
                partdict["Isc"]=float(params['Jsc']*partdict["CellSurface"])
                partdict["FF"]=float(params['FF']) #%
                partdict["Eff"]=float(params['Pmax'])#%
                partdict["Pmpp"]=float(partdict["Eff"])*10#W/cm2
                partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
                partdict["Roc"]=float(params['Roc'])
                partdict["Rsc"]=float(params['Rsc'])
                partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
                partdict["Vmpp"]=float(params['Vmpp'])
                partdict["Jmpp"]=float(params['Jmpp'])
                # for item in range(len(filerawdata)):
                #     if "#IV results" in filerawdata[item]:
                #         partdict["Voc"]=float(filerawdata[item+2][4:-1]) #mV
                #         partdict["Jsc"]=float(filerawdata[item+4][4:-1]) #mA/cm2
                #         partdict["Isc"]=float(filerawdata[item+5][4:-1])
                #         partdict["FF"]=float(filerawdata[item+3][3:-1]) #%
                #         partdict["Eff"]=float(filerawdata[item+1][4:-1])#%
                #         partdict["Pmpp"]=float(filerawdata[item+6][5:-1]) #W/cm2
                #         partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
                #         partdict["Roc"]=float(filerawdata[item+9][4:-1])
                #         partdict["Rsc"]=float(filerawdata[item+10][4:-1])
                #         partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
                #         partdict["Vmpp"]=float(filerawdata[item+7][5:-1])
                #         partdict["Jmpp"]=float(filerawdata[item+8][5:-1])
                #         break
                
                partdict["ImaxComp"]=-1
                partdict["Isenserange"]=-1
                
                for item in range(len(filerawdata)):
                    if "UserName:" in filerawdata[item]:
                        partdict["Operator"]=str(filerawdata[item][10:-1])
                        break
                
                try:
                    if partdict["Illumination"]=="Light" and max(ivpartdat[0])>0.001*float(partdict["Voc"]):
                        f = interp1d(ivpartdat[0], ivpartdat[1], kind='cubic')
                        x2 = lambda x: f(x)
                        partdict["AreaJV"] = integrate.quad(x2,0,0.001*float(partdict["Voc"]))[0]
                    else:
                        partdict["AreaJV"] =""
                except ValueError:
                    print("there is a ValueError on sample ",i)
                
                
                partdict["Group"]="Default group"
                partdict["Group2"]=""
                partdict["Setup"]="SERISpythonIV"              
                for item in range(len(filerawdata)):
                    if "Diode nominal current:" in filerawdata[item]:
                        partdict["RefNomCurr"]=float(filerawdata[item][23:-1])
                        break
                for item in range(len(filerawdata)):
                    if "Diode measured current:" in filerawdata[item]:
                        partdict["RefMeasCurr"]=float(filerawdata[item][24:-1])
                        break
                for item in range(len(filerawdata)):
                    if "temperature:" in filerawdata[item]:
                        partdict["AirTemp"]=float(filerawdata[item][13:-1])
                        break
                partdict["ChuckTemp"]=999
                partdict["Celltype"]=''#M2-244.32
                partdict["SerialNumber"]=''#forSERIS
                partdict["IVlinestyle"]=[partdict["SampleName"],"-",colors[len(DATA.keys())],2]
#                DATA.append(partdict)
                
                if partdict["Illumination"]=="Light":
                    # DATA.append(partdict)
                    partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(partdict["Isc"]))+'_'+str(float(partdict["FF"]))
                    
                    DATA[partdict["SampleNameID"]]=partdict
                    numbLightfiles+=1
                else:
                    partdict["SampleName"]=partdict["SampleName"]+'_D'
                    partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(partdict["Isc"]))+'_'+str(float(partdict["FF"]))
                    DATA[partdict["SampleNameID"]]=partdict
                    DATAdark.append(partdict)
                    numbDarkfiles+=1
                
            elif filetype ==2 : #mpp files of SERF C215 labview program
                #assumes file name: batch_samplenumber_cellLetter_mpp
                partdict = {}
                partdict["filepath"]=self.file_path[i]
                filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
                if 'MPPT' in filename:
                    partdict["DepID"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+'MPPT'
                    partdict["SampleName"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+'MPPT'+'_'+filename.split('_')[2][2:]
                    partdict["Cellletter"]=filename.split('_')[2][2:]
                    partdict["batchname"]=filename.split('_')[0]
                elif 'FixedCurrent' in filename:
                    partdict["DepID"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+'FC'
                    partdict["SampleName"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+'FC'+'_'+filename.split('_')[2][2:]
                    partdict["Cellletter"]=filename.split('_')[2][2:]
                    partdict["batchname"]=filename.split('_')[0]
                elif 'FixedVoltage' in filename:
                    partdict["DepID"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+'FV'
                    partdict["SampleName"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+'FV'+'_'+filename.split('_')[2][2:]
                    partdict["Cellletter"]=filename.split('_')[2][2:]
                    partdict["batchname"]=filename.split('_')[0]
                
                found=0
                for item in range(len(filerawdata)):
                    if "Comment: " in filerawdata[item]:
                        partdict["MeasComment"]=filerawdata[item][9:-1]
                        found=1
                        break
                if found==0:
                    partdict["MeasComment"]=''
                for item in range(len(filerawdata)):
                    if "PixArea:" in filerawdata[item]:
                        partdict["CellSurface"]=float(filerawdata[item][9:-1])
                        break
                for item in range(len(filerawdata)):
                    if "DateTime:" in filerawdata[item]:
                        partdict["MeasDayTime2"]=parser.parse(filerawdata[item][10:-1])
                        partdict["MeasDayTime"]=filerawdata[item][10:-1]
                        # print(partdict["MeasDayTime2"])
#                        print(partdict["MeasDayTime"].split(' ')[-2])
                        break
                # partdict["MeasDayTime"]=modification_date(self.file_path[i])
                for item in range(len(filerawdata)):
                    if "PixArea:" in filerawdata[item]:
                        partdict["CellSurface"]=float(filerawdata[item][9:-1])
                        break
                for item in range(len(filerawdata)):
                    if "initialdelay:" in filerawdata[item]:
                        partdict["Delay"]=float(filerawdata[item][14:-1])
                        break
                partdict["IntegTime"]=0
                for item in range(len(filerawdata)):
                    if "initialstep:" in filerawdata[item]:
                        partdict["Vstep"]=float(filerawdata[item][13:-1])
                        break
                for item in range(len(filerawdata)):
                    if "InitialVoltage:" in filerawdata[item]:
                        partdict["Vstart"]=float(filerawdata[item][16:-1])
                        break
                partdict["Vend"]=0
                partdict["ExecTime"]=0
                for item in range(len(filerawdata)):
                    if "UserName:" in filerawdata[item]:
                        partdict["Operator"]=str(filerawdata[item][10:-1])
                        break
                partdict["Group"]="Default group"
                partdict["Group2"]=""
                partdict["Setup"]="SERISpythonIV"
                partdict["sunintensity"]=-1
                for item in range(len(filerawdata)):
                    if "#sun:" in filerawdata[item]:
                        partdict["sunintensity"]=float(filerawdata[item][6:-1])
                        break
                
                if partdict["sunintensity"]>0 and partdict["sunintensity"]!=-1:
                    partdict["Illumination"]="Light"
                else:
                    partdict["Illumination"]="Dark"
                partdict["SerialNumber"]=''#forSERIS
                for item in range(len(filerawdata)):
                    if "#MPPT data" in filerawdata[item]:
                            pos=item+2
                            break
                        
                mpppartdat = [[],[],[],[],[]]#[voltage,current,time,power,vstep]
                for item in range(pos,len(filerawdata),1):
                    mpppartdat[0].append(float(filerawdata[item].split("\t")[2]))
                    mpppartdat[1].append(float(filerawdata[item].split("\t")[3]))
                    mpppartdat[2].append(float(filerawdata[item].split("\t")[0]))
                    mpppartdat[3].append(float(filerawdata[item].split("\t")[1]))
                    mpppartdat[4].append(float(filerawdata[item].split("\t")[5]))
                partdict["PowerEnd"]=mpppartdat[3][-1]
                partdict["PowerAvg"]=sum(mpppartdat[3])/float(len(mpppartdat[3]))
                partdict["trackingduration"]=mpppartdat[2][-1]
                partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(partdict["PowerEnd"])

                partdict["MppData"]=mpppartdat
                partdict["SampleName"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(':','').replace(' ','-')
                partdict["MPPlinestyle"]=[partdict["SampleName"],"-",colors[len(DATAMPP.keys())],2]
                
                DATAMPP[partdict["SampleNameID"]]=partdict
                
            self.change_value.emit(100*(i+1)/len(self.file_path))
        self.finished.emit()
        

class Thread_getdatalistsfromSERISsinusGUI(QThread):
    finished = pyqtSignal()
    change_value = pyqtSignal(int)
    def __init__(self, file_path, parent=None):
        QThread.__init__(self, parent)
        self.file_path=file_path
        
    def run(self):
        global DATA, DATAdark, DATAFFloss
        global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
        print('threadstart')
        
        dictmeaspos=[]
        for i in range(len(self.file_path)):
            wb = xlrd.open_workbook(self.file_path[i])
            IVsummarySheet = wb.sheet_by_name('IV-Summary')
            IVrawSheet = wb.sheet_by_name('IV-Raw')
            LEsummarySheet = wb.sheet_by_name('LE-Summary')
            
            measpositions=[[[],[],[]],[[],[],[]],[[],[],[]]]#positions of (light,sunsvoc,dark) in sheets (IV-summary,IV-raw,LE-summary)
            col=3
            while(True):
                try:
                    meastype=str(IVsummarySheet.cell(0,col+1).value)
                    if 'Input' in str(IVsummarySheet.cell(1,col).value):
                        if ('Light' in meastype or '1sun forward' in meastype or '1sun reverse' in meastype or 'Reverse' in meastype or 'Forward' in meastype or 'AM1.5G' in meastype) and 'Lightsoak' not in meastype:
                            measpositions[0][0].append(col)
                        elif 'Suns-Voc' in meastype:
                            measpositions[0][1].append(col)
                        elif 'Dark' in meastype:
                            measpositions[0][2].append(col)
                    col+=1
                except IndexError:
                    break
            col=3
            while(True):
                try:
                    meastype=str(IVrawSheet.cell(0,col+1).value)
                    if 'Curve' in str(IVrawSheet.cell(1,col+1).value):
                        if ('Light' in meastype or '1sun forward' in meastype or '1sun reverse' in meastype or 'Reverse' in meastype or 'Forward' in meastype or 'AM1.5G' in meastype) and 'Lightsoak' not in meastype:
                            measpositions[1][0].append(col)
                        elif 'Suns-Voc' in meastype :
                            measpositions[1][1].append(col)
                        elif 'Dark' in meastype :
                            measpositions[1][2].append(col)
                    col+=1
                except IndexError:
                    break
            col=3
            while(True):
                try:
                    meastype=str(LEsummarySheet.cell(0,col+1).value)
                    if 'Input' in str(LEsummarySheet.cell(1,col).value):
                        if ('Light' in meastype or '1sun forward' in meastype or '1sun reverse' in meastype or 'Reverse' in meastype or 'Forward' in meastype or 'AM1.5G' in meastype) and 'Lightsoak' not in meastype:
                            measpositions[2][0].append(col)
                        elif 'Suns-Voc' in meastype:
                            measpositions[2][1].append(col)
                        elif 'Dark' in meastype :
                            measpositions[2][2].append(col)
                    col+=1
                except IndexError:
                    break
            # print(measpositions[0][0])
            dictmeaspos.append(measpositions)
        numbnewIVfiles=0
        for i in range(len(dictmeaspos)):
            numbnewIVfiles=numbnewIVfiles+len(dictmeaspos[i][0][0])+len(dictmeaspos[i][0][1])+len(dictmeaspos[i][0][2])
        cmap = plt.get_cmap(colormapname)
        num_plotsIV=len(DATA.keys())+numbnewIVfiles
        colorsIV = cmap(np.linspace(0, 1.0, numbnewIVfiles))
        colorsIV=[matplotlib.colors.to_hex(item) for item in colorsIV]
        
        # print(len(colorsIV))
        filenumb=0
        for i in range(len(dictmeaspos)):
            wb = xlrd.open_workbook(self.file_path[i])
            
            IVsummarySheet = wb.sheet_by_name('IV-Summary')
            IVrawSheet = wb.sheet_by_name('IV-Raw')
            LEsummarySheet = wb.sheet_by_name('LE-Summary')
            
            isItforFFloss=0
            if len(dictmeaspos[i][0][0])!=0 and len(dictmeaspos[i][0][1])!=0 and len(dictmeaspos[i][0][2])!=0:
                isItforFFloss=1
                partdictFFloss=[]
            
            #dealing with light
            for ii in range(len(dictmeaspos[i][0][0])):
                ivsumpos=dictmeaspos[i][0][0][ii]
                ivrawpos=dictmeaspos[i][1][0][ii]
                # lesumpos=dictmeaspos[i][2][0][ii]
                # print(ivsumpos)
                partdict = {}
                partdict["filepath"]=self.file_path[i]
                        
                # print(str(IVsummarySheet.cell(3,1).value))
                if len(str(IVsummarySheet.cell(3,1).value).split('_'))==3:
                    partdict["Cellletter"]=str(IVsummarySheet.cell(3,1).value).split('_')[2]
                else:
                    partdict["Cellletter"]='S'
                partdict["batchname"]=str(IVsummarySheet.cell(3,1).value).split('_')[0]
                partdict["SampleName"]=str(IVsummarySheet.cell(3,1).value)
                try:
                    partdict["DepID"]=partdict["batchname"]+'_'+str(IVsummarySheet.cell(3,1).value).split('_')[1]
                except:
                    partdict["DepID"]=partdict["batchname"]
                
                partdict["MeasDayTime"]=xldate_to_datetime(IVsummarySheet.cell(2,1).value)
                partdict["MeasDayTime2"]=parser.parse(partdict["MeasDayTime"])
                
                partdict["Illumination"]="Light"
                
                try:
                    partdict["sunintensity"]=float(IVsummarySheet.cell(17,ivsumpos+1).value)*1/100 # percent to numb of sun
                except:
                    partdict["sunintensity"]=1
                    
                partdict["MeasComment"]="-"
                partdict["aftermpp"]=0
                
                partdict["Operator"]=str(IVsummarySheet.cell(6,1).value)
                partdict["Group"]="Default group"
                partdict["Setup"]=str(IVsummarySheet.cell(10,1).value)
                partdict["AirTemp"]=IVsummarySheet.cell(19,ivsumpos+5).value
                
                partdict["Celltype"]=str(IVsummarySheet.cell(12,1).value)
                partdict["SerialNumber"]=str(IVsummarySheet.cell(1,1).value)
                partdict["IVlinestyle"]=[partdict["SampleName"],"-",colorsIV[filenumb],2]

                partdict["CellSurface"]=float(IVsummarySheet.cell(17,ivsumpos+5).value)
                # print(partdict["CellSurface"])
                
                # partdict["Voc"]=float(IVsummarySheet.cell(5,ivsumpos+5).value)*1000 #mV
                # partdict["Isc"]=float(IVsummarySheet.cell(4,ivsumpos+5).value)*1000 #mA
                # partdict["Jsc"]=partdict["Isc"]/partdict["CellSurface"] #mA/cm2
                # partdict["FF"]=float(IVsummarySheet.cell(6,ivsumpos+5).value) #%
                # partdict["Eff"]=float(IVsummarySheet.cell(11,ivsumpos+5).value)#%
                # partdict["Pmpp"]=float(IVsummarySheet.cell(10,ivsumpos+5).value)/partdict["CellSurface"] #W/cm2
                # partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
                # partdict["Roc"]=float(IVsummarySheet.cell(7,ivsumpos+5).value)*partdict["CellSurface"] #ohm.cm2
                # partdict["Rsc"]=-1
                # partdict["RscJsc"]=-1
                # partdict["Vmpp"]=float(IVsummarySheet.cell(8,ivsumpos+5).value)*1000
                # partdict["Jmpp"]=float(IVsummarySheet.cell(9,ivsumpos+5).value)*1000/partdict["CellSurface"]
                
                #scan the IVRaw sheet
                ivpartdat = [[],[]]#[voltage,current]
                row=4
                while(True):
                    try:
                        ivpartdat[0].append(float(IVrawSheet.cell(row,ivrawpos).value))
                        ivpartdat[1].append(float(IVrawSheet.cell(row,ivrawpos+1).value)*1000/partdict["CellSurface"]) # mA/cm2
                        row+=1
                    except:
                        break
                partdict["IVData"]=ivpartdat
                
                params=extract_jv_params(ivpartdat)
            
                partdict["Voc"]= float(params['Voc'])*1000 #mV
                partdict["Jsc"]= float(params['Jsc'])#mA/cm2
                partdict["Isc"]=float(params['Jsc']*partdict["CellSurface"])
                partdict["FF"]=float(params['FF']) #%
                partdict["Eff"]=float(params['Pmax'])#%
                partdict["Pmpp"]=float(partdict["Eff"])*10#W/cm2
                partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
                partdict["Roc"]=float(params['Roc'])
                partdict["Rsc"]=float(params['Rsc'])
                partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
                partdict["Vmpp"]=float(params['Vmpp'])
                partdict["Jmpp"]=float(params['Jmpp'])
                
                
                
                partdict["Vstart"]=ivpartdat[0][0]
                partdict["Vend"]=ivpartdat[0][-1]
                if partdict["Vstart"]>partdict["Vend"]:
                    partdict["ScanDirection"]="Reverse"
                else:
                    partdict["ScanDirection"]="Forward"
                partdict["NbPoints"]=len(ivpartdat[0])
                
                #scan the LEsummary sheet
                
                # partdict["ChuckTemp"]=LEsummarySheet.cell(7,lesumpos+5).value
                partdict["ChuckTemp"]=999
                partdict["Delay"]=999
                partdict["IntegTime"]=999
                partdict["RefNomCurr"]=999
                partdict["RefMeasCurr"]=999
                
                # ivpartdat=[np.array([x for x,y in sorted(zip(ivpartdat[0],ivpartdat[1]))]),np.array([y for x,y in sorted(zip(ivpartdat[0],ivpartdat[1]))])]
                # print(len(ivpartdat[0]))
                # print(len(set(ivpartdat[0])))
                # it generates a ValueError because there is a duplicate x value for some reason...
                # try:
                # if partdict["Illumination"]=="Light" and max(ivpartdat[0])>0.001*float(partdict["Voc"]):
                #     f = interp1d(ivpartdat[0], ivpartdat[1], kind='cubic', assume_sorted = False)
                #     x2 = lambda x: f(x)
                #     partdict["AreaJV"] = integrate.quad(x2,0,0.001*float(partdict["Voc"]))[0]
                # else:
                partdict["AreaJV"] =""
                # except ValueError:
                    # print("there is a ValueError on sample ",i)

                
                partdict["ImaxComp"]=-1
                partdict["Isenserange"]=-1
                
                partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(partdict["Isc"]))+'_'+str(float(partdict["FF"]))
                DATA[partdict["SampleNameID"]]=partdict
                numbLightfiles+=1
                filenumb+=1
                if isItforFFloss and DATA[partdict["SampleNameID"]]["sunintensity"]==1:
                    partdictFFloss.append(partdict["SampleNameID"])
                    DATAFFloss[partdict["SampleNameID"]]={'Eff':DATA[partdict["SampleNameID"]]["Eff"],
                                                          'Voc':DATA[partdict["SampleNameID"]]["Voc"],
                                                          'Jsc':DATA[partdict["SampleNameID"]]["Jsc"],
                                                          'Vmpp':DATA[partdict["SampleNameID"]]["Vmpp"],
                                                          'Jmpp':DATA[partdict["SampleNameID"]]["Jmpp"],
                                                          'FF':DATA[partdict["SampleNameID"]]["FF"],
                                                          'Roc':DATA[partdict["SampleNameID"]]["Roc"],
                                                          'Rsh':0,
                                                          'pFF':0,'pVoc':0,'pJsc':0,'pVmpp':0,'pJmpp':0,
                                                          'idFF':0,'idVmpp':0,'idJmpp':0,
                                                          'VmppPlusJmppRs':0,'SerisResistanceAbs':0,'Shuntterm':0,'FFlossSeries':0,'FFlossShunt':0,'FFlossJo2':0,
                                                          }
                
                
            #dealing with sunsvoc
            for ii in range(len(dictmeaspos[i][0][2])):
                ivsumpos=dictmeaspos[i][0][1][ii]
                ivrawpos=dictmeaspos[i][1][1][ii]
                lesumpos=dictmeaspos[i][2][1][ii]
                # print(ivsumpos)
                partdict = {}
                partdict["filepath"]=self.file_path[i]
                        
                # print(str(IVsummarySheet.cell(3,1).value))
                if len(str(IVsummarySheet.cell(3,1).value).split('_'))==3:
                    partdict["Cellletter"]=str(IVsummarySheet.cell(3,1).value).split('_')[2]
                else:
                    partdict["Cellletter"]='S'
                partdict["batchname"]=str(IVsummarySheet.cell(3,1).value).split('_')[0]
                partdict["SampleName"]=str(IVsummarySheet.cell(3,1).value)+'_SunsVoc'
                partdict["DepID"]=partdict["batchname"]+'_'+str(IVsummarySheet.cell(3,1).value).split('_')[1]
                
                partdict["MeasDayTime"]=xldate_to_datetime(IVsummarySheet.cell(2,1).value)
                partdict["MeasDayTime2"]=parser.parse(partdict["MeasDayTime"])
                
                partdict["Illumination"]="Light"
                partdict["sunintensity"]=float(IVsummarySheet.cell(12,ivsumpos+1).value)*1/100 # percent to numb of sun
                
                partdict["MeasComment"]="-"
                partdict["aftermpp"]=0
                
                partdict["Operator"]=str(IVsummarySheet.cell(6,1).value)
                partdict["Group"]="Default group"
                partdict["Setup"]=str(IVsummarySheet.cell(10,1).value)
                partdict["AirTemp"]=IVsummarySheet.cell(14,ivsumpos+5).value
                
                partdict["Celltype"]=str(IVsummarySheet.cell(12,1).value)
                partdict["SerialNumber"]=str(IVsummarySheet.cell(1,1).value)
                partdict["IVlinestyle"]=[partdict["SampleName"],"-",colorsIV[filenumb],2]
                
                #here we take "pseudo" param
                partdict["CellSurface"]=float(IVsummarySheet.cell(13,ivsumpos+5).value)
                partdict["Voc"]=float(IVsummarySheet.cell(7,ivsumpos+5).value)*1000 #mV
                partdict["Isc"]=float(IVsummarySheet.cell(6,ivsumpos+5).value)*1000 #mA
                partdict["Jsc"]=partdict["Isc"]/partdict["CellSurface"] #mA/cm2
                partdict["FF"]=float(IVsummarySheet.cell(8,ivsumpos+5).value) #%
                partdict["Eff"]=float(IVsummarySheet.cell(12,ivsumpos+5).value)#%
                partdict["Pmpp"]=float(IVsummarySheet.cell(11,ivsumpos+5).value)/partdict["CellSurface"] #W/cm2
                partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
                partdict["Roc"]=-1
                partdict["Rsc"]=-1
                partdict["RscJsc"]=-1
                partdict["Vmpp"]=float(IVsummarySheet.cell(9,ivsumpos+5).value)*1000
                partdict["Jmpp"]=float(IVsummarySheet.cell(10,ivsumpos+5).value)*1000/partdict["CellSurface"]
                
                #scan the IVRaw sheet
                ivpartdat = [[],[]]#[voltage,current]
                row=4
                while(True):
                    try:
                        ivpartdat[0].append(float(IVrawSheet.cell(row,ivrawpos).value))# V
                        ivpartdat[1].append(float(IVrawSheet.cell(row,ivrawpos+1).value)*1000/partdict["CellSurface"]) # mA/cm2
                        row+=1
                    except:
                        break
                
                partdict["IVData"]=[[x for x,y in sorted(zip(ivpartdat[0],ivpartdat[1]))],[y for x,y in sorted(zip(ivpartdat[0],ivpartdat[1]))]]
                partdict["Vstart"]=ivpartdat[0][0]
                partdict["Vend"]=ivpartdat[0][-1]
                if partdict["Vstart"]>partdict["Vend"]:
                    partdict["ScanDirection"]="Reverse"
                else:
                    partdict["ScanDirection"]="Forward"
                partdict["NbPoints"]=len(ivpartdat[0])
                
                #scan the LEsummary sheet
                
                partdict["ChuckTemp"]=LEsummarySheet.cell(7,lesumpos+5).value
                
                partdict["Delay"]=999
                partdict["IntegTime"]=999
                partdict["RefNomCurr"]=999
                partdict["RefMeasCurr"]=999
                
                partdict["AreaJV"] =""
                
                partdict["ImaxComp"]=-1
                partdict["Isenserange"]=-1
                
                partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(partdict["Isc"]))+'_'+str(float(partdict["FF"]))
                DATA[partdict["SampleNameID"]]=partdict
                numbLightfiles+=1
                filenumb+=1
                
                if isItforFFloss :
                    for item in partdictFFloss:
                        DATAFFloss[item]['pFF']=DATA[partdict["SampleNameID"]]['FF']
                        DATAFFloss[item]['pVoc']=DATA[partdict["SampleNameID"]]['Voc']
                        DATAFFloss[item]['pJsc']=DATA[partdict["SampleNameID"]]['Jsc']
                        DATAFFloss[item]['pVmpp']=DATA[partdict["SampleNameID"]]['Vmpp']
                        DATAFFloss[item]['pJmpp']=DATA[partdict["SampleNameID"]]['Jmpp']
                
            #dealing with dark
            for ii in range(len(dictmeaspos[i][0][2])):
                ivsumpos=dictmeaspos[i][0][2][ii]
                ivrawpos=dictmeaspos[i][1][2][ii]
                lesumpos=dictmeaspos[i][2][2][ii]
                # print(ivsumpos)
                partdict = {}
                partdict["filepath"]=self.file_path[i]
                
                # print(str(IVsummarySheet.cell(3,1).value))
                if len(str(IVsummarySheet.cell(3,1).value).split('_'))==3:
                    partdict["Cellletter"]=str(IVsummarySheet.cell(3,1).value).split('_')[2]
                else:
                    partdict["Cellletter"]='S'
                partdict["batchname"]=str(IVsummarySheet.cell(3,1).value).split('_')[0]
                partdict["SampleName"]=str(IVsummarySheet.cell(3,1).value)
                partdict["DepID"]=partdict["batchname"]+'_'+str(IVsummarySheet.cell(3,1).value).split('_')[1]
                
                partdict["MeasDayTime"]=xldate_to_datetime(IVsummarySheet.cell(2,1).value)
                partdict["MeasDayTime2"]=parser.parse(partdict["MeasDayTime"])
                
                partdict["Illumination"]="Dark"
                partdict["sunintensity"]=0 # percent to numb of sun
                
                partdict["MeasComment"]="-"
                partdict["aftermpp"]=0
                
                partdict["Operator"]=str(IVsummarySheet.cell(6,1).value)
                partdict["Group"]="Default group"
                partdict["Setup"]=str(IVsummarySheet.cell(10,1).value)
                partdict["AirTemp"]=IVsummarySheet.cell(20,ivsumpos+5).value
                
                partdict["Celltype"]=str(IVsummarySheet.cell(12,1).value)
                partdict["SerialNumber"]=str(IVsummarySheet.cell(1,1).value)
                partdict["IVlinestyle"]=[partdict["SampleName"],"-",colorsIV[filenumb],2]

                partdict["CellSurface"]=float(IVsummarySheet.cell(19,ivsumpos+5).value)
                partdict["Voc"]=-1 #mV
                partdict["Isc"]=float(IVsummarySheet.cell(4,ivsumpos+5).value)*1000 #mA
                partdict["Jsc"]=partdict["Isc"]/partdict["CellSurface"] #mA/cm2
                partdict["FF"]=-1 #%
                partdict["Eff"]=-1#%
                partdict["Pmpp"]=-1 #W/cm2
                partdict["VocFF"]=-1
                partdict["Roc"]=-1
                partdict["Rsc"]=float(IVsummarySheet.cell(9,ivsumpos+5).value)*partdict["CellSurface"] #ohm.cm2
                partdict["RscJsc"]=partdict["Jsc"]*partdict["Rsc"]
                partdict["Vmpp"]=-1
                partdict["Jmpp"]=-1
                
                #scan the IVRaw sheet
                ivpartdat = [[],[]]#[voltage,current]
                row=4
                while(True):
                    try:
                        ivpartdat[0].append(float(IVrawSheet.cell(row,ivrawpos).value))# V
                        ivpartdat[1].append(float(IVrawSheet.cell(row,ivrawpos+1).value)*1000/partdict["CellSurface"]) # mA/cm2
                        row+=1
                    except:
                        break
                partdict["IVData"]=ivpartdat
                partdict["Vstart"]=ivpartdat[0][0]
                partdict["Vend"]=ivpartdat[0][-1]
                if partdict["Vstart"]>partdict["Vend"]:
                    partdict["ScanDirection"]="Reverse"
                else:
                    partdict["ScanDirection"]="Forward"
                partdict["NbPoints"]=len(ivpartdat[0])
                
                #scan the LEsummary sheet
                
                partdict["ChuckTemp"]=LEsummarySheet.cell(7,lesumpos+5).value
                
                partdict["Delay"]=999
                partdict["IntegTime"]=999
                partdict["RefNomCurr"]=999
                partdict["RefMeasCurr"]=999
                
                partdict["AreaJV"] =""
                partdict["ImaxComp"]=-1
                partdict["Isenserange"]=-1
                
                partdict["SampleName"]=partdict["SampleName"]+'_D'
                partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+str(float(partdict["Isc"]))+'_'+str(float(partdict["FF"]))
                DATA[partdict["SampleNameID"]]=partdict
                DATAdark.append(partdict)
                numbDarkfiles+=1
                filenumb+=1
                if isItforFFloss :
                    for item in partdictFFloss:
                        DATAFFloss[item]['Rsh']=DATA[partdict["SampleNameID"]]['Rsc']
                        
                        xideal=[x*DATAFFloss[item]['Voc']/1000 for x in range(1000)]
                        yideal=[DATAFFloss[item]['Jsc']-(DATAFFloss[item]['Jsc']/math.exp(DATAFFloss[item]['Voc']/25.7))*(math.exp(item1/25.7)-1) for item1 in xideal]
                        powerideal=[xideal[item1]*yideal[item1]/1000 for item1 in range(1000)]
                        DATAFFloss[item]['xideal']=xideal
                        DATAFFloss[item]['yideal']=yideal
                        DATAFFloss[item]['powerideal']=powerideal
                        
                        DATAFFloss[item]['idFF']=100*max(powerideal)/(DATAFFloss[item]['Voc']*DATAFFloss[item]['Jsc']/1000)
                        DATAFFloss[item]['idVmpp']=xideal[powerideal.index(max(powerideal))]
                        DATAFFloss[item]['idJmpp']=yideal[powerideal.index(max(powerideal))]
                        VmppJmppRs=DATAFFloss[item]['Vmpp']+DATAFFloss[item]['Jmpp']*DATAFFloss[item]['Roc']
                        DATAFFloss[item]['VmppPlusJmppRs']=VmppJmppRs
                        DATAFFloss[item]['SerisResistanceAbs']=100*(DATAFFloss[item]['Jmpp']*DATAFFloss[item]['Jmpp']*DATAFFloss[item]['Roc'])/(DATAFFloss[item]['Voc']*DATAFFloss[item]['Jsc'])
                        DATAFFloss[item]['Shuntterm']=100*VmppJmppRs*VmppJmppRs/(DATAFFloss[item]['Voc']*DATAFFloss[item]['Jsc']*DATAFFloss[item]['Rsh'])
                        if DATAFFloss[item]['pFF']<DATAFFloss[item]['FF']:
                            DATAFFloss[item]['FFlossSeries']=DATAFFloss[item]['SerisResistanceAbs']
                        else:
                            DATAFFloss[item]['FFlossSeries']=DATAFFloss[item]['pFF']-DATAFFloss[item]['FF']
                        
                        DATAFFloss[item]['FFlossShunt']=DATAFFloss[item]['Shuntterm']
                        DATAFFloss[item]['FFlossJo2']=DATAFFloss[item]['idFF']-DATAFFloss[item]['FFlossSeries']-DATAFFloss[item]['FFlossShunt']-DATAFFloss[item]['FF']
                        
                        
            self.change_value.emit(100*(i+1)/len(self.file_path))
        self.finished.emit()
        
# class Thread_getdatalistsfromCUBfiles(QThread):
#     finished = pyqtSignal()
#     change_value = pyqtSignal(int)
#     def __init__(self, file_path, parent=None):
#         QThread.__init__(self, parent)
#         self.file_path=file_path
        
#     def run(self):
#         global DATA, DATAdark
#         global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
#         print('threadstart')
        
#         num_plots=len(DATA.keys())+len(file_path)
#         cmap = plt.get_cmap(colormapname)
#         colors = cmap(np.linspace(0, 1.0, num_plots))
#         colors=[tuple(item) for item in colors]
        
#         for i in range(len(self.file_path)):
#             filetoread = open(self.file_path[i],"r", encoding='ISO-8859-1')
#             filerawdata = filetoread.readlines()
                              
#             partdict = {}
#             partdict["filepath"]=self.file_path[i]
            
#             filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]                
            
#             partdict["Cellletter"]=filename.split('_')[2][2:]
#             partdict["batchname"]=filename.split('_')[0]
#             partdict["DepID"]=partdict["batchname"]+"_"+filename.split('_')[1]
#             partdict["SampleName"]=partdict["DepID"]+"_"+partdict["Cellletter"] #+"_"+aftername.split('_')[4]
            
#             if "light" in filename:
#                 partdict["Illumination"]="Light"
#             else:
#                 partdict["Illumination"]="Dark"
                
#             if "rev" in filename:
#                 partdict["ScanDirection"]="Reverse"
#             else:
#                 partdict["ScanDirection"]="Forward" 
            
            
#             partdict["MeasDayTime2"]=parser.parse(filerawdata[0])
#             partdict["MeasDayTime"]=filerawdata[0]
# #                print(partdict["MeasDayTime2"])
# #                print(partdict["MeasDayTime"])
                    
#             partdict["MeasComment"]="-"
#             for item in range(len(filerawdata)):
#                 if "Notes = " in filerawdata[item]:
#                     partdict["MeasComment"]=filerawdata[item][8:-1]
#                     break
#             if "aftermpp" in partdict["MeasComment"]:
#                 partdict["aftermpp"]=1
#             else:
#                 partdict["aftermpp"]=0
                
#             for item in range(len(filerawdata)):
#                 if "Device Area = " in filerawdata[item]:
#                     partdict["CellSurface"]=float(filerawdata[item][14:-5])
# #                        print(partdict["CellSurface"])
#                     break
#             for item in range(len(filerawdata)):
#                 if "Delay = " in filerawdata[item]:
#                     partdict["Delay"]=float(filerawdata[item][8:-3])
# #                        print(partdict["Delay"])
#                     break
#             for item in range(len(filerawdata)):
#                 if "NPLC = " in filerawdata[item]:
#                     partdict["IntegTime"]=float(filerawdata[item][7:-1])
#                     break     
            
#             for item in range(len(filerawdata)):
#                 if "Voltage" in filerawdata[item]:
#                         pos=item+1
#                         break
                    
#             ivpartdat = [[],[]]#[voltage,current]
#             for item in range(pos,len(filerawdata),1):
#                 try:
#                     ivpartdat[0].append(float(filerawdata[item].split("\t")[0]))
#                     ivpartdat[1].append(float(filerawdata[item].split("\t")[1]))
#                 except:
#                     break
#             partdict["IVData"]=ivpartdat
#             partdict["NbPoints"]=len(ivpartdat[0])
#             partdict["Vstart"]=ivpartdat[0][-1]
#             partdict["Vend"]=ivpartdat[0][0]
                    
#             params=extract_jv_params(partdict["IVData"])
#             partdict["Voc"]=params['Voc']*1000 #mV
#             partdict["Jsc"]=params['Jsc'] #mA/cm2
#             partdict["Isc"]=params['Jsc']*partdict["CellSurface"]
#             partdict["FF"]=params['FF'] #%
#             partdict["Eff"]=params['Pmax'] #%
#             partdict["Pmpp"]=partdict["Eff"]*10 #W/cm2
#             partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
#             partdict["Roc"]=params['Roc'] 
#             partdict["Rsc"]=params['Rsc'] 
#             partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
            
#             partdict["Vmpp"]=params['Vmpp']
#             partdict["Jmpp"]=params['Jmpp']
#             partdict["ImaxComp"]=-1
#             partdict["Isenserange"]=-1
            
#             partdict["Operator"]=-1
                          
#             try:
#                 if partdict["Illumination"]=="Light" and max(ivpartdat[0])>0.001*float(partdict["Voc"]):
#                     f = interp1d(ivpartdat[0], ivpartdat[1], kind='cubic')
#                     x2 = lambda x: f(x)
#                     partdict["AreaJV"] = integrate.quad(x2,0,0.001*float(partdict["Voc"]))[0]
#                 else:
#                     partdict["AreaJV"] =""
#             except ValueError:
#                 print("there is a ValueError on sample ",i)
            
            
#             partdict["Group"]="Default group"
#             partdict["Setup"]="CUBoulder"              
#             partdict["RefNomCurr"]=999
#             partdict["RefMeasCurr"]=999
#             partdict["AirTemp"]=999
#             partdict["ChuckTemp"]=999
#             partdict["IVlinestyle"]=[partdict["SampleName"],"-",colors[len(DATA.keys())],2]
            
#             if partdict["Illumination"]=="Light":
#                 # DATA.append(partdict)
#                 partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])
#                 DATA[partdict["SampleNameID"]]=partdict
#                 numbLightfiles+=1
#             else:
#                 partdict["SampleName"]=partdict["SampleName"]+'_D'
#                 partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])
#                 DATA[partdict["SampleNameID"]]=partdict
#                 DATAdark.append(partdict)
#                 numbDarkfiles+=1
                
#             self.change_value.emit(100*(i+1)/len(self.file_path))
            
#         self.finished.emit()
        
# class Thread_getdatalistsfromNRELcigssetup(QThread):
#     finished = pyqtSignal()
#     change_value = pyqtSignal(int)
#     def __init__(self, file_path, parent=None):
#         QThread.__init__(self, parent)
#         self.file_path=file_path
        
#     def run(self):
#         global DATA, DATAdark
#         global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
#         print('threadstart')
        
#         num_plots=len(DATA.keys())+len(file_path)
#         cmap = plt.get_cmap(colormapname)
#         colors = cmap(np.linspace(0, 1.0, num_plots))
#         colors=[tuple(item) for item in colors]
        
#         for i in range(len(self.file_path)):
#             filetoread = open(self.file_path[i],"r", encoding='ISO-8859-1')
#             filerawdata = filetoread.readlines()
#             if os.path.splitext(file_path[i])[1]=='.txt':
# #                print("txt mpp file")
#                 partdict = {}
#                 partdict["filepath"]=file_path[i]
#                 filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
#                 partdict["DepID"]=filename.split('.')[0]+'_'+filename.split('.')[1]
#                 partdict["SampleName"]=filename.split('.')[0]+'_'+filename.split('.')[1]+'_'+filename.split('.')[2]
#                 partdict["Cellletter"]='Single'
#                 partdict["batchname"]=filename.split('.')[0]
#                 partdict["MeasComment"]=filerawdata[0].split('\t')[-1]

#                 partdict["MeasDayTime"]=modification_date(file_path[i])

#                 partdict["CellSurface"]= 1

#                 partdict["Delay"]=0
#                 partdict["IntegTime"]=0
#                 partdict["Vstep"]=0
#                 partdict["Vstart"]=0
#                 partdict["Vend"]=0
#                 partdict["ExecTime"]=0
#                 partdict["Operator"]='unknown'
#                 partdict["Group"]="Default group"
                
#                 mpppartdat = [[],[],[],[],[]]#[voltage,current,time,power,vstep,delay]
#                 for item in range(2,len(filerawdata),1):
#                     mpppartdat[0].append(float(filerawdata[item].split("\t")[0]))
#                     mpppartdat[1].append(float(filerawdata[item].split("\t")[1]))
#                     mpppartdat[2].append(float(filerawdata[item].split("\t")[2]))
#                     mpppartdat[3].append(float(filerawdata[item].split("\t")[3]))
#                     mpppartdat[4].append(float(filerawdata[item].split("\t")[4]))
#                 partdict["PowerEnd"]=mpppartdat[3][-1]
#                 partdict["PowerAvg"]=sum(mpppartdat[3])/float(len(mpppartdat[3]))
#                 partdict["trackingduration"]=mpppartdat[2][-1]
#                 partdict["MppData"]=mpppartdat
#                 partdict["MPPlinestyle"]=[partdict["SampleName"],"-",colors[len(DATAMPP.keys())],2]
#                 DATAMPP.append(partdict)   
                
                
#             elif os.path.splitext(file_path[i])[1]=='.itx':
# #                print("cigs iv file")
#                 partdict = {}
#                 partdict["filepath"]=file_path[i]
                
#                 filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
# #                print(filename)
#                 if 'Reverse' in filename:
#                     partdict["DepID"]=filename[:filename.index('Reverse')-1]
#                     aftername=filename[filename.index('Reverse'):]
#                     partdict["ScanDirection"]="Reverse"
#                 elif 'Forward' in filename:
#                     partdict["DepID"]=filename[:filename.index('Forward')-1]
#                     aftername=filename[filename.index('Forward'):]
#                     partdict["ScanDirection"]="Forward" 
                
#                 partdict["Cellletter"]='Single'
#                 partdict["batchname"]=partdict["DepID"].split('.')[0]
#                 partdict["SampleName"]=partdict["DepID"]+"_"+aftername.split('.')[1]+"_"+aftername.split('.')[2]
                
# #                print(partdict["SampleName"])
                
#                 if 'LIV' in aftername:
#                     partdict["Illumination"]="Light"
#                     partdict["sunintensity"]=1
#                 elif 'DIV' in aftername:
#                     partdict["Illumination"]="Dark"
#                     partdict["sunintensity"]=0
                    
                    
#                 partdict["MeasDayTime2"]=modification_date(file_path[i])#'2020-01-29 12:55:00'
#                 partdict["MeasDayTime"]='Mon, Jan 01, 0000 0:00:00'
                
#                 for item in range(len(filerawdata)):
#                     if "X Note" in filerawdata[item]:
# #                        print(filerawdata[item].index('\\r'))
# #                        print(filerawdata[item][filerawdata[item].index('\\r')+2:filerawdata[item].index('\\rArea')-1])
#                         partdict["MeasDayTime2"]=parser.parse(filerawdata[item][filerawdata[item].index('\\r')+2:filerawdata[item].index('\\rArea')-1])
#                         partdict["MeasDayTime"]=filerawdata[item][filerawdata[item].index('\\r')+2:filerawdata[item].index('\\rArea')-1]
# #                        print(partdict["MeasDayTime2"])
#                         break
                
# #                partdict["MeasComment"]=filerawdata[-1][filerawdata[-1].index('"')+1:-3]
#                 partdict["MeasComment"]=''
# #                if "aftermpp" in partdict["MeasComment"]:
# #                    partdict["aftermpp"]=1
# #                else:
# #                    partdict["aftermpp"]=0
                
#                 for item in range(len(filerawdata)):
#                     if "X SetScale" in filerawdata[item]:
#                         partdict["Vstart"]=float(filerawdata[item][15:filerawdata[item].index(',')])
#                         break
#                 #vstep
#                 for item in range(len(filerawdata)):
#                     if "X SetScale" in filerawdata[item]:
#                         partdict["Vstep"]=float(filerawdata[item].split(',')[1])
#                         break
# #                print(partdict["Vstart"])
# #                print(partdict["Vstep"])
#                 ivpartdat = [[],[]]#[voltage,current]
#                 increm=0
#                 for item in range(3,len(filerawdata),1):
#                     if 'END' not in filerawdata[item]:
#                         ivpartdat[0].append(partdict["Vstart"]+increm*partdict["Vstep"])
# #                        print(item)
# #                        if partdict["ScanDirection"]=="Forward":
# #                            ivpartdat[0].append(partdict["Vstart"]+increm*partdict["Vstep"])
# #                        else:
# #                            ivpartdat[0].append(partdict["Vstart"]+increm*partdict["Vstep"])
#                         ivpartdat[1].append(float(filerawdata[item].split('\t')[2][:-2])) 
#                         increm+=1
#                     else:
#                         break
#                 partdict["IVData"]=ivpartdat
                
#                 partdict["Vend"]=ivpartdat[0][-1]
                
# #                if partdict["ScanDirection"]=="Reverse":
# #                    if partdict["Vstart"]<partdict["Vend"]:
# #                        vend=partdict["Vend"]
# #                        partdict["Vend"]=partdict["Vstart"]
# #                        partdict["Vstart"]=vend
# #                else:
# #                    if partdict["Vstart"]>partdict["Vend"]:
# #                        vend=partdict["Vend"]
# #                        partdict["Vend"]=partdict["Vstart"]
# #                        partdict["Vstart"]=vend 
                        
#                 partdict["NbPoints"]=len(ivpartdat[0])
#                 partdict["CellSurface"]=1
#                 for item in range(len(filerawdata)):
#                     if "X Note" in filerawdata[item]:
#                         # print(filerawdata[item])
#                         partdict["CellSurface"]=float(filerawdata[item][filerawdata[item].index('\\rArea')+18:filerawdata[item].index('\\rVoc')-1])
#                         break
                
#                 partdict["Delay"]=-1
#                 partdict["IntegTime"]=-1                        

#                 params=extract_jv_params(partdict["IVData"])
#                 partdict["Voc"]=params['Voc']*1000 #mV
#                 partdict["Jsc"]=params['Jsc'] #mA/cm2
#                 partdict["Isc"]=params['Jsc']*partdict["CellSurface"]
#                 partdict["FF"]=params['FF'] #%
#                 partdict["Eff"]=params['Pmax'] #%
#                 partdict["Pmpp"]=partdict["Eff"]*10 #W/cm2
#                 partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
#                 partdict["Roc"]=params['Roc'] 
#                 partdict["Rsc"]=params['Rsc'] 
#                 partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
                
#                 partdict["Vmpp"]=params['Vmpp']
#                 partdict["Jmpp"]=params['Jmpp']
#                 partdict["ImaxComp"]=-1
#                 partdict["Isenserange"]=-1
                
#                 partdict["Operator"]=-1
                              
#                 try:
#                     if partdict["Illumination"]=="Light" and max(ivpartdat[0])>0.001*float(partdict["Voc"]):
#                         f = interp1d(ivpartdat[0], ivpartdat[1], kind='cubic')
#                         x2 = lambda x: f(x)
#                         partdict["AreaJV"] = integrate.quad(x2,0,0.001*float(partdict["Voc"]))[0]
#                     else:
#                         partdict["AreaJV"] =""
#                 except ValueError:
#                     print("there is a ValueError on sample ",i)
                
                
#                 partdict["Group"]="Default group"
#                 partdict["Setup"]="SSIgorC215"              
#                 partdict["RefNomCurr"]=999
#                 partdict["RefMeasCurr"]=999
#                 partdict["AirTemp"]=999
#                 partdict["ChuckTemp"]=999
#                 partdict["IVlinestyle"]=[partdict["SampleName"],"-",colors[len(DATA.keys())],2]
# #                DATA.append(partdict)

#                 if partdict["Illumination"]=="Light":
#                     partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])

#                     DATA[partdict["SampleNameID"]]=partdict
#                     numbLightfiles+=1
#                 else:
#                     partdict["SampleName"]=partdict["SampleName"]+'_D'
#                     partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])
#                     DATA[partdict["SampleNameID"]]=partdict
#                     DATAdark.append(partdict)
#                     numbDarkfiles+=1
                
#             self.change_value.emit(100*(i+1)/len(self.file_path))
            
#         self.finished.emit()
        
        
# class Thread_getdatalistsfromIIIVsetupfiles(QThread):
#     finished = pyqtSignal()
#     change_value = pyqtSignal(int)
#     def __init__(self, file_path, parent=None):
#         QThread.__init__(self, parent)
#         self.file_path=file_path
        
#     def run(self):
#         global DATA, DATAdark
#         global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
#         print('threadstart')
        

        
#         for i in range(len(self.file_path)):
#             filetoread = open(self.file_path[i],"r", encoding='ISO-8859-1')
#             filerawdata = filetoread.readlines()
                              

            
#             measnames=filerawdata[0].split('\t\t\t')
#             measnames[-1]=measnames[-1][:-3]
#             # print(measnames)
#             # print(len(filerawdata[0].split('\t\t\t')))
            
#             num_plots=len(DATA.keys())+len(measnames)
#             cmap = plt.get_cmap(colormapname)
#             colors = cmap(np.linspace(0, 1.0, num_plots))
#             colors=[tuple(item) for item in colors]
#             for item in range(len(measnames)):
#                 partdict = {}
#                 partdict["filepath"]=self.file_path[i]
#                 filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
#                 ivpartdat = [[],[]]#[voltage,current]
#                 for row in range(1, len(filerawdata)):
#                     if filerawdata[row].split('\t')[item*3+1] != '':
#                         ivpartdat[0].append(float(filerawdata[row].split('\t')[item*3+1]))
#                         ivpartdat[1].append(float(filerawdata[row].split('\t')[item*3+2]))
                
#                 partdict["Cellletter"]='Z' #could search for the 'n' in name + number after
#                 partdict["batchname"]='X'
#                 partdict["DepID"]=partdict["batchname"]
#                 partdict["SampleName"]=measnames[item]
                
#                 if "L" in measnames[item]:
#                     partdict["Illumination"]="Light"
#                 else:
#                     partdict["Illumination"]="Dark"
                
#                 if ivpartdat[0][0]>ivpartdat[0][-1]:
#                     partdict["ScanDirection"]="Reverse"
#                 else:
#                     partdict["ScanDirection"]="Forward" 
            
            
#                 partdict["MeasDayTime2"]=''
#                 partdict["MeasDayTime"]=''
                        
#                 partdict["MeasComment"]="-"
                    
#                 partdict["CellSurface"]=0.09
#                 partdict["Delay"]=999
#                 partdict["IntegTime"]=999
#                 partdict["IVData"]=ivpartdat
#                 partdict["NbPoints"]=len(ivpartdat[0])
#                 partdict["Vstart"]=ivpartdat[0][0]
#                 partdict["Vend"]=ivpartdat[0][-1]
                
#                 params=extract_jv_params(partdict["IVData"])
#                 partdict["Voc"]=params['Voc']*1000 #mV
#                 partdict["Jsc"]=params['Jsc'] #mA/cm2
#                 partdict["Isc"]=params['Jsc']*partdict["CellSurface"]
#                 partdict["FF"]=params['FF'] #%
#                 partdict["Eff"]=params['Pmax'] #%
#                 partdict["Pmpp"]=partdict["Eff"]*10 #W/cm2
#                 partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
#                 partdict["Roc"]=params['Roc'] 
#                 partdict["Rsc"]=params['Rsc'] 
#                 partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
                
#                 partdict["Vmpp"]=params['Vmpp']
#                 partdict["Jmpp"]=params['Jmpp']
#                 partdict["ImaxComp"]=-1
#                 partdict["Isenserange"]=-1
                
#                 partdict["Operator"]=-1
                              
#                 try:
#                     if partdict["Illumination"]=="Light" and max(ivpartdat[0])>0.001*float(partdict["Voc"]):
#                         f = interp1d(ivpartdat[0], ivpartdat[1], kind='cubic')
#                         x2 = lambda x: f(x)
#                         partdict["AreaJV"] = integrate.quad(x2,0,0.001*float(partdict["Voc"]))[0]
#                     else:
#                         partdict["AreaJV"] =""
#                 except ValueError:
#                     print("there is a ValueError on sample ",i)
                
#                 partdict["sunintensity"]=1
#                 partdict["Group"]="Default group"
#                 partdict["Setup"]="IIIV"              
#                 partdict["RefNomCurr"]=999
#                 partdict["RefMeasCurr"]=999
#                 partdict["AirTemp"]=999
#                 partdict["ChuckTemp"]=999
#                 partdict["IVlinestyle"]=[partdict["SampleName"],"-",colors[len(DATA.keys())],2]
                
#                 if partdict["Illumination"]=="Light":
#                     # DATA.append(partdict)
#                     partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])
#                     DATA[partdict["SampleNameID"]]=partdict
#                     numbLightfiles+=1
#                 else:
#                     partdict["SampleName"]=partdict["SampleName"]+'_D'
#                     partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])
#                     DATA[partdict["SampleNameID"]]=partdict
#                     DATAdark.append(partdict)
#                     numbDarkfiles+=1
                    
#                 self.change_value.emit(100*(i+1)/len(self.file_path))
            
#         self.finished.emit()

# class Thread_getdatalistsfromNRELfiles(QThread):
#     finished = pyqtSignal()
#     change_value = pyqtSignal(int)
#     def __init__(self, file_path, parent=None):
#         QThread.__init__(self, parent)
#         self.file_path=file_path

#     def run(self):
#         global DATA, DATAdark
#         global DATAMPP, numbLightfiles, numbDarkfiles, colormapname
#         print('threadstart')
        
#         for i in range(len(self.file_path)):
#             filetoread = open(self.file_path[i],"r", encoding='ISO-8859-1')
#             filerawdata = filetoread.readlines()
#             # print(i)
#             filetype = 0
#             if "HEADER START" in filerawdata[0]:
#                 if 'SPO' in self.file_path[i]:
#                     filetype = 11
#                     num_plots=len(DATA.keys())+len(file_path)
#                     cmap = plt.get_cmap(colormapname)
#                     colors = cmap(np.linspace(0, 1.0, num_plots))
#                     colors=[tuple(item) for item in colors]
#                 else:
#                     filetype = 1 #JV file from solar simulator in SERF C215
#                     num_plots=len(DATA.keys())+len(file_path)
#                     cmap = plt.get_cmap(colormapname)
#                     colors = cmap(np.linspace(0, 1.0, num_plots))
#                     colors=[tuple(item) for item in colors]
#                 # print('jv')
#             elif "Power (mW/cm2)" in filerawdata[0]:
#                 filetype = 2
#                 num_plots=len(DATAMPP.keys())+len(file_path)
#                 cmap = plt.get_cmap(colormapname)
#                 colors = cmap(np.linspace(0, 1.0, num_plots))
#                 colors=[tuple(item) for item in colors]
#                 # print('mpp')
#             elif "V\tI" in filerawdata[0]:
#                 filetype = 3
#                 num_plots=len(DATA.keys())+len(file_path)
#                 cmap = plt.get_cmap(colormapname)
#                 colors = cmap(np.linspace(0, 1.0, num_plots))
#                 colors=[tuple(item) for item in colors]
#                 # print("JVT")
            
            
#             if filetype ==1 : #J-V files of SERF C215
                              
#                 partdict = {}
#                 partdict["filepath"]=self.file_path[i]
                
#                 filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
# #                print(filename)
#                 if 'rev' in filename:
#                     partdict["DepID"]=filename[:filename.index('rev')-1]
#                     aftername=filename[filename.index('rev'):]
#                 elif 'fwd' in filename:
#                     partdict["DepID"]=filename[:filename.index('fwd')-1]
#                     aftername=filename[filename.index('fwd'):]
                
#                 partdict["Cellletter"]=aftername.split('_')[3][2:]
#                 partdict["batchname"]=partdict["DepID"].split('_')[0]
#                 partdict["SampleName"]=partdict["DepID"]+"_"+partdict["Cellletter"]+"_"+aftername.split('_')[4]
                
#                 if aftername.split('_')[1]=="lt":
#                     partdict["Illumination"]="Light"
#                 else:
#                     partdict["Illumination"]="Dark"
                    
#                 if aftername.split('_')[0]=="rev":
#                     partdict["ScanDirection"]="Reverse"
#                 else:
#                     partdict["ScanDirection"]="Forward" 
                
#                 for item in range(len(filerawdata)):
#                     if "Date/Time:" in filerawdata[item]:
#                         partdict["MeasDayTime2"]=parser.parse(filerawdata[item][11:-1])
#                         partdict["MeasDayTime"]=filerawdata[item][11:-1]
#                         # print(partdict["MeasDayTime2"])
# #                        print(partdict["MeasDayTime"].split(' ')[-2])
#                         break
#                 for item in range(len(filerawdata)):
#                     if "Intensity:" in filerawdata[item]:
#                         partdict["sunintensity"]=float(filerawdata[item][11:-1])
#                         break
#                 partdict["MeasComment"]="-"
#                 for item in range(len(filerawdata)):
#                     if "Comments: " in filerawdata[item]:
#                         partdict["MeasComment"]=filerawdata[item][10:-1]
#                         break
#                 if "aftermpp" in partdict["MeasComment"]:
#                     partdict["aftermpp"]=1
#                 else:
#                     partdict["aftermpp"]=0
#                 for item in range(len(filerawdata)):
#                     if "Start V:" in filerawdata[item]:
#                         partdict["Vstart"]=float(filerawdata[item][9:-1])
#                         break
#                 for item in range(len(filerawdata)):
#                     if "End V:" in filerawdata[item]:
#                         partdict["Vend"]=float(filerawdata[item][7:-1])
#                         break
#                 if partdict["ScanDirection"]=="Reverse":
#                     if partdict["Vstart"]<partdict["Vend"]:
#                         vend=partdict["Vend"]
#                         partdict["Vend"]=partdict["Vstart"]
#                         partdict["Vstart"]=vend
#                 else:
#                     if partdict["Vstart"]>partdict["Vend"]:
#                         vend=partdict["Vend"]
#                         partdict["Vend"]=partdict["Vstart"]
#                         partdict["Vstart"]=vend 
#                 for item in range(len(filerawdata)):
#                     if "Number Points:" in filerawdata[item]:
#                         partdict["NbPoints"]=float(filerawdata[item][15:-1])
#                         break    
#                 for item in range(len(filerawdata)):
#                     if "Pixel Size:" in filerawdata[item]:
#                         partdict["CellSurface"]=float(filerawdata[item][12:-5])
#                         #print(partdict["CellSurface"])
#                         break
#                 for item in range(len(filerawdata)):
#                     if "Source Delay:" in filerawdata[item]:
#                         partdict["Delay"]=float(filerawdata[item][14:-1])
#                         break
#                 for item in range(len(filerawdata)):
#                     if "NPLC:" in filerawdata[item]:
#                         partdict["IntegTime"]=float(filerawdata[item][6:-1])
#                         break
#                 for item in range(len(filerawdata)):
#                     if "HEADER END" in filerawdata[item]:
#                             pos=item+3
#                             break
                        
#                 ivpartdat = [[],[]]#[voltage,current]
#                 for item in range(pos,len(filerawdata),1):
#                     try:
#                         ivpartdat[0].append(float(filerawdata[item].split("\t")[0]))
#                         ivpartdat[1].append(-float(filerawdata[item].split("\t")[1]))
#                     except:
#                         break
#                 partdict["IVData"]=ivpartdat
#                 params=extract_jv_params(partdict["IVData"])
#                 partdict["Voc"]=params['Voc']*1000 #mV
#                 partdict["Jsc"]=params['Jsc'] #mA/cm2
#                 partdict["Isc"]=params['Jsc']*partdict["CellSurface"]
#                 partdict["FF"]=params['FF'] #%
#                 partdict["Eff"]=params['Pmax'] #%
#                 partdict["Pmpp"]=partdict["Eff"]*10 #W/cm2
#                 partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
#                 partdict["Roc"]=params['Roc'] 
#                 partdict["Rsc"]=params['Rsc'] 
#                 partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
                
#                 partdict["Vmpp"]=params['Vmpp']
#                 partdict["Jmpp"]=params['Jmpp']
#                 partdict["ImaxComp"]=-1
#                 partdict["Isenserange"]=-1
                
#                 partdict["Operator"]=-1
                
                
                
#                 try:
#                     if partdict["Illumination"]=="Light" and max(ivpartdat[0])>0.001*float(partdict["Voc"]):
#                         f = interp1d(ivpartdat[0], ivpartdat[1], kind='cubic')
#                         x2 = lambda x: f(x)
#                         partdict["AreaJV"] = integrate.quad(x2,0,0.001*float(partdict["Voc"]))[0]
#                     else:
#                         partdict["AreaJV"] =""
#                 except ValueError:
#                     print("there is a ValueError on sample ",i)
                
                
#                 partdict["Group"]="Default group"
#                 partdict["Setup"]="SSIgorC215"              
#                 partdict["RefNomCurr"]=999
#                 partdict["RefMeasCurr"]=999
#                 partdict["AirTemp"]=999
#                 partdict["ChuckTemp"]=999
#                 partdict["IVlinestyle"]=[partdict["SampleName"],"-",colors[len(DATA.keys())],2]
# #                DATA.append(partdict)

#                 if partdict["Illumination"]=="Light":
#                     # DATA.append(partdict)
#                     partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])

#                     DATA[partdict["SampleNameID"]]=partdict
#                     numbLightfiles+=1
#                 else:
#                     partdict["SampleName"]=partdict["SampleName"]+'_D'
#                     partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])
#                     DATA[partdict["SampleNameID"]]=partdict
#                     DATAdark.append(partdict)
#                     numbDarkfiles+=1
#             elif filetype ==3 : #JVT files from Taylor
#                 partdict = {}
#                 partdict["filepath"]=self.file_path[i]
                
#                 filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
# #                print(filename)
                
#                 partdict["DepID"]=filename
#                 aftername=filename
    
                
#                 partdict["Cellletter"]='A'
#                 partdict["batchname"]='batch'
#                 partdict["SampleName"]=partdict["DepID"]
                
#                 partdict["Illumination"]="Light"
                    
#                 partdict["ScanDirection"]="Reverse"
                
                
#                 partdict["MeasDayTime2"]=modification_date(self.file_path[i])#'2020-01-29 12:55:00'
#                 partdict["MeasDayTime"]='Wed, Jan 29, 2020 0:00:00'
#                 partdict["MeasComment"]="-"
#                 partdict["aftermpp"]=1
#                 partdict["Vstart"]=0
#                 partdict["Vend"]=0
#                 partdict["NbPoints"]=0      
#                 partdict["CellSurface"]=0.1  
#                 partdict["Delay"]=0    
#                 partdict["IntegTime"]=0
#                 partdict["sunintensity"]=1
                        
#                 ivpartdat = [[],[]]#[voltage,current]
#                 for item in range(1,len(filerawdata),1):
#                     try:
#                         ivpartdat[0].append(float(filerawdata[item].split("\t")[0]))
#                         ivpartdat[1].append(1000*float(filerawdata[item].split("\t")[1])/partdict["CellSurface"])
#                     except:
#                         break
#                 partdict["IVData"]=ivpartdat
#                 params=extract_jv_params(partdict["IVData"])
#                 partdict["Voc"]=params['Voc']*1000 #mV
#                 partdict["Jsc"]=params['Jsc'] #mA/cm2
#                 partdict["Isc"]=params['Jsc']*partdict["CellSurface"]
#                 partdict["FF"]=params['FF'] #%
#                 partdict["Eff"]=params['Pmax'] #%
#                 partdict["Pmpp"]=partdict["Eff"]*10 #W/cm2
#                 partdict["VocFF"]=partdict["Voc"]*partdict["FF"]
#                 partdict["Roc"]=params['Roc'] 
#                 partdict["Rsc"]=params['Rsc'] 
#                 partdict["RscJsc"]=partdict["Rsc"]*partdict["Jsc"]
                
#                 partdict["Vmpp"]=params['Vmpp']
#                 partdict["Jmpp"]=params['Jmpp']
#                 partdict["ImaxComp"]=-1
#                 partdict["Isenserange"]=-1
                
#                 partdict["Operator"]=-1
#                 partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime2"]).replace(' ','_').replace(':','-')+'_'+'%.2f'%float(partdict["Eff"])

#                 try:
#                     if partdict["Illumination"]=="Light" and max(ivpartdat[0])>0.001*float(partdict["Voc"]):
#                         f = interp1d(ivpartdat[0], ivpartdat[1], kind='cubic')
#                         x2 = lambda x: f(x)
#                         partdict["AreaJV"] = integrate.quad(x2,0,0.001*float(partdict["Voc"]))[0]
#                     else:
#                         partdict["AreaJV"] =""
#                 except ValueError:
#                     print("there is a ValueError on sample ",i)
                
#                 partdict["IVlinestyle"]=[partdict["SampleName"],"-",colors[len(DATA.keys())],2]

#                 partdict["Group"]="Default group"
#                 partdict["Setup"]="JVT"              
#                 partdict["RefNomCurr"]=999
#                 partdict["RefMeasCurr"]=999
#                 partdict["AirTemp"]=999
#                 partdict["ChuckTemp"]=999
                    
# #                DATA.append(partdict)
#                 DATA[partdict["SampleNameID"]]=partdict
#                 numbLightfiles+=1
                
#             elif filetype ==2 : #mpp files of SERF C215 labview program
#                 #assumes file name: batch_samplenumber_cellLetter_mpp
#                 partdict = {}
#                 partdict["filepath"]=self.file_path[i]
#                 filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
#                 partdict["DepID"]=filename.split('_')[0]+'_'+filename.split('_')[1]
#                 partdict["SampleName"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+filename.split('_')[2]
#                 partdict["Cellletter"]=filename.split('_')[2]
#                 partdict["batchname"]=filename.split('_')[0]
#                 partdict["MeasComment"]=filename[filename.index('_')+1:]
                
#                 partdict["MeasDayTime"]=modification_date(self.file_path[i])
#                 # print(partdict["MeasDayTime"])
#                 partdict["CellSurface"]= float(filerawdata[0].split('\t')[-1])

#                 partdict["Delay"]=0
#                 partdict["IntegTime"]=0
#                 partdict["Vstep"]=0
#                 partdict["Vstart"]=0
#                 partdict["Vend"]=0
#                 partdict["ExecTime"]=0
#                 partdict["Operator"]='unknown'
#                 partdict["Group"]="Default group"
#                 partdict["sunintensity"]=1
                
#                 mpppartdat = [[],[],[],[],[]]#[voltage,current,time,power,vstep]
#                 for item in range(1,len(filerawdata),1):
#                     mpppartdat[0].append(float(filerawdata[item].split("\t")[2]))
#                     mpppartdat[1].append(float(filerawdata[item].split("\t")[3]))
#                     mpppartdat[2].append(float(filerawdata[item].split("\t")[0]))
#                     mpppartdat[3].append(float(filerawdata[item].split("\t")[1]))
#                     mpppartdat[4].append(-1)
#                 partdict["PowerEnd"]=mpppartdat[3][-1]
#                 partdict["PowerAvg"]=sum(mpppartdat[3])/float(len(mpppartdat[3]))
#                 partdict["trackingduration"]=mpppartdat[2][-1]
#                 partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime"]).replace(' ','_').replace(':','-')+'_'+str(partdict["PowerEnd"])

#                 partdict["MppData"]=mpppartdat
#                 partdict["SampleName"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime"]).replace(':','').replace(' ','-')
#                 partdict["MPPlinestyle"]=[partdict["SampleName"],"-",colors[len(DATAMPP.keys())],2]
                
#                 DATAMPP[partdict["SampleNameID"]]=partdict                

#             elif filetype ==11 : #SPO files of SERF C215 igor program
#                 #assumes file name: batch_samplenumber_cellLetter_mpp
#                 partdict = {}
#                 partdict["filepath"]=self.file_path[i]
#                 filename=os.path.splitext(os.path.basename(partdict["filepath"]))[0]
#                 partdict["DepID"]=filename.split('_')[0]+'_'+filename.split('_')[1]
#                 partdict["SampleName"]=filename.split('_')[0]+'_'+filename.split('_')[1]+'_'+filename.split('_')[2]
#                 partdict["Cellletter"]=filename.split('_')[2]
#                 partdict["batchname"]=filename.split('_')[0]
#                 partdict["MeasComment"]=filename[filename.index('_')+1:]
                
#                 partdict["MeasDayTime"]=modification_date(self.file_path[i])
#                 # print(partdict["MeasDayTime"])
#                 partdict["CellSurface"]= -1

#                 partdict["Delay"]=0
#                 partdict["IntegTime"]=0
#                 partdict["Vstep"]=0
#                 partdict["Vstart"]=0
#                 partdict["Vend"]=0
#                 partdict["ExecTime"]=0
#                 partdict["Operator"]='unknown'
#                 partdict["Group"]="Default group"
#                 partdict["sunintensity"]=1
                
#                 for item in range(len(filerawdata)):
#                     if "HEADER END" in filerawdata[item]:
#                             pos=item+3
#                             break
#                 mpppartdat = [[],[],[],[],[]]#[voltage,current,time,power,vstep]
#                 for item in range(pos,len(filerawdata),1):
#                     mpppartdat[0].append(float(filerawdata[item].split("\t")[1]))
#                     mpppartdat[1].append(float(filerawdata[item].split("\t")[0]))
#                     mpppartdat[2].append(float(filerawdata[item].split("\t")[3]))
#                     mpppartdat[3].append(float(filerawdata[item].split("\t")[2]))
#                     mpppartdat[4].append(-1)
#                 partdict["PowerEnd"]=mpppartdat[3][-1]
#                 partdict["PowerAvg"]=sum(mpppartdat[3])/float(len(mpppartdat[3]))
#                 partdict["trackingduration"]=mpppartdat[2][-1]
#                 partdict["SampleNameID"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime"]).replace(' ','_').replace(':','-')+'_'+str(partdict["PowerEnd"])

#                 partdict["MppData"]=mpppartdat
#                 partdict["SampleName"]=partdict["SampleName"]+'_'+str(partdict["MeasDayTime"]).replace(':','').replace(' ','-')
#                 partdict["MPPlinestyle"]=[partdict["SampleName"],"-",colors[len(DATAMPP.keys())],2]
                
#                 DATAMPP[partdict["SampleNameID"]]=partdict                                
#             self.change_value.emit(100*(i+1)/len(self.file_path))
        
# #         DATA = sorted(DATA, key=itemgetter('SampleName')) 
# #         names=[d["SampleName"] for d in DATA if "SampleName" in d]
# #         groupednames=[list(j) for i, j in groupby(names)]
# #         # print(groupednames)
# #         for item in range(len(groupednames)):
# #             if len(groupednames[item])>1 and groupednames[item][0][-1]!='D':
# #                 positions=[]
# #                 effrev=0
# #                 efffor=0
# #                 for item2 in range(len(DATA)):
# #                     if DATA[item2]['SampleName']==groupednames[item][0]:
# #                         positions.append(item2)
# #                         if DATA[item2]["ScanDirection"]=="Reverse":
# #                             effrev=DATA[item2]['Eff']
# #                         else:
# #                             efffor=DATA[item2]['Eff']
# #                     if len(positions)==len(groupednames[item]):
# #                         break
# #                 try:
# #                     hyste=100*(effrev-efffor)/effrev
# #                     for item2 in range(len(positions)):
# #                         DATA[positions[item2]]['HI']=hyste
# # #                        print(hyste)
# #                 except:
# #                     print("except HI")
        
# #         for item in range(len(groupednames)):
# #             if len(groupednames[item])!=1:
# #                 k=1
# #                 for item0 in range(1,len(groupednames[item])):
                    
# # #                    groupednames2=copy.deepcopy(groupednames)
# # #                    groupednames[item][item0]+= "_"+str(k)
# # #                    print(groupednames[item][item0])
# #                     while(1):
# #                         groupednames2=list(chain.from_iterable(groupednames))
# # #                        print(groupednames2)
                        
# #                         if groupednames[item][item0]+"_"+str(k) in groupednames2:
# #                             k+=1
# #                             groupednames[item][item0]+= "_"+str(k)
# # #                            print(groupednames[item][item0])
# # #                            print('')
# #                         else:
# #                             groupednames[item][item0]+= "_"+str(k)
# # #                            print('notin')
# #                             break
                        
# #         groupednames=list(chain.from_iterable(groupednames))
# # #        print("")
# # #        print(groupednames)
# #         for item in range(len(DATA)):
# #             DATA[item]['SampleName']=groupednames[item]
        
#         # DATAMPP = sorted(DATAMPP, key=itemgetter('SampleName')) 
#         # names=[d["SampleName"] for d in DATAMPP if "SampleName" in d]
#         # groupednames=[list(j) for i, j in groupby(names)]
#         # for item in range(len(groupednames)):
#         #     if len(groupednames[item])!=1:
#         #         for item0 in range(1,len(groupednames[item]),1):
#         #             groupednames[item][item0]+= "_"+str(item0)
#         # groupednames=list(chain.from_iterable(groupednames))
#         # for item in range(len(DATAMPP)):
#         #     DATAMPP[item]['SampleName']=groupednames[item]
        
#         self.finished.emit()

def savitzky_golay(y, window_size, order, deriv=0, rate=1):
    r"""Smooth (and optionally differentiate) data with a Savitzky-Golay filter.
    The Savitzky-Golay filter removes high frequency noise from data.
    It has the advantage of preserving the original shape and
    features of the signal better than other types of filtering
    approaches, such as moving averages techniques.
    Parameters
    ----------
    y : array_like, shape (N,)
        the values of the time history of the signal.
    window_size : int
        the length of the window. Must be an odd integer number.
    order : int
        the order of the polynomial used in the filtering.
        Must be less then `window_size` - 1.
    deriv: int
        the order of the derivative to compute (default = 0 means only smoothing)
    Returns
    -------
    ys : ndarray, shape (N)
        the smoothed signal (or it's n-th derivative).
    Notes
    -----
    The Savitzky-Golay is a type of low-pass filter, particularly
    suited for smoothing noisy data. The main idea behind this
    approach is to make for each point a least-square fit with a
    polynomial of high order over a odd-sized window centered at
    the point.
    Examples
    --------
    t = np.linspace(-4, 4, 500)
    y = np.exp( -t**2 ) + np.random.normal(0, 0.05, t.shape)
    ysg = savitzky_golay(y, window_size=31, order=4)
    import matplotlib.pyplot as plt
    plt.plot(t, y, label='Noisy signal')
    plt.plot(t, np.exp(-t**2), 'k', lw=1.5, label='Original signal')
    plt.plot(t, ysg, 'r', label='Filtered signal')
    plt.legend()
    plt.show()
    References
    ----------
    .. [1] A. Savitzky, M. J. E. Golay, Smoothing and Differentiation of
       Data by Simplified Least Squares Procedures. Analytical
       Chemistry, 1964, 36 (8), pp 1627-1639.
    .. [2] Numerical Recipes 3rd Edition: The Art of Scientific Computing
       W.H. Press, S.A. Teukolsky, W.T. Vetterling, B.P. Flannery
       Cambridge University Press ISBN-13: 9780521880688
    """
    

    try:
        window_size = npabs(npint(window_size))
        order = npabs(npint(order))
    except ValueError:
        raise ValueError("window_size and order have to be of type int")
    if window_size % 2 != 1 or window_size < 1:
        raise TypeError("window_size size must be a positive odd number")
    if window_size < order + 2:
        raise TypeError("window_size is too small for the polynomials order")
    order_range = range(order+1)
    half_window = (window_size -1) // 2
    # print(half_window)
    # precompute coefficients
    b = npmat([[k**i for i in order_range] for k in range(-half_window, half_window+1)])
    m = nplinalg.pinv(b).A[deriv] * rate**deriv * factorial(deriv)
    # pad the signal at the extremes with
    # values taken from the signal itself
    firstvals = y[0] - npabs( y[1:half_window+1][::-1] - y[0] )
    lastvals = y[-1] + npabs(y[-half_window-1:-1][::-1] - y[-1])
    y = npconcatenate((firstvals, y, lastvals))
    return npconvolve( m[::-1], y, mode='valid')

def extract_jv_params(jv):
    '''
    Extract Voc, Jsc, FF, Pmax from a given JV curve
        * Assume given JV curve is in volts and mA/cm2
    '''
    resample_step_size = 0.00001 # Voltage step size to use while resampling JV curve to find Pmax

    from scipy.interpolate import interp1d

    # Create a dict to store the parameters. Default values are -1 indicating failure to extract parameter
    params = {'Voc': -1., 'Jsc': -1., 'FF': -1., 'Pmax': -1., 'Roc':-1., 'Rsc':-1., 'Jmpp':-1, 'Vmpp':-1, 'Rshunt':-1, 'Rseries':-1}
    
    try:
        # Extract Jsc by interpolating wrt V
        jv_interp_V = interp1d(jv[0], jv[1], bounds_error=False, fill_value=0.)
        Jsc = jv_interp_V(0.)
        params['Jsc'] = abs(np.around(Jsc, decimals=8))
#            print(Jsc)
#            print(params['Jsc'])
    
        # Extract Voc by interpolating wrt J
        jv_interp_J = interp1d(jv[1], jv[0], bounds_error=False, fill_value=0.)
        Voc = jv_interp_J(0.)
#            print(Voc)
        params['Voc'] = np.around(Voc, decimals=4)
    
        # Resample JV curve over standard interval and find Pmax
        Vrange_new = np.arange(0., Voc, resample_step_size)
#            print(Vrange_new)
        jv_resampled = np.zeros((len(Vrange_new), 3))
        jv_resampled[:,0] = np.copy(Vrange_new)
        jv_resampled[:,1] = jv_interp_V(jv_resampled[:,0])
        jv_resampled[:,2] = np.abs(np.multiply(jv_resampled[:,0], jv_resampled[:,1]))
#            print(jv_resampled)
        pmax=np.max(np.abs(np.multiply(jv_resampled[:,0], jv_resampled[:,1])))
        params['Pmax'] = np.around(np.max(np.abs(np.multiply(jv_resampled[:,0], jv_resampled[:,1]))), decimals=4)
        indPmax=list(jv_resampled[:,2]).index(pmax)
        params['Jmpp']=abs(list(jv_resampled[:,1])[indPmax])
#            print(list(jv_resampled[:,1])[indPmax])
#            print(indPmax)
#            print(jv_interp_J(list(jv_resampled[:,1])[indPmax]))
        params['Vmpp']=1000*abs(list(jv_resampled[:,0])[indPmax])
#            print(params['Vmpp'])
    
        # Calculate fill factor
        params['FF'] = abs(100*np.around(pmax/(Jsc*Voc), decimals=4))
    except:
        print("error with param. calc")
        
    try:
    # Calculate Rsc&Roc 
    # x= [x0 for x0,y0 in sorted(zip(params['Voltage'],params['CurrentDensity']))]
    # y= [0.001*y0 for x0,y0 in sorted(zip(params['Voltage'],params['CurrentDensity']))]
    
        x= [x0 for x0,y0 in sorted(zip(jv[0], jv[1]))]
        y= [0.001*y0 for x0,y0 in sorted(zip(jv[0], jv[1]))]
        
        xSC=[]
        ySC=[]
        for i in range(len(x)):
            if x[i]>=0:
                xSC.append(x[i-3])
                xSC.append(x[i-2])
                xSC.append(x[i-1])
                xSC.append(x[i])
                xSC.append(x[i+1])
                xSC.append(x[i+2])
                ySC.append(y[i-3])
                ySC.append(y[i-2])
                ySC.append(y[i-2])
                ySC.append(y[i])
                ySC.append(y[i+1])
                ySC.append(y[i+2])
                break
    
        xSC=np.array(xSC)
        ySC=np.array(ySC)    
              
        xy=[xi*yi for xi, yi in zip(xSC,ySC)]
        xSC2=[xi**2 for xi in xSC]
        
        params['Rsc'] =abs( 1/(((sum(xSC)*sum(ySC)) - len(xSC)*sum(xy)) / ((sum(xSC)*sum(xSC)) - len(xSC)*sum(xSC2))))  
        # print(AllDATA[sample]['Rsc'])
        
        if params['Jsc']>1:
            xSC=[]
            ySC=[]
            for i in range(len(x)):
                if x[i]>=params['Voc']:
                    xSC.append(x[i-2])
                    xSC.append(x[i-1])
                    xSC.append(x[i])
                    xSC.append(x[i+1])
                    
                    ySC.append(y[i-2])
                    ySC.append(y[i-1])
                    ySC.append(y[i])
                    ySC.append(y[i+1])
                    break
    #                plt.plot(xSC,ySC,'bo')
            xSC=np.array(xSC)
            ySC=np.array(ySC)
            
            xy=[xi*yi for xi, yi in zip(xSC,ySC)]
            xSC2=[xi**2 for xi in xSC]
            params['Roc'] =abs( 1/(((sum(xSC)*sum(ySC)) - len(xSC)*sum(xy)) / ((sum(xSC)*sum(xSC)) - len(xSC)*sum(xSC2))))
        else:
            xSC=x[-3:]
            ySC=y[-3:]
            xSC=np.array(xSC)
            ySC=np.array(ySC)      
            xy=[xi*yi for xi, yi in zip(xSC,ySC)]
            xSC2=[xi**2 for xi in xSC]
            
            params['Roc'] = abs( 1/(((sum(xSC)*sum(ySC)) - len(xSC)*sum(xy)) / ((sum(xSC)*sum(xSC)) - len(xSC)*sum(xSC2))))   
        # print(AllDATA[sample]['Roc'])
#             x= [x0 for x0,y0 in sorted(zip(jv[0],jv[1]))]
#             y= [0.001*y0 for x0,y0 in sorted(zip(jv[0],jv[1]))]


#             xSC=[]
#             ySC=[]
#             for i in range(len(x)):
#                 if x[i]>=0:
#                     xSC.append(x[i-3])
#                     xSC.append(x[i-2])
#                     xSC.append(x[i-1])
#                     xSC.append(x[i])
#                     xSC.append(x[i+1])
#                     xSC.append(x[i+2])
#                     ySC.append(y[i-3])
#                     ySC.append(y[i-2])
#                     ySC.append(y[i-2])
#                     ySC.append(y[i])
#                     ySC.append(y[i+1])
#                     ySC.append(y[i+2])
#                     break
#     #        print(xSC)
#     #        print(ySC)
#     #        plt.plot(xSC,ySC,'bo')
#             xSC=np.array(xSC)
#             ySC=np.array(ySC)    
            
#     #        slope = stats.linregress(xSC,ySC)   
        
#             params['Rsc'] =abs( 1/(((mean(xSC)*mean(ySC)) - mean(xSC*ySC)) / ((mean(xSC)**2) - mean(xSC**2))))    
        
#             if params['Jsc']>1:
#                 xSC=[]
#                 ySC=[]
#                 for i in range(len(x)):
#                     if x[i]>=params['Voc']:
#                         xSC.append(x[i-2])
#                         xSC.append(x[i-1])
#                         xSC.append(x[i])
#                         xSC.append(x[i+1])
                    
#                         ySC.append(y[i-2])
#                         ySC.append(y[i-1])
#                         ySC.append(y[i])
#                         ySC.append(y[i+1])
#                         break
# #                plt.plot(xSC,ySC,'bo')
#                 xSC=np.array(xSC)
#                 ySC=np.array(ySC)      
            
#                 params['Roc'] =abs( 1/(((mean(xSC)*mean(ySC)) - mean(xSC*ySC)) / ((mean(xSC)**2) - mean(xSC**2))))
#             else:
#                 xSC=x[-3:]
#                 ySC=y[-3:]
# #                plt.plot(xSC,ySC,'bo')
#                 xSC=np.array(xSC)
#                 ySC=np.array(ySC)      
            
#                 params['Roc'] = abs(1/(((mean(xSC)*mean(ySC)) - mean(xSC*ySC)) / ((mean(xSC)**2) - mean(xSC**2))) )   
        
        
        
        
#        plt.show()
#        print(params['Rsc'])
#        print(params['Roc'])
#        print(params['Jsc'])
    except:
        print("error with Roc or Rsc calc")
    return  params


#%%#############
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = IVapp()
    window.show()
    sys.exit(app.exec())













