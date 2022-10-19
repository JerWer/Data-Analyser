from sys import argv
from sys import exit as Exit
from datetime import datetime as datetimefct
import os
from os import path as ospath
from os import makedirs as osmakedirs
from os import chdir as oschdir
from pathlib import Path
import pickle
from numpy import array, linspace, asarray, arange, abs as npabs, int as npint, mat as npmat, linalg as nplinalg, concatenate as npconcatenate, convolve as npconvolve
from math import factorial
# from sqlite3 import connect as sqlite3connect
# from sqlite3 import IntegrityError as sqlite3IntegrityError
# import pip

# def import_or_install(package):
#     try:
#         __import__(package)
#     except ImportError:
#         pip.main(['install', package]) 

# import_or_install('tkcolorpicker')
# import_or_install('peakutils')
#%%######################################################################################################
# import matplotlib
from matplotlib import colors as mplcolors
from matplotlib import pyplot as plt
# import matplotlib.transforms as mtransforms
# matplotlib.use("Qt5Agg")
#%%######################################################################################################
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QAction, QTableWidgetItem, QColorDialog
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
import copy
import xlsxwriter
import openpyxl as Opxl
import xlrd
from copy import deepcopy
import csv
# import xlsxwriter
from xlrd import open_workbook as xlrdopen_workbook
from scipy import integrate
# from operator import itemgetter
# from itertools import groupby, chain
# from PIL import Image as ImageTk
# from matplotlib.ticker import MaxNLocator
# from matplotlib.transforms import Bbox
# import pickle
from six import moves as sixmoves
from functools import partial
# import darktolight as DtoL
# import os.path
# import shutil
# from dateutil import parser
from scipy import stats
from scipy.interpolate import interp1d, UnivariateSpline
import math

# from PyQt5.uic import loadUiType
# Ui_MainWindow, QMainWindow = loadUiType('EQEgui.ui')
# Ui_MainWindow, QMainWindow = loadUiType(r'C:\Users\serjw\Documents\nBox\PythonScripts\SERIS-pythonscripts\executables\All\EQEgui.ui')
from EQEgui import Ui_MainWindow

exedirectory=str(Path(os.path.abspath(__file__)).parent)


LARGE_FONT= ("Verdana", 12)

echarge = 1.60218e-19
planck = 6.62607e-34
lightspeed = 299792458
spectrumName='AM15G'
file = open(os.path.join(exedirectory,'spectratxtfiles\AM15G.txt'), encoding='ISO-8859-1')
# file = open(r'C:\Users\serjw\Documents\nBox\PythonScripts\SERIS-pythonscripts\executables\All\AM15G.txt', encoding='ISO-8859-1')
# file = open(ospath.join(ospath.dirname(ospath.dirname(ospath.abspath(__file__))),'spectratxtfiles','AM15G.txt'), encoding='ISO-8859-1')
am15g = file.readlines()
file.close()
dataWave = []
dataInt = []
for i in range(len(am15g)):
    pos = am15g[i].find('\t')
    dataWave.append(float(am15g[i][:pos]))
    dataInt.append(float(am15g[i][pos+1:-1]))
  
SpectIntensFct = interp1d(dataWave,dataInt)

def modification_date(path_to_file):
    return datetimefct.fromtimestamp(ospath.getmtime(path_to_file)).strftime("%Y-%m-%d %H:%M:%S")

titEQE=0
firstimport=1
EQElegendMod=[]
DATA={} #{samplename, batchnumb, samplenumb, file_path, meastype, JscfromFile, JscCalc, datetime,
        #Eg0, EgIP, EgTauc, EgTauc2, lnDat, EgLn, EuLn, stderrEgLn, dataXwave, dataXEV, dataY, tangent
        #tangentLn, comment, Vbias, filterbias, ledbias}
DATAforexport=[]
takenforplot=[]
listofanswer={}
listoflinestyle={}
listofcolorstyle={}
listoflinewidthstyle={}
colorstylelist = ['black', 'red', 'blue', 'brown', 'green','cyan','magenta','olive','navy','orange','gray','aliceblue','antiquewhite','aqua','aquamarine','azure','beige','bisque','blanchedalmond','blue','blueviolet','brown','burlywood','cadetblue','chartreuse','chocolate','coral','cornflowerblue','cornsilk','crimson','darkblue','darkcyan','darkgoldenrod','darkgray','darkgreen','darkkhaki','darkmagenta','darkolivegreen','darkorange','darkorchid','darkred','darksalmon','darkseagreen','darkslateblue','darkslategray','darkturquoise','darkviolet','deeppink','deepskyblue','dimgray','dodgerblue','firebrick','floralwhite','forestgreen','fuchsia','gainsboro','ghostwhite','gold','goldenrod','greenyellow','honeydew','hotpink','indianred','indigo','ivory','khaki','lavender','lavenderblush','lawngreen','lemonchiffon','lightblue','lightcoral','lightcyan','lightgoldenrodyellow','lightgreen','lightgray','lightpink','lightsalmon','lightseagreen','lightskyblue','lightslategray','lightsteelblue','lightyellow','lime','limegreen','linen','magenta','maroon','mediumaquamarine','mediumblue','mediumorchid','mediumpurple','mediumseagreen','mediumslateblue','mediumspringgreen','mediumturquoise','mediumvioletred','midnightblue','mintcream','mistyrose','moccasin','navajowhite','navy','oldlace','olive','olivedrab','orange','orangered','orchid','palegoldenrod','palegreen','paleturquoise','palevioletred','papayawhip','peachpuff','peru','pink','plum','powderblue','purple','red','rosybrown','royalblue','saddlebrown','salmon','sandybrown','seagreen','seashell','sienna','silver','skyblue','slateblue','slategray','snow','springgreen','steelblue','tan','teal','thistle','tomato','turquoise','violet','wheat','white','whitesmoke','yellow','yellowgreen']
stitching=0

def savitzky_golay(y, window_size, order, deriv=0, rate=1):
    r"""Smooth (and optionally differentiate) data with a Savitzky-Golay filter.
    The Savitzky-Golay filter removes high frequency noise from data.
    It has the advantage of preserving the original shape and
    features of the signal better than other types of filtering
    approaches, such as moving averages techniques.
    Parameters
    ----------
    y : array_like, shape (N,)
        the values of the time history of the signal.
    window_size : int
        the length of the window. Must be an odd integer number.
    order : int
        the order of the polynomial used in the filtering.
        Must be less then `window_size` - 1.
    deriv: int
        the order of the derivative to compute (default = 0 means only smoothing)
    Returns
    -------
    ys : ndarray, shape (N)
        the smoothed signal (or it's n-th derivative).
    Notes
    -----
    The Savitzky-Golay is a type of low-pass filter, particularly
    suited for smoothing noisy data. The main idea behind this
    approach is to make for each point a least-square fit with a
    polynomial of high order over a odd-sized window centered at
    the point.
    Examples
    --------
    t = np.linspace(-4, 4, 500)
    y = np.exp( -t**2 ) + np.random.normal(0, 0.05, t.shape)
    ysg = savitzky_golay(y, window_size=31, order=4)
    import matplotlib.pyplot as plt
    plt.plot(t, y, label='Noisy signal')
    plt.plot(t, np.exp(-t**2), 'k', lw=1.5, label='Original signal')
    plt.plot(t, ysg, 'r', label='Filtered signal')
    plt.legend()
    plt.show()
    References
    ----------
    .. [1] A. Savitzky, M. J. E. Golay, Smoothing and Differentiation of
       Data by Simplified Least Squares Procedures. Analytical
       Chemistry, 1964, 36 (8), pp 1627-1639.
    .. [2] Numerical Recipes 3rd Edition: The Art of Scientific Computing
       W.H. Press, S.A. Teukolsky, W.T. Vetterling, B.P. Flannery
       Cambridge University Press ISBN-13: 9780521880688
    """
    

    try:
        window_size = npabs(npint(window_size))
        order = npabs(npint(order))
    except ValueError:
        raise ValueError("window_size and order have to be of type int")
    if window_size % 2 != 1 or window_size < 1:
        raise TypeError("window_size size must be a positive odd number")
    if window_size < order + 2:
        raise TypeError("window_size is too small for the polynomials order")
    order_range = range(order+1)
    half_window = (window_size -1) // 2
    # precompute coefficients
    b = npmat([[k**i for i in order_range] for k in range(-half_window, half_window+1)])
    m = nplinalg.pinv(b).A[deriv] * rate**deriv * factorial(deriv)
    # pad the signal at the extremes with
    # values taken from the signal itself
    firstvals = y[0] - npabs( y[1:half_window+1][::-1] - y[0] )
    lastvals = y[-1] + npabs(y[-half_window-1:-1][::-1] - y[-1])
    y = npconcatenate((firstvals, y, lastvals))
    return npconvolve( m[::-1], y, mode='valid')

class FixFigureCanvas(FigureCanvas):
    def resizeEvent(self, event):
        if event.size().width() <= 0 or event.size().height() <= 0:
            return
        super(FixFigureCanvas, self).resizeEvent(event)
        
class EQEapp(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        global spectrumName
        QtWidgets.QMainWindow.__init__(self, parent)
        
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        
        finish = QAction("Quit", self)
        finish.triggered.connect(lambda: self.closeEvent(1))
        
        self.figloss = Figure(figsize=(1, 2))
        self.EQElossgraph = self.figloss.add_subplot(111)
        self.addmpl(self.figloss,self.ui.gridLayout_mplwidgetLoss, self.ui.mplwidgetLoss)
        
        self.fig = Figure(figsize=(2, 2))
        self.EQEgraph = self.fig.add_subplot(111)
        self.EQEgraphY2 = self.EQEgraph.twinx()
        self.addmpl(self.fig,self.ui.verticalLayout_mplwidget, self.ui.mplwidget)
        
        # self.ui.actionHelp.triggered.connect(self.Helpcall)
        self.ui.actionImport_DATA.triggered.connect(self.onOpenEQE)
        # self.ui.actionExport_All_DATA.triggered.connect(self.sortandexportEQEdat)
        self.ui.actionExport_Graph_DATA.triggered.connect(self.ExportEQEGraph)
        self.ui.actionExport_Table_data.triggered.connect(self.ExportTableData)
        self.ui.actionSave.triggered.connect(self.SaveSession)
        self.ui.actionLoad.triggered.connect(self.LoadSession)

        
        # self.ui.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.ui.listWidget.itemClicked.connect(self.select)
        
        self.ui.checkBox_AutoScale.toggled.connect(self.UpdateEQEGraph)
        
        self.ui.checkBox_legend.toggled.connect(self.UpdateEQEGraph)
        self.ui.radioButton_topleft.toggled.connect(self.UpdateEQEGraph)
        self.ui.radioButton_topright.toggled.connect(self.UpdateEQEGraph)
        self.ui.radioButton_bottomleft.toggled.connect(self.UpdateEQEGraph)
        self.ui.radioButton_bottomright.toggled.connect(self.UpdateEQEGraph)
        self.ui.radioButton_outside.toggled.connect(self.UpdateEQEGraph)
        self.ui.radioButton_best.toggled.connect(self.UpdateEQEGraph)
        self.ui.spinBox_fontsize.valueChanged.connect(self.UpdateEQEGraph)
        
        self.ui.doubleSpinBox_Xmin.valueChanged.connect(self.UpdateEQEGraph)
        self.ui.doubleSpinBox_Xmax.valueChanged.connect(self.UpdateEQEGraph)
        self.ui.doubleSpinBox_Ymin.valueChanged.connect(self.UpdateEQEGraph)
        self.ui.doubleSpinBox_Ymax.valueChanged.connect(self.UpdateEQEGraph)
        
        self.ui.checkBox_Jsc.toggled.connect(self.UpdateEQEGraph)
        self.ui.checkBox_Eg.toggled.connect(self.UpdateEQEGraph)
        
        self.ui.comboBox.currentTextChanged.connect(self.UpdateEQEGraph)
        # self.ui.checkBox_integrJsc.toggled.connect(self.UpdateEQEGraph)
        self.ui.checkBox_showsecreteg.toggled.connect(self.UpdateEQEGraph)
        # self.ui.pushButton_CalcIQE.clicked.connect(self.CalculateIQE)
        self.ui.pushButton_stitchEQEs.clicked.connect(self.StitchEQE)
        self.ui.pushButton_CalcCurrent.clicked.connect(self.CalcCurrent)
        
        self.ui.pushButton_reorder.clicked.connect(self.reorder)
        
        self.ui.tabWidget_right.currentChanged.connect(self.onclicklegendtab)

        # self.ui.pushButton_updatefromlistwidget.clicked.connect(self.select)
        
        self.ui.pushButton_SGFilter.clicked.connect(self.SavitzkyGolayFiltering)
        self.ui.pushButton_goback.clicked.connect(self.backtoOriginal)
        
        self.ui.pushButton_lossupdate.clicked.connect(self.updateLossAnalysis)
        self.ui.comboBox_R.currentTextChanged.connect(lambda: self.loadR_LA('combo'))
        self.ui.pushButton_loadR.clicked.connect(self.chooseRfile)
        self.ui.pushButton_loadOtherEQE.clicked.connect(self.chooseAlternEQEfile)
        self.ui.comboBox_T.currentTextChanged.connect(lambda: self.loadT_LA('combo'))
        self.ui.pushButton_loadT.clicked.connect(self.chooseTfile)
        # self.ui.zonesoverlays.toggled.connect(self.updateLossAnalysis)
        
        self.ui.lineEdit_spectrumloaded.setText(spectrumName)
        self.ui.pushButton_loadspectrum.clicked.connect(self.loadspectrum)
    
    
    def loadspectrum(self):
        global spectrumName, dataWave,dataInt, SpectIntensFct
        
        file_path = QFileDialog.getOpenFileName(caption = 'Please select the spectrum file')[0]
        file=open(file_path, encoding='ISO-8859-1')
        am15g = file.readlines()
        file.close()
        dataWave = []
        dataInt = []
        for i in range(len(am15g)):
            pos = am15g[i].find('\t')
            dataWave.append(float(am15g[i][:pos]))
            dataInt.append(float(am15g[i][pos+1:-1]))
          
        SpectIntensFct = interp1d(dataWave,dataInt)
        
        self.ui.lineEdit_spectrumloaded.setText(os.path.basename(file_path).split('.')[0])
        
        
    
    def closeEvent(self, event):
        """ what happens when close the program"""
        close = QMessageBox.question(self,
                                     "QUIT",
                                     "Are you sure?",
                                      QMessageBox.Yes | QMessageBox.No)
        if close == QMessageBox.Yes:        
            event.accept()
            try:
                app.quit()
            except:
                pass
        else:
            event.ignore()
            
    def SaveSession(self):
        global dataWave, dataInt, SpectIntensFct, titEQE, firstimport,EQElegendMod,DATA,DATAforexport,takenforplot
        
        current_path = os.getcwd()
        directory=QFileDialog.getExistingDirectory(self, 'Select directory')
        os.chdir(directory)
        
        listofglobalvariables= [dataWave, dataInt, SpectIntensFct, titEQE, firstimport,EQElegendMod,DATA,DATAforexport,takenforplot]

        for item in range(len(listofglobalvariables)):
            print(item)
            # print(listofglobalvariables[item])
            pickle.dump(listofglobalvariables[item],open(str(item)+'.pkl','wb'), protocol=pickle.HIGHEST_PROTOCOL)
        
        print("dumped")

    def LoadSession(self):
        global dataWave, dataInt, SpectIntensFct, titEQE, firstimport,EQElegendMod,DATA,DATAforexport,takenforplot

        current_path = os.getcwd()
        path=QFileDialog.getExistingDirectory(self, 'Select directory')
        os.chdir(path)
        
        listofglobalvariables= ["dataWave", "dataInt", "SpectIntensFct", "titEQE", "firstimport","EQElegendMod","DATA","DATAforexport","takenforplot"]
            
        for item in range(len(listofglobalvariables)):
            globals()[listofglobalvariables[item]] = pickle.load(open(str(item)+'.pkl','rb'))
        
        print("loaded")
            
            
        if DATA!={}:
            # self.onOpenEQE()
            self.initlistbox()
            self.UpdateEQEGraph
    
    def addmpl(self, fig, whereLayout, whereWidget):
        self.canvas = FixFigureCanvas(fig)
        whereLayout.addWidget(self.canvas)
        self.canvas.draw()
        self.toolbar = NavigationToolbar(self.canvas, 
                whereWidget, coordinates=True)
        whereLayout.addWidget(self.toolbar)
        
    def choiceYtype(self,a):
        self.UpdateEQEGraph()
        
    def onOpenEQE(self):
        self.GetEQEDATA()
        self.initlistbox()
        
    def reorder(self):
        self.w=reorderwin()
        self.w.show()
        
    def ncolumneqe(self, n):
        twocolumneqe = [['Wavelength','EQE'],['nm','-']]
        for i in range(1,n-1):
            twocolumneqe[0].append('EQE')
            twocolumneqe[1].append('-')
        return twocolumneqe
    def ncolumneqeJsc(self, n):
        twocolumneqe = [['Wavelength','Current density'],['nm','mA/cm2']]
        for i in range(1,n-1):
            twocolumneqe[0].append('Current density')
            twocolumneqe[1].append('mA/cm2')
        return twocolumneqe
    
    def AM15GParticlesinnm(self, x):
        return (x*10**(-9))*SpectIntensFct(x)/(planck*lightspeed)

    def clearLayout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clearLayout(item.layout())
                    
    def onclicklegendtab(self,indexoftab):
        # print(indexoftab)   
        if indexoftab==2 and self.ui.checkBox_legend.isChecked():
            self.populate()
            
    def populate(self):
        global numberofLayer
        global DATA
        global takenforplot
        global colorstylelist
        global listofanswer
        global listoflinestyle
        global listofcolorstyle, listoflinewidthstyle
        DATAx=DATA
        listofanswer={}
        sampletotake=takenforplot
        
        listoflinestyle={}
        listofcolorstyle={}
        listoflinewidthstyle={}
        for item in DATAx.keys():
            listoflinestyle[item]=DATAx[item]['linestyle']
            listofcolorstyle[item]=DATAx[item]['linecolor']
            listofanswer[item]=DATAx[item]['NameMod']
            listoflinewidthstyle[item]=str(DATAx[item]['linewidth'])
            
        self.clearLayout(self.ui.gridLayout_5)
        self.ui.scrollArea_legend = QtWidgets.QScrollArea(self.ui.EditLegend)
        self.ui.scrollArea_legend.setWidgetResizable(True)
        self.ui.scrollArea_legend.setObjectName("scrollArea_legend")
        self.ui.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.ui.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 500, 400))
        self.ui.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.ui.scrollArea_legend.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ui.scrollAreaWidgetContents.sizePolicy().hasHeightForWidth())
        self.ui.scrollAreaWidgetContents.setSizePolicy(sizePolicy)
        self.ui.verticalLayout_legend = QtWidgets.QVBoxLayout(self.ui.scrollAreaWidgetContents)
        self.ui.verticalLayout_legend.setObjectName("verticalLayout_legend")
        
        forbiddenrange=[]
        for itemm in sampletotake:
            i=0
            for item1 in DATAx.keys(): 
                if item1 not in forbiddenrange:
                    if DATAx[item1]['ID'] == itemm:
                        #print(itemm)
                        self.frame = QtWidgets.QFrame(self.ui.scrollAreaWidgetContents)
                        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
                        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
                        self.frame.setObjectName("frame")
                        self.horizontalLayout = QtWidgets.QHBoxLayout(self.frame)
                        self.horizontalLayout.setObjectName("horizontalLayout")
                        
                        forbiddenrange.append(item1)
                        label = QtWidgets.QLabel(self.frame)
                        label.setText(DATAx[item1]['samplename'])
                        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
                        sizePolicy.setHorizontalStretch(0)
                        sizePolicy.setVerticalStretch(0)
                        sizePolicy.setHeightForWidth(label.sizePolicy().hasHeightForWidth())
                        label.setSizePolicy(sizePolicy)
                        self.horizontalLayout.addWidget(label)
                        
                        listofanswer[item1]=QtWidgets.QLineEdit(self.frame)
                        listofanswer[item1].setText(DATAx[item1]['NameMod'])
                        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Maximum)
                        sizePolicy.setHorizontalStretch(0)
                        sizePolicy.setVerticalStretch(0)
                        sizePolicy.setHeightForWidth(listofanswer[item1].sizePolicy().hasHeightForWidth())
                        listofanswer[item1].setSizePolicy(sizePolicy)
                        self.horizontalLayout.addWidget(listofanswer[item1])
                        listofanswer[item1].textChanged.connect(self.UpdateEQELegMod)
                        
                        listoflinestyle[item1] = QtWidgets.QComboBox(self.frame)
                        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
                        sizePolicy.setHorizontalStretch(0)
                        sizePolicy.setVerticalStretch(0)
                        sizePolicy.setHeightForWidth(listoflinestyle[item1].sizePolicy().hasHeightForWidth())
                        listoflinestyle[item1].setObjectName("comboBox_matname"+str(i))
                        listoflinestyle[item1].addItems(["-","--","-.",":"])
                        listoflinestyle[item1].setCurrentText(DATAx[item1]['linestyle'])
                        self.horizontalLayout.addWidget(listoflinestyle[item1])
                        listoflinestyle[item1].currentTextChanged.connect(self.UpdateEQELegMod)
                        
                        colstyle = QtWidgets.QPushButton('Select Color', self.frame)
                        colstyle.setStyleSheet("color:"+str(listofcolorstyle[item1])+";")
                        self.horizontalLayout.addWidget(colstyle)
                        colstyle.clicked.connect(partial(self.getColor,item1))
                        
                        listoflinewidthstyle[item1] = QtWidgets.QSpinBox(self.frame)
                        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Maximum)
                        sizePolicy.setHorizontalStretch(0)
                        sizePolicy.setVerticalStretch(0)
                        sizePolicy.setHeightForWidth(listoflinewidthstyle[item1].sizePolicy().hasHeightForWidth())
                        listoflinewidthstyle[item1].setSizePolicy(sizePolicy)
                        listoflinewidthstyle[item1].setMaximum(9999999)
                        listoflinewidthstyle[item1].setObjectName("spinBox_"+str(i))
                        listoflinewidthstyle[item1].setValue(DATAx[item1]['linewidth'])
                        self.horizontalLayout.addWidget(listoflinewidthstyle[item1])
                        listoflinewidthstyle[item1].valueChanged.connect(self.UpdateEQELegMod)
                        
                        self.ui.verticalLayout_legend.addWidget(self.frame)
                        break
                    
                    else:
                        listofanswer[item1]=str(DATAx[item1]['NameMod'])
                        listoflinestyle[item1]=str(DATAx[item1]['linestyle'])
                        listofcolorstyle[item1]=str(DATAx[item1]['linecolor'])
                        listoflinewidthstyle[item1]=str(DATAx[item1]['linewidth'])
        
        self.ui.scrollArea_legend.setWidget(self.ui.scrollAreaWidgetContents)
        self.ui.gridLayout_5.addWidget(self.ui.scrollArea_legend, 0, 0, 1, 1)
    
    def getColor(self,rowitem):
        global listofcolorstyle
        color = QColorDialog.getColor()
        listofcolorstyle[rowitem]=color.name()
        self.UpdateEQELegMod()
        self.populate()
        
    def UpdateEQELegMod(self):
        global DATA
        global listofanswer
        global listoflinestyle
        global listofcolorstyle,listoflinewidthstyle
        # print('')
        # print(listofanswer)
        
        leglist={}
        for e in listofanswer.keys():
            if type(listofanswer[e])!=str:
                leglist[e]=listofanswer[e].text()
            else:
                leglist[e]=listofanswer[e]
        # print(leglist)
        for item in DATA.keys():
            DATA[item]['NameMod']=leglist[item]
            DATA[item]['Name_Jsc']= leglist[item]+'_'+'Jsc: %.2f' % DATA[item]['JscCalc']
            DATA[item]['Name_Eg']= leglist[item]+'_'+'EgIP: %.2f' % DATA[item]['EgIP']
            DATA[item]['Name_Jsc_Eg']= leglist[item]+'_'+'Jsc: %.2f' % DATA[item]['JscCalc']+'_'+'EgIP: %.2f' % DATA[item]['EgIP']
        leglist={}
        for e in listoflinestyle.keys():
            if type(listoflinestyle[e])!=str:
                leglist[e]=listoflinestyle[e].currentText()
            else:
                leglist[e]=listoflinestyle[e]
        for item in DATA.keys():
            DATA[item]['linestyle']=leglist[item]
        leglist={}
        for e in listofcolorstyle.keys():
            if type(listofcolorstyle[e])!=str:
                leglist[e]=listofcolorstyle[e].get()
            else:
                leglist[e]=listofcolorstyle[e]
        for item in DATA.keys():
            DATA[item]['linecolor']=leglist[item]
        leglist={}
        for e in listoflinewidthstyle.keys():
            if type(listoflinewidthstyle[e])!=str:
                leglist[e]=listoflinewidthstyle[e].value()
            else:
                leglist[e]=listoflinewidthstyle[e]
        for item in DATA.keys():
            DATA[item]['linewidth']=int(leglist[item]) 
                
        
        self.UpdateEQEGraph()

    def SavitzkyGolayFiltering(self):
        global DATA
        DATAx=DATA
        
        if self.ui.spinBox_SG1.value()>self.ui.spinBox_SG2.value() and self.ui.spinBox_SG1.value()%2==1:
            if self.ui.listWidget.selectedItems()!=():
                # samplestakenforplot = [str(self.ui.listWidget.selectedItems()[i].text()) for i in range(len(self.ui.listWidget.selectedItems()))]
                samplestakenforplot=self.ui.listWidget.selectedItems()
                samplestakenforplot=[x.text() for x in samplestakenforplot]
                
                if samplestakenforplot!=[]:
                    for item in samplestakenforplot:
                        x = DATAx[item]['DATAmod'][0]
                        y = DATAx[item]['DATAmod'][2]
                        y=array(y)
                        # print(DATA[item]['DATAmod'][2][10])
                        # print(DATA[item]['DATAorig'][2][10])
                        DATAx[item]['DATAmod'][2] = list(savitzky_golay(y, window_size=self.ui.spinBox_SG1.value(), order=self.ui.spinBox_SG2.value()))
                        # print(DATA[item]['DATAmod'][2][10])
                        # print(DATA[item]['DATAorig'][2][10])
                        xE=[]
                        yln=[]
                        for xi in range(len(x)):
                            if y[xi]>0:
                                xE.append(1239.8/x[xi])
                                yln.append(math.log(100*y[xi]))
                        DATAx[item]['lnDat'][0]=xE
                        DATAx[item]['DATAmod'][1]=xE
                        DATAx[item]['lnDat'][1]=yln
                        
                        try:
                            xtauc=[1239.8/xm for xm in x]
                            ytauc=[(y[m]*xtauc[m])**2 for m in range(len(y)) ]
                            xtauc=xtauc[::-1]
                            ytauc=ytauc[::-1]
                            spl = UnivariateSpline(xtauc, ytauc, s=0)
                            splder = spl.derivative(n=1)
                            splderlist = []
                            newx=[]
                            for itemtauc in xtauc :
                                if itemtauc <2:
                                    splderlist.append(splder(itemtauc))
                                    newx.append(itemtauc)
                            
                            maxder=splderlist.index(max(splderlist))
                            xhighslope = newx[maxder]
                            yhighslope = spl(newx[maxder]).tolist()
                            yprimehighslope = splder(newx[maxder]).tolist()
                            Eg= (xhighslope - yhighslope/yprimehighslope)
                            
                            m=yprimehighslope
                            h=yhighslope-yprimehighslope*xhighslope
                            x2=Eg
                            x=linspace(x2,x2+0.1,10)
                            y=eval('m*x+h')
                            DATAx[item]['EgTauc']=Eg
                            DATAx[item]['xtauc']=xtauc
                            DATAx[item]['ytauc']=ytauc
                            DATAx[item]['mtauc']=m
                            DATAx[item]['htauc']=h
                        except:
                            DATAx[item]['EgTauc']=-1
                            DATAx[item]['xtauc']=[]
                            DATAx[item]['ytauc']=[]
                            DATAx[item]['mtauc']=-1
                            DATAx[item]['htauc']=-1
                                
                self.UpdateEQEGraph()
        else:
            QMessageBox.information(self,'Information', "the SG window-size must be larger than the SG order, positive and odd.")

    def backtoOriginal(self):
        global DATA
        samplestakenforplot = self.ui.listWidget.selectedItems()
        samplestakenforplot=[x.text() for x in samplestakenforplot]
        # print('here')
        if samplestakenforplot!=[]:
            # print('here')
            for item in samplestakenforplot:
                # print(DATA[item]['DATAmod'][2][10])
                # print(DATA[item]['DATAorig'][2][10])
                DATA[item]['DATAmod'][2]=DATA[item]['DATAorig'][2]
                # print(DATA[item]['DATAmod'][2][10])
                # print(DATA[item]['DATAorig'][2][10])
                x=DATA[item]['DATAmod'][0]
                y=DATA[item]['DATAmod'][2]
                xE=[]
                yln=[]
                for xi in range(len(x)):
                    if y[xi]>0:
                        xE.append(1239.8/x[xi])
                        yln.append(math.log(100*y[xi]))
                DATA[item]['lnDat'][0]=xE
                DATA[item]['lnDat'][1]=yln
                
                try:
                    xtauc=[1239.8/xm for xm in x]
                    ytauc=[(y[m]*xtauc[m])**2 for m in range(len(y)) ]
                    xtauc=xtauc[::-1]
                    ytauc=ytauc[::-1]
                    spl = UnivariateSpline(xtauc, ytauc, s=0)
                    splder = spl.derivative(n=1)
                    splderlist = []
                    newx=[]
                    for itemtauc in xtauc :
                        if itemtauc <2:
                            splderlist.append(splder(itemtauc))
                            newx.append(itemtauc)
                    
                    maxder=splderlist.index(max(splderlist))
                    xhighslope = newx[maxder]
                    yhighslope = spl(newx[maxder]).tolist()
                    yprimehighslope = splder(newx[maxder]).tolist()
                    Eg= (xhighslope - yhighslope/yprimehighslope)
                    
                    m=yprimehighslope
                    h=yhighslope-yprimehighslope*xhighslope
                    x2=Eg
                    x=linspace(x2,x2+0.1,10)
                    y=eval('m*x+h')
                    DATA[item]['EgTauc']=Eg
                    DATA[item]['xtauc']=xtauc
                    DATA[item]['ytauc']=ytauc
                    DATA[item]['mtauc']=m
                    DATA[item]['htauc']=h
                except:
                    DATA[item]['EgTauc']=-1
                    DATA[item]['xtauc']=[]
                    DATA[item]['ytauc']=[]
                    DATA[item]['mtauc']=-1
                    DATA[item]['htauc']=-1
        
        self.UpdateEQEGraph()
        
    #%%###########        
    def GetEQEDATA(self):
        global DATAFORGRAPH, DATA
        global colorstylelist,firstimport,stitching
        
        if stitching==0:
            file_path = QFileDialog.getOpenFileNames(caption = 'Please select the eqe files')[0]
        else:
            file_path=[stitching]
            stitching=0
#        print(file_path[0])
#        print(modification_date(file_path[0]))
#        
#        print(ospath.basename(file_path[0].split('.')[0]))
#        print("")
        integrationJscYes=0
#        MsgBox = messagebox.askquestion("IntegrJsc?", "Do you want to calculate the Integrated Jsc curve?\nWarning: it will slow a bit down the importation")
#        if MsgBox == 'yes':
#            integrationJscYes=1
        
            
        try:
            directory = str(Path(file_path[0]).parent.parent)+'\\resultFilesEQE'
        except IndexError:
            return
        
        
#        print(directory)
        if not ospath.exists(directory):
            osmakedirs(directory)
            oschdir(directory)
        else :
            oschdir(directory)
        
        colormapname="jet"
        num_plots=len(file_path)
        cmap = plt.get_cmap(colormapname)
        colors = cmap(linspace(0, 1.0, num_plots))
        colors=[mplcolors.to_hex(item) for item in colors]

        if firstimport:
            DATA={}
            firstimport=0

        for k in range(len(file_path)):
            datadict = {'ID':'','samplename':'', 'batchnumb':'', 'samplenumb':'', 'file_path':'', 'meastype':'', 'JscfromFile':-1, 'JscCalc':-1, 'datetime':'',
                                 'NameMod':'', 'Name_Jsc':'','Name_Eg':'', 'Name_Jsc_Eg':'','Name_Egln':'', 'Name_Jsc_Egln':'', 'Name_Egtauc':'', 'Name_Jsc_Egtauc':'', 
                                'Eg0':-1, 'EgIP':-1, 'tangent':[-1, -1],'slope':-1, 'h':-1,
                                'EgTauc':-1, 'xtauc':[], 'ytauc':[], 'mtauc':-1, 'htauc':-1,
                                'lnDat':[[-1],[-1]], 'EgLn':-1,  'stderrEgLn':[-1,-1], 'tangentLn':[-1, -1,[-1],[-1]], 'slopeLn':-1, 'hln':-1,'dataEnergyLn':[-1], 'dataIntLn':[-1], 'ptstgtLnX':[-1], 'ptstgtLnY':[-1],
                                'EuLn':-1,
                                'DATAorig':[[],[],[]],'DATAmod':[[],[],[]], 
                                'comment':'', 'Vbias':'', 'filterbias':'', 'ledbias':'','integJsclist':[],
                                 'linestyle':'-', 'linecolor':colors[k], 'linewidth':int(2)
                                }
            if ospath.splitext(file_path[k])[1]==".txt":
                print('this is a text file')
                # Batch#_Sample#_EQE, Batch#_Sample#_TR, Batch#_Sample#_SR, 
                #2 columns, no headers, 1st= wavelength in nm, 2nd= data in %, tabulated \t
                samplename=file_path[k].replace('\\','/') 
                samplename=samplename.split('/')[-1].replace('-','_').split('.')[0]
                
                datadict['samplename']=samplename
                batchnumb=samplename.split('_')[0]
                datadict['batchnumb']=batchnumb
                samplenumb=samplename.split('_')[1]
                datadict['samplenumb']=samplenumb
                
                datadict['file_path']=file_path[k]
                
                file=open(file_path[k])
                dat=file.readlines()
                file.close()
                dataWave = []
                dataInt = []
                for i in range(len(dat)):
                    # print(dat[i])
                    pos = dat[i].find('\t')
                    dataWave.append(float(dat[i][:pos]))
                    # print(dat[i][pos+1:-1])
                    dataInt.append(float(dat[i][pos+1:-1])/100)
                
                #EQE
                # print(samplename.split('_'))
                if 'EQE'in samplename.split('_')[2]:
                    datadict['meastype']='EQE'
                    datadict['JscfromFile']=-1
                    datetime=modification_date(file_path[k])
                    datadict['datetime']=datetime
                    
                    datadict['DATAorig'][0]=dataWave
                    datadict['DATAorig'][1]=[1239.8/item for item in dataWave]
                    datadict['DATAorig'][2]=dataInt
                    datadict['DATAmod']=deepcopy(datadict['DATAorig'])
                    
                    if integrationJscYes:#not implemented at the moment, quite slow as needs integration 
                            datadict['integJsclist']=[datadict['DATAorig'][0]]
                    else:
                        datadict['integJsclist']=[[]]
                            
                    #jsc calculation
                    x = datadict['DATAorig'][0]
                    y = datadict['DATAorig'][2]
                    if len(x)>3:
                        spl = UnivariateSpline(x, y, s=0)
                        f = interp1d(x, y, kind='cubic')
                        x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                        integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],datadict['DATAorig'][0][-1])[0]
                        datadict['JscCalc']=integral
                        print(datadict['JscCalc'])
                        integlist=[]
                        if integrationJscYes:
                            for item in x:
                                integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],item)[0]
                                integlist.append(integral)
                        datadict['integJsclist'].append(integlist)
                        #Eg calculation from linear normal curve
                        splder = spl.derivative(n=1)
                        splderlist = []
                        newx=[]
                        for item in x :
                            if item >400:
                                splderlist.append(splder(item))
                                newx.append(item)
                        if splderlist==[]:
                            datadict['Eg0']=0
                            datadict['EgIP']=0
                            datadict['tangent']=[0, 0]
                            datadict['slope']=0
                            datadict['h']=0
                        else:
                            minder=splderlist.index(min(splderlist))
                            xhighslope = newx[minder]
                            datadict['EgIP']=1239.8/xhighslope
                            yhighslope = spl(newx[minder]).tolist()
                            yprimehighslope = splder(newx[minder]).tolist()
                            Eg= 1239.8/(xhighslope - yhighslope/yprimehighslope)
                            datadict['Eg0']=Eg
                            datadict['tangent']=[yprimehighslope, yhighslope-yprimehighslope*xhighslope]#[pente,ordonnee a l'origine]
                            datadict['slope']=datadict['tangent'][0]
                            datadict['h']=datadict['tangent'][1]

                        #Eg calculation from ln(EQE) curve
                        xE=[]
                        yln=[]
                        for xi in range(len(x)):
                            if y[xi]>0:
                                xE.append(1239.8/x[xi])
                                yln.append(math.log(100*y[xi]))

                        datadict['lnDat']=[xE,yln]
                        datadict['dataEnergyLn']=datadict['lnDat'][0]
                        datadict['dataIntLn']=datadict['lnDat'][1]
                        
                        xErestricted=[]
                        ylnrestricted=[]
                        
                        for xi in range(len(xE)-1,-1,-1):
                            if yln[xi]<3 and yln[xi]>-2:
                                xErestricted.append(xE[xi])
                                ylnrestricted.append(yln[xi])
                        xErestricted.append(999)
                        ylnrestricted.append(999)
                        xErestricted2=[]
                        ylnrestricted2=[]
                        for xi in range(len(xErestricted)-1):
                            xErestricted2.append(xErestricted[xi])
                            ylnrestricted2.append(ylnrestricted[xi])
                            if abs(xErestricted[xi]-xErestricted[xi+1])>0.3:
                                break
                        if len(xErestricted2)>1:
                            slope, intercept, r_value, p_value, std_err = stats.linregress(xErestricted2,ylnrestricted2)
                            print('slope')
                            print(slope)
                            print(1000/slope)
                            datadict['EgLn']=-intercept/slope
                            datadict['EuLn']=1000/slope #Eu calculation from ln(EQE) curve slope at band edge
                            datadict['tangentLn']=[slope, intercept,xErestricted2,ylnrestricted2]#[pente,ordonnee a l'origine]
                            datadict['slopeLn']=datadict['tangentLn'][0]
                            datadict['hln']=datadict['tangentLn'][1]
                            datadict['ptstgtLnX']=datadict['tangentLn'][2]
                            datadict['ptstgtLnY']=datadict['tangentLn'][3]
                            datadict['stderrEgLn']=[std_err,len(xErestricted2)]
                        else:
                            print("EgLn not found enough points...")
                        
                        try:
                            xtauc=[1239.8/xm for xm in x]
                            ytauc=[(y[m]*xtauc[m])**2 for m in range(len(y)) ]
                            xtauc=xtauc[::-1]
                            ytauc=ytauc[::-1]
                            spl = UnivariateSpline(xtauc, ytauc, s=0)
                            splder = spl.derivative(n=1)
                            splderlist = []
                            newx=[]
                            for item in xtauc :
                                if item <2:
                                    splderlist.append(splder(item))
                                    newx.append(item)
                            
                            maxder=splderlist.index(max(splderlist))
                            xhighslope = newx[maxder]
                            yhighslope = spl(newx[maxder]).tolist()
                            yprimehighslope = splder(newx[maxder]).tolist()
                            Eg= (xhighslope - yhighslope/yprimehighslope)
                            
                            m=yprimehighslope
                            h=yhighslope-yprimehighslope*xhighslope
                            x2=Eg
                            x=linspace(x2,x2+0.1,10)
                            y=eval('m*x+h')
                            datadict['EgTauc']=[Eg,xtauc,ytauc,m,h]
                            datadict['xtauc']=datadict['EgTauc'][1]
                            datadict['ytauc']=datadict['EgTauc'][2]
                            datadict['mtauc']=datadict['EgTauc'][3]
                            datadict['htauc']=datadict['EgTauc'][4]
                            datadict['EgTauc']=datadict['EgTauc'][0]
                        except:
                            pass
                            
                    datadict['ID']=datadict['samplename']+str(datadict['datetime'])+str(datadict['JscCalc'])
                    index1=1
                    while(1):
                        if datadict['ID'] in DATA.keys():
                            datadict['ID']=datadict['ID']+'_'+str(index1)
                            index1+=1
                        else:
                            break
                    datadict['NameMod']=datadict['samplename']
                    datadict['Name_Jsc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    datadict['Name_Eg']=datadict['samplename']+'_'+'EgIP: %.2f' % datadict['EgIP']
                    datadict['Name_Jsc_Eg']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']+'_'+'EgIP: %.2f' % datadict['EgIP']
                    datadict['Name_Egln']=datadict['samplename']+'_'+'EgLn: %.2f' % datadict['EgLn']
                    datadict['Name_Jsc_Egln']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']+'_'+'EgLn: %.2f' % datadict['EgLn']
                    datadict['Name_Egtauc']=datadict['samplename']+'_'+'EgTauc: %.2f' % datadict['EgTauc']
                    datadict['Name_Jsc_Egtauc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']+'_'+'EgTauc: %.2f' % datadict['EgTauc']
                    
                    DATA[datadict['ID']]=datadict
                
                #Reflectance
                if 'TR'in samplename.split('_')[2]:
                    datadict['meastype']='R'
                    datetime=modification_date(file_path[k])
                    datadict['datetime']=datetime
                    datadict['DATAorig'][0]=dataWave
                    datadict['DATAorig'][1]=[1239.8/item for item in dataWave]
                    datadict['DATAorig'][2]=dataInt
                    datadict['DATAmod']=deepcopy(datadict['DATAorig'])
                    if integrationJscYes:#not implemented at the moment, quite slow as needs integration 
                        datadict['integJsclist']=[datadict['DATAorig'][0]]
                    else:
                        datadict['integJsclist']=[[]]
                    #jsc calculation
                    x = datadict['DATAorig'][0]
                    y = datadict['DATAorig'][2]
                    if len(x)>3:
                        spl = UnivariateSpline(x, y, s=0)
                        f = interp1d(x, y, kind='cubic')
                        x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                        integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],datadict['DATAorig'][0][-1])[0]
                        datadict['JscCalc']=integral
                        # print(datadict['JscCalc'])
                        
                        integlist=[]
                        if integrationJscYes:
                            for item in x:
                                integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],item)[0]
                                integlist.append(integral)
                        datadict['integJsclist'].append(integlist)
                        xE=[]
                        yln=[]
                        for xi in range(len(x)):
                            if y[xi]>0:
                                xE.append(1239.8/x[xi])
                                yln.append(math.log(100*y[xi]))
                        datadict['lnDat']=[xE,yln]
                        datadict['dataEnergyLn']=datadict['lnDat'][0]
                        datadict['dataIntLn']=datadict['lnDat'][1]
                    
                    datadict['ID']=datadict['samplename']+str(datadict['datetime'])+str(datadict['JscCalc'])
                    index1=1
                    while(1):
                        if datadict['ID'] in DATA.keys():
                            datadict['ID']=datadict['ID']+'_'+str(index1)
                            index1+=1
                        else:
                            break
                    datadict['NameMod']=datadict['samplename']
                    datadict['Name_Jsc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    datadict['Name_Eg']=datadict['samplename']+'_'+'EgIP: %.2f' % datadict['EgIP']
                    datadict['Name_Jsc_Eg']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    datadict['Name_Egln']=datadict['samplename']
                    datadict['Name_Jsc_Egln']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    datadict['Name_Egtauc']=datadict['samplename']
                    datadict['Name_Jsc_Egtauc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    
                    DATA[datadict['ID']]=datadict
                    
                #SR
                if 'SR'in samplename.split('_')[2]:
                    datadict['meastype']='SR'
                    datetime=modification_date(file_path[k])
                    datadict['JscfromFile']=-1
                    datadict['datetime']=datetime
                    datadict['DATAorig'][0]=dataWave
                    datadict['DATAorig'][1]=[1239.8/item for item in dataWave]
                    datadict['DATAorig'][2]=dataInt
                    datadict['DATAmod']=deepcopy(datadict['DATAorig'])
                    if integrationJscYes:#not implemented at the moment, quite slow as needs integration 
                        datadict['integJsclist']=[datadict['DATAorig'][0]]
                    else:
                        datadict['integJsclist']=[[]]
                    
                    #jsc calculation
                    x = datadict['DATAorig'][0]
                    y = datadict['DATAorig'][2]
                    if len(x)>3:
                        spl = UnivariateSpline(x, y, s=0)
                        f = interp1d(x, y, kind='cubic')
                        x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                        integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],datadict['DATAorig'][0][-1])[0]
                        datadict['JscCalc']=integral
                        print(datadict['JscCalc'])
                        integlist=[]
                        if integrationJscYes:
                            for item in x:
                                integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],item)[0]
                                integlist.append(integral)
                        datadict['integJsclist'].append(integlist)
                        
                        xE=[]
                        yln=[]
                        for xi in range(len(x)):
                            if y[xi]>0:
                                xE.append(1239.8/x[xi])
                                yln.append(math.log(100*y[xi]))
                        datadict['lnDat']=[xE,yln]
                        datadict['dataEnergyLn']=datadict['lnDat'][0]
                        datadict['dataIntLn']=datadict['lnDat'][1]
                    
                    datadict['ID']=datadict['samplename']+str(datadict['datetime'])+str(datadict['JscCalc'])
                    index1=1
                    while(1):
                        if datadict['ID'] in DATA.keys():
                            datadict['ID']=datadict['ID']+'_'+str(index1)
                            index1+=1
                        else:
                            break
                    datadict['NameMod']=datadict['samplename']
                    datadict['Name_Jsc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    datadict['Name_Eg']=datadict['samplename']+'_'+'EgIP: %.2f' % datadict['EgIP']
                    datadict['Name_Jsc_Eg']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    datadict['Name_Egln']=datadict['samplename']
                    datadict['Name_Jsc_Egln']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    datadict['Name_Egtauc']=datadict['samplename']
                    datadict['Name_Jsc_Egtauc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                    
                    DATA[datadict['ID']]=datadict
                
            elif ospath.splitext(file_path[k])[1]==".xls":
    #            samplename=ospath.basename(file_path[k].split('.')[0]).replace('-','_')  
                samplename=file_path[k].replace('\\','/') 
                samplename=samplename.split('/')[-1].replace('-','_').split('.')[0]
                datadict['samplename']=samplename
                print(samplename)
                try:
                    batchnumb=samplename.split('_')[0]
                except IndexError:
                    batchnumb='0'
                datadict['batchnumb']=batchnumb
                try:
                    samplenumb=samplename.split('_')[1]
                except IndexError:
                    samplenumb='0'
                
                datadict['samplenumb']=samplenumb
                
                datadict['file_path']=file_path[k]
                
                wb = xlrdopen_workbook(file_path[k])
                sheet_names = wb.sheet_names()
                if sheet_names == ['Data','Miscellaneous']: #files from SERIS c-Si cluster
                    print('Sample: %2f' % float(k+1))
                    xlsheet = wb.sheet_by_index(0)
                    if xlsheet.cell(1,1).value == 'External Quantum Efficiency':
                        print('EQE')
                        datadict['meastype']='EQE'
                        #read Jsc in xls file
                        i=0
                        Jsc=-1
                        while(1):#current calculated by setup and written in xls file. strangly not the same as the one calculated here below?
                            if xlsheet.cell(i,0).value == 'Integral / mA cm-2' or xlsheet.cell(i,0).value == 'Integral / A W-1':
                                try:
                                    Jsc=xlsheet.cell(i,1).value.replace('*','')
                                except AttributeError:
                                    Jsc=xlsheet.cell(i,1).value
                                datadict['JscfromFile']=float(Jsc)
                                break
                            i+=1
                        
                        datetime=modification_date(file_path[k])
                        datadict['datetime']=datetime
                        
                        i=3
                        while(1):
                            if xlsheet.cell(i,0).value !='':
                                datadict['DATAorig'][0].append(float(xlsheet.cell(i,0).value))
                                datadict['DATAorig'][1].append(1239.8/float(xlsheet.cell(i,0).value))
                                datadict['DATAorig'][2].append(float(xlsheet.cell(i,1).value)/100)
                                i+=1
                            else:
                                break
                        datadict['DATAmod']=deepcopy(datadict['DATAorig'])
                        
                        if integrationJscYes:#not implemented at the moment, quite slow as needs integration 
                            datadict['integJsclist']=[datadict['DATAorig'][0]]
                        else:
                            datadict['integJsclist']=[[]]
                        
                        #jsc calculation
                        x = datadict['DATAorig'][0]
                        y = datadict['DATAorig'][2]
                        if len(x)>3:
                            spl = UnivariateSpline(x, y, s=0)
                            f = interp1d(x, y, kind='cubic')
                            x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                            integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],datadict['DATAorig'][0][-1])[0]
                            datadict['JscCalc']=integral
                            print(datadict['JscCalc'])
                            integlist=[]
                            if integrationJscYes:
                                for item in x:
                                    integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],item)[0]
                                    integlist.append(integral)
                            datadict['integJsclist'].append(integlist)
                            #Eg calculation from linear normal curve
                            splder = spl.derivative(n=1)
                            splderlist = []
                            newx=[]
                            for item in x :
                                if item >400:
                                    splderlist.append(splder(item))
                                    newx.append(item)
                            if splderlist==[]:
                                datadict['Eg0']=0
                                datadict['EgIP']=0
                                datadict['tangent']=[0, 0]
                                datadict['slope']=0
                                datadict['h']=0
                            else:
                                minder=splderlist.index(min(splderlist))
                                xhighslope = newx[minder]
                                datadict['EgIP']=1239.8/xhighslope
                                yhighslope = spl(newx[minder]).tolist()
                                yprimehighslope = splder(newx[minder]).tolist()
                                Eg= 1239.8/(xhighslope - yhighslope/yprimehighslope)
                                datadict['Eg0']=Eg
                                datadict['tangent']=[yprimehighslope, yhighslope-yprimehighslope*xhighslope]#[pente,ordonnee a l'origine]
                                datadict['slope']=datadict['tangent'][0]
                                datadict['h']=datadict['tangent'][1]

                            #Eg calculation from ln(EQE) curve
                            xE=[]
                            yln=[]
                            for xi in range(len(x)):
                                if y[xi]>0:
                                    xE.append(1239.8/x[xi])
                                    yln.append(math.log(100*y[xi]))

                            datadict['lnDat']=[xE,yln]
                            datadict['dataEnergyLn']=datadict['lnDat'][0]
                            datadict['dataIntLn']=datadict['lnDat'][1]
                            
                            xErestricted=[]
                            ylnrestricted=[]
                            
                            for xi in range(len(xE)-1,-1,-1):
                                if yln[xi]<3 and yln[xi]>-2:
                                    xErestricted.append(xE[xi])
                                    ylnrestricted.append(yln[xi])
                            xErestricted.append(999)
                            ylnrestricted.append(999)
                            xErestricted2=[]
                            ylnrestricted2=[]
                            for xi in range(len(xErestricted)-1):
                                xErestricted2.append(xErestricted[xi])
                                ylnrestricted2.append(ylnrestricted[xi])
                                if abs(xErestricted[xi]-xErestricted[xi+1])>0.3:
                                    break
                            if len(xErestricted2)>1:
                                slope, intercept, r_value, p_value, std_err = stats.linregress(xErestricted2,ylnrestricted2)
                                print('slope')
                                print(slope)
                                print(1000/slope)
                                datadict['EgLn']=-intercept/slope
                                datadict['EuLn']=1000/slope #Eu calculation from ln(EQE) curve slope at band edge
                                datadict['tangentLn']=[slope, intercept,xErestricted2,ylnrestricted2]#[pente,ordonnee a l'origine]
                                datadict['slopeLn']=datadict['tangentLn'][0]
                                datadict['hln']=datadict['tangentLn'][1]
                                datadict['ptstgtLnX']=datadict['tangentLn'][2]
                                datadict['ptstgtLnY']=datadict['tangentLn'][3]
                                datadict['stderrEgLn']=[std_err,len(xErestricted2)]
                            else:
                                print("EgLn not found enough points...")
                            
                            try:
                                xtauc=[1239.8/xm for xm in x]
                                ytauc=[(y[m]*xtauc[m])**2 for m in range(len(y)) ]
                                xtauc=xtauc[::-1]
                                ytauc=ytauc[::-1]
                                spl = UnivariateSpline(xtauc, ytauc, s=0)
                                splder = spl.derivative(n=1)
                                splderlist = []
                                newx=[]
                                for item in xtauc :
                                    if item <2:
                                        splderlist.append(splder(item))
                                        newx.append(item)
                                
                                maxder=splderlist.index(max(splderlist))
                                xhighslope = newx[maxder]
                                yhighslope = spl(newx[maxder]).tolist()
                                yprimehighslope = splder(newx[maxder]).tolist()
                                Eg= (xhighslope - yhighslope/yprimehighslope)
                                
                                m=yprimehighslope
                                h=yhighslope-yprimehighslope*xhighslope
                                x2=Eg
                                x=linspace(x2,x2+0.1,10)
                                y=eval('m*x+h')
                                datadict['EgTauc']=[Eg,xtauc,ytauc,m,h]
                                datadict['xtauc']=datadict['EgTauc'][1]
                                datadict['ytauc']=datadict['EgTauc'][2]
                                datadict['mtauc']=datadict['EgTauc'][3]
                                datadict['htauc']=datadict['EgTauc'][4]
                                datadict['EgTauc']=datadict['EgTauc'][0]
                            except:
                                pass
                                
                        datadict['ID']=datadict['samplename']+'_'+str(datadict['datetime'])+'_'+str(datadict['JscCalc'])
                        index1=1
                        while(1):
                            if datadict['ID'] in DATA.keys():
                                datadict['ID']=datadict['ID']+'_'+str(index1)
                                index1+=1
                            else:
                                break
                        datadict['NameMod']=datadict['samplename']
                        datadict['Name_Jsc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        datadict['Name_Eg']=datadict['samplename']+'_'+'EgIP: %.2f' % datadict['EgIP']
                        datadict['Name_Jsc_Eg']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']+'_'+'EgIP: %.2f' % datadict['EgIP']
                        datadict['Name_Egln']=datadict['samplename']+'_'+'EgLn: %.2f' % datadict['EgLn']
                        datadict['Name_Jsc_Egln']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']+'_'+'EgLn: %.2f' % datadict['EgLn']
                        datadict['Name_Egtauc']=datadict['samplename']+'_'+'EgTauc: %.2f' % datadict['EgTauc']
                        datadict['Name_Jsc_Egtauc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']+'_'+'EgTauc: %.2f' % datadict['EgTauc']
                        
                        DATA[datadict['ID']]=datadict
                        
                    elif xlsheet.cell(1,1).value == 'Reflectance':
                        print('R')
                        datadict['meastype']='R'
                        datetime=modification_date(file_path[k])
                        datadict['datetime']=datetime
                        i=3
                        while(1):
                            if xlsheet.cell(i,0).value !='':
                                datadict['DATAorig'][0].append(float(xlsheet.cell(i,0).value))
                                datadict['DATAorig'][1].append(1239.8/float(xlsheet.cell(i,0).value))
                                datadict['DATAorig'][2].append(float(xlsheet.cell(i,1).value)/100)
                                i+=1
                            else:
                                break
                        datadict['DATAmod']=deepcopy(datadict['DATAorig'])
                        if integrationJscYes:#not implemented at the moment, quite slow as needs integration 
                            datadict['integJsclist']=[datadict['DATAorig'][0]]
                        else:
                            datadict['integJsclist']=[[]]
                        #jsc calculation
                        x = datadict['DATAorig'][0]
                        y = datadict['DATAorig'][2]
                        if len(x)>3:
                            spl = UnivariateSpline(x, y, s=0)
                            f = interp1d(x, y, kind='cubic')
                            x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                            integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],datadict['DATAorig'][0][-1])[0]
                            datadict['JscCalc']=integral
                            print(datadict['JscCalc'])
                            
                            integlist=[]
                            if integrationJscYes:
                                for item in x:
                                    integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],item)[0]
                                    integlist.append(integral)
                            datadict['integJsclist'].append(integlist)
                            xE=[]
                            yln=[]
                            for xi in range(len(x)):
                                if y[xi]>0:
                                    xE.append(1239.8/x[xi])
                                    yln.append(math.log(100*y[xi]))
                            datadict['lnDat']=[xE,yln]
                            datadict['dataEnergyLn']=datadict['lnDat'][0]
                            datadict['dataIntLn']=datadict['lnDat'][1]
                        
                        datadict['ID']=datadict['samplename']+str(datadict['datetime'])+str(datadict['JscCalc'])
                        index1=1
                        while(1):
                            if datadict['ID'] in DATA.keys():
                                datadict['ID']=datadict['ID']+'_'+str(index1)
                                index1+=1
                            else:
                                break
                        datadict['NameMod']=datadict['samplename']
                        datadict['Name_Jsc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        datadict['Name_Eg']=datadict['samplename']+'_'+'EgIP: %.2f' % datadict['EgIP']
                        datadict['Name_Jsc_Eg']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        datadict['Name_Egln']=datadict['samplename']
                        datadict['Name_Jsc_Egln']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        datadict['Name_Egtauc']=datadict['samplename']
                        datadict['Name_Jsc_Egtauc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        
                        DATA[datadict['ID']]=datadict
                        

                    elif xlsheet.cell(1,1).value == 'Detector Responsivity':
                        print('SR')
                        datadict['meastype']='SR'
                        i=0
                        Jsc=-1
                        while(1):#current calculated by setup and written in xls file. strangly not the same as the one calculated here below?
                            if xlsheet.cell(i,0).value == 'Integral / mA cm-2':
                                Jsc=float(xlsheet.cell(i,1).value[:-1])
                                datadict['JscfromFile']=float(Jsc)
                                break
                            i+=1
                            
                        datetime=modification_date(file_path[k])
                        datadict['datetime']=datetime
                        
                        i=3
                        while(1):
                            if xlsheet.cell(i,0).value !='':
                                datadict['DATAorig'][0].append(float(xlsheet.cell(i,0).value))
                                datadict['DATAorig'][1].append(1239.8/float(xlsheet.cell(i,0).value))
                                datadict['DATAorig'][2].append(float(xlsheet.cell(i,1).value)/100)
                                i+=1
                            else:
                                break
                        datadict['DATAmod']=deepcopy(datadict['DATAorig'])
                        if integrationJscYes:#not implemented at the moment, quite slow as needs integration 
                            datadict['integJsclist']=[datadict['DATAorig'][0]]
                        else:
                            datadict['integJsclist']=[[]]
                        
                        #jsc calculation
                        x = datadict['DATAorig'][0]
                        y = datadict['DATAorig'][2]
                        if len(x)>3:
                            spl = UnivariateSpline(x, y, s=0)
                            f = interp1d(x, y, kind='cubic')
                            x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                            integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],datadict['DATAorig'][0][-1])[0]
                            datadict['JscCalc']=integral
                            print(datadict['JscCalc'])
                            integlist=[]
                            if integrationJscYes:
                                for item in x:
                                    integral = echarge/10*integrate.quad(x2,datadict['DATAorig'][0][0],item)[0]
                                    integlist.append(integral)
                            datadict['integJsclist'].append(integlist)
                            
                            xE=[]
                            yln=[]
                            for xi in range(len(x)):
                                if y[xi]>0:
                                    xE.append(1239.8/x[xi])
                                    yln.append(math.log(100*y[xi]))
                            datadict['lnDat']=[xE,yln]
                            datadict['dataEnergyLn']=datadict['lnDat'][0]
                            datadict['dataIntLn']=datadict['lnDat'][1]
                        
                        datadict['ID']=datadict['samplename']+str(datadict['datetime'])+str(datadict['JscCalc'])
                        index1=1
                        while(1):
                            if datadict['ID'] in DATA.keys():
                                datadict['ID']=datadict['ID']+'_'+str(index1)
                                index1+=1
                            else:
                                break
                        datadict['NameMod']=datadict['samplename']
                        datadict['Name_Jsc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        datadict['Name_Eg']=datadict['samplename']+'_'+'EgIP: %.2f' % datadict['EgIP']
                        datadict['Name_Jsc_Eg']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        datadict['Name_Egln']=datadict['samplename']
                        datadict['Name_Jsc_Egln']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        datadict['Name_Egtauc']=datadict['samplename']
                        datadict['Name_Jsc_Egtauc']=datadict['samplename']+'_'+'Jsc: %.2f' % datadict['JscCalc']
                        
                        DATA[datadict['ID']]=datadict
                        
        # QMessageBox.information(self, 'Information', "It's done")
        # print("It's done")
        # self.initlistbox()
        
#%%#############        
    def initlistbox(self):
        global DATA
        
        self.names = ()
        self.names=DATA.keys()
        self.ui.listWidget.clear()
        self.ui.comboBox_LAsamples.clear()
        self.ui.comboBox_R.clear()
        self.ui.comboBox_T.clear()
        self.ui.listWidget.addItems(self.names)
        
        self.ui.comboBox_calccurrent.addItems(self.names)
        isEQE=0
        isR=0
        isT=0
        for item in self.names:
            if DATA[item]['meastype']=='EQE':
                self.ui.comboBox_LAsamples.addItem(item)
                isEQE=1
            elif DATA[item]['meastype']=='R':
                self.ui.comboBox_R.addItem(item)
                isR=1
            elif DATA[item]['meastype']=='T':
                self.ui.comboBox_T.addItem(item)
                isT=1
        if not isEQE:
            self.ui.comboBox_LAsamples.addItem('')
        if not isR:
            self.ui.comboBox_R.addItem('')
        if not isT:
            self.ui.comboBox_T.addItem('')
        
        #update Table
        self.updateTable(DATA)
        
        
    def updateTable(self, dictdata):
        try:
            self.ui.tableWidget.setRowCount(0)
            self.ui.tableWidget.setSortingEnabled(False)
            self.ui.tableWidget.setRowCount(len(dictdata.keys()))
            self.ui.tableWidget.setColumnCount(12)
            self.ui.tableWidget.setHorizontalHeaderLabels(
                ['sample','batch','sample#','date/time','meas. type','Jsc from file', 'Jsc recalc.','Eg0','EgIP','EgTauc','EgLn','EuLn'])
            i=0
            
            
            for key in dictdata.keys():
                self.ui.tableWidget.setItem(i,0,QTableWidgetItem(dictdata[key]['samplename']))
                self.ui.tableWidget.setItem(i,1,QTableWidgetItem(dictdata[key]['batchnumb']))
                self.ui.tableWidget.setItem(i,2,QTableWidgetItem(dictdata[key]['samplenumb']))
                self.ui.tableWidget.setItem(i,3,QTableWidgetItem(str(dictdata[key]['datetime'])))
                self.ui.tableWidget.setItem(i,4,QTableWidgetItem(dictdata[key]['meastype']))
                item3=QtWidgets.QTableWidgetItem()
                item3.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['JscfromFile']))
                self.ui.tableWidget.setItem(i,5,item3)
                item4=QtWidgets.QTableWidgetItem()
                item4.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['JscCalc']))
                self.ui.tableWidget.setItem(i,6,item4)
                item5=QtWidgets.QTableWidgetItem()
                item5.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['Eg0']))
                self.ui.tableWidget.setItem(i,7,item5)
                item6=QtWidgets.QTableWidgetItem()
                item6.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['EgIP']))
                self.ui.tableWidget.setItem(i,8,item6)
                item7=QtWidgets.QTableWidgetItem()
                item7.setData(QtCore.Qt.EditRole, QtCore.QVariant(dictdata[key]['EgTauc']))
                self.ui.tableWidget.setItem(i,9,item7)
                item8=QtWidgets.QTableWidgetItem()
                item8.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(dictdata[key]['EgLn'])))
                self.ui.tableWidget.setItem(i,10,item8)
                item9=QtWidgets.QTableWidgetItem()
                item9.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(dictdata[key]['EuLn'])))
                self.ui.tableWidget.setItem(i,11,item9)
                i+=1
            self.ui.tableWidget.setSortingEnabled(True)
        except RuntimeError:
            pass
        
#     def sortandexportEQEdat(self):
#         global DATAFORGRAPH
#         DATA=self.DATA
        
# #        selectedtoexport=list(self.listboxsamples.curselection())
# ##        print(selectedtoexport)
# #        selectedtoexport=[self.listboxsamples.get(i) for i in selectedtoexport]
# ##        print(selectedtoexport)

        
#         #creating excel summary file
#         Summary = []
#         for i in range(len(DATA)):
#             for j in range(len(DATA[i]['Jsc'])):
# #                print(DATA[i]['Name']+'_'+str(j)+'_'+ '%.2f' % DATA[i]['Jsc'][j])
# #                if DATA[i]['Name']+'_'+str(j)+'_'+ '%.2f' % DATA[i]['Jsc'][j] in selectedtoexport:
#                 Summary.append([DATA[i]['Name'],DATA[i]['Jsc'][j],DATA[i]['Eg0'][j],DATA[i]['EgLn'][j],
#                                 DATA[i]['EuLn'][j],DATA[i]['stderrEgLn'][j][0],DATA[i]['stderrEgLn'][j][1],
#                                 DATA[i]['EgTauc'][j][0],DATA[i]['EgIP'][j],DATA[i]['comment'],DATA[i]['Vbias'][j],
#                                 DATA[i]['filterbias'][j],DATA[i]['ledbias'][j],DATA[i]['dateTime']])
#         Summary.insert(0,['Sample Name','Jsc','Eg0','EgLn','EuLn','stderrEgLn','nbptsEgLn','EgTauc','EgIP','comment','Vbias','filterbias','ledbias','datetimeMod'])
#         workbook = xlsxwriter.Workbook('Summary.xlsx')
#         worksheet = workbook.add_worksheet()
#         row=0
#         for name, jsc, eg, egln, euln, stderr, nbptEgLn, EgTauc, EgIP, comment, Vbias, filterbias, ledbias, dateandtime in Summary:
#             worksheet.write(row, 0, name)
#             if jsc!=999:
#                 worksheet.write(row, 1, jsc)
#             else:
#                 worksheet.write(row, 1, ' ')
#             if eg!=999:
#                 worksheet.write(row, 2, eg)
#             else:
#                 worksheet.write(row, 2, ' ')
#             if egln!=999:
#                 worksheet.write(row, 3, egln)
#             else:
#                 worksheet.write(row, 3, ' ')
#             if euln!=999:
#                 worksheet.write(row, 4, euln)
#             else:
#                 worksheet.write(row, 4, ' ')
#             if stderr!=999:
#                 worksheet.write(row, 5, stderr)
#             else:
#                 worksheet.write(row, 5, ' ')
#             if nbptEgLn!=999:
#                 worksheet.write(row, 6, nbptEgLn)
#             else:
#                 worksheet.write(row, 6, ' ')
#             if EgTauc!=999:
#                 worksheet.write(row, 7, EgTauc)
#             else:
#                 worksheet.write(row, 7, ' ')
#             if EgIP!=999:
#                 worksheet.write(row, 8, EgIP)
#             else:
#                 worksheet.write(row, 8, ' ')    
                
#             worksheet.write(row, 9, comment)
#             worksheet.write(row, 10, Vbias)
#             worksheet.write(row, 11, filterbias)
#             worksheet.write(row, 12, ledbias)
#             worksheet.write(row, 13, dateandtime)
#             row += 1
#         workbook.close()
        
#         #creating text files with eqe data and currents
#         for i in range(len(DATA)):
#             listeqe=self.ncolumneqe(int(DATA[i]['NbColumn']))
#             listeqe += asarray(DATA[i]['DATA']).T.tolist()
#             content1=[]
#             for j in range(len(listeqe)):
#                 strr=''
#                 for k in range(len(listeqe[j])):
#                     strr = strr + str(listeqe[j][k])+'\t'
#                 strr = strr[:-1]+'\n'
#                 content1.append(strr)
            
#             namerow =DATA[i]['Name']+'\t'
#             for k in range(len(DATA[i]['Jsc'])):
#                 namerow +='J = '+'%.2f' % DATA[i]['Jsc'][k]+' mA/cm2\t'
#             namerow=namerow[:-1]+'\n'   
#             content1.insert(2,namerow)    
                
#             file = open(DATA[i]['Name'] + '.txt','w', encoding='ISO-8859-1')
#             file.writelines("%s" % item for item in content1)
#             file.close()
        
#         #creating text files with eqe integrated total Jsc
#         for i in range(len(DATA)):
#             if DATA[i]['integJsclist'][0]!=[]:
#                 listeqe=self.ncolumneqeJsc(int(DATA[i]['NbColumn']))
#                 listeqe += asarray(DATA[i]['integJsclist']).T.tolist()
#                 content1=[]
#                 for j in range(len(listeqe)):
#                     strr=''
#                     for k in range(len(listeqe[j])):
#                         strr = strr + str(listeqe[j][k])+'\t'
#                     strr = strr[:-1]+'\n'
#                     content1.append(strr)
                
#                 namerow =DATA[i]['Name']+'\t'
#                 for k in range(len(DATA[i]['Jsc'])):
#                     namerow +='J = '+'%.2f' % DATA[i]['Jsc'][k]+' mA/cm2\t'
#                 namerow=namerow[:-1]+'\n'   
#                 content1.insert(2,namerow)    
                    
#                 file = open(DATA[i]['Name'] + '_integJsc.txt','w', encoding='ISO-8859-1')
#                 file.writelines("%s" % item for item in content1)
#                 file.close()
            
#         #creating graphs
#         plt.clf()
#         for i in range(len(DATA)):
#             for k in range(1,int(DATA[i]['NbColumn']),1):
#                 x=DATA[i]['DATA'][0]
# #                print(x[0])
# #                print(x[-1])
#                 plt.plot(x,DATA[i]['DATA'][k])
#                 plt.axis([x[0],x[-1],0,1])
#             plt.xlabel('Wavelength (nm)')
#             plt.ylabel('EQE (-)')
#             text='Jsc= '
#             for m in range(len(DATA[i]['Jsc'])):
#                 text += '%.2f' % DATA[i]['Jsc'][m]+ '; '
#             text+='Eg= '
#             for m in range(len(DATA[i]['Eg0'])):
#                 text += '%.2f' % DATA[i]['Eg0'][m]+ '; '
#             text=text[:-2]
#             plt.annotate(DATA[i]['Name']+' - '+text, xy=(0.1,1.01), xycoords='axes fraction', fontsize=12,
#                                                 horizontalalignment='left', verticalalignment='bottom')
#             plt.savefig(DATA[i]['Name']+'.png')
#             plt.clf()
        
#         #creating graphs for integJsc
#         if DATA[0]['integJsclist'][0]!=[]:
#             plt.clf()
#             maxyvalues=[]
#             for i in range(len(DATA)):
#                 for k in range(1,int(DATA[i]['NbColumn']),1):
#                     x=DATA[i]['integJsclist'][0]
#                     plt.plot(x,DATA[i]['integJsclist'][k],label=DATA[i]['Name']+'_'+'%.2f'% DATA[i]['Jsc'][k-1])
#                     maxyvalues.append(DATA[i]['integJsclist'][k][-1])
#             plt.legend()
#             plt.axis([x[0],x[-1],0,math.ceil(max(maxyvalues))])
#             plt.xlabel('Wavelength (nm)')
#             plt.ylabel('Integrated Jsc (mA/cm2)')
#             plt.savefig('integJsc.png')
#             plt.clf()
        
        
#         plt.close()
#         self.destroy()
#         self.__init__()
#         self.UpdateEQEGraph()
# #        self.initlistbox()
     
#     def WriteEQEtoDatabase(self):
#         global DATA
        
#         #connection to DB
#         path = QFileDialog.getOpenFileNames(caption = 'Please select the DB file')[0]
#         self.db_conn=sqlite3connect(path)
#         self.theCursor=self.db_conn.cursor()
        
#         self.theCursor.execute("SELECT batchname FROM batch")
#         batchnamesdb=self.theCursor.fetchall()
#         print(batchnamesdb)
#         self.theCursor.execute("SELECT samplename FROM samples")
#         batchnamesdb=self.theCursor.fetchall()
#         print(batchnamesdb)
        
#         print("EQEs...")
        
#         for i in DATA.keys():
#             batchname=DATA[i]["batchnumb"]
# #            print(batchname)
#             samplenumber=DATA[i]["samplenumb"]
# #            print(samplenumber)
#             samplenumber = batchname+'_'+samplenumber
# #            samplenumber = DATA[i]["Name"]
# #            print(samplenumber)
            
#             self.theCursor.execute("SELECT id FROM batch WHERE batchname=?",(batchname,))
#             batch_id_exists = self.theCursor.fetchone()[0]
#             self.theCursor.execute("SELECT id FROM samples WHERE samplename=?",(samplenumber,))            
#             sample_id_exists = self.theCursor.fetchone()[0]
#             self.theCursor.execute("SELECT id FROM cells WHERE samples_id=? AND batch_id=?",(sample_id_exists, batch_id_exists))            
#             cellletter_id_exists = self.theCursor.fetchall()[0][0]
# #            print(cellletter_id_exists)

#             if batch_id_exists and sample_id_exists and cellletter_id_exists:
#                 for j in range(len(DATA[i]['JscCalc'])):
#                     uniquedatentry=DATA[i]['samplename']+str(DATA[i]['datetime'])+str(DATA[i]['JscCalc'][j])
#                     uniquedatentry=uniquedatentry.replace(' ','')
#                     uniquedatentry=uniquedatentry.replace(':','')
#                     uniquedatentry=uniquedatentry.replace('.','')
#                     uniquedatentry=uniquedatentry.replace('-','')
#                     uniquedatentry=uniquedatentry.replace('_','')
#                     try:
#                         self.db_conn.execute("""INSERT INTO eqemeas (
#                                 EQEmeasname,
#                                 EQEmeasnameDateTimeEQEJsc,
#                                 commenteqe,
#                                 DateTimeEQE,
#                                 Vbias,
#                                 filter,
#                                 LEDbias,
#                                 integJsc,
#                                 Eg0,
#                                 EgIP,
#                                 EgTauc,
#                                 EgLn,
#                                 linktofile,
#                                 samples_id,
#                                 batch_id,
#                                 cells_id
#                             ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
#                             (DATA[i]['samplename'],
#                              uniquedatentry,
#                              DATA[i]['comment'],
#                              DATA[i]['datetime'],
#                              DATA[i]['Vbias'][j],
#                              DATA[i]['filterbias'][j],
#                              DATA[i]['ledbias'][j],
#                              DATA[i]['JscCalc'][j],
#                              DATA[i]['Eg0'][j],
#                              DATA[i]['EgIP'][j],
#                              DATA[i]['EgTauc'][j][0],
#                              DATA[i]['EgLn'][j],
#                              DATA[i]['file_path'],
#                              sample_id_exists,
#                              batch_id_exists,
#                              cellletter_id_exists))
#                         self.db_conn.commit()
#                     except sqlite3IntegrityError:
#                         print("the file already exists in the DB")
        
#         #disconnect from DB
#         self.theCursor.close()
#         self.db_conn.close()
        
#         #exit window
# #        print("it's in the DB!")
#         QMessageBox.information(self,'Information', "it's in the DB!")
        
#%%#############         
    # def SampleNames(self, DATAx):#for DATAFORGRAPH, obsolete function
    #     Names = list(self.names)
    #     for item in range(len(DATAx)):
    #         Names.append(DATAx[item][0]+'_'+ '%.2f' % DATAx[item][1])
    #     return tuple(Names)
    
    def UpdateEQEGraph(self):
        global titEQE
        global takenforplot
        global DATA
        global DATAforexport
        global colorstylelist
        
#        takenforplot=list(self.listboxsamples.curselection())
        
        DATAx=DATA
        
        sampletotake=[]
        DATAforexport=[]
        
        if takenforplot!=[]:
            sampletotake=takenforplot
            # print(sampletotake)
            
            if self.ui.comboBox.currentText()=="linear":
                if 1: #not self.ui.checkBox_integrJsc.isChecked():
                    self.EQEgraph.clear()
                    self.EQEgraphY2.clear()
                    self.EQEgraphY2.get_yaxis().set_visible(False)
                    EQEfig=self.EQEgraph
                    for i in range(len(sampletotake)):
                        # print(sampletotake[i])
                        x = DATAx[sampletotake[i]]['DATAmod'][0]
                        y = DATAx[sampletotake[i]]['DATAmod'][2]
                        
                        colx=["Wavelength","nm"," "]+x
                        coly=["EQE","-",DATAx[sampletotake[i]]['NameMod']]+y
                        DATAforexport.append(colx)
                        DATAforexport.append(coly)
                        
                        if self.ui.checkBox_legend.isChecked():
                            if not self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['NameMod'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                            elif self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                            elif not self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Eg'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                            elif self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc_Eg'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        else:
                            EQEfig.plot(x,y,linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        
                        if self.ui.checkBox_showsecreteg.isChecked():
                            if 'EQE' in DATAx[sampletotake[i]]['meastype'] or 'IQE' in DATAx[sampletotake[i]]['meastype']:
                                m=DATAx[sampletotake[i]]['slope']
                                h=DATAx[sampletotake[i]]['h']
                                x2=1239.8/DATAx[sampletotake[i]]['Eg0']
                                x=linspace(x2-100,x2,10)
                                #x=array(range(int(round(x2-100)),int(round(x2))))
                                y=eval('m*x+h')
                                EQEfig.plot(x,y,linewidth=DATAx[sampletotake[i]]['linewidth'])
                                
                    self.EQEgraph.set_ylabel('EQE (-)', fontsize=self.ui.spinBox_fontsize.value())
                    self.EQEgraph.set_xlabel('Wavelength (nm)', fontsize=self.ui.spinBox_fontsize.value())
                    
                    
                # else:#if wants the Jsc integrated curve in 2-y-axis graph
                #     self.EQEgraph.clear()
                #     self.EQEgraphY2.clear()
                #     self.EQEgraphY2.get_yaxis().set_visible(True)
                #     EQEfig=self.EQEgraph
                #     ax2=self.EQEgraphY2
                #     maxyvalues=[]
                #     for i in range(len(sampletotake)):
                #         try:
                #             x = DATAx[sampletotake[i]][2][1]
                #             y = DATAx[sampletotake[i]][3][1]
                #             y2= DATAx[sampletotake[i]][28]
                #             maxyvalues.append(y2[-1])
                #             colx=["Wavelength","nm"," "]+x
                #             coly=["EQE","-",DATAx[sampletotake[i]][5]]+y
                #             coly2=["Jsc","mA/cm2",DATAx[sampletotake[i]][5]]+y2
                #             DATAforexport.append(colx)
                #             DATAforexport.append(coly)
                #             DATAforexport.append(coly2)
                            
                #             if self.ui.checkBox_legend.isChecked():
                #                 if not self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                #                     EQEfig.plot(x,y,label=DATAx[sampletotake[i]][5],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
                #                     ax2.plot(x,y2,label=DATAx[sampletotake[i]][5],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])                      
                #                 elif self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                #                     EQEfig.plot(x,y,label=DATAx[sampletotake[i]][6],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
                #                     ax2.plot(x,y2,label=DATAx[sampletotake[i]][6],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
                #                 elif not self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                #                     EQEfig.plot(x,y,label=DATAx[sampletotake[i]][7],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
                #                     ax2.plot(x,y2,label=DATAx[sampletotake[i]][7],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])                        
                #                 elif self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                #                     EQEfig.plot(x,y,label=DATAx[sampletotake[i]][8],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
                #                     ax2.plot(x,y2,label=DATAx[sampletotake[i]][8],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])                    
                #             else:
                #                 EQEfig.plot(x,y,linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
                #                 ax2.plot(x,y2,linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
                                
                #             if self.ui.checkBox_showsecreteg.isChecked():
                #                 m=DATAx[sampletotake[i]][11]
                #                 h=DATAx[sampletotake[i]][12]
                #                 x2=1239.8/DATAx[sampletotake[i]][4]
                #                 x=linspace(x2-100,x2,10)
                #                 #x=array(range(int(round(x2-100)),int(round(x2))))
                #                 y=eval('m*x+h')
                #                 EQEfig.plot(x,y,linewidth=DATAx[sampletotake[i]][29])
                #         except:
                #             pass
                                
                #     self.EQEgraph.set_ylabel('EQE (-)', fontsize=self.ui.spinBox_fontsize.value())
                #     self.EQEgraphY2.set_ylabel('Jsc',fontsize=self.ui.spinBox_fontsize.value())
                #     self.EQEgraph.set_xlabel('Wavelength (nm)', fontsize=self.ui.spinBox_fontsize.value())
                    
                
            elif self.ui.comboBox.currentText()=="log":
                self.EQEgraph.clear()
                self.EQEgraphY2.clear()
                self.EQEgraphY2.get_yaxis().set_visible(False)
    #            self.EQEgraphY2.get_xaxis().set_visible(False)
                EQEfig=self.EQEgraph
                # minxlist=[]
                # maxxlist=[]
                # minylist=[]
                # maxylist=[]
                for i in range(len(sampletotake)):
                    x = DATAx[sampletotake[i]]['lnDat'][0]
                    y = DATAx[sampletotake[i]]['lnDat'][1]
                    # minxlist.append(x[0])
                    # maxxlist.append(x[-1])
                    # minylist.append(min(y))
                    # maxylist.append(max(y))
                    colx=["Energy","eV"," "]+x
                    coly=["Ln(EQE)","-",DATAx[sampletotake[i]]['NameMod']]+y
                    DATAforexport.append(colx)
                    DATAforexport.append(coly)
                    
                    if self.ui.checkBox_legend.isChecked():
                        if not self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['NameMod'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif not self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Egln'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc_Egln'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                    else:
                        EQEfig.plot(x,y,linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                    
                    if self.ui.checkBox_showsecreteg.isChecked():
                        m=DATAx[sampletotake[i]]['slopeLn']
                        h=DATAx[sampletotake[i]]['hln']
                        x2=DATAx[sampletotake[i]]['Eg0']
                        x=linspace(x2-0.1,x2+0.1,10)
                        y=eval('m*x+h')
                        EQEfig.plot(x,y,linewidth=DATAx[sampletotake[i]]['linewidth'])
                        EQEfig.plot(DATAx[sampletotake[i]]['ptstgtLnX'],DATAx[sampletotake[i]]['ptstgtLnY'],'ro')
                        
                self.EQEgraph.set_ylabel('Ln(EQE) (-)', fontsize=self.ui.spinBox_fontsize.value())
                self.EQEgraph.set_xlabel('Energy (eV)', fontsize=self.ui.spinBox_fontsize.value())
                
            elif self.ui.comboBox.currentText()=="Tauc1":
                self.EQEgraph.clear()
                self.EQEgraphY2.clear()
                self.EQEgraphY2.get_yaxis().set_visible(False)
                EQEfig=self.EQEgraph
                # minxlist=[]
                # maxxlist=[]
                # minylist=[]
                # maxylist=[]
                for i in range(len(sampletotake)):
                    x = DATAx[sampletotake[i]]['xtauc']
                    y = DATAx[sampletotake[i]]['ytauc']
                    # minxlist.append(x[0])
                    # maxxlist.append(x[-1])
                    # minylist.append(min(y))
                    # maxylist.append(max(y))
                    colx=["Energy","eV"," "]+x
                    coly=["(EQE*E)^2","(eV)^2",DATAx[sampletotake[i]]['NameMod']]+y
                    DATAforexport.append(colx)
                    DATAforexport.append(coly)
                    
                    if self.ui.checkBox_legend.isChecked():
                        if not self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['NameMod'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif not self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Egtauc'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc_Egtauc'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                    else:
                        EQEfig.plot(x,y,linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                    
                    if self.ui.checkBox_showsecreteg.isChecked():
                        m=DATAx[sampletotake[i]]['mtauc']
                        h=DATAx[sampletotake[i]]['htauc']
                        x2=DATAx[sampletotake[i]]['EgTauc']
                        x=linspace(x2,x2+0.1,10)
                        y=eval('m*x+h')
                        EQEfig.plot(x,y,linewidth=DATAx[sampletotake[i]]['linewidth'])
                    
                self.EQEgraph.set_ylabel("(EQE*E)^2 (eV)^2", fontsize=self.ui.spinBox_fontsize.value())
                self.EQEgraph.set_xlabel('Energy (eV)', fontsize=self.ui.spinBox_fontsize.value())
                
    #        elif self.ui.comboBox.currentText()=="Tauc2":
    #            self.EQEgraph.clear()
    #            self.EQEgraphY2.clear()
    #            self.EQEgraphY2.get_yaxis().set_visible(False)
    #            EQEfig=self.EQEgraph
    #            minxlist=[]
    #            maxxlist=[]
    #            minylist=[]
    #            maxylist=[]
    #            for i in range(len(sampletotake)):
    #                x = DATAx[sampletotake[i]][20]
    #                y = DATAx[sampletotake[i]][21]                
    #                minxlist.append(x[0])
    #                maxxlist.append(x[-1])
    #                minylist.append(min(y))
    #                maxylist.append(max(y))
    #                colx=["Energy","eV"," "]+x
    #                coly=["Ln(1-EQE)^2 * E^2","a.u.",DATAx[sampletotake[i]][5]]+y
    #                DATAforexport.append(colx)
    #                DATAforexport.append(coly)
    #                
    #                if self.CheckLegend.get()==1:
    #                    if self.CheckLegJsc.get()==0 and self.CheckLegEg.get()==0:
    #                        EQEfig.plot(x,y,label=DATAx[sampletotake[i]][5],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
    #                    elif self.CheckLegJsc.get()==1 and self.CheckLegEg.get()==0:
    #                        EQEfig.plot(x,y,label=DATAx[sampletotake[i]][6],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
    #                    elif self.CheckLegJsc.get()==0 and self.CheckLegEg.get()==1:
    #                        EQEfig.plot(x,y,label=DATAx[sampletotake[i]][26],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
    #                    elif self.CheckLegJsc.get()==1 and self.CheckLegEg.get()==1:
    #                        EQEfig.plot(x,y,label=DATAx[sampletotake[i]][27],linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
    #                else:
    #                    EQEfig.plot(x,y,linestyle=DATAx[sampletotake[i]][9],color=DATAx[sampletotake[i]][10],linewidth=DATAx[sampletotake[i]][29])
    #                
    #                if self.CheckTangent.get()==1:
    #                    m=DATAx[sampletotake[i]][22]
    #                    h=DATAx[sampletotake[i]][23]
    #                    x2=DATAx[sampletotake[i]][19]
    #                    x=linspace(x2,x2+0.1,10)
    #                    y=eval('m*x+h')
    #                    EQEfig.plot(x,y)
    #                
    #            self.EQEgraph.set_ylabel('Ln(1-EQE)^2 * E^2 (a.u.)', fontsize=14)
    #            self.EQEgraph.set_xlabel('Energy (eV)', fontsize=14)
    #            if titEQE:
    #                self.EQEgraph.set_title(self.titleEQE.get())
    #            if self.CheckLegend.get()==1:
    #                if self.pos1.get()==5:
    #                    self.leg=EQEfig.legend(bbox_to_anchor=(1, 0.5), loc=2, ncol=1)
    #                elif self.pos1.get()==1 or self.pos1.get()==2  or self.pos1.get()==3 or self.pos1.get()==4:   
    #                    self.leg=EQEfig.legend(loc=self.pos1.get())
    #                else:
    #                    self.leg=EQEfig.legend(loc=0)
    ##            if self.CheckAutoscale.get()==0:        
    ##                self.EQEgraph.axis([self.minx.get(),self.maxx.get(),self.miny.get(),self.maxy.get()])
    #            self.EQEgraph.axis([min(minxlist),max(maxxlist),min(minylist),math.ceil(max(maxylist))])
    #            plt.gcf().canvas.draw()    
                
            elif self.ui.comboBox.currentText()=="NormalizedBySingle":
                self.EQEgraph.clear()
                self.EQEgraphY2.clear()
                self.EQEgraphY2.get_yaxis().set_visible(False)
                EQEfig=self.EQEgraph
                for i in range(len(sampletotake)):
                    x = DATAx[sampletotake[i]]['DATAmod'][0]
                    y1 = DATAx[sampletotake[i]]['DATAmod'][2]
                    
                    y=[(m-min(y1))/(max(y1)-min(y1)) for m in y1]
                    
                    colx=["Wavelength","nm"," "]+x
                    coly=["Norm. EQE","-",DATAx[sampletotake[i]]['NameMod']]+y
                    DATAforexport.append(colx)
                    DATAforexport.append(coly)
                    
                    if self.ui.checkBox_legend.isChecked():
                        if not self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['NameMod'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif not self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Eg'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        elif self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                            EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc_Eg'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                    else:
                        EQEfig.plot(x,y,linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                self.EQEgraph.set_ylabel('Norm. EQE (-)', fontsize=self.ui.spinBox_fontsize.value())
                self.EQEgraph.set_xlabel('Wavelength (nm)', fontsize=self.ui.spinBox_fontsize.value())
                
            elif self.ui.comboBox.currentText()=="NormalizedByAll":
                self.EQEgraph.clear()
                self.EQEgraphY2.clear()
                self.EQEgraphY2.get_yaxis().set_visible(False)
                EQEfig=self.EQEgraph
                AllY=[]
                for i in range(len(sampletotake)):
                    AllY+=DATAx[sampletotake[i]]['DATAmod'][2]
                if AllY!=[]:
                    miny=min(AllY)
                    maxy=max(AllY)
                    
                    for i in range(len(sampletotake)):
                        x = DATAx[sampletotake[i]]['DATAmod'][0]
                        y1 = DATAx[sampletotake[i]]['DATAmod'][2]
                        
                        y=[(m-miny)/(maxy-miny) for m in y1]
                        
                        colx=["Wavelength","nm"," "]+x
                        coly=["Norm. EQE","-",DATAx[sampletotake[i]]['NameMod']]+y
                        DATAforexport.append(colx)
                        DATAforexport.append(coly)
                        
                        if self.ui.checkBox_legend.isChecked():
                            if not self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['NameMod'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                            elif self.ui.checkBox_Jsc.isChecked() and not self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                            elif not self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Eg'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                            elif self.ui.checkBox_Jsc.isChecked() and self.ui.checkBox_Eg.isChecked():
                                EQEfig.plot(x,y,label=DATAx[sampletotake[i]]['Name_Jsc_Eg'],linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                        else:
                            EQEfig.plot(x,y,linestyle=DATAx[sampletotake[i]]['linestyle'],color=DATAx[sampletotake[i]]['linecolor'],linewidth=DATAx[sampletotake[i]]['linewidth'])
                self.EQEgraph.set_ylabel('Norm. EQE (-)', fontsize=self.ui.spinBox_fontsize.value())
                self.EQEgraph.set_xlabel('Wavelength (nm)', fontsize=self.ui.spinBox_fontsize.value())
            
            if titEQE:
                self.EQEgraph.set_title(self.titleEQE.get())
            if self.ui.checkBox_legend.isChecked():
                if self.ui.radioButton_topleft.isChecked():
                    self.leg=self.EQEgraph.legend(loc=2, fontsize = self.ui.spinBox_fontsize.value())
                elif self.ui.radioButton_topright.isChecked():
                    self.leg=self.EQEgraph.legend(loc=1, fontsize = self.ui.spinBox_fontsize.value())
                elif self.ui.radioButton_bottomleft.isChecked():
                    self.leg=self.EQEgraph.legend(loc=3, fontsize = self.ui.spinBox_fontsize.value())
                elif self.ui.radioButton_bottomright.isChecked():
                    self.leg=self.EQEgraph.legend(loc=4, fontsize = self.ui.spinBox_fontsize.value())
                elif self.ui.radioButton_outside.isChecked():
                    self.leg=self.EQEgraph.legend(bbox_to_anchor=(1, 1), loc='upper left', ncol=1, fontsize = self.ui.spinBox_fontsize.value())
                elif self.ui.radioButton_best.isChecked():
                    self.leg=self.EQEgraph.legend(loc=0, fontsize = self.ui.spinBox_fontsize.value())
            if self.ui.checkBox_AutoScale.isChecked():
                self.EQEgraph.autoscale()
            else:
                self.EQEgraph.axis([self.ui.doubleSpinBox_Xmin.value(),self.ui.doubleSpinBox_Xmax.value(),self.ui.doubleSpinBox_Ymin.value(),self.ui.doubleSpinBox_Ymax.value()])
    
            for item in ([self.EQEgraph.title, self.EQEgraph.xaxis.label, self.EQEgraph.yaxis.label] +
                         self.EQEgraph.get_xticklabels() + self.EQEgraph.get_yticklabels()):
                item.set_fontsize(self.ui.spinBox_fontsize.value())
                    
            self.fig.canvas.draw_idle()
            
        

#%%#############             
    def ExportTableData(self):
        path = QFileDialog.getSaveFileName(self, 'Save table data',"tabledata.csv", "CSV Files(*.csv *.txt)")[0]
     
        with open(path, 'w') as fd:
            writer = csv.writer(fd, lineterminator='\n')
            
            # write headers
            headers = []
            for column in range(self.ui.tableWidget.columnCount()):
                header = self.ui.tableWidget.horizontalHeaderItem(column)
                if header is not None:
                     headers.append(header.text()+'\t')
                else:
                    headers.append('\t')
            writer.writerow(headers)
             
            # write data
            for row in range(self.ui.tableWidget.rowCount()):
                rowdata = []
                for column in range(self.ui.tableWidget.columnCount()):
                    item = self.ui.tableWidget.item(row, column)
                    if item is not None:
                        rowdata.append(item.text())
                    else:
                        rowdata.append('')
                writer.writerow(rowdata)
        
    def ExportEQEGraph(self):
        global DATAforexport
        #graphname = self.entrytext.get()
        #plt.savefig(graphname +'.png', bbox_extra_artists=(self.leg,), bbox_inches='tight') 
        try:
            path = QFileDialog.getSaveFileName(self, 'Save graph', ".png", "graph file (*.png);; All Files (*)")[0]
            
            if path!='':
                if self.ui.checkBox_legend.isChecked():
                    self.fig.savefig(path, dpi=300, bbox_extra_artists=(self.leg,))#, transparent=True)
                else:
                    self.fig.savefig(path, dpi=300)#, transparent=True)
                        
                DATAforexport=map(list, sixmoves.zip_longest(*DATAforexport, fillvalue=' '))
    
                DATAforexport1=[]
                for item in DATAforexport:
                    line=""
                    for item1 in item:
                        line=line+str(item1)+"\t"
                    line=line[:-1]+"\n"
                    DATAforexport1.append(line)
                    
                file = open(str(path[:-4]+"_dat.txt"),'w', encoding='ISO-8859-1')
                file.writelines("%s" % item for item in DATAforexport1)
                file.close() 
        except:
            QMessageBox.information(self,'Information', "there is an exception...check legend maybe...")
             

    def CalcCurrent(self):
        global DATA
        itempos=self.ui.comboBox_calccurrent.currentText()
        try:
            x = DATA[itempos]['DATAmod'][0]
            y = DATA[itempos]['DATAmod'][2]
            f = interp1d(x, y, kind='cubic')
            x2 = lambda x: self.AM15GParticlesinnm(x)*f(x)
            integral = echarge/10*integrate.quad(x2,self.ui.spinBox_from.value(),self.ui.spinBox_to.value())[0]
            self.ui.doubleSpinBox_calcjsc.setValue(integral)
        except ValueError:
            QMessageBox.information(self,'Information', "a limit value is outside of interpolation range,\nmin: "+str(x[0])+", max: "+str(x[-1]))



    def select(self):
        global takenforplot
        # takenforplot = [str(self.ui.listWidget.selectedItems()[i].text()) for i in range(len(self.ui.listWidget.selectedItems()))]
        takenforplot = self.ui.listWidget.selectedItems()
        takenforplot=[x.text() for x in takenforplot]
        # print('\nhere')
        # print(takenforplot)
        self.UpdateEQEGraph()

#%%#############             
        
    def chooseRfile(self):
        #ask to open file
        file_path = QFileDialog.getOpenFileName(caption = 'Please select the R files')[0]
        # print(file_path)
        #send the path to label
        self.loadR_LA(file_path)
        
    def loadR_LA(self,A):
        global DATA, tempDATAlossA

        if A=='combo':
            self.updatePathtoR(self.ui.comboBox_R.currentText())
        else:
            self.updatePathtoR(A)
        
    def updatePathtoR(self, pathtoR):#the path is either the ID/key in DATA or a new path to a file to be read. 
        global DATA
        self.ui.label_pathToR.setText(pathtoR)

    def chooseTfile(self):
        #ask to open file
        file_path = QFileDialog.getOpenFileName(caption = 'Please select the T files')[0]
        # print(file_path)
        #send the path to label
        self.loadT_LA(file_path)
        
    def chooseAlternEQEfile(self):
        #ask to open file
        file_path = QFileDialog.getOpenFileName(caption = 'Please select the Alternative EQE files')[0]
        self.ui.label_pathToOtherEQE.setText(file_path)
        
    def loadT_LA(self,A):
        global DATA, tempDATAlossA

        if A=='combo':
            self.updatePathtoT(self.ui.comboBox_T.currentText())
        else:
            self.updatePathtoT(A)
        
    def updatePathtoT(self, pathtoT):#the path is either the ID/key in DATA or a new path to a file to be read. 
        global DATA
        self.ui.label_pathToT.setText(pathtoT)
        
        
    def updateLossAnalysis(self):
        global DATA
        
        tempDATAlossA={}
        
        #load EQE
        sampleselected=self.ui.comboBox_LAsamples.currentText()
        if sampleselected!='':
            
            if self.ui.radioButton_fullareaEQE.isChecked():
                tempDATAlossA['EQExfullarea']=DATA[sampleselected]['DATAmod'][0]
                tempDATAlossA['EQEyfullarea']=DATA[sampleselected]['DATAmod'][2]
                tempDATAlossA['interpEQEfullarea']=interp1d(tempDATAlossA['EQExfullarea'], tempDATAlossA['EQEyfullarea'], bounds_error=False, fill_value=0.)
            else:
                tempDATAlossA['EQExfullarea']=[]
                tempDATAlossA['EQEyfullarea']=[]
                tempDATAlossA['interpEQEfullarea']=[]
            if self.ui.radioButton_spotEQE.isChecked():
                tempDATAlossA['EQExspotarea']=DATA[sampleselected]['DATAmod'][0]
                tempDATAlossA['EQEyspotarea']=DATA[sampleselected]['DATAmod'][2]
                tempDATAlossA['interpEQEspotarea']=interp1d(tempDATAlossA['EQExspotarea'], tempDATAlossA['EQEyspotarea'], bounds_error=False, fill_value=0.)
            else:
                tempDATAlossA['EQExspotarea']=[]
                tempDATAlossA['EQEyspotarea']=[]
                tempDATAlossA['interpEQEspotarea']=[]
            
            if ':/' in self.ui.label_pathToOtherEQE.text():#if true then something was loaded by user
                # print('there is a path')
                alternativeEQEx=[]
                alternativeEQEy=[]
                if '.txt' in self.ui.label_pathToOtherEQE.text():
                    # print('its a path and txt')

                    file = open(self.ui.label_pathToOtherEQE.text(), encoding='ISO-8859-1')
                    Edata = file.readlines()
                    file.close()
                    for i in range(len(Edata)):
                        pos = Edata[i].find('\t')
                        alternativeEQEx.append(float(Edata[i][:pos]))
                        alternativeEQEy.append(float(Edata[i][pos+1:-1])/100)
                    
                elif '.xls' in self.ui.label_pathToOtherEQE.text():
                    # print('its a path and xls')
                    wb = xlrdopen_workbook(self.ui.label_pathToOtherEQE.text())
                    xlsheet = wb.sheet_by_index(0)
                    i=3
                    while(1):
                        if xlsheet.cell(i,0).value !='':
                            alternativeEQEx.append(float(xlsheet.cell(i,0).value))
                            alternativeEQEy.append(float(xlsheet.cell(i,1).value)/100)
                            i+=1
                        else:
                            break
                if self.ui.radioButton_fullareaEQE.isChecked():
                    tempDATAlossA['EQExspotarea']=alternativeEQEx
                    tempDATAlossA['EQEyspotarea']=alternativeEQEy
                    tempDATAlossA['interpEQEspotarea']=interp1d(tempDATAlossA['EQExspotarea'], tempDATAlossA['EQEyspotarea'], bounds_error=False, fill_value=0.)
                else:
                    tempDATAlossA['EQExfullarea']=alternativeEQEx
                    tempDATAlossA['EQEyfullarea']=alternativeEQEy
                    tempDATAlossA['interpEQEfullarea']=interp1d(tempDATAlossA['EQExfullarea'], tempDATAlossA['EQEyfullarea'], bounds_error=False, fill_value=0.)
                
                
            #load R
            if self.ui.checkBox_R0.isChecked():
                if self.ui.radioButton_fullareaEQE.isChecked():
                    minx=min(tempDATAlossA['EQExfullarea'])
                    maxx=max(tempDATAlossA['EQExfullarea'])
                else:
                    minx=min(tempDATAlossA['EQExspotarea'])
                    maxx=max(tempDATAlossA['EQExspotarea'])
                    
                tempDATAlossA['Rx']=[minx,maxx]
                tempDATAlossA['Ry']=[0,0]
                tempDATAlossA['interpR']=interp1d(tempDATAlossA['Rx'], tempDATAlossA['Ry'], bounds_error=False, fill_value=0.)
            else:
                if ':/' in self.ui.label_pathToR.text():
                    if '.txt' in self.ui.label_pathToR.text():
                        # print('its a path and txt')
                        tempDATAlossA['Rx']=[]
                        tempDATAlossA['Ry']=[]
                        file = open(self.ui.label_pathToR.text(), encoding='ISO-8859-1')
                        Rdata = file.readlines()
                        file.close()
                        for i in range(len(Rdata)):
                            pos = Rdata[i].find('\t')
                            tempDATAlossA['Rx'].append(float(Rdata[i][:pos]))
                            tempDATAlossA['Ry'].append(float(Rdata[i][pos+1:-1])/100)
                        tempDATAlossA['interpR']=interp1d(tempDATAlossA['Rx'], tempDATAlossA['Ry'], bounds_error=False, fill_value=0.)
                        
                    elif '.xls' in self.ui.label_pathToR.text():
                        # print('its a path and xls')
                        wb = xlrdopen_workbook(self.ui.label_pathToR.text())
                        xlsheet = wb.sheet_by_index(0)
                        i=3
                        tempDATAlossA['Rx']=[]
                        tempDATAlossA['Ry']=[]
                        while(1):
                            if xlsheet.cell(i,0).value !='':
                                tempDATAlossA['Rx'].append(float(xlsheet.cell(i,0).value))
                                tempDATAlossA['Ry'].append(float(xlsheet.cell(i,1).value)/100)
                                i+=1
                            else:
                                break
                        tempDATAlossA['interpR']=interp1d(tempDATAlossA['Rx'], tempDATAlossA['Ry'], bounds_error=False, fill_value=0.)
                else:
                    # print('take from data')
                    tempDATAlossA['Rx'] = DATA[self.ui.label_pathToR.text()]['DATAmod'][0]
                    tempDATAlossA['Ry'] = DATA[self.ui.label_pathToR.text()]['DATAmod'][2]
                    tempDATAlossA['interpR']=interp1d(tempDATAlossA['Rx'], tempDATAlossA['Ry'], bounds_error=False, fill_value=0.)
            #load T
            if self.ui.checkBox_T0.isChecked():
                if self.ui.radioButton_fullareaEQE.isChecked():
                    minx=min(tempDATAlossA['EQExfullarea'])
                    maxx=max(tempDATAlossA['EQExfullarea'])
                else:
                    minx=min(tempDATAlossA['EQExspotarea'])
                    maxx=max(tempDATAlossA['EQExspotarea'])
                    
                tempDATAlossA['Tx']=[minx,maxx]
                tempDATAlossA['Ty']=[0,0]
                tempDATAlossA['interpT']=interp1d(tempDATAlossA['Tx'], tempDATAlossA['Ty'], bounds_error=False, fill_value=0.)
            else:
                if ':/' in self.ui.label_pathToT.text():
                    if '.txt' in self.ui.label_pathToT.text():
                        # print('its a path and txt')
                        tempDATAlossA['Tx']=[]
                        tempDATAlossA['Ty']=[]
                        file = open(self.ui.label_pathToT.text(), encoding='ISO-8859-1')
                        Tdata = file.readlines()
                        file.close()
                        for i in range(len(Tdata)):
                            pos = Tdata[i].find('\t')
                            tempDATAlossA['Tx'].append(float(Tdata[i][:pos]))
                            tempDATAlossA['Ty'].append(float(Tdata[i][pos+1:-1])/100)
                        tempDATAlossA['interpT']=interp1d(tempDATAlossA['Tx'], tempDATAlossA['Ty'], bounds_error=False, fill_value=0.)
                        
                    elif '.xls' in self.ui.label_pathToT.text():
                        # print('its a path and xls')
                        wb = xlrdopen_workbook(self.ui.label_pathToT.text())
                        xlsheet = wb.sheet_by_index(0)
                        i=3
                        tempDATAlossA['Tx']=[]
                        tempDATAlossA['Ty']=[]
                        while(1):
                            if xlsheet.cell(i,0).value !='':
                                tempDATAlossA['Tx'].append(float(xlsheet.cell(i,0).value))
                                tempDATAlossA['Ty'].append(float(xlsheet.cell(i,1).value)/100)
                                i+=1
                            else:
                                break
                        tempDATAlossA['interpT']=interp1d(tempDATAlossA['Tx'], tempDATAlossA['Ty'], bounds_error=False, fill_value=0.)
                else:
                    # print('take from data')
                    tempDATAlossA['Tx'] = DATA[self.ui.label_pathToT.text()]['DATAmod'][0]
                    tempDATAlossA['Ty'] = DATA[self.ui.label_pathToT.text()]['DATAmod'][2]
                    tempDATAlossA['interpT']=interp1d(tempDATAlossA['Tx'], tempDATAlossA['Ty'], bounds_error=False, fill_value=0.)
            
            if tempDATAlossA['EQExspotarea']!=[]:
                #calc IQE=EQE/(1-R-T)
                tempDATAlossA['IQEx']=tempDATAlossA['EQExspotarea']
                eqedat=[tempDATAlossA['interpEQEspotarea'](i) for i in tempDATAlossA['EQExspotarea']]
                Rdat=[tempDATAlossA['interpR'](i) for i in tempDATAlossA['EQExspotarea']]
                Tdat=[tempDATAlossA['interpT'](i) for i in tempDATAlossA['EQExspotarea']]
                tempDATAlossA['IQEy']=[eqedat[i]/(1-Rdat[i]-Tdat[i]) for i in range(len(eqedat))]
            else:
                tempDATAlossA['IQEx']=[]
                tempDATAlossA['IQEy']=[]
            tempDATAlossA['interpIQE']=interp1d(tempDATAlossA['IQEx'], tempDATAlossA['IQEy'], bounds_error=False, fill_value=0.)
            
            Rdat=[1-tempDATAlossA['interpR'](i) for i in tempDATAlossA['EQExspotarea']]
            tempDATAlossA['interp1R']=interp1d(tempDATAlossA['EQExspotarea'], Rdat, bounds_error=False, fill_value=0.)
            Rdat=[1-tempDATAlossA['interpR'](i)-tempDATAlossA['interpT'](i) for i in tempDATAlossA['EQExspotarea']]
            tempDATAlossA['interp1RT']=interp1d(tempDATAlossA['EQExspotarea'], Rdat, bounds_error=False, fill_value=0.)
            
            #calc losses and fill table
            
            interpfctlist=[tempDATAlossA['interpEQEfullarea'],tempDATAlossA['interpEQEspotarea'],tempDATAlossA['interpIQE'],tempDATAlossA['interp1RT'],tempDATAlossA['interp1R']]
            
            for i in range(len(interpfctlist)):
                if interpfctlist[i]!=[]:
                    f = interpfctlist[i]
                    x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                    for j in range(4):
                        integral = echarge/10*integrate.quad(x2,float(self.ui.tableWidget_zonesdefinition.item(j,0).text()), float(self.ui.tableWidget_zonesdefinition.item(j,1).text()))[0]
                        tableitem=QtWidgets.QTableWidgetItem()
                        tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(integral))
                        self.ui.tableWidget_detailsoflosses.setItem(j,i,tableitem)
                        
            if tempDATAlossA['EQExfullarea']!=[]:
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(0,0).text())))
                self.ui.tableWidget_Losses.setItem(0,0,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(1,0).text())))
                self.ui.tableWidget_Losses.setItem(0,1,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(2,0).text())))
                self.ui.tableWidget_Losses.setItem(0,2,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(3,0).text())))
                self.ui.tableWidget_Losses.setItem(0,3,tableitem)

            if tempDATAlossA['EQExspotarea']!=[]:
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(0,1).text())))
                self.ui.tableWidget_Losses.setItem(1,0,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(1,1).text())))
                self.ui.tableWidget_Losses.setItem(1,1,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(2,1).text())))
                self.ui.tableWidget_Losses.setItem(1,2,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(3,1).text())))
                self.ui.tableWidget_Losses.setItem(1,3,tableitem)
                
            if tempDATAlossA['EQExspotarea']!=[] and tempDATAlossA['EQExfullarea']!=[]:
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(1,0).text())-float(self.ui.tableWidget_Losses.item(0,0).text())))
                self.ui.tableWidget_Losses.setItem(2,0,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(1,1).text())-float(self.ui.tableWidget_Losses.item(0,1).text())))
                self.ui.tableWidget_Losses.setItem(2,1,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(1,2).text())-float(self.ui.tableWidget_Losses.item(0,2).text())))
                self.ui.tableWidget_Losses.setItem(2,2,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(1,3).text())-float(self.ui.tableWidget_Losses.item(0,3).text())))
                self.ui.tableWidget_Losses.setItem(2,3,tableitem)
                
            interpfctlist=[tempDATAlossA['interpT'],tempDATAlossA['interpR']]
            
            if tempDATAlossA['interpT']!=[]:
                f = tempDATAlossA['interpT']
                x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                for j in range(4):
                    integral = echarge/10*integrate.quad(x2,float(self.ui.tableWidget_zonesdefinition.item(j,0).text()), float(self.ui.tableWidget_zonesdefinition.item(j,1).text()))[0]
                    tableitem=QtWidgets.QTableWidgetItem()
                    tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(integral))
                    self.ui.tableWidget_Losses.setItem(3,j,tableitem)
            if tempDATAlossA['interpR']!=[]:
                f = tempDATAlossA['interpR']
                x2 = lambda x0: self.AM15GParticlesinnm(x0)*f(x0)
                for j in range(4):
                    integral = echarge/10*integrate.quad(x2,float(self.ui.tableWidget_zonesdefinition.item(j,0).text()), float(self.ui.tableWidget_zonesdefinition.item(j,1).text()))[0]
                    tableitem=QtWidgets.QTableWidgetItem()
                    tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(integral))
                    self.ui.tableWidget_Losses.setItem(4,j,tableitem)
                    
            if tempDATAlossA['IQEx']!=[]:
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(0,2).text())-float(self.ui.tableWidget_detailsoflosses.item(0,1).text())))
                self.ui.tableWidget_Losses.setItem(5,0,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(1,2).text())-float(self.ui.tableWidget_detailsoflosses.item(1,1).text())))
                self.ui.tableWidget_Losses.setItem(5,1,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(2,2).text())-float(self.ui.tableWidget_detailsoflosses.item(2,1).text())))
                self.ui.tableWidget_Losses.setItem(5,2,tableitem)
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_detailsoflosses.item(3,2).text())-float(self.ui.tableWidget_detailsoflosses.item(3,1).text())))
                self.ui.tableWidget_Losses.setItem(5,3,tableitem)
            
            tableitem=QtWidgets.QTableWidgetItem()
            tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(3,0).text())+float(self.ui.tableWidget_Losses.item(4,0).text())+float(self.ui.tableWidget_Losses.item(5,0).text())))
            self.ui.tableWidget_Losses.setItem(6,0,tableitem)
            tableitem=QtWidgets.QTableWidgetItem()
            tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(3,1).text())+float(self.ui.tableWidget_Losses.item(4,1).text())+float(self.ui.tableWidget_Losses.item(5,1).text())))
            self.ui.tableWidget_Losses.setItem(6,1,tableitem)
            tableitem=QtWidgets.QTableWidgetItem()
            tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(3,2).text())+float(self.ui.tableWidget_Losses.item(4,2).text())+float(self.ui.tableWidget_Losses.item(5,2).text())))
            self.ui.tableWidget_Losses.setItem(6,2,tableitem)
            tableitem=QtWidgets.QTableWidgetItem()
            tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(float(self.ui.tableWidget_Losses.item(3,3).text())+float(self.ui.tableWidget_Losses.item(4,3).text())+float(self.ui.tableWidget_Losses.item(5,3).text())))
            self.ui.tableWidget_Losses.setItem(6,3,tableitem)
            
            x2 = lambda x0: self.AM15GParticlesinnm(x0)*1
            for j in range(4):
                integral = echarge/10*integrate.quad(x2,float(self.ui.tableWidget_zonesdefinition.item(j,0).text()), float(self.ui.tableWidget_zonesdefinition.item(j,1).text()))[0]
                tableitem=QtWidgets.QTableWidgetItem()
                tableitem.setData(QtCore.Qt.EditRole, QtCore.QVariant(integral))
                self.ui.tableWidget_detailsoflosses.setItem(j,5,tableitem)
                
            if self.ui.checkBox_addtoMain.isChecked():
                #for IQE
                datadict = {'ID':DATA[sampleselected]['ID']+'IQE','samplename':DATA[sampleselected]['samplename']+'IQE', 'batchnumb':DATA[sampleselected]['batchnumb'], 'samplenumb':DATA[sampleselected]['samplenumb'], 'file_path':'', 'meastype':'IQE', 'JscfromFile':-1, 'JscCalc':-1, 'datetime':'',
                                'NameMod':DATA[sampleselected]['NameMod']+'_IQE', 'Name_Jsc':DATA[sampleselected]['NameMod']+'_IQE','Name_Eg':DATA[sampleselected]['NameMod']+'_IQE', 'Name_Jsc_Eg':DATA[sampleselected]['NameMod']+'_IQE','Name_Egln':DATA[sampleselected]['NameMod']+'_IQE', 'Name_Jsc_Egln':DATA[sampleselected]['NameMod']+'_IQE', 'Name_Egtauc':DATA[sampleselected]['NameMod']+'_IQE', 'Name_Jsc_Egtauc':DATA[sampleselected]['NameMod']+'_IQE', 
                                'Eg0':-1, 'EgIP':-1, 'tangent':[-1, -1],'slope':-1, 'h':-1,
                                'EgTauc':-1, 'xtauc':[], 'ytauc':[], 'mtauc':-1, 'htauc':-1,
                                'lnDat':[[-1],[-1]], 'EgLn':-1,  'stderrEgLn':[-1,-1], 'tangentLn':[-1, -1,[-1],[-1]], 'slopeLn':-1, 'hln':-1,'dataEnergyLn':[-1], 'dataIntLn':[-1], 'ptstgtLnX':[-1], 'ptstgtLnY':[-1],
                                'EuLn':-1,
                                'DATAorig':[tempDATAlossA['IQEx'],[1239.8/i for i in tempDATAlossA['IQEx']],tempDATAlossA['IQEy']],'DATAmod':[tempDATAlossA['IQEx'],[1239.8/i for i in tempDATAlossA['IQEx']],tempDATAlossA['IQEy']], 
                                'comment':'', 'Vbias':'', 'filterbias':'', 'ledbias':'','integJsclist':[],
                                 'linestyle':'-', 'linecolor':'black', 'linewidth':int(2)
                                }
                DATA[datadict['ID']]=datadict
                
                #for R, if not already
                if not self.ui.checkBox_R0.isChecked() and ':/' in self.ui.label_pathToR.text():
                    datadict = {'ID':DATA[sampleselected]['ID']+'R','samplename':DATA[sampleselected]['samplename']+'R', 'batchnumb':DATA[sampleselected]['batchnumb'], 'samplenumb':DATA[sampleselected]['samplenumb'], 'file_path':'', 'meastype':'R', 'JscfromFile':-1, 'JscCalc':-1, 'datetime':'',
                                    'NameMod':DATA[sampleselected]['NameMod']+'_R', 'Name_Jsc':DATA[sampleselected]['NameMod']+'_R','Name_Eg':DATA[sampleselected]['NameMod']+'_R', 'Name_Jsc_Eg':DATA[sampleselected]['NameMod']+'_R','Name_Egln':DATA[sampleselected]['NameMod']+'_R', 'Name_Jsc_Egln':DATA[sampleselected]['NameMod']+'_R', 'Name_Egtauc':DATA[sampleselected]['NameMod']+'_R', 'Name_Jsc_Egtauc':DATA[sampleselected]['NameMod']+'_R', 
                                    'Eg0':-1, 'EgIP':-1, 'tangent':[-1, -1],'slope':-1, 'h':-1,
                                    'EgTauc':-1, 'xtauc':[], 'ytauc':[], 'mtauc':-1, 'htauc':-1,
                                    'lnDat':[[-1],[-1]], 'EgLn':-1,  'stderrEgLn':[-1,-1], 'tangentLn':[-1, -1,[-1],[-1]], 'slopeLn':-1, 'hln':-1,'dataEnergyLn':[-1], 'dataIntLn':[-1], 'ptstgtLnX':[-1], 'ptstgtLnY':[-1],
                                    'EuLn':-1,
                                    'DATAorig':[tempDATAlossA['Rx'],[1239.8/i for i in tempDATAlossA['Rx']],tempDATAlossA['Ry']],'DATAmod':[tempDATAlossA['Rx'],[1239.8/i for i in tempDATAlossA['Rx']],tempDATAlossA['Ry']], 
                                    'comment':'', 'Vbias':'', 'filterbias':'', 'ledbias':'','integJsclist':[],
                                     'linestyle':'-', 'linecolor':'black', 'linewidth':int(2)
                                    }
                    DATA[datadict['ID']]=datadict
                
                #for T, if not already
                if not self.ui.checkBox_T0.isChecked() and ':/' in self.ui.label_pathToT.text():
                    datadict = {'ID':DATA[sampleselected]['ID']+'T','samplename':DATA[sampleselected]['samplename']+'T', 'batchnumb':DATA[sampleselected]['batchnumb'], 'samplenumb':DATA[sampleselected]['samplenumb'], 'file_path':'', 'meastype':'T', 'JscfromFile':-1, 'JscCalc':-1, 'datetime':'',
                                    'NameMod':DATA[sampleselected]['NameMod']+'_T', 'Name_Jsc':DATA[sampleselected]['NameMod']+'_T','Name_Eg':DATA[sampleselected]['NameMod']+'_T', 'Name_Jsc_Eg':DATA[sampleselected]['NameMod']+'_T','Name_Egln':DATA[sampleselected]['NameMod']+'_T', 'Name_Jsc_Egln':DATA[sampleselected]['NameMod']+'_T', 'Name_Egtauc':DATA[sampleselected]['NameMod']+'_T', 'Name_Jsc_Egtauc':DATA[sampleselected]['NameMod']+'_T', 
                                    'Eg0':-1, 'EgIP':-1, 'tangent':[-1, -1],'slope':-1, 'h':-1,
                                    'EgTauc':-1, 'xtauc':[], 'ytauc':[], 'mtauc':-1, 'htauc':-1,
                                    'lnDat':[[-1],[-1]], 'EgLn':-1,  'stderrEgLn':[-1,-1], 'tangentLn':[-1, -1,[-1],[-1]], 'slopeLn':-1, 'hln':-1,'dataEnergyLn':[-1], 'dataIntLn':[-1], 'ptstgtLnX':[-1], 'ptstgtLnY':[-1],
                                    'EuLn':-1,
                                    'DATAorig':[tempDATAlossA['Tx'],[1239.8/i for i in tempDATAlossA['Tx']],tempDATAlossA['Ty']],'DATAmod':[tempDATAlossA['Tx'],[1239.8/i for i in tempDATAlossA['Tx']],tempDATAlossA['Ty']], 
                                    'comment':'', 'Vbias':'', 'filterbias':'', 'ledbias':'','integJsclist':[],
                                     'linestyle':'-', 'linecolor':'black', 'linewidth':int(2)
                                    }
                    DATA[datadict['ID']]=datadict
                
                #for other EQE, if any imported
                if ':/' in self.ui.label_pathToOtherEQE.text():
                    if self.ui.radioButton_fullareaEQE.isChecked():
                        exten='_spotEQE'
                    else:
                        exten='_fullareaEQE'
                    datadict = {'ID':DATA[sampleselected]['ID']+exten,'samplename':DATA[sampleselected]['samplename']+exten, 'batchnumb':DATA[sampleselected]['batchnumb'], 'samplenumb':DATA[sampleselected]['samplenumb'], 'file_path':'', 'meastype':exten[1:], 'JscfromFile':-1, 'JscCalc':-1, 'datetime':'',
                                'NameMod':DATA[sampleselected]['NameMod']+exten, 'Name_Jsc':DATA[sampleselected]['NameMod']+exten,'Name_Eg':DATA[sampleselected]['NameMod']+exten, 'Name_Jsc_Eg':DATA[sampleselected]['NameMod']+exten,'Name_Egln':DATA[sampleselected]['NameMod']+exten, 'Name_Jsc_Egln':DATA[sampleselected]['NameMod']+exten, 'Name_Egtauc':DATA[sampleselected]['NameMod']+exten, 'Name_Jsc_Egtauc':DATA[sampleselected]['NameMod']+exten, 
                                'Eg0':-1, 'EgIP':-1, 'tangent':[-1, -1],'slope':-1, 'h':-1,
                                'EgTauc':-1, 'xtauc':[], 'ytauc':[], 'mtauc':-1, 'htauc':-1,
                                'lnDat':[[-1],[-1]], 'EgLn':-1,  'stderrEgLn':[-1,-1], 'tangentLn':[-1, -1,[-1],[-1]], 'slopeLn':-1, 'hln':-1,'dataEnergyLn':[-1], 'dataIntLn':[-1], 'ptstgtLnX':[-1], 'ptstgtLnY':[-1],
                                'EuLn':-1,
                                'DATAorig':[alternativeEQEx,[1239.8/i for i in alternativeEQEx],alternativeEQEy],'DATAmod':[alternativeEQEx,[1239.8/i for i in alternativeEQEx],alternativeEQEy], 
                                'comment':'', 'Vbias':'', 'filterbias':'', 'ledbias':'','integJsclist':[],
                                 'linestyle':'-', 'linecolor':'black', 'linewidth':int(2)
                                 }
                    DATA[datadict['ID']]=datadict
                #update the listwidget on plotting page
                self.initlistbox()
            
            #plotting
            self.EQElossgraph.clear()
            if tempDATAlossA['EQExfullarea']!=[]:
                self.EQElossgraph.plot(tempDATAlossA['EQExfullarea'],tempDATAlossA['EQEyfullarea'], label='EQEfullarea')
            if tempDATAlossA['EQExspotarea']!=[]:
                self.EQElossgraph.plot(tempDATAlossA['EQExspotarea'],tempDATAlossA['EQEyspotarea'], label='EQEspotarea')
            if tempDATAlossA['IQEx']!=[]:
                self.EQElossgraph.plot(tempDATAlossA['IQEx'],tempDATAlossA['IQEy'],label='IQEspot')
            Ry = [1-i for i in tempDATAlossA['Ry']]
            self.EQElossgraph.plot(tempDATAlossA['Rx'],Ry,label='1-R')
            self.EQElossgraph.plot(tempDATAlossA['Tx'],tempDATAlossA['Ty'],label='T')
            self.EQElossgraph.set_ylabel('EQE; 1-R; T; IQE (-)')
            self.EQElossgraph.set_xlabel('Wavelength (nm)')
            self.EQElossgraph.legend()
            # #plotting
            # self.EQElossgraph.clear()
            # if tempDATAlossA['EQExfullarea']!=[]:
            #     self.EQElossgraph.plot(tempDATAlossA['EQExfullarea'],tempDATAlossA['interpEQEfullarea'], label='EQEfullarea')
            # if tempDATAlossA['EQExspotarea']!=[]:
            #     ydat=[tempDATAlossA['interpEQEspotarea'](i) for i in tempDATAlossA['EQExspotarea']]
            #     self.EQElossgraph.plot(tempDATAlossA['EQExspotarea'],ydat, label='EQEspotarea')
            # ydat=[tempDATAlossA['interpR'](i) for i in tempDATAlossA['EQExspotarea']]
            # Ry = [1-i for i in ydat]
            # self.EQElossgraph.plot(tempDATAlossA['Rx'],Ry,label='1-R')
            # # self.EQElossgraph.plot(tempDATAlossA['Tx'],tempDATAlossA['interpT'],label='T')
            # self.EQElossgraph.set_ylabel('EQE; 1-R; T; IQE (-)')
            # self.EQElossgraph.set_xlabel('Wavelength (nm)')
            # self.EQElossgraph.legend()            
            if self.ui.zonesoverlays.isChecked():
                self.EQElossgraph.axvspan(float(self.ui.tableWidget_zonesdefinition.item(0,0).text()), float(self.ui.tableWidget_zonesdefinition.item(0,1).text()), facecolor='b', alpha=0.1)
                self.EQElossgraph.axvspan(float(self.ui.tableWidget_zonesdefinition.item(1,0).text()), float(self.ui.tableWidget_zonesdefinition.item(1,1).text()), facecolor='g', alpha=0.1)
                self.EQElossgraph.axvspan(float(self.ui.tableWidget_zonesdefinition.item(2,0).text()), float(self.ui.tableWidget_zonesdefinition.item(2,1).text()), facecolor='r', alpha=0.1)
            
            self.figloss.canvas.draw_idle()
        
        
    # def CalculateIQE(self):
    #     global takenforplot, stitching
    #     global DATAFORGRAPH
        
        
    #     if len(takenforplot) != 2: 
    #         QMessageBox.information(self, 'Information', "Please select only 2 files at a time: 1 EQE + 1 Refl")
    #     else:
    #         goodtogo=1
    #         if 'EQE' in DATAFORGRAPH[takenforplot[0]][0] and 'Reflectance' in DATAFORGRAPH[takenforplot[1]][0]:
    #             eqex=DATAFORGRAPH[takenforplot[0]][2][1]
    #             eqey=DATAFORGRAPH[takenforplot[0]][3][1]
    #             Rx=DATAFORGRAPH[takenforplot[1]][2][1]
    #             Ry=DATAFORGRAPH[takenforplot[1]][3][1]
    #         elif 'EQE' in DATAFORGRAPH[takenforplot[1]][0] and 'Reflectance' in DATAFORGRAPH[takenforplot[0]][0]:
    #             eqex=DATAFORGRAPH[takenforplot[1]][2][1]
    #             eqey=DATAFORGRAPH[takenforplot[1]][3][1]
    #             Rx=DATAFORGRAPH[takenforplot[0]][2][1]
    #             Ry=DATAFORGRAPH[takenforplot[0]][3][1]
    #         else:
    #             QMessageBox.information(self, 'Information', "Please select only 2 files at a time: 1 EQE + 1 Refl")
    #             goodtogo=0
            
    #         if goodtogo:
    #             EQEfct = interp1d(eqex, eqey, kind='cubic')
    #             Reflfct = interp1d(Rx, Ry, kind='cubic')
                
    #             xnew = arange(max([min(eqex),min(Rx)]), min([max(eqex),max(Rx)]), eqex[1]-eqex[0])
    #             ynew = EQEfct(xnew)/(1-Reflfct(xnew))
                
    #             #export txt file with new data where the original file was imported from
    #             datexport=[]
    #             for i in range(len(xnew)):
    #                 datexport.append(str(xnew[i])+'\t'+str(ynew[i])+'\n')
    #             # stitching = filedialog.asksaveasfilename(defaultextension=".txt")
    #             # stitching = QFileDialog.getSaveFileName()[0]
    #             stitching=QFileDialog.getSaveFileName(self,"Save file","","All Files (*);;Text Files (*.txt)")[0]
    #             # print(stitching)
    #             stitching=stitching[:-4]+'_IQE_data.txt'
    #             # print(stitching)
    #             file = open(stitching,'w', encoding='ISO-8859-1')
    #             file.writelines("%s" % item for item in datexport)
    #             file.close()
    #             #reimport that file with import function
    #             self.onOpenEQE()
                
                
    def StitchEQE(self):
        print("stitching not implemented")
        global takenforplot
        global DATA, stitching
        
#         DATAx=DATAFORGRAPH
#         sampletotake=takenforplot
        
#         newDatlistx=[]
#         newDatlisty=[]
        
#         for i in range(len(sampletotake)):
# #            newDatlistx.append(DATAx[sampletotake[i]][2])
# #            newDatlisty.append(DATAx[sampletotake[i]][3])
#             newDatlistx+=DATAx[sampletotake[i]][2][1]
#             newDatlisty+=DATAx[sampletotake[i]][3][1]
#         overlapx=Repeat(newDatlistx)
# #        print(overlapx)
#         newx=[]
#         newy=[]
#         seen=[]
#         for item in range(len(newDatlistx)):
#             if newDatlistx[item] not in overlapx[0]:
#                 newx.append(newDatlistx[item])
#                 newy.append(newDatlisty[item])
#             elif newDatlistx[item] in overlapx[0] and newDatlistx[item] not in seen:
#                 newx.append(newDatlistx[item])
#                 seen.append(newDatlistx[item])
#                 indexlist=overlapx[0].index(newDatlistx[item])
#                 newy.append((newDatlisty[overlapx[1][indexlist][0]]+newDatlisty[overlapx[1][indexlist][1]])/2)

#         #export txt file with new data where the original file was imported from
#         datexport=[]
#         for i in range(len(newx)):
#             datexport.append(str(newx[i])+'\t'+str(newy[i])+'\n')
#         # stitching = filedialog.asksaveasfilename(defaultextension=".txt")
#         stitching = QFileDialog.getOpenFileNames()[0]


#         file = open(stitching,'w', encoding='ISO-8859-1')
#         file.writelines("%s" % item for item in datexport)
#         file.close()
#         #reimport that file with import function
#         self.onOpenEQE()
        
        
        
def Repeat(x): 
    _size = len(x) 
    repeated = [] 
    repindices=[]
    for i in range(_size): 
        k = i + 1
        for j in range(k, _size): 
            if x[i] == x[j] and x[i] not in repeated: 
                repeated.append(x[i]) 
                repindices.append([i,j])
                
    return [repeated, repindices]
        

reordered=0
class reorderwin(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        global DATA
        global takenforplot
        # DATAx=DATA
        sampletotake=takenforplot
                
        self.resize(300, 300)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.sizePolicy().hasHeightForWidth())
        self.setSizePolicy(sizePolicy)
        self.setMinimumSize(QtCore.QSize(200, 200))
        self.setWindowTitle("Reorder stack by drag and drop")
        
        self.widget_layout = QtWidgets.QVBoxLayout()

        # Create ListWidget and add 10 items to move around.
        self.list_widget = QtWidgets.QListWidget()
        self.list_widget.addItems(sampletotake)

        # Enable drag & drop ordering of items.
        self.list_widget.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)

        self.pushButton_reorder = QtWidgets.QPushButton("Validate", self)
        self.widget_layout.addWidget(self.pushButton_reorder)
        self.pushButton_reorder.clicked.connect(self.validate)

        self.widget_layout.addWidget(self.list_widget)
        self.setLayout(self.widget_layout)
        
    def validate(self):
        global DATA
        global takenforplot
        global reordered, window
        takenforplot=[]
        sampletotake=[self.list_widget.item(x).text() for x in range(self.list_widget.count())]
        for item in sampletotake:
            for i in DATA.keys():
                if DATA[i]['ID']==item:
                    takenforplot.append(i)
        
        # self.UpdateEQELegMod()
        
        window.UpdateEQEGraph()
        self.hide()
        
#%%#############         
#%%#############
if __name__ == "__main__":
    app = QtWidgets.QApplication(argv)
    window = EQEapp()
    window.show()
    Exit(app.exec())

